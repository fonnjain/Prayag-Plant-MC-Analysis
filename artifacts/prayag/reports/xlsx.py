"""Render a ``ReportModel`` into a styled openpyxl ``Workbook``.

Styling (per the management-report spec):
- Arial throughout.
- Navy ``#1F3864`` title + table headers (white bold text).
- Terracotta ``#C55A11`` totals/subtotal rows.
- Dates dd-mm-yyyy; zeros shown as "-".
- A provenance/validation footer (source workbooks, tie-out result, flags).

Every workbook automatically receives:
  • A **Cover** tab as the first sheet (report name, period, timestamp, source,
    recomputed-not-copied statement).
  • A **Notes** tab as the last sheet — one row per Flag (rule, section, month,
    our figure, source figure, difference, explanation).

Values approach: every figure is the authoritative Python-recomputed VALUE.
AWAITING SOURCE DATA / n/a / IDLE / ⚠ unavailable are written as text strings
and never converted to 0 or blank.
"""
from __future__ import annotations

import datetime
import io

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .model import Flag, ReportModel, ReportSheet, Section

NAVY  = "1F3864"
TERRA = "C55A11"
WHITE = "FFFFFF"
GREY  = "595959"
LIGHT = "F2F2F2"
FONT  = "Arial"
GOLD  = "C9A227"

_NAVY_FILL  = PatternFill("solid", fgColor=NAVY)
_TERRA_FILL = PatternFill("solid", fgColor=TERRA)
_LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
_GOLD_FILL  = PatternFill("solid", fgColor=GOLD)
_THIN   = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _num_fmt(kind: str) -> str:
    """Number format string; the third clause renders a real 0 as "-".

    Kinds:
      "int"  — whole counts, run hours, headcount         → no decimal
      "kg"   — output / reject KG totals                  → no decimal
      "cur"  — rupee totals (wages, power cost, …)        → no decimal
      "rate" — cost/kg, cost/hr, devot/person, avg/hr     → 2 dp
      "pct"  — percentage                                  → 1 dp + %
      "num"  — generic numeric (paid hours, kWh, …)       → 1 dp
    """
    if kind in ("int", "kg", "cur"):
        return '#,##0;-#,##0;"-"'
    if kind == "rate":
        return '#,##0.00;-#,##0.00;"-"'
    if kind == "pct":
        return '0.0"%";-0.0"%";"-"'
    if kind == "num":
        return '#,##0.0;-#,##0.0;"-"'
    return "General"


def _safe_tab(name: str) -> str:
    bad = set(r'[]:*?/\\')
    cleaned = "".join(c for c in name if c not in bad).strip()
    return (cleaned or "Report")[:31]


def _write_cell(ws, row, col, value, kind, *, bold=False, color=None,
                fill=None, align="right", border=True, comment_text=None):
    c = ws.cell(row=row, column=col)
    if isinstance(value, str):
        c.value = value
        c.number_format = "General"
    elif value is None:
        c.value = None              # genuinely-missing → blank cell
    else:
        c.value = value
        c.number_format = _num_fmt(kind)
    c.font = Font(name=FONT, bold=bold, color=color or "000000", size=10)
    if fill is not None:
        c.fill = fill
    if kind == "text" or isinstance(value, str):
        align = "left" if align == "right" else align
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if border:
        c.border = _BORDER
    if comment_text:
        c.comment = Comment(str(comment_text)[:300], "Prayag Analytics")
    return c


# ---------------------------------------------------------------------------
# Main sheet renderer
# ---------------------------------------------------------------------------
def _render_sheet(ws, sheet: ReportSheet) -> dict:
    """Render one ReportSheet into *ws*.

    Returns a dict ``{(first_col_value, col_key): cell}`` for comment
    placement — the caller matches Flag.cell_row_label + Flag.cell_col_key.
    """
    cell_registry: dict = {}

    ncols = 1
    for sec in sheet.sections:
        ncols = max(ncols, len(sec.columns))
    ncols = max(ncols, 4)
    last_col = get_column_letter(ncols)

    # --- Title ---
    r = 1
    ws.merge_cells(f"A{r}:{last_col}{r}")
    t = ws.cell(row=r, column=1, value=sheet.title)
    t.font = Font(name=FONT, bold=True, size=14, color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    # --- Subtitle ---
    if sheet.subtitle:
        ws.merge_cells(f"A{r}:{last_col}{r}")
        s = ws.cell(row=r, column=1, value=sheet.subtitle)
        s.font = Font(name=FONT, italic=True, size=9, color=GREY)
        s.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1  # spacer

    if not sheet.sections:
        ws.merge_cells(f"A{r}:{last_col}{r}")
        n = ws.cell(row=r, column=1, value=sheet.note or "Awaiting source data.")
        n.font = Font(name=FONT, bold=True, size=11, color=TERRA)
        n.alignment = Alignment(horizontal="left", vertical="center")
        r += 2

    col_widths: dict = {}
    for sec in sheet.sections:
        if sec.heading:
            ws.merge_cells(f"A{r}:{last_col}{r}")
            h = ws.cell(row=r, column=1, value=sec.heading)
            h.font = Font(name=FONT, bold=True, size=10, color=NAVY)
            r += 1
        # header row
        for j, col in enumerate(sec.columns, start=1):
            c = _write_cell(ws, r, j, col.label, "text",
                            bold=True, color=WHITE, fill=_NAVY_FILL,
                            align="center")
            col_widths[j] = max(col_widths.get(j, 0),
                                len(str(col.label)) + 2, col.width or 0)
        ws.row_dimensions[r].height = 28
        for cc in ws[r]:
            cc.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
        r += 1
        # data rows
        for ri, row in enumerate(sec.rows):
            fill = _LIGHT_FILL if ri % 2 else None
            first_col_val = None
            for j, col in enumerate(sec.columns, start=1):
                val = row.get(col.key)
                cell = _write_cell(ws, r, j, val, col.kind, fill=fill,
                                   align="left" if col.kind == "text" else "right")
                if j == 1:
                    first_col_val = str(val) if val is not None else ""
                if first_col_val is not None:
                    cell_registry[(first_col_val, col.key)] = cell
                if isinstance(val, str):
                    col_widths[j] = max(col_widths.get(j, 0), len(val) + 2)
            r += 1
        # total row
        if sec.total_row is not None:
            first_col_val = None
            for j, col in enumerate(sec.columns, start=1):
                val = sec.total_row.get(col.key)
                cell = _write_cell(ws, r, j, val, col.kind,
                                   bold=True, color=WHITE, fill=_TERRA_FILL,
                                   align="left" if col.kind == "text" else "right")
                if j == 1:
                    first_col_val = str(val) if val is not None else ""
                if first_col_val is not None:
                    cell_registry[(first_col_val, col.key)] = cell
            r += 1
        r += 1  # spacer between sections

    # --- Provenance / validation footer ---
    if sheet.provenance:
        r += 1
        for line in sheet.provenance:
            ws.merge_cells(f"A{r}:{last_col}{r}")
            f = ws.cell(row=r, column=1, value=line)
            f.font = Font(name=FONT, italic=True, size=8, color=GREY)
            f.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True)
            r += 1

    for j, w in col_widths.items():
        ws.column_dimensions[get_column_letter(j)].width = min(max(w, 10), 40)
    ws.sheet_view.showGridLines = False
    return cell_registry


def _apply_comments(cell_registry: dict, sheet: ReportSheet):
    """Place Excel cell comments described in sheet.cell_comments.

    sheet.cell_comments: {(row_label_value, col_key): comment_text}
    """
    for (row_label, col_key), comment_text in sheet.cell_comments.items():
        c = cell_registry.get((row_label, col_key))
        if c is not None and comment_text:
            try:
                c.comment = Comment(str(comment_text)[:300], "Prayag Analytics")
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Cover tab
# ---------------------------------------------------------------------------
def _render_cover(ws, model: ReportModel, fy_label: str = ""):
    """Write a styled cover sheet — always the first tab of every workbook."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50

    def _hdr(r, text, size=12, bold=False, color=NAVY, merge_to=4):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=FONT, bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        ws.row_dimensions[r].height = size + 8
        return r + 1

    def _kv(r, key, value, vcolor="000000"):
        k_cell = ws.cell(row=r, column=1, value=key)
        k_cell.font = Font(name=FONT, bold=True, size=10, color=GREY)
        k_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(f"B{r}:D{r}")
        v_cell = ws.cell(row=r, column=2, value=value)
        v_cell.font = Font(name=FONT, size=10, color=vcolor)
        v_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
        ws.row_dimensions[r].height = 16
        return r + 1

    r = 1
    ws.row_dimensions[r].height = 10
    r += 1

    # Report name — large navy
    ws.merge_cells(f"A{r}:D{r}")
    title_cell = ws.cell(row=r, column=1,
                         value=model.label or "Management Report")
    title_cell.font = Font(name=FONT, bold=True, size=18, color=NAVY)
    title_cell.fill = PatternFill("solid", fgColor="EBF0F8")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 32
    r += 1

    ws.row_dimensions[r].height = 6
    r += 1

    # Metadata
    r = _kv(r, "FY Period", fy_label or model.ym or "")
    r = _kv(r, "Plant / Location", model.plant or "All plants")
    r = _kv(r, "Report ID", model.rid or "—")

    ts = model.cover_source  # reuse field for timestamp if set by generator
    generated_ts = datetime.datetime.utcnow().strftime("%d-%b-%Y %H:%M UTC")
    r = _kv(r, "Generated at", generated_ts, vcolor=GREY)

    ws.row_dimensions[r].height = 10
    r += 1

    # Source basis
    r = _hdr(r, "Source & Basis", size=11, bold=True)
    source_text = (
        model.cover_source
        if model.cover_source and not model.cover_source.startswith("20")
        else "Live Google Sheets — same source workbooks the management-report pages use."
    )
    ws.merge_cells(f"A{r}:D{r}")
    sc = ws.cell(row=r, column=1, value=source_text)
    sc.font = Font(name=FONT, size=10, color="000000")
    sc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 36
    r += 1

    ws.row_dimensions[r].height = 10
    r += 1

    # Compliance statement
    stmt = (
        "RECOMPUTED — NOT COPIED. Every figure in this workbook is recomputed "
        "from the live source data at generation time using the same calculation "
        "engine as the web-based management-report pages. No cell contains a "
        "formula linked to an external workbook. Figures may differ from a "
        "previously downloaded copy if the underlying source workbooks have been "
        "updated since that copy was generated. See the Notes tab for annotated "
        "discrepancies and data caveats."
    )
    ws.merge_cells(f"A{r}:D{r}")
    sc = ws.cell(row=r, column=1, value=stmt)
    sc.font = Font(name=FONT, italic=True, size=9, color=GREY)
    sc.fill = PatternFill("solid", fgColor="FFFBF0")
    sc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 60
    r += 1

    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Notes / Flags tab
# ---------------------------------------------------------------------------
_NOTES_COLS = [
    ("Rule",          20),
    ("Section",       22),
    ("Month",         10),
    ("Our Figure",    16),
    ("Source Figure", 16),
    ("Difference",    14),
    ("Note",          60),
]


def _render_notes(ws, flags: list):
    """Write the Notes tab — one row per Flag, always the last tab."""
    ws.sheet_view.showGridLines = False

    # Title
    r = 1
    ws.merge_cells(f"A{r}:G{r}")
    t = ws.cell(row=r, column=1, value="Flags & Disclosures")
    t.font = Font(name=FONT, bold=True, size=14, color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 24
    r += 1

    ws.merge_cells(f"A{r}:G{r}")
    st = ws.cell(row=r, column=1,
                 value=("Annotated discrepancies, basis differences, and "
                        "data-quality caveats. One row per flagged figure."))
    st.font = Font(name=FONT, italic=True, size=9, color=GREY)
    st.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 2

    # Header row
    for j, (lbl, w) in enumerate(_NOTES_COLS, start=1):
        c = ws.cell(row=r, column=j, value=lbl)
        c.font = Font(name=FONT, bold=True, color=WHITE, size=10)
        c.fill = _NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[r].height = 24
    r += 1

    if not flags:
        ws.merge_cells(f"A{r}:G{r}")
        nc = ws.cell(row=r, column=1, value="No flags for this period.")
        nc.font = Font(name=FONT, italic=True, size=10, color=GREY)
        nc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 18
        return

    for i, flag in enumerate(flags):
        if not isinstance(flag, Flag):
            continue
        fill = _LIGHT_FILL if i % 2 else None
        vals = [
            flag.rule,
            flag.section,
            flag.month,
            flag.our_figure,
            flag.source_figure,
            flag.difference,
            flag.note,
        ]
        for j, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=val or "")
            c.font = Font(name=FONT, size=10)
            if fill:
                c.fill = fill
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=(j == 7),  # wrap Note column
            )
        ws.row_dimensions[r].height = 18
        r += 1


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def _fy_label_from_ym(ym: str) -> str:
    """'2026-04' → 'FY 2026–27';  '2025-12' → 'FY 2025–26'."""
    import re as _re
    try:
        yr, mo = int(ym[:4]), int(ym[5:7])
        fs = yr if mo >= 4 else yr - 1
        return f"FY {fs}–{str(fs + 1)[-2:]}"
    except Exception:
        return ym


def render_workbook(model: ReportModel) -> Workbook:
    """Render model → Workbook.

    Tab order: Cover | <content sheets> | Notes
    The Cover and Notes tabs are always present (even on failed builds — the
    Notes tab then records the build failure explicitly).
    """
    wb = Workbook()
    used: set = set()

    # Derive FY label from model.ym (e.g. "2026-04" → "FY 2026–27")
    fy_label = _fy_label_from_ym(model.ym or "")

    # ---- Cover (always first) ----
    ws_cover = wb.active
    ws_cover.title = "Cover"
    used.add("Cover")
    _render_cover(ws_cover, model, fy_label=fy_label)

    # ---- Content sheets ----
    content = model.sheets or []
    if not content and not model.available:
        # Failed build — show one sheet with the failure note
        content = [ReportSheet(
            name="Report",
            title=model.label or "Management Report",
            note=(model.headline or "Build failed — see Notes tab for details."),
        )]

    for sh in content:
        raw_name = _safe_tab(sh.name)
        name = raw_name
        k = 2
        while name in used:
            name = f"{raw_name[:28]}-{k}"
            k += 1
        used.add(name)
        ws = wb.create_sheet(title=name)
        cell_registry = _render_sheet(ws, sh)
        _apply_comments(cell_registry, sh)

    # ---- Notes (always last) ----
    notes_name = "Notes"
    while notes_name in used:
        notes_name = "Notes-2"
    ws_notes = wb.create_sheet(title=notes_name)
    _render_notes(ws_notes, model.flags or [])

    return wb


def workbook_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

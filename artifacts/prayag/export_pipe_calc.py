"""One-off: export PIPE M/C output & rejection calculation to an Excel workbook.

Faithfully REUSES the app's own pure functions (no re-implementation):
  * parsers.parse_daily_matrix  -> Report-5 per-(machine,date) rows
  * pipe_reconcile.parse_report11 -> Report-11 per-(machine,date) + type split
  * pipe_reconcile.reconcile      -> date-wise MAX + pro-rata type allocation

Output: <repo>/exports/pipe_mc_calculation_<YYYY-MM>.xlsx
"""
from __future__ import annotations

import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import parsers
import pipe_reconcile
import sheets
import sources

NAVY = "1F3864"
TERRA = "C55A11"
LIGHT = "EAEDF3"

PIPE_FILES = sources.DAILY_SOURCES["PIPE"]["files"]
FY2627 = ["2026-07", "2026-06", "2026-05", "2026-04"]  # newest first


def _pick_month(token: str):
    """Return (ym, r5, r11, corrected, audit, raw) for the newest month w/ output."""
    for ym in FY2627:
        if ym not in PIPE_FILES or ("PIPE", ym) in sources.EMPTY_SOURCES:
            continue
        file_id = PIPE_FILES[ym]
        seg, unit = sheets._daily_seg_unit("PIPE")
        v5 = sheets.read_values(file_id, "Report-5", token)
        raw = parsers.parse_daily_matrix(
            v5, plant="PIPE", segment=seg, unit=unit, year_month=ym,
            source_file=file_id, source_tab="Report-5",
            mc_header_spec=("eq", "MACHINE"),
        )
        raw = [r for r in raw if sheets._mc_key(r.machine) is not None]
        r5: dict = {}
        label_for: dict = {}
        for r in raw:
            k = sheets._mc_key(r.machine)
            d = r5.setdefault((k, r.date), {"out": 0.0, "rej": 0.0})
            d["out"] += r.total_count
            d["rej"] += r.reject_count
            label_for.setdefault(k, r.machine)
        v11 = sheets.read_values(file_id, "Report-11", token)
        r11 = pipe_reconcile.parse_report11(v11, ym, sheets._mc_key)
        corrected, audit = pipe_reconcile.reconcile(r5, r11)
        if audit["out_total"] > 0:
            return ym, r5, r11, corrected, audit, raw, label_for, unit
    raise SystemExit("No PIPE month with output found in FY2026-27.")


def _machine_rollup(corrected, raw, label_for):
    """Aggregate the per-(machine,date) corrected cells up to per-machine rows."""
    hours = defaultdict(float)
    for r in raw:
        hours[sheets._mc_key(r.machine)] += r.actual_hours or 0.0
    agg: dict = {}
    for (k, _date), c in corrected.items():
        a = agg.setdefault(k, {
            "r5_out": 0.0, "r11_out": 0.0, "out": 0.0,
            "r5_rej": 0.0, "r11_rej": 0.0, "rej": 0.0,
            "types": defaultdict(float), "untyped": 0.0,
            "cells": 0, "r5_only": 0, "r11_only": 0, "both": 0,
        })
        a["r5_out"] += c["r5_out"]; a["r11_out"] += c["r11_out"]; a["out"] += c["out"]
        a["r5_rej"] += c["r5_rej"]; a["r11_rej"] += c["r11_rej"]; a["rej"] += c["rej"]
        for t, v in c["types"].items():
            a["types"][t] += v
        a["untyped"] += c["untyped"]
        a["cells"] += 1
        r5_has = c["r5_out"] > 0 or c["r5_rej"] > 0
        r11_has = c["r11_out"] > 0 or c["r11_rej"] > 0
        if r5_has and r11_has:
            a["both"] += 1
        elif r5_has:
            a["r5_only"] += 1
        else:
            a["r11_only"] += 1
    for k, a in agg.items():
        a["hours"] = hours.get(k, 0.0)
        a["label"] = label_for.get(k, f"M/C-{k}")
    return agg


# ---- styling helpers -------------------------------------------------------
_thin = Side(style="thin", color="C9CFDD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _cell(c, val, num=False, bold=False, fill=None, color=None):
    c.value = val
    c.border = BORDER
    c.font = Font(bold=bold, color=color or "1A1A1A")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if num:
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
    else:
        c.alignment = Alignment(horizontal="left")


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build(ym, r5, r11, corrected, audit, agg, unit):
    wb = openpyxl.Workbook()
    machines = sorted(agg)
    all_types = sorted({t for a in agg.values() for t in a["types"]})

    # ---------------- Sheet 1: How it works --------------------------------
    ws = wb.active
    ws.title = "How it works"
    ws.sheet_view.showGridLines = False
    t = ws.cell(1, 1, "PIPE — How M/C output & rejection are calculated")
    t.font = Font(bold=True, size=15, color=NAVY)
    ws.cell(2, 1, f"Plant: Pipe & Fitting   ·   Period: {ym}   ·   Output unit: {unit}").font = Font(italic=True, color=TERRA, size=11)
    lines = [
        ("Two independent sources", None),
        ("Report-5", "Per-machine daily matrix — Run Hours / Output / Rejection per date. Carries run hours, but NO pipe type."),
        ("Report-11", "Item-wise actual production journal — carries the pipe TYPE (CPVC/UPVC/SWR/AGRI) per row, but NO run hours."),
        ("", ""),
        ("Output & rejection (per machine)", None),
        ("Rule", "For EVERY (machine, date) cell the corrected figure = MAX(Report-5, Report-11), taken over the UNION of all cells either source reports — separately for output and for rejection."),
        ("Why MAX", "Neither report is complete: each misses machine-days the other records. The date-wise maximum recovers the true total without double-counting."),
        ("Run hours", "Always taken from Report-5 only (Report-11 has none). A Report-11-only machine-date has output but no run hours, so its efficiency is understated."),
        ("", ""),
        ("Category split (CPVC / UPVC / SWR / AGRI)", None),
        ("Rule", "Report-11's per-type proportions for a cell are scaled pro-rata to that cell's corrected OUTPUT. When Report-5 is the higher figure, the extra output is spread across the same proportions."),
        ("Untyped pickup", "A machine-date present only in Report-5 (no Report-11 type signal) is reported as 'untyped pickup' — its type is never guessed."),
        ("Audit-only", "The headline per-machine figure is type-agnostic; the category split is a reconciliation/audit view and never changes the headline total."),
        ("Rejection by category", "NOT split by type — the source only records rejection per machine-date, not per pipe type. Rejection is therefore shown per machine only."),
    ]
    row = 4
    for head, body in lines:
        if body is None:
            c = ws.cell(row, 1, head)
            c.font = Font(bold=True, size=12, color=NAVY)
            c.fill = PatternFill("solid", fgColor=LIGHT)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        else:
            c1 = ws.cell(row, 1, head)
            c1.font = Font(bold=True, color=TERRA)
            c1.alignment = Alignment(vertical="top")
            c2 = ws.cell(row, 2, body)
            c2.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            if body:
                ws.row_dimensions[row].height = 42
        row += 1
    _autosize(ws, [26, 22, 18, 18, 18, 18])

    # ---------------- Sheet 2: By Machine ----------------------------------
    ws = wb.create_sheet("By Machine")
    ws.sheet_view.showGridLines = False
    heads = ["M/C", "Report-5\nOutput", "Report-11\nOutput", "Corrected\nOutput (MAX)",
             "Report-5\nRejection", "Report-11\nRejection", "Corrected\nRejection (MAX)",
             "Run\nHours", "Cells\n(R5/R11/both)"]
    for j, h in enumerate(heads, 1):
        _hdr(ws.cell(1, j, h))
    tot = defaultdict(float)
    r = 2
    for k in machines:
        a = agg[k]
        _cell(ws.cell(r, 1, a["label"]), a["label"], bold=True)
        _cell(ws.cell(r, 2), round(a["r5_out"]), num=True)
        _cell(ws.cell(r, 3), round(a["r11_out"]), num=True)
        _cell(ws.cell(r, 4), round(a["out"]), num=True, bold=True, fill=LIGHT)
        _cell(ws.cell(r, 5), round(a["r5_rej"]), num=True)
        _cell(ws.cell(r, 6), round(a["r11_rej"]), num=True)
        _cell(ws.cell(r, 7), round(a["rej"]), num=True, bold=True, fill=LIGHT)
        _cell(ws.cell(r, 8), round(a["hours"]), num=True)
        _cell(ws.cell(r, 9), f'{a["r5_only"]}/{a["r11_only"]}/{a["both"]}')
        for key, col in (("r5_out", 2), ("r11_out", 3), ("out", 4),
                         ("r5_rej", 5), ("r11_rej", 6), ("rej", 7), ("hours", 8)):
            tot[col] += a[key]
        r += 1
    _cell(ws.cell(r, 1, "TOTAL"), "TOTAL", bold=True, fill=NAVY, color="FFFFFF")
    for col in range(2, 9):
        _cell(ws.cell(r, col), round(tot[col]), num=True, bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 9), "", fill=NAVY)
    ws.freeze_panes = "A2"
    _autosize(ws, [16, 12, 12, 14, 12, 12, 14, 10, 14])

    # ---------------- Sheet 3: Machine x Category (output) -----------------
    ws = wb.create_sheet("Machine x Category")
    ws.sheet_view.showGridLines = False
    cols = ["M/C"] + all_types + ["Untyped pickup", "Total output"]
    for j, h in enumerate(cols, 1):
        _hdr(ws.cell(1, j, h))
    coltot = defaultdict(float)
    r = 2
    for k in machines:
        a = agg[k]
        _cell(ws.cell(r, 1, a["label"]), a["label"], bold=True)
        rowtot = 0.0
        for j, t in enumerate(all_types, 2):
            v = a["types"].get(t, 0.0)
            _cell(ws.cell(r, j), round(v), num=True)
            coltot[j] += v; rowtot += v
        ju = 2 + len(all_types)
        _cell(ws.cell(r, ju), round(a["untyped"]), num=True)
        coltot[ju] += a["untyped"]; rowtot += a["untyped"]
        _cell(ws.cell(r, ju + 1), round(rowtot), num=True, bold=True, fill=LIGHT)
        r += 1
    _cell(ws.cell(r, 1, "TOTAL"), "TOTAL", bold=True, fill=NAVY, color="FFFFFF")
    grand = 0.0
    for j in range(2, 2 + len(all_types) + 1):
        _cell(ws.cell(r, j), round(coltot[j]), num=True, bold=True, fill=NAVY, color="FFFFFF")
        grand += coltot[j]
    _cell(ws.cell(r, 2 + len(all_types) + 1), round(grand), num=True, bold=True, fill=NAVY, color="FFFFFF")
    ws.freeze_panes = "A2"
    _autosize(ws, [16] + [12] * (len(all_types)) + [15, 14])

    # ---------------- Sheet 4: By Category (totals) ------------------------
    ws = wb.create_sheet("By Category")
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "Category output split (audit-only — does not change headline)").font = Font(bold=True, size=12, color=NAVY)
    for j, h in enumerate(["Category", f"Output ({unit})", "% of output"], 1):
        _hdr(ws.cell(3, j, h))
    tt = dict(audit["type_totals"])
    tt_untyped = audit["untyped_kg"]
    grand = audit["out_total"] or 1.0
    r = 4
    for t in all_types:
        v = tt.get(t, 0.0)
        _cell(ws.cell(r, 1, t), t, bold=True)
        _cell(ws.cell(r, 2), round(v), num=True)
        pc = ws.cell(r, 3, v / grand); pc.number_format = "0.0%"; pc.border = BORDER
        pc.alignment = Alignment(horizontal="right")
        r += 1
    _cell(ws.cell(r, 1, "Untyped pickup"), "Untyped pickup", bold=True)
    _cell(ws.cell(r, 2), round(tt_untyped), num=True)
    pc = ws.cell(r, 3, tt_untyped / grand); pc.number_format = "0.0%"; pc.border = BORDER
    pc.alignment = Alignment(horizontal="right"); r += 1
    _cell(ws.cell(r, 1, "TOTAL OUTPUT"), "TOTAL OUTPUT", bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 2), round(audit["out_total"]), num=True, bold=True, fill=NAVY, color="FFFFFF")
    pc = ws.cell(r, 3, 1.0); pc.number_format = "0.0%"; pc.border = BORDER
    pc.font = Font(bold=True, color="FFFFFF"); pc.fill = PatternFill("solid", fgColor=NAVY)
    pc.alignment = Alignment(horizontal="right")
    r += 2
    ws.cell(r, 1, "Note: rejection is reconciled per machine (MAX of R5/R11) and is NOT").font = Font(italic=True, color="666666")
    ws.cell(r + 1, 1, "split by category — the source records rejection per machine-date, not per type.").font = Font(italic=True, color="666666")
    ws.cell(r + 3, 1, f"Total corrected rejection this period ({unit}): {round(audit['rej_total']):,}").font = Font(bold=True, color=TERRA)
    _autosize(ws, [22, 16, 12])

    return wb


def main():
    token = sheets._get_access_token()
    if not token:
        raise SystemExit("Google Sheets connection not authorized.")
    ym, r5, r11, corrected, audit, raw, label_for, unit = _pick_month(token)
    agg = _machine_rollup(corrected, raw, label_for)
    wb = build(ym, r5, r11, corrected, audit, agg, unit)
    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "..", "exports")
    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"pipe_mc_calculation_{ym}.xlsx")
    wb.save(path)
    print(f"PERIOD={ym}")
    print(f"OUT_TOTAL={round(audit['out_total']):,}  REJ_TOTAL={round(audit['rej_total']):,}")
    print(f"R5_OUT={round(sum(d['out'] for d in r5.values())):,}  R11_OUT={round(sum(d['out'] for d in r11.values())):,}")
    print(f"TYPES={ {t: round(v) for t, v in audit['type_totals'].items()} }  UNTYPED={round(audit['untyped_kg']):,}")
    print(f"MACHINES={len(agg)}")
    print(f"SAVED={path}")


if __name__ == "__main__":
    main()

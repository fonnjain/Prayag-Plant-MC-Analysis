"""One-off: export per-machine OUTPUT & REJECTION calculation for the remaining
plants (HDPE, Injection Moulding, PTMT, Tanks-KH) to Excel workbooks.

Faithfully REUSES the app's own compute layer: sheets.get_daily_records returns
the fully reconciled daily Records exactly as the dashboard uses them (MOULDING
run-hours joined from Report-5, PTMT process-group segmentation, TANK per-item
rows, HDPE self-published baselines). We only roll them up by machine and by the
plant's natural sub-dimension — no figure is re-derived here.

One workbook per plant -> <repo>/exports/<plant>_mc_calculation_<YYYY-MM>.xlsx
"""
from __future__ import annotations

import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sheets
import sources

NAVY = "1F3864"
TERRA = "C55A11"
LIGHT = "EAEDF3"

# Newest month with real (non-idle) data per plant — see probe.
PLANT_MONTH = {"HDPE": "2026-05", "MOULDING": "2026-06", "PTMT": "2026-06", "TANK": "2026-06"}

_thin = Side(style="thin", color="C9CFDD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _cell(c, val, num=False, bold=False, fill=None, color=None, pct=False):
    c.value = val
    c.border = BORDER
    c.font = Font(bold=bold, color=color or "1A1A1A")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if pct:
        c.number_format = "0.0%"
        c.alignment = Alignment(horizontal="right")
    elif num:
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
    else:
        c.alignment = Alignment(horizontal="left")


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _method_sheet(wb, title, ym, unit, lines):
    ws = wb.active
    ws.title = "How it works"
    ws.sheet_view.showGridLines = False
    t = ws.cell(1, 1, title)
    t.font = Font(bold=True, size=15, color=NAVY)
    ws.cell(2, 1, f"Period: {ym}   ·   Output unit: {unit}").font = Font(italic=True, color=TERRA, size=11)
    row = 4
    for head, body in lines:
        if body is None:
            c = ws.cell(row, 1, head)
            c.font = Font(bold=True, size=12, color=NAVY)
            c.fill = PatternFill("solid", fgColor=LIGHT)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        else:
            c1 = ws.cell(row, 1, head)
            c1.font = Font(bold=True, color=TERRA)
            c1.alignment = Alignment(vertical="top")
            c2 = ws.cell(row, 2, body)
            c2.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            if body:
                ws.row_dimensions[row].height = max(30, 15 * (len(body) // 60 + 1))
        row += 1
    _autosize(ws, [26, 20, 18, 18, 18, 18])
    return ws


def _group(records, keyfn):
    g = defaultdict(lambda: {"out": 0.0, "rej": 0.0, "hrs": 0.0, "ideal_hrs": 0.0,
                             "rate": 0.0, "n": 0, "unit": "", "fin": False})
    for r in records:
        k = keyfn(r)
        if k is None:
            continue
        d = g[k]
        d["out"] += r.total_count
        d["rej"] += r.reject_count
        d["hrs"] += r.actual_hours or 0.0
        d["ideal_hrs"] += r.ideal_hours or 0.0
        d["rate"] = r.ideal_rate or d["rate"]
        d["n"] += 1
        d["unit"] = r.unit or d["unit"]
        d["fin"] = d["fin"] or r.is_finishing
    return g


def _machine_sheet(wb, records, unit, title, hours=True, extra_hdpe=False):
    ws = wb.create_sheet("By Machine")
    ws.sheet_view.showGridLines = False
    heads = ["M/C", f"Output ({unit})", f"Rejection ({unit})", "Reject %"]
    if hours:
        heads.append("Run Hours")
    if extra_hdpe:
        heads += ["Ideal Hours", "Ideal Output (kg/hr)"]
    heads.append("Note")
    for j, h in enumerate(heads, 1):
        _hdr(ws.cell(1, j, h))
    g = _group(records, lambda r: r.machine or "(unlabelled)")
    tot = defaultdict(float)
    r = 2
    for k in sorted(g):
        d = g[k]
        _cell(ws.cell(r, 1, k), k, bold=True)
        _cell(ws.cell(r, 2), round(d["out"]), num=True, fill=LIGHT if not d["fin"] else None)
        _cell(ws.cell(r, 3), round(d["rej"]), num=True)
        rp = (d["rej"] / d["out"]) if d["out"] else None
        _cell(ws.cell(r, 4), rp if rp is not None else "", pct=(rp is not None))
        col = 5
        if hours:
            _cell(ws.cell(r, col), round(d["hrs"]), num=True); col += 1
        if extra_hdpe:
            _cell(ws.cell(r, col), round(d["ideal_hrs"]), num=True); col += 1
            _cell(ws.cell(r, col), round(d["rate"], 1), num=True); col += 1
        _cell(ws.cell(r, col), "regrind / auxiliary — excluded from headline" if d["fin"] else "")
        if not d["fin"]:
            tot["out"] += d["out"]; tot["rej"] += d["rej"]; tot["hrs"] += d["hrs"]
        r += 1
    _cell(ws.cell(r, 1, "PLANT TOTAL (excl. auxiliaries)"), "PLANT TOTAL (excl. auxiliaries)", bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 2), round(tot["out"]), num=True, bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 3), round(tot["rej"]), num=True, bold=True, fill=NAVY, color="FFFFFF")
    rp = (tot["rej"] / tot["out"]) if tot["out"] else None
    _cell(ws.cell(r, 4), rp if rp is not None else "", pct=(rp is not None), bold=True, fill=NAVY, color="FFFFFF")
    col = 5
    if hours:
        _cell(ws.cell(r, col), round(tot["hrs"]), num=True, bold=True, fill=NAVY, color="FFFFFF"); col += 1
    if extra_hdpe:
        _cell(ws.cell(r, col), "", fill=NAVY); col += 1
        _cell(ws.cell(r, col), "", fill=NAVY); col += 1
    _cell(ws.cell(r, col), "", fill=NAVY)
    ws.freeze_panes = "A2"
    widths = [26, 14, 14, 10]
    if hours:
        widths.append(11)
    if extra_hdpe:
        widths += [12, 18]
    widths.append(42)
    _autosize(ws, widths)
    return ws


def _category_sheet(wb, records, unit, sheet_name, key_label, keyfn):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    for j, h in enumerate([key_label, f"Output ({unit})", "% of output", f"Rejection ({unit})", "Daily rows", "Note"], 1):
        _hdr(ws.cell(1, j, h))
    g = _group(records, keyfn)
    grand = sum(d["out"] for d in g.values() if not d["fin"]) or 1.0
    r = 2
    tout = trej = 0.0
    for k in sorted(g, key=lambda x: -g[x]["out"]):
        d = g[k]
        _cell(ws.cell(r, 1, k), k, bold=True)
        _cell(ws.cell(r, 2), round(d["out"]), num=True)
        _cell(ws.cell(r, 3), (d["out"] / grand) if not d["fin"] else "", pct=(not d["fin"]))
        _cell(ws.cell(r, 4), round(d["rej"]), num=True)
        _cell(ws.cell(r, 5), d["n"], num=True)
        _cell(ws.cell(r, 6), "regrind / auxiliary — excluded from headline" if d["fin"] else "")
        if not d["fin"]:
            tout += d["out"]; trej += d["rej"]
        r += 1
    _cell(ws.cell(r, 1, "TOTAL (excl. auxiliaries)"), "TOTAL (excl. auxiliaries)", bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 2), round(tout), num=True, bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 3), 1.0, pct=True, bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 4), round(trej), num=True, bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 5), "", fill=NAVY)
    _cell(ws.cell(r, 6), "", fill=NAVY)
    ws.freeze_panes = "A2"
    _autosize(ws, [30, 14, 12, 14, 12, 42])
    return ws


def _tank_item_sheet(wb, records, unit):
    ws = wb.create_sheet("By Item")
    ws.sheet_view.showGridLines = False
    for j, h in enumerate(["Item code", "Colour / size", f"Output ({unit})", "% of output", f"Rejection ({unit})"], 1):
        _hdr(ws.cell(1, j, h))
    g = defaultdict(lambda: {"out": 0.0, "rej": 0.0})
    for r in records:
        item = (r.mould or "").strip()
        colour = (r.material or "").strip()
        if not item or item.lower() == "item":  # header artefact
            continue
        d = g[(item, colour)]
        d["out"] += r.total_count
        d["rej"] += r.reject_count
    grand = sum(d["out"] for d in g.values()) or 1.0
    rows = sorted(g.items(), key=lambda kv: -kv[1]["out"])
    r = 2
    tout = trej = 0.0
    for (item, colour), d in rows:
        if d["out"] == 0 and d["rej"] == 0:
            continue
        _cell(ws.cell(r, 1, item), item, bold=True)
        _cell(ws.cell(r, 2, colour), colour)
        _cell(ws.cell(r, 3), round(d["out"]), num=True)
        _cell(ws.cell(r, 4), d["out"] / grand, pct=True)
        _cell(ws.cell(r, 5), round(d["rej"]), num=True)
        tout += d["out"]; trej += d["rej"]
        r += 1
    _cell(ws.cell(r, 1, "TOTAL"), "TOTAL", bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 2), "", fill=NAVY)
    _cell(ws.cell(r, 3), round(tout), num=True, bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 4), 1.0, pct=True, bold=True, fill=NAVY, color="FFFFFF")
    _cell(ws.cell(r, 5), round(trej), num=True, bold=True, fill=NAVY, color="FFFFFF")
    ws.freeze_panes = "A2"
    _autosize(ws, [18, 20, 14, 12, 14])
    return ws


# --------------------------- per-plant builders ----------------------------
def build_hdpe(recs, ym, unit):
    wb = openpyxl.Workbook()
    _method_sheet(wb, "HDPE — How M/C output & rejection are calculated", ym, unit, [
        ("Source", None),
        ("Workbook / tab", "HDPE daily workbook -> 'Daily Report' tab, read with the per-machine MATRIX parser."),
        ("", ""),
        ("Output & rejection (per machine)", None),
        ("Rule", "Each machine row carries repeating per-date triplets (Run Hours / Output / Rejection in kg). Output and rejection are summed per machine across every date in the month."),
        ("Run hours", "Taken from the same 'Daily Report' matrix (HDPE records its own run hours per date)."),
        ("Utilisation & efficiency", "HDPE is unique: it publishes its own Ideal Output rate (kg/hr) and monthly M/C Run Hours in the sheet, so utilisation and output-efficiency compute directly from the daily matrix — no external baseline is needed."),
        ("", ""),
        ("Categories", None),
        ("None below machine", "HDPE has no pipe-type / product sub-breakdown in the source — output is tracked per machine only."),
        ("Note on this month", "HDPE ran only lightly in this period; other months in the workbook are empty templates (idle), never a real zero-output month."),
    ])
    _machine_sheet(wb, recs, unit, "HDPE", hours=True, extra_hdpe=True)
    return wb


def build_moulding(recs, ym, unit):
    wb = openpyxl.Workbook()
    _method_sheet(wb, "Injection Moulding — How M/C output & rejection are calculated", ym, unit, [
        ("Source", None),
        ("Workbook / tab", "Moulding data lives INSIDE the Pipe & Fitting workbook. Output is read from 'Report-12' (the 'Wt in Kgs' column) with the LONG parser; rejection from 'Actual Rejection'; runner produce is tracked separately."),
        ("", ""),
        ("Output & rejection (per machine)", None),
        ("Rule", "One row per machine per date; output (kg) and rejection (kg) are summed per machine across the month."),
        ("Run hours", "Report-12 has NO run hours, so they are joined in from the same workbook's 'Report-5' moulding rows (matched by machine label, e.g. 'A01(NU-200)'). The baseline is Report-5's Ideal Run Hour/Day x Total Run Days."),
        ("Efficiency", "Not available — Moulding publishes no in-sheet ideal-output rate, so output-efficiency is left blank rather than shown against a placeholder."),
        ("", ""),
        ("Categories", None),
        ("Segment split", "The only sub-dimension is the process segment: 'Moulding' (the plant headline) vs 'Grinding' / 'Pulverizing' (regrind / finishing). The auxiliary segments are EXCLUDED from the plant headline — they are reprocessing, not new production."),
        ("No product split", "The source does not break Moulding output down by product/type below the machine."),
    ])
    _machine_sheet(wb, recs, unit, "MOULDING", hours=True)
    _category_sheet(wb, recs, unit, "By Segment", "Process segment", lambda r: r.segment or "(none)")
    return wb


def build_ptmt(recs, ym, unit):
    wb = openpyxl.Workbook()
    _method_sheet(wb, "PTMT — How M/C output & rejection are calculated", ym, unit, [
        ("Source", None),
        ("Workbook / tab", "PTMT daily workbook -> 'Report-5' per-machine MATRIX. Roster is the authoritative 55-machine register."),
        ("", ""),
        ("Output & rejection (per machine)", None),
        ("Rule", "Per-date triplets (Run Hours / Output / Rejection in kg) summed per machine across the month."),
        ("Rejection timing", "The matrix books the WHOLE-MONTH rejection onto the last day, so rejection is only valid at machine-month AGGREGATE grain (exactly what this roll-up sums) — never per single day."),
        ("Run hours & baseline", "Run hours from the matrix; utilisation baseline from the in-sheet flat 572 h/machine/month (Col E) and the in-sheet IDEAL HOUR column."),
        ("", ""),
        ("Categories", None),
        ("Process groups", "PTMT machines are grouped into process segments: Injection (standard), Injection (N-line), Blow Moulding, Corrugator, and Grinding. See the 'By Process Group' sheet."),
        ("Grinding excluded", "The 'Grinding' group is regrind and is EXCLUDED from the plant headline total."),
    ])
    _machine_sheet(wb, recs, unit, "PTMT", hours=True)
    _category_sheet(wb, recs, unit, "By Process Group", "Process group", lambda r: r.segment or "(none)")
    return wb


def build_tank(recs, ym, unit):
    wb = openpyxl.Workbook()
    _method_sheet(wb, "Tanks (KH) — How output & rejection are calculated", ym, unit, [
        ("Source", None),
        ("Workbook / tab", "TANK daily workbook -> 'PROD. REPORT' tab, read with the tank parser."),
        ("", ""),
        ("Output & rejection", None),
        ("Per ITEM, not per machine", "Tanks are logged per ITEM with no machine identity, so there is no per-machine breakdown and no per-machine OEE. Output and rejection are summed per item."),
        ("Unit", f"The headline unit is chosen by data presence (Ltr -> pcs -> kg). This period is recorded in {unit}."),
        ("Run hours", "None are recorded, so utilisation / efficiency are SUPPRESSED (left blank) — never shown as a fake 0%. Only output (and rejection) are reported."),
        ("", ""),
        ("Categories", None),
        ("Item + colour", "The natural sub-dimension is the item code and its colour/size — see the 'By Item' sheet. This is Tanks' equivalent of the per-category breakdown."),
    ])
    _tank_item_sheet(wb, recs, unit)
    return wb


BUILDERS = {"HDPE": build_hdpe, "MOULDING": build_moulding, "PTMT": build_ptmt, "TANK": build_tank}


def main():
    months = sorted(set(PLANT_MONTH.values()), reverse=True)
    recs, reports, warns = sheets.get_daily_records(months)
    failed_pairs = next(
        (report["_failed_pairs"] for report in reports
         if isinstance(report, dict) and "_failed_pairs" in report),
        [],
    )

    def _withheld(plant: str, ym: str) -> bool:
        # Moulding is emitted by the PIPE workbook; a failed PIPE pair means
        # Moulding's calculation workbook would be incomplete too.
        return (
            (plant, ym) in failed_pairs
            or (plant == "MOULDING" and ("PIPE", ym) in failed_pairs)
        )

    for w in warns:
        print("WARN:", w)
    out_dir = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "exports"))
    os.makedirs(out_dir, exist_ok=True)
    saved = []
    for plant, ym in PLANT_MONTH.items():
        if _withheld(plant, ym):
            print(
                f"WITHHELD: {plant} {ym} daily source is incomplete; "
                "no calculation workbook was produced."
            )
            continue
        pr = [r for r in recs if r.plant == plant and (r.date or r.period)[:7] == ym]
        unit = ""
        for r in pr:
            if r.total_count:
                unit = r.unit or unit
                break
        unit = unit or (pr[0].unit if pr else "kg")
        wb = BUILDERS[plant](pr, ym, unit)
        fname = f"{sources.PLANT_NAMES.get(plant, plant).replace(' ', '_').replace('(', '').replace(')', '')}_calculation_{ym}.xlsx"
        path = os.path.join(out_dir, fname)
        wb.save(path)
        tot = sum(r.total_count for r in pr if not r.is_finishing)
        rej = sum(r.reject_count for r in pr if not r.is_finishing)
        print(f"{plant} {ym}: out={round(tot):,} {unit} rej={round(rej):,} recs={len(pr)} -> {path}")
        saved.append(path)
    print("SAVED:", len(saved))


if __name__ == "__main__":
    main()

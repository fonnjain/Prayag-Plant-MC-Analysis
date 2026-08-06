"""
Machine Planning Report generators — Phase MP-2 + MP-3.

Generates:
  Report-11 (all pipe machines) + Report-11A–D (machine-group splits)
  Report-12 (fittings / moulding)
  Consolidated plan report (7-tab summary workbook)

Column order for Report-11 (exact, 16 cols):
  DATE | WEEK | SHIFT | MACHINE NAME | MACHINE NO. | TYPES | ITEM CODE |
  Running Hours | Production Wt (KG) | Material Req (KG) | Pcs | Wt./Pc. |
  Ideal Output Per Hour | Actual Output Per Hour | Output Efficiency |
  Compound Cost (Rs)

  When a ScheduleResult is supplied (FIX 1):
    DATE  = day number (1..N)
    WEEK  = "W1".."W4"
    SHIFT = "DAY" | "NIGHT"
    Running Hours = planned_hours − excess_hours  (net production hours)
  Without a ScheduleResult (monthly aggregate fallback):
    DATE  = month label (e.g. "Jul-2026")
    WEEK  = "" (blank)
    SHIFT = "" (blank)
    Running Hours = aggregate assignment hours

  Production Wt (KG) = qty_pcs × wt_per_pc              (no waste — FIX 3)
  Material Req (KG)  = production_wt × (1 + waste_pct%)  (with waste — FIX 3)

Header at row 5 (rows 1–4 = title block). Data from row 6 onward.

ADDITIVE / ISOLATED: reads only mp_engine.EngineResult / ScheduleResult,
never touches the production pipeline.
"""
from __future__ import annotations

import io
import re
from collections import defaultdict
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from mp_engine import AssignedPortion, EngineResult, ItemResult, REPORT_11_GROUPS

# ── Style constants ───────────────────────────────────────────────────────────
NAVY  = "1F3864"
TERRA = "C55A11"
WHITE = "FFFFFF"
LIGHT = "F2F2F2"
_FONT = "Arial"

_NAVY_FILL  = PatternFill("solid", fgColor=NAVY)
_TERRA_FILL = PatternFill("solid", fgColor=TERRA)
_LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
_THIN  = Side(style="thin",   color="D9D9D9")
_MED   = Side(style="medium", color="9E9E9E")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_HEADER = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)

# ── Column spec for Report-11 (16 columns) ────────────────────────────────────
# (header_text, col_width, num_format, align)
# Positional index is referenced by _COL_* constants.
_COLUMNS: List[Tuple[str, int, str, str]] = [
    # 0  DATE
    ("DATE",                     9,  "General",                     "center"),
    # 1  WEEK (W1..W4)  — FIX 1
    ("WEEK",                     7,  "General",                     "center"),
    # 2  SHIFT (DAY/NIGHT)  — FIX 1
    ("SHIFT",                    8,  "General",                     "center"),
    # 3  MACHINE NAME
    ("MACHINE NAME",            14,  "General",                     "center"),
    # 4  MACHINE NO.
    ("MACHINE NO.",             10,  "General",                     "center"),
    # 5  TYPES (material)
    ("TYPES",                   10,  "General",                     "center"),
    # 6  ITEM CODE
    ("ITEM CODE",               14,  "General",                     "left"),
    # 7  Running Hours
    ("Running Hours",           13,  '#,##0.00;-#,##0.00;"-"',      "right"),
    # 8  Production Wt (KG) = qty_pcs × wt/pc, NO waste  — FIX 3
    ("Production Wt (KG)",      16,  '#,##0.0;-#,##0.0;"-"',        "right"),
    # 9  Material Req (KG) = production × (1+waste%), WITH waste  — FIX 3
    ("Material Req (KG)",       16,  '#,##0.0;-#,##0.0;"-"',        "right"),
    # 10 Pcs
    ("Pcs",                     11,  '#,##0;-#,##0;"-"',            "right"),
    # 11 Wt./Pc.
    ("Wt./Pc.",                 10,  '0.0000;-0.0000;"-"',          "right"),
    # 12 Ideal Output Per Hour
    ("Ideal Output Per Hour",   18,  '#,##0.0;-#,##0.0;"-"',        "right"),
    # 13 Actual Output Per Hour (blank for plan)
    ("Actual Output Per Hour",  18,  "General",                     "center"),
    # 14 Output Efficiency (blank for plan)
    ("Output Efficiency",       16,  "General",                     "center"),
    # 15 Compound Cost (Rs)
    ("Compound Cost (Rs)",      18,  '#,##0.0;-#,##0.0;"-"',        "right"),
]

# Columns to sum in the TOTALS row (0-indexed into _COLUMNS)
_TOTAL_COLS = {7, 8, 9, 10, 15}   # Running Hrs, Prod Wt, Mat Req, Pcs, Cost

_PLAN_PLACEHOLDER = ""
_HDR_ROW  = 5
_DATA_ROW = 6


# ── General helpers ───────────────────────────────────────────────────────────

def _mc_num(label: str) -> str:
    """'M/C-3' → '3'."""
    m = re.search(r"(\d+)$", label or "")
    return m.group(1) if m else label


def _month_label(em: str) -> str:
    """'2026-07' → 'Jul-2026'."""
    try:
        import datetime
        y, mo = int(em[:4]), int(em[5:7])
        return datetime.date(y, mo, 1).strftime("%b-%Y")
    except Exception:
        return em


def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_cell(ws, row: int, col: int, value, num_fmt: str, align: str,
                bold: bool = False, fill=None, font_color: str = "000000",
                border=None, wrap: bool = False):
    c = ws.cell(row=row, column=col)
    c.value = value
    if num_fmt and num_fmt != "General" and not isinstance(value, str):
        c.number_format = num_fmt
    c.font = Font(name=_FONT, bold=bold, color=font_color, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        c.fill = fill
    c.border = border or _BORDER
    return c


# ── Report-11 title + header ──────────────────────────────────────────────────

def _write_title_block(ws, report_label: str, em: str, machine_filter: str = "",
                       n_cols: Optional[int] = None):
    """Write rows 1–4 (title block above the header row)."""
    if n_cols is None:
        n_cols = len(_COLUMNS)
    month_str = _month_label(em)
    scheduled_note = ""
    titles = [
        f"REPORT - {report_label}",
        f"Machine Planning — Pipe Production Plan   [{month_str}]{scheduled_note}",
        machine_filter or "All Extrusion Machines",
        "",
    ]
    for r, text in enumerate(titles, start=1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        c = ws.cell(row=r, column=1)
        c.value = text
        if r == 1:
            c.font = Font(name=_FONT, bold=True, size=13, color=WHITE)
            c.fill = _NAVY_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif r == 2:
            c.font = Font(name=_FONT, bold=True, size=11, color=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.font = Font(name=_FONT, size=10, color="595959")
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18


def _write_header_row(ws, columns: list = None):
    """Write header row at _HDR_ROW."""
    if columns is None:
        columns = _COLUMNS
    for col, (hdr, width, _, _align) in enumerate(columns, start=1):
        c = ws.cell(row=_HDR_ROW, column=col)
        c.value = hdr
        c.font = Font(name=_FONT, bold=True, color=WHITE, size=10)
        c.fill = _NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER_HEADER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[_HDR_ROW].height = 36


# ── Row-building — schedule-based (FIX 1) ────────────────────────────────────

def _build_rows_from_schedule(
    result: EngineResult,
    schedule,                          # ScheduleResult
    machine_filter: Optional[Set[str]] = None,
) -> List[tuple]:
    """
    Build data rows from ShiftBlock data (shift-level granularity).

    Each non-idle ShiftBlock → one row.  DATE = day number, WEEK = "W{n}",
    SHIFT = "DAY" | "NIGHT".

    Returns list of 16-tuples matching _COLUMNS order.
    """
    item_by_code: Dict[str, ItemResult] = {it.item_code: it for it in result.items}
    cost_map: Dict[str, float] = getattr(result, "effective_costs", {}) or {}
    waste_pct = float((result.params_used or {}).get("waste_pct", 4.0))
    waste_factor = 1.0 + waste_pct / 100.0

    rows: List[tuple] = []

    for block in schedule.blocks:
        if block.is_idle or not block.item_code:
            continue
        if machine_filter and block.machine not in machine_filter:
            continue

        item = item_by_code.get(block.item_code)
        if item is None or not item.has_weight:
            continue

        production_hrs = max(0.0, block.planned_hours - block.excess_hours)
        # material_req_kg = net production hours × rate (rate is kg-material/hr)
        material_req_kg = production_hrs * item.rate_kg_per_hr
        # production weight = material without waste
        prod_wt_kg = material_req_kg / waste_factor
        # pcs from production weight ÷ weight-per-piece
        wt_pc = item.weight_per_pc_kg or 0.0
        qty_pcs = prod_wt_kg / wt_pc if wt_pc > 0 else 0.0

        # Compound cost (proportional)
        mat_upper = item.material.upper()
        cost_per_kg = cost_map.get(mat_upper)
        if cost_per_kg is not None and item.material_kg > 0:
            fresh_frac = item.fresh_compound_kg / item.material_kg
            a_cost: Any = round(material_req_kg * fresh_frac * cost_per_kg, 1)
        else:
            a_cost = _PLAN_PLACEHOLDER

        # Sort key: (machine_num, item_code, day, shift_order)
        try:
            mc_n = int(block.machine.split("-")[-1]) if "-" in block.machine else 0
        except ValueError:
            mc_n = 0
        shift_order = 0 if block.shift == "DAY" else 1

        rows.append((
            # sort keys (stripped later)
            mc_n, block.item_code, block.day, shift_order,
            # data columns (16)
            block.day,                         # 0  DATE
            f"W{block.week}",                  # 1  WEEK
            block.shift,                       # 2  SHIFT
            block.machine,                     # 3  MACHINE NAME
            _mc_num(block.machine),            # 4  MACHINE NO.
            item.material,                     # 5  TYPES
            item.item_code,                    # 6  ITEM CODE
            round(production_hrs, 2),          # 7  Running Hours
            round(prod_wt_kg, 1),              # 8  Production Wt (KG)
            round(material_req_kg, 1),         # 9  Material Req (KG)
            round(qty_pcs, 0),                 # 10 Pcs
            wt_pc,                             # 11 Wt./Pc.
            round(item.rate_kg_per_hr, 1),     # 12 Ideal Output Per Hour
            _PLAN_PLACEHOLDER,                 # 13 Actual Output Per Hour
            _PLAN_PLACEHOLDER,                 # 14 Output Efficiency
            a_cost,                            # 15 Compound Cost (Rs)
        ))

    rows.sort(key=lambda r: r[:4])
    return [r[4:] for r in rows]  # strip sort keys


# ── Row-building — monthly aggregate fallback ─────────────────────────────────

def _build_rows_from_assignments(
    result: EngineResult,
    machine_filter: Optional[Set[str]] = None,
) -> List[tuple]:
    """
    Build data rows from LPT-optimised AssignedPortion data (monthly totals).

    DATE = month label, WEEK = blank, SHIFT = blank.
    Returns list of 16-tuples matching _COLUMNS order.
    """
    date_label = _month_label(result.effective_month)
    cost_map: Dict[str, float] = getattr(result, "effective_costs", {}) or {}
    rows: List[tuple] = []

    for item in sorted(result.items, key=lambda x: x.item_code):
        for a in item.assignments:
            if machine_filter and a.machine not in machine_filter:
                continue
            if not item.has_weight or not item.has_machine:
                continue

            # FIX 3: separate production weight (no waste) from material req (with waste)
            wt_pc = item.weight_per_pc_kg or 0.0
            prod_wt_kg = a.qty_pcs * wt_pc          # no waste
            material_req_kg = a.material_kg          # includes waste (from engine)

            # Compound cost
            mat_upper = item.material.upper()
            cost_per_kg = cost_map.get(mat_upper)
            if cost_per_kg is not None and item.material_kg > 0:
                fresh_frac = item.fresh_compound_kg / item.material_kg
                a_cost: Any = round(material_req_kg * fresh_frac * cost_per_kg, 1)
            else:
                a_cost = _PLAN_PLACEHOLDER

            try:
                mc_n = int(a.machine.split("-")[-1]) if "-" in a.machine else 0
            except ValueError:
                mc_n = 0

            rows.append((
                mc_n, item.item_code,   # sort keys
                date_label,             # 0  DATE
                "",                     # 1  WEEK
                "",                     # 2  SHIFT
                a.machine,              # 3  MACHINE NAME
                _mc_num(a.machine),     # 4  MACHINE NO.
                item.material,          # 5  TYPES
                item.item_code,         # 6  ITEM CODE
                round(a.hrs, 2),        # 7  Running Hours
                round(prod_wt_kg, 1),   # 8  Production Wt (KG)
                round(material_req_kg, 1),  # 9  Material Req (KG)
                round(a.qty_pcs, 0),    # 10 Pcs
                wt_pc,                  # 11 Wt./Pc.
                round(item.rate_kg_per_hr, 1),  # 12 Ideal Output Per Hour
                _PLAN_PLACEHOLDER,      # 13 Actual Output Per Hour
                _PLAN_PLACEHOLDER,      # 14 Output Efficiency
                a_cost,                 # 15 Compound Cost (Rs)
            ))

    rows.sort(key=lambda r: r[:2])
    return [r[2:] for r in rows]  # strip sort keys


def _build_rows(
    result: EngineResult,
    machine_filter: Optional[Set[str]] = None,
    schedule=None,                     # Optional[ScheduleResult]
) -> List[tuple]:
    """Dispatch to schedule-based or assignment-based row builder."""
    if schedule is not None:
        return _build_rows_from_schedule(result, schedule, machine_filter)
    return _build_rows_from_assignments(result, machine_filter)


# ── Data + totals writers ─────────────────────────────────────────────────────

def _write_data_rows(ws, data_rows: List[tuple]):
    for ri, row_vals in enumerate(data_rows):
        r = _DATA_ROW + ri
        fill = _LIGHT_FILL if ri % 2 == 0 else None
        for ci, (val, (_, _, nfmt, align)) in enumerate(zip(row_vals, _COLUMNS), start=1):
            _write_cell(ws, r, ci, val, nfmt, align, fill=fill)


def _add_totals_row(ws, data_rows: List[tuple], n_data: int):
    """Append a TOTAL row summing numeric columns in _TOTAL_COLS."""
    r = _DATA_ROW + n_data
    totals = [None] * len(_COLUMNS)
    totals[0] = "TOTAL"
    for col_idx in _TOTAL_COLS:
        vals = []
        for row_vals in data_rows:
            v = row_vals[col_idx]
            try:
                vals.append(float(v))
            except (TypeError, ValueError):
                pass
        totals[col_idx] = round(sum(vals), 2) if vals else None

    for ci, (val, (_, _, nfmt, align)) in enumerate(zip(totals, _COLUMNS), start=1):
        c = ws.cell(row=r, column=ci)
        c.value = val
        if val is not None and not isinstance(val, str):
            c.number_format = nfmt
        c.font = Font(name=_FONT, bold=True, color=WHITE, size=10)
        c.fill = _TERRA_FILL
        c.alignment = Alignment(
            horizontal="right" if ci > 1 else "left", vertical="center"
        )
        c.border = _BORDER_HEADER


def _add_rate_fallback_note(ws, result: EngineResult, n_data: int):
    """
    FIX 4: Append a small rate-fallback summary block below the totals row.

    Counts items by rate_fallback_tier per material and reports the
    total machine-hours contributed by each fallback tier so users can
    see how much of the plan relies on estimated rates.
    """
    from collections import defaultdict

    # Aggregate: tier → {material: (count, hours)}
    tier_mat: Dict[str, Dict[str, Tuple[int, float]]] = defaultdict(
        lambda: defaultdict(lambda: [0, 0.0])
    )
    overall_avg_val: Optional[float] = None

    for it in result.items:
        if not it.has_weight:
            continue
        tier = getattr(it, "rate_fallback_tier", "item")
        mat  = it.material
        tier_mat[tier][mat][0] += 1
        tier_mat[tier][mat][1] += it.machine_hrs

    has_estimated = any(t != "item" for t in tier_mat)
    if not has_estimated:
        return

    n_cols = len(_COLUMNS)
    r = _DATA_ROW + n_data + 2   # one blank row gap

    # Section title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    tc = ws.cell(row=r, column=1)
    tc.value = "Rate Fallback Summary (FIX 4 — items using estimated throughput rates)"
    tc.font  = Font(name=_FONT, bold=True, size=9, color=WHITE)
    tc.fill  = _NAVY_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    r += 1

    # Sub-header
    hdrs = ["Fallback Tier", "Material", "Items", "Machine Hours", "Note"]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=r, column=ci)
        c.value = h
        c.font  = Font(name=_FONT, bold=True, size=9, color=WHITE)
        c.fill  = PatternFill("solid", fgColor="2E4C91")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
    r += 1

    tier_labels = {
        "item":        "Per-item (seeded)",
        "mat_avg":     "Per-material average",
        "overall_avg": "Overall pipe average",
    }
    tier_notes = {
        "item":        "Exact seeded throughput rate — most accurate",
        "mat_avg":     "Average of seeded items for this material — check adequacy",
        "overall_avg": "Last-resort average — review seeded data for this material",
    }

    fill_warn = PatternFill("solid", fgColor="FFF3CD")
    fill_err  = PatternFill("solid", fgColor="FFE0E0")

    for tier in ("item", "mat_avg", "overall_avg"):
        mat_data = tier_mat.get(tier, {})
        if not mat_data:
            continue
        row_fill = None if tier == "item" else (
            fill_warn if tier == "mat_avg" else fill_err
        )
        for mat, (cnt, hrs) in sorted(mat_data.items()):
            vals = [tier_labels.get(tier, tier), mat, cnt, round(hrs, 1),
                    tier_notes.get(tier, "")]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=ci)
                c.value = v
                c.font  = Font(name=_FONT, size=9)
                c.alignment = Alignment(horizontal="left" if ci in (1, 5) else "center",
                                        vertical="center")
                c.border = _BORDER
                if row_fill:
                    c.fill = row_fill
            r += 1


# ── Workbook builder ─────────────────────────────────────────────────────────

def _build_workbook(
    result: EngineResult,
    report_label: str,
    machine_filter: Optional[Set[str]] = None,
    filter_label: str = "",
    schedule=None,                     # Optional[ScheduleResult]
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Report-{report_label}"
    ws.freeze_panes = f"A{_DATA_ROW}"

    _write_title_block(ws, report_label, result.effective_month, filter_label)
    _write_header_row(ws)
    data_rows = _build_rows(result, machine_filter, schedule)
    _write_data_rows(ws, data_rows)
    n = len(data_rows)
    _add_totals_row(ws, data_rows, n)
    _add_rate_fallback_note(ws, result, n + 1)   # FIX 4

    ws.auto_filter.ref = (
        f"A{_HDR_ROW}:{get_column_letter(len(_COLUMNS))}{_HDR_ROW + n}"
    )
    return wb


# ── Public API — Report-11 / 11A-D ───────────────────────────────────────────

def report_11_bytes(result: EngineResult, schedule=None) -> bytes:
    """Report-11: all pipe machines combined.

    If ``schedule`` (ScheduleResult) is provided the rows reflect the
    day/shift schedule (FIX 1).  Otherwise falls back to monthly LPT totals.
    """
    wb = _build_workbook(result, "11", machine_filter=None,
                         filter_label="All Extrusion Machines",
                         schedule=schedule)
    return _wb_bytes(wb)


def report_11x_bytes(result: EngineResult, group: str, schedule=None) -> bytes:
    """Report-11A / 11B / 11C / 11D: machine-group split.

    ``group`` is one of 'A' | 'B' | 'C' | 'D'.
    Machine membership comes from REPORT_11_GROUPS (config-driven — FIX 2).
    ``schedule`` is an optional ScheduleResult for shift-level rows (FIX 1).
    """
    group = group.upper()
    machines = REPORT_11_GROUPS.get(group, [])
    wb = _build_workbook(
        result,
        f"11{group}",
        machine_filter=set(machines),
        filter_label=", ".join(machines),
        schedule=schedule,
    )
    return _wb_bytes(wb)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT-12 — MOULDING PLAN (MP-3 fittings, unchanged from original)
# ═══════════════════════════════════════════════════════════════════════════════

_R12_COLS = [
    "DATE",
    "MATERIAL",
    "ITEM CODE",
    "Moulding Machine",
    "Mould Cavity",
    "Run Cavity",
    "No. of Cycle",
    "Pcs",
    "Wt in Kgs",
    "Cycle Time",
    "Running Hours",
    "Ideal Output Per Hour",
    "Actual Output Per Hour",
    "Output Efficiency",
    "Rejection Pcs",
    "Rejection Kg",
    "Compound Cost (Rs)",
]
_R12_N_COLS = len(_R12_COLS)   # 17


def report_12_bytes(result: "FittingEngineResult") -> bytes:  # type: ignore[name-defined]
    """Return Report-12 planning .xlsx for the given fitting engine result."""
    from mp_engine import FittingEngineResult, FittingItemResult, FittingAssignedPortion

    wb = Workbook()
    ws = wb.active
    ws.title = "Report-12"

    month = result.effective_month
    seg   = result.segment

    title_font  = Font(bold=True, size=13, color="FFFFFF")
    header_font = Font(bold=True, size=10)
    plan_font   = Font(color="1F3864")
    plan_fill   = PatternFill("solid", fgColor="EBF0FA")
    sub_fill    = PatternFill("solid", fgColor="D9E1F2")

    thin = Side(style="thin")
    med  = Side(style="medium")
    bord_header = Border(left=med, right=med, top=med, bottom=med)
    bord_data   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_aln    = Alignment(horizontal="left",   vertical="center")

    def _hfmt(cell, val):
        cell.value     = val
        cell.font      = header_font
        cell.fill      = sub_fill
        cell.border    = bord_header
        cell.alignment = center_aln

    def _dfmt(cell, val, bold=False):
        cell.value     = val
        cell.font      = Font(bold=bold, size=10, color="1F3864")
        cell.fill      = plan_fill
        cell.border    = bord_data
        cell.alignment = left_aln

    _last_col = get_column_letter(_R12_N_COLS)

    ws.merge_cells(f"A1:{_last_col}1")
    t1 = ws["A1"]
    t1.value     = "PRAYAG PIPES & FITTINGS — MOULDING SECTION"
    t1.font      = title_font
    t1.fill      = PatternFill("solid", fgColor="1F3864")
    t1.alignment = center_aln

    ws.merge_cells(f"A2:{_last_col}2")
    t2 = ws["A2"]
    t2.value     = "REPORT-12  ·  PRODUCTION PLAN (FITTINGS)"
    t2.font      = Font(bold=True, size=11, color="1F3864")
    t2.alignment = center_aln

    ws.merge_cells(f"A3:{_last_col}3")
    t3 = ws["A3"]
    t3.value     = f"Month: {month}   |   Segment: {seg}"
    t3.font      = Font(size=10, italic=True)
    t3.alignment = center_aln

    ws.merge_cells(f"A4:{_last_col}4")
    t4 = ws["A4"]
    t4.value     = "Status: MACHINE PLAN (actuals-only columns left blank)"
    t4.font      = Font(size=9, italic=True, color="C55A11")
    t4.alignment = center_aln

    ws.merge_cells(f"A5:{_last_col}5")

    for col_idx, col_name in enumerate(_R12_COLS, start=1):
        cell = ws.cell(row=6, column=col_idx)
        _hfmt(cell, col_name)
    ws.row_dimensions[6].height = 36

    _widths = [12, 9, 14, 18, 12, 12, 13, 10, 12, 12, 14, 20, 22, 16, 14, 12, 18]
    for i, w in enumerate(_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 7
    machine_totals: dict = {}

    routable_items = [
        it for it in result.items
        if it.has_weight and it.has_machine and it.assignments
    ]
    routable_items = sorted(routable_items, key=lambda x: (x.material, x.item_code))

    fit_cost_map: Dict[str, float] = getattr(result, "effective_costs", {}) or {}

    for it in routable_items:
        if it.material_kg > 0:
            fresh_frac = it.fresh_compound_kg / it.material_kg
        else:
            fresh_frac = 0.0
        item_cost_per_kg = fit_cost_map.get(it.material.upper())

        for a in it.assignments:
            mc          = a.machine
            qty_a       = a.qty_pcs
            mat_kg_a    = a.material_kg
            hrs_a       = a.hrs
            cavity      = it.cavity
            cycle       = it.cycle_time_sec
            ideal_pps   = it.pcs_per_hr
            wt_a        = round(qty_a * (it.weight_per_pc_kg or 0.0), 3)
            num_cycles_a = round(qty_a / cavity) if (cavity and cavity > 0) else None

            if item_cost_per_kg is not None:
                a_cost_r12 = round(a.material_kg * fresh_frac * item_cost_per_kg, 1)
            else:
                a_cost_r12 = None

            row_vals = [
                month,
                it.material,
                it.item_code,
                mc,
                cavity,
                cavity,
                num_cycles_a,
                round(qty_a),
                wt_a,
                cycle,
                round(hrs_a, 3),
                round(ideal_pps, 2) if ideal_pps else None,
                None,
                None,
                None,
                None,
                a_cost_r12,
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                _dfmt(ws.cell(row=row, column=col_idx), val)

            if mc not in machine_totals:
                machine_totals[mc] = {"hrs": 0.0, "pcs": 0.0, "wt_kg": 0.0}
            machine_totals[mc]["hrs"]   += hrs_a
            machine_totals[mc]["pcs"]   += qty_a
            machine_totals[mc]["wt_kg"] += wt_a
            row += 1

    row += 1
    ws.merge_cells(f"A{row}:{_last_col}{row}")
    gap_cell = ws.cell(row=row, column=1)
    no_wt    = len(result.coverage_gaps.no_weight)
    no_mc    = len(result.coverage_gaps.no_machine)
    gap_cell.value = (
        f"Coverage: {len(routable_items)} items planned  |  "
        f"No-weight flagged: {no_wt}  |  "
        f"No-machine (unroutable): {no_mc}  |  "
        f"Material-level fallback used: {result.n_route_estimated} items"
    )
    gap_cell.font      = Font(bold=True, italic=True, size=9, color="C55A11")
    gap_cell.alignment = left_aln
    row += 2

    ws.merge_cells(f"A{row}:{_last_col}{row}")
    ws.cell(row=row, column=1).value = "MACHINE SUMMARY"
    ws.cell(row=row, column=1).font  = Font(bold=True, size=10, color="FFFFFF")
    ws.cell(row=row, column=1).fill  = PatternFill("solid", fgColor="1F3864")
    ws.cell(row=row, column=1).alignment = center_aln
    row += 1

    sum_headers = ["Machine", "Planned Hours", "Total Pcs", "Total Wt (kg)"]
    for ci, h in enumerate(sum_headers, start=1):
        c = ws.cell(row=row, column=ci)
        _hfmt(c, h)
    row += 1

    for mc in sorted(machine_totals):
        t = machine_totals[mc]
        _dfmt(ws.cell(row=row, column=1), mc, bold=True)
        _dfmt(ws.cell(row=row, column=2), round(t["hrs"], 3))
        _dfmt(ws.cell(row=row, column=3), round(t["pcs"]))
        _dfmt(ws.cell(row=row, column=4), round(t["wt_kg"], 2))
        row += 1

    return _wb_bytes(wb)


# ═══════════════════════════════════════════════════════════════════════════════
# CONSOLIDATED PLAN REPORT — 7 tabs
# ═══════════════════════════════════════════════════════════════════════════════
#
# Tabs:
#  1. Summary          — plan parameters, headline totals (pipe + fitting)
#  2. Machine Load     — per-machine scheduled vs capacity (pipe extrusion)
#  3. Weekly Fill      — per (machine, week) fill from ScheduleResult
#  4. Shift Schedule   — block-level day/shift table
#  5. Item Assignment  — per-item detail from engine
#  6. Material & Cost  — compound cost breakdown by material × type
#  7. Coverage & Bottleneck — gaps, rate fallback (FIX 4), unfinished

def consolidated_plan_bytes(
    engine_result: Optional[EngineResult] = None,
    fitting_result=None,                # Optional[FittingEngineResult]
    schedule_result=None,               # Optional[ScheduleResult]
) -> bytes:
    """
    Build the 7-tab consolidated plan workbook.

    Any of the three inputs may be None; tabs that require unavailable data
    are labelled "(no data)" rather than raising.
    """
    wb = Workbook()
    # Remove default sheet
    for sh in list(wb.sheetnames):
        del wb[sh]

    em = (
        (engine_result.effective_month if engine_result else None)
        or (fitting_result.effective_month if fitting_result else None)
        or "—"
    )
    month_label = _month_label(em) if em != "—" else "—"

    # ── Shared style helpers ──────────────────────────────────────────────────
    def _ws_header(ws, title: str, subtitle: str = ""):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        c = ws["A1"]
        c.value     = title
        c.font      = Font(name=_FONT, bold=True, size=13, color=WHITE)
        c.fill      = _NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
        c2 = ws["A2"]
        c2.value     = subtitle or f"Month: {month_label}"
        c2.font      = Font(name=_FONT, size=10, italic=True, color=NAVY)
        c2.alignment = Alignment(horizontal="center", vertical="center")

    def _hdr_cell(ws, row: int, col: int, val: str, w: int = 15):
        c = ws.cell(row=row, column=col)
        c.value     = val
        c.font      = Font(name=_FONT, bold=True, color=WHITE, size=9)
        c.fill      = _NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _BORDER_HEADER
        ws.column_dimensions[get_column_letter(col)].width = w
        return c

    def _data_cell(ws, row: int, col: int, val, num_fmt: str = "General",
                   align: str = "left", bold: bool = False, fill=None):
        c = ws.cell(row=row, column=col)
        c.value = val
        if num_fmt != "General" and not isinstance(val, str) and val is not None:
            c.number_format = num_fmt
        c.font      = Font(name=_FONT, size=9, bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = _BORDER
        if fill:
            c.fill = fill
        return c

    def _terra_total(ws, row: int, n_cols: int, label_col: int,
                     label: str, col_sums: Dict[int, float]):
        for ci in range(1, n_cols + 1):
            c = ws.cell(row=row, column=ci)
            if ci == label_col:
                c.value = label
            elif ci in col_sums:
                c.value = round(col_sums[ci], 2)
            c.font      = Font(name=_FONT, bold=True, color=WHITE, size=9)
            c.fill      = _TERRA_FILL
            c.alignment = Alignment(
                horizontal="left" if ci == label_col else "right",
                vertical="center",
            )
            c.border = _BORDER_HEADER

    # ── TAB 1: Summary ────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("1. Summary")
    _ws_header(ws1, "Consolidated Plan — Summary", f"Month: {month_label}")
    r = 4

    def _kv(label, value, row):
        c1 = ws1.cell(row=row, column=1)
        c1.value     = label
        c1.font      = Font(name=_FONT, bold=True, size=9)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c1.border    = _BORDER
        c1.fill      = PatternFill("solid", fgColor="EEF2FA")
        ws1.column_dimensions["A"].width = 32

        c2 = ws1.cell(row=row, column=2)
        c2.value     = value
        c2.font      = Font(name=_FONT, size=9)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border    = _BORDER
        ws1.column_dimensions["B"].width = 28

    # Plan parameters
    ws1.cell(row=r, column=1).value = "PLAN PARAMETERS"
    ws1.cell(row=r, column=1).font  = Font(name=_FONT, bold=True, size=10, color=NAVY)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    if engine_result:
        wp = engine_result.params_used or {}
        _kv("Waste pct",           f"{wp.get('waste_pct', '—')} %",          r); r += 1
        _kv("Pulverizer pct",      f"{wp.get('pulverizer_pct', '—')} %",     r); r += 1
    if schedule_result:
        sp = schedule_result.params_used or {}
        _kv("Min run block (hrs)", sp.get("min_run_block_hours", "—"),        r); r += 1
        _kv("Week days (W1..W4)",  str(sp.get("week_days", schedule_result.week_days)), r); r += 1

    r += 1
    ws1.cell(row=r, column=1).value = "PIPE PLAN TOTALS"
    ws1.cell(row=r, column=1).font  = Font(name=_FONT, bold=True, size=10, color=NAVY)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    if engine_result:
        t = engine_result.totals
        _kv("Total demand (pcs)",        f"{t.total_qty_pcs:,.0f}",              r); r += 1
        _kv("Total material req (kg)",   f"{t.total_material_kg:,.1f}",          r); r += 1
        _kv("  — Fresh compound (kg)",   f"{t.total_fresh_compound_kg:,.1f}",    r); r += 1
        _kv("  — Pulverizer (kg)",       f"{t.total_pulverizer_kg:,.1f}",        r); r += 1
        _kv("Routable material (kg)",    f"{t.routable_material_kg:,.1f}",       r); r += 1
        total_pipe_hrs = sum(ml.assigned_hrs for ml in engine_result.machine_loads)
        _kv("Total machine hours",       f"{total_pipe_hrs:,.1f}",               r); r += 1
        n_est = sum(1 for it in engine_result.items
                    if it.has_weight and it.rate_estimated)
        n_all = sum(1 for it in engine_result.items if it.has_weight)
        _kv("Items with estimated rate", f"{n_est} of {n_all}",                  r); r += 1
        if engine_result.effective_costs:
            pipe_cost = sum(engine_result.cost_by_material.values())
            _kv("Pipe compound cost (Rs)", f"{pipe_cost:,.0f}",                  r); r += 1
    else:
        _kv("Pipe engine result",        "(not available)",                       r); r += 1

    r += 1
    ws1.cell(row=r, column=1).value = "FITTING PLAN TOTALS"
    ws1.cell(row=r, column=1).font  = Font(name=_FONT, bold=True, size=10, color=NAVY)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    if fitting_result:
        ft = fitting_result.totals
        _kv("Total demand (pcs)",        f"{ft.total_qty_pcs:,.0f}",             r); r += 1
        _kv("Total material req (kg)",   f"{ft.total_material_kg:,.1f}",         r); r += 1
        _kv("Routable material (kg)",    f"{ft.routable_material_kg:,.1f}",      r); r += 1
        total_fit_hrs = sum(ml.assigned_hrs for ml in fitting_result.machine_loads)
        _kv("Total machine hours",       f"{total_fit_hrs:,.1f}",                r); r += 1
        n_re = getattr(fitting_result, "n_route_estimated", 0)
        _kv("Material-fallback routed",  f"{n_re} items",                        r); r += 1
        if getattr(fitting_result, "effective_costs", {}):
            fit_cost = sum(getattr(fitting_result, "cost_by_material", {}).values())
            _kv("Fitting compound cost (Rs)", f"{fit_cost:,.0f}",                r); r += 1
    else:
        _kv("Fitting engine result",     "(not available)",                       r); r += 1

    r += 1
    ws1.cell(row=r, column=1).value = "SCHEDULE TOTALS"
    ws1.cell(row=r, column=1).font  = Font(name=_FONT, bold=True, size=10, color=NAVY)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    if schedule_result:
        _kv("Total capacity (hrs)",   f"{schedule_result.total_capacity_hrs:,.1f}",   r); r += 1
        _kv("Scheduled (hrs)",        f"{schedule_result.total_scheduled_hrs:,.1f}",  r); r += 1
        _kv("Idle (hrs)",             f"{schedule_result.total_idle_hrs:,.1f}",       r); r += 1
        if schedule_result.total_capacity_hrs > 0:
            util = (schedule_result.total_scheduled_hrs
                    / schedule_result.total_capacity_hrs * 100)
            _kv("Overall utilisation",    f"{util:.1f} %",                            r); r += 1
        _kv("Changeovers",            schedule_result.total_changeovers,               r); r += 1
        _kv("Excess stock (kg)",      f"{schedule_result.total_excess_kg:,.1f}",      r); r += 1
        _kv("Unfinished items",       len(schedule_result.unfinished),                 r); r += 1
    else:
        _kv("Shift schedule",         "(not run — upload a demand file)",             r); r += 1

    # ── TAB 2: Machine Load ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("2. Machine Load")
    _ws_header(ws2, "Machine Load — Pipe Extrusion")
    col_hdrs2 = [
        ("Machine", 14), ("Capacity (hrs)", 14), ("Assigned (hrs)", 14),
        ("Utilisation %", 14), ("Machine Days", 13), ("Material (kg)", 14),
        ("Fresh Comp (kg)", 15), ("Pulverizer (kg)", 15),
        ("Staffing OK", 12), ("OT Operators", 13), ("Support (W)", 12),
    ]
    r2 = 4
    for ci, (h, w) in enumerate(col_hdrs2, start=1):
        _hdr_cell(ws2, r2, ci, h, w)
    ws2.row_dimensions[r2].height = 32
    r2 += 1

    if engine_result:
        for ml in engine_result.machine_loads:
            over = ml.utilisation_pct > 100
            row_fill = PatternFill("solid", fgColor="FFE0E0") if over else (
                _LIGHT_FILL if r2 % 2 == 0 else None
            )
            vals2 = [
                ml.machine, ml.capacity_hrs, round(ml.assigned_hrs, 1),
                round(ml.utilisation_pct, 1), round(ml.machine_days, 1),
                round(ml.material_kg, 0), round(ml.fresh_compound_kg, 0),
                round(ml.pulverizer_kg, 0), "Yes" if ml.staffing_ok else "No",
                ml.operators_ot, ml.support_w,
            ]
            fmts2 = ["General", '#,##0.0', '#,##0.0', '0.0', '0.00',
                     '#,##0', '#,##0', '#,##0', "General", "General", "General"]
            alns2 = ["left"] + ["right"] * 7 + ["center", "center", "center"]
            for ci, (v, fmt, aln) in enumerate(zip(vals2, fmts2, alns2), start=1):
                _data_cell(ws2, r2, ci, v, fmt, aln, fill=row_fill)
            r2 += 1

        # Totals
        ml_list = engine_result.machine_loads
        _terra_total(ws2, r2, len(col_hdrs2), 1, "TOTAL", {
            2: sum(ml.capacity_hrs for ml in ml_list),
            3: sum(ml.assigned_hrs for ml in ml_list),
            6: sum(ml.material_kg for ml in ml_list),
            7: sum(ml.fresh_compound_kg for ml in ml_list),
            8: sum(ml.pulverizer_kg for ml in ml_list),
        })
    else:
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=11)
        ws2["A" + str(r2)].value = "(pipe engine result not available)"

    # ── TAB 3: Weekly Fill ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("3. Weekly Fill")
    _ws_header(ws3, "Weekly Fill — Shift Schedule")
    col_hdrs3 = [
        ("Week", 7), ("Machine", 13), ("Capacity (hrs)", 14),
        ("Scheduled (hrs)", 15), ("Idle (hrs)", 12), ("Utilisation %", 13),
        ("Changeovers", 13), ("Excess kg", 12),
        ("Origin W1 (hrs)", 14), ("Origin W2 (hrs)", 14),
        ("Origin W3 (hrs)", 14), ("Origin W4 (hrs)", 14),
    ]
    r3 = 4
    for ci, (h, w) in enumerate(col_hdrs3, start=1):
        _hdr_cell(ws3, r3, ci, h, w)
    ws3.row_dimensions[r3].height = 32
    r3 += 1

    if schedule_result:
        for wf in schedule_result.weekly_fill:
            over = wf.utilisation_pct > 100
            row_fill = PatternFill("solid", fgColor="FFE0E0") if over else (
                _LIGHT_FILL if r3 % 2 == 0 else None
            )
            ob = wf.origin_breakdown
            vals3 = [
                f"W{wf.week}", wf.machine, wf.capacity_hrs,
                wf.scheduled_hrs, wf.idle_hrs, round(wf.utilisation_pct, 1),
                wf.changeovers, wf.excess_kg,
                ob.get(1, 0.0), ob.get(2, 0.0), ob.get(3, 0.0), ob.get(4, 0.0),
            ]
            fmts3 = ["General", "General"] + ['#,##0.0'] * 4 + ['#,##0', '#,##0.0'] + ['#,##0.0'] * 4
            alns3 = ["center", "left"] + ["right"] * 10
            for ci, (v, fmt, aln) in enumerate(zip(vals3, fmts3, alns3), start=1):
                _data_cell(ws3, r3, ci, v, fmt, aln, fill=row_fill)
            r3 += 1
    else:
        ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=12)
        ws3["A" + str(r3)].value = "(schedule not available — upload a demand file)"

    # ── TAB 4: Shift Schedule ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("4. Shift Schedule")
    _ws_header(ws4, "Shift Schedule — Block Detail")
    col_hdrs4 = [
        ("Week", 7), ("Day", 6), ("Machine", 13), ("Shift", 8),
        ("Item Code", 14), ("Material", 10),
        ("Planned Hrs", 12), ("Prod Hrs", 11), ("Excess Hrs", 11),
        ("Origin Wk", 10),
    ]
    r4 = 4
    for ci, (h, w) in enumerate(col_hdrs4, start=1):
        _hdr_cell(ws4, r4, ci, h, w)
    ws4.row_dimensions[r4].height = 28
    r4 += 1

    if schedule_result:
        for blk in schedule_result.blocks:
            if blk.is_idle:
                continue
            prod_hrs = round(blk.planned_hours - blk.excess_hours, 2)
            night_fill = PatternFill("solid", fgColor="E8EAF6") if blk.shift == "NIGHT" else None
            row_fill = night_fill or (_LIGHT_FILL if r4 % 2 == 0 else None)
            vals4 = [
                f"W{blk.week}", blk.day, blk.machine, blk.shift,
                blk.item_code, blk.material,
                round(blk.planned_hours, 2), prod_hrs, round(blk.excess_hours, 2),
                f"W{blk.origin_week}" if blk.origin_week else "—",
            ]
            fmts4 = ["General", "General", "General", "General", "General", "General",
                     '#,##0.00', '#,##0.00', '#,##0.00', "General"]
            alns4 = ["center", "center", "left", "center", "left", "center",
                     "right", "right", "right", "center"]
            for ci, (v, fmt, aln) in enumerate(zip(vals4, fmts4, alns4), start=1):
                _data_cell(ws4, r4, ci, v, fmt, aln, fill=row_fill)
            r4 += 1
    else:
        ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=10)
        ws4["A" + str(r4)].value = "(schedule not available)"

    # ── TAB 5: Item Assignment ────────────────────────────────────────────────
    ws5 = wb.create_sheet("5. Item Assignment")
    _ws_header(ws5, "Item Assignment — Pipe + Fitting")
    col_hdrs5 = [
        ("Type", 8), ("Material", 10), ("Item Code", 15),
        ("Qty (pcs)", 12), ("Wt/pc (kg)", 12),
        ("Machine(s)", 20), ("Machine Hrs", 12),
        ("Prod Wt (kg)", 13), ("Material Req (kg)", 16),
        ("Rate (kg/hr)", 13), ("Rate Tier", 13),
        ("Compound Cost (Rs)", 18),
    ]
    r5 = 4
    for ci, (h, w) in enumerate(col_hdrs5, start=1):
        _hdr_cell(ws5, r5, ci, h, w)
    ws5.row_dimensions[r5].height = 32
    r5 += 1

    waste_pct5 = float((engine_result.params_used or {}).get("waste_pct", 4.0)) if engine_result else 4.0
    waste_factor5 = 1.0 + waste_pct5 / 100.0
    cost_map5 = (engine_result.effective_costs or {}) if engine_result else {}

    if engine_result:
        for it in sorted(engine_result.items, key=lambda x: (x.material, x.item_code)):
            if not it.has_weight:
                continue
            assigned_mcs = ", ".join(
                sorted({a.machine for a in it.assignments})
            ) if it.assignments else ("—" if not it.has_machine else "unrouted")
            prod_wt = it.qty_pcs * (it.weight_per_pc_kg or 0.0)
            mat_req = it.material_kg
            cost_per_kg = cost_map5.get(it.material.upper())
            if cost_per_kg and it.material_kg > 0:
                fresh_frac = it.fresh_compound_kg / it.material_kg
                item_cost = round(it.fresh_compound_kg * cost_per_kg, 1)
            else:
                item_cost = None
            tier = getattr(it, "rate_fallback_tier", "item")
            tier_label = {"item": "seeded", "mat_avg": "mat-avg ⚠", "overall_avg": "overall ⚠⚠"}.get(tier, tier)
            no_mc_fill = PatternFill("solid", fgColor="FFF3CD") if not it.has_machine else (
                _LIGHT_FILL if r5 % 2 == 0 else None
            )
            vals5 = [
                "Pipe", it.material, it.item_code,
                round(it.qty_pcs, 0), it.weight_per_pc_kg or 0.0,
                assigned_mcs, round(it.machine_hrs, 1),
                round(prod_wt, 1), round(mat_req, 1),
                round(it.rate_kg_per_hr, 2), tier_label,
                item_cost,
            ]
            fmts5 = ["General"] * 3 + ['#,##0', '0.0000', "General",
                     '#,##0.0', '#,##0.0', '#,##0.0', '#,##0.0', "General", '#,##0.0']
            alns5 = ["center", "center", "left", "right", "right", "left",
                     "right", "right", "right", "right", "center", "right"]
            for ci, (v, fmt, aln) in enumerate(zip(vals5, fmts5, alns5), start=1):
                _data_cell(ws5, r5, ci, v, fmt, aln, fill=no_mc_fill)
            r5 += 1

    if fitting_result:
        fit_cost5 = getattr(fitting_result, "effective_costs", {}) or {}
        for it in sorted(fitting_result.items, key=lambda x: (x.material, x.item_code)):
            if not it.has_weight:
                continue
            assigned_mcs = ", ".join(
                sorted({a.machine for a in it.assignments})
            ) if it.assignments else ("—" if not it.has_machine else "unrouted")
            prod_wt = it.qty_pcs * (it.weight_per_pc_kg or 0.0)
            mat_req = it.material_kg
            cost_per_kg = fit_cost5.get(it.material.upper())
            if cost_per_kg and it.material_kg > 0:
                ffc = it.fresh_compound_kg
                item_cost = round(ffc * cost_per_kg, 1)
            else:
                item_cost = None
            fit_fill = _LIGHT_FILL if r5 % 2 == 0 else None
            pcs_per_hr = getattr(it, "pcs_per_hr", None)
            rate_display = round(pcs_per_hr, 2) if pcs_per_hr else "—"
            vals5 = [
                "Fitting", it.material, it.item_code,
                round(it.qty_pcs, 0), it.weight_per_pc_kg or 0.0,
                assigned_mcs, round(it.machine_hrs, 1),
                round(prod_wt, 1), round(mat_req, 1),
                rate_display, "pcs/hr",
                item_cost,
            ]
            fmts5 = ["General"] * 3 + ['#,##0', '0.0000', "General",
                     '#,##0.0', '#,##0.0', '#,##0.0', "General", "General", '#,##0.0']
            alns5 = ["center", "center", "left", "right", "right", "left",
                     "right", "right", "right", "right", "center", "right"]
            for ci, (v, fmt, aln) in enumerate(zip(vals5, fmts5, alns5), start=1):
                _data_cell(ws5, r5, ci, v, fmt, aln, fill=fit_fill)
            r5 += 1

    # ── TAB 6: Material & Cost ────────────────────────────────────────────────
    ws6 = wb.create_sheet("6. Material & Cost")
    _ws_header(ws6, "Material & Compound Cost Breakdown")
    col_hdrs6 = [
        ("Type", 8), ("Material", 11), ("Items", 7),
        ("Total Pcs", 12), ("Prod Wt (kg)", 13), ("Material Req (kg)", 16),
        ("Fresh Comp (kg)", 16), ("Pulverizer (kg)", 15),
        ("Cost/kg (Rs)", 13), ("Compound Cost (Rs)", 18),
    ]
    r6 = 4
    for ci, (h, w) in enumerate(col_hdrs6, start=1):
        _hdr_cell(ws6, r6, ci, h, w)
    ws6.row_dimensions[r6].height = 32
    r6 += 1

    def _mat_rows(items_list, item_type: str, cost_map_local: dict, waste_factor_local: float):
        from collections import defaultdict
        acc: Dict[str, dict] = defaultdict(lambda: {
            "cnt": 0, "pcs": 0.0, "prod_wt": 0.0, "mat_req": 0.0,
            "fresh": 0.0, "pulv": 0.0, "cost": 0.0, "cost_rate": None,
        })
        for it in items_list:
            if not it.has_weight:
                continue
            mat = it.material.upper()
            a = acc[mat]
            a["cnt"]     += 1
            a["pcs"]     += it.qty_pcs
            a["prod_wt"] += it.qty_pcs * (it.weight_per_pc_kg or 0.0)
            a["mat_req"] += it.material_kg
            a["fresh"]   += it.fresh_compound_kg
            a["pulv"]    += it.pulverizer_kg
            crate = cost_map_local.get(mat)
            if crate:
                a["cost"]      += it.fresh_compound_kg * crate
                a["cost_rate"]  = crate
        return [(item_type, mat, d) for mat, d in sorted(acc.items())]

    pipe_rows6 = _mat_rows(
        engine_result.items if engine_result else [],
        "Pipe",
        engine_result.effective_costs if engine_result else {},
        waste_factor5,
    )
    fit_rows6 = _mat_rows(
        fitting_result.items if fitting_result else [],
        "Fitting",
        getattr(fitting_result, "effective_costs", {}) if fitting_result else {},
        1.04,
    )

    for item_type, mat, d in pipe_rows6 + fit_rows6:
        row_fill = _LIGHT_FILL if r6 % 2 == 0 else None
        cost_rate_str = f"{d['cost_rate']:.4f}" if d["cost_rate"] else "—"
        vals6 = [
            item_type, mat, d["cnt"],
            round(d["pcs"], 0), round(d["prod_wt"], 1), round(d["mat_req"], 1),
            round(d["fresh"], 1), round(d["pulv"], 1),
            cost_rate_str, round(d["cost"], 0) if d["cost"] > 0 else None,
        ]
        fmts6 = ["General", "General", '#,##0', '#,##0', '#,##0.0', '#,##0.0',
                 '#,##0.0', '#,##0.0', "General", '#,##0']
        alns6 = ["center", "center", "right", "right", "right", "right",
                 "right", "right", "right", "right"]
        for ci, (v, fmt, aln) in enumerate(zip(vals6, fmts6, alns6), start=1):
            _data_cell(ws6, r6, ci, v, fmt, aln, fill=row_fill)
        r6 += 1

    # ── TAB 7: Coverage & Bottleneck ─────────────────────────────────────────
    ws7 = wb.create_sheet("7. Coverage & Bottleneck")
    _ws_header(ws7, "Coverage & Bottleneck — Rate Fallback (FIX 4)")
    r7 = 4

    def _section_title7(label):
        nonlocal r7
        ws7.merge_cells(start_row=r7, start_column=1, end_row=r7, end_column=10)
        c = ws7.cell(row=r7, column=1)
        c.value     = label
        c.font      = Font(name=_FONT, bold=True, size=10, color=WHITE)
        c.fill      = _NAVY_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws7.column_dimensions["A"].width = 18
        r7 += 1

    # 7a: Rate fallback by item (FIX 4)
    _section_title7("7A — Rate Fallback Summary (pipe items)")
    tier_hdrs7a = [("Tier", 20), ("Material", 12), ("Item Count", 11),
                   ("Machine Hrs", 13), ("Note", 50)]
    for ci, (h, w) in enumerate(tier_hdrs7a, start=1):
        _hdr_cell(ws7, r7, ci, h, w)
    r7 += 1

    tier_labels7 = {
        "item":        "seeded (per-item rate)",
        "mat_avg":     "material average ⚠",
        "overall_avg": "overall average ⚠⚠",
    }
    tier_notes7 = {
        "item":        "Most accurate — exact seeded throughput rate",
        "mat_avg":     "Uses average of other seeded items for this material. Review adequacy.",
        "overall_avg": "Last resort — no seeded data for this material. Seed mp_per_hour table.",
    }
    tier_fills7 = {
        "item":        None,
        "mat_avg":     PatternFill("solid", fgColor="FFF3CD"),
        "overall_avg": PatternFill("solid", fgColor="FFE0E0"),
    }

    if engine_result:
        from collections import defaultdict
        tier_acc: Dict[str, Dict[str, list]] = defaultdict(
            lambda: defaultdict(lambda: [0, 0.0])
        )
        for it in engine_result.items:
            if not it.has_weight:
                continue
            tier = getattr(it, "rate_fallback_tier", "item")
            tier_acc[tier][it.material][0] += 1
            tier_acc[tier][it.material][1] += it.machine_hrs

        for tier in ("item", "mat_avg", "overall_avg"):
            mat_data = tier_acc.get(tier, {})
            for mat, (cnt, hrs) in sorted(mat_data.items()):
                fill7 = tier_fills7.get(tier)
                if fill7 is None and r7 % 2 == 0:
                    fill7 = _LIGHT_FILL
                vals7a = [tier_labels7.get(tier, tier), mat, cnt,
                          round(hrs, 1), tier_notes7.get(tier, "")]
                for ci, v in enumerate(vals7a, start=1):
                    _data_cell(ws7, r7, ci, v, "General",
                               "left" if ci in (1, 5) else "center", fill=fill7)
                r7 += 1
    else:
        ws7.cell(row=r7, column=1).value = "(pipe engine result not available)"
        r7 += 1

    r7 += 1

    # 7b: No-weight items
    _section_title7("7B — Items with No BOM Weight (unplannable)")
    if engine_result and engine_result.coverage_gaps.no_weight:
        _hdr_cell(ws7, r7, 1, "Item Code", 15)
        _hdr_cell(ws7, r7, 2, "Material", 12)
        r7 += 1
        # Build a map item_code → material from demand
        mat_map = {it.item_code: it.material for it in engine_result.items}
        for ic in sorted(engine_result.coverage_gaps.no_weight):
            _data_cell(ws7, r7, 1, ic, "General", "left")
            _data_cell(ws7, r7, 2, mat_map.get(ic, "—"), "General", "center",
                       fill=PatternFill("solid", fgColor="FFE0E0"))
            r7 += 1
    else:
        ws7.cell(row=r7, column=1).value = "None" if engine_result else "(not available)"
        r7 += 1

    r7 += 1

    # 7c: No-machine items
    _section_title7("7C — Pipe Items with No Routing (unroutable)")
    if engine_result and engine_result.coverage_gaps.no_machine:
        _hdr_cell(ws7, r7, 1, "Item Code", 15)
        _hdr_cell(ws7, r7, 2, "Material", 12)
        r7 += 1
        mat_map_c = {it.item_code: it.material for it in engine_result.items}
        for ic in sorted(engine_result.coverage_gaps.no_machine):
            _data_cell(ws7, r7, 1, ic, "General", "left")
            _data_cell(ws7, r7, 2, mat_map_c.get(ic, "—"), "General", "center",
                       fill=PatternFill("solid", fgColor="FFF3CD"))
            r7 += 1
    else:
        ws7.cell(row=r7, column=1).value = "None" if engine_result else "(not available)"
        r7 += 1

    r7 += 1

    # 7d: Unfinished items (from schedule)
    _section_title7("7D — Unfinished Items (capacity overrun in scheduler)")
    if schedule_result and schedule_result.unfinished:
        uf_hdrs = [("Item Code", 14), ("Material", 11), ("Rem Hrs", 11),
                   ("Rem kg", 11), ("Capable Machines", 30), ("Origin Wk", 10)]
        for ci, (h, w) in enumerate(uf_hdrs, start=1):
            _hdr_cell(ws7, r7, ci, h, w)
        r7 += 1
        for uf in schedule_result.unfinished:
            vals7d = [
                uf.item_code, uf.material, round(uf.remaining_hours, 1),
                round(uf.remaining_kg, 1), ", ".join(uf.capable_machines),
                f"W{uf.origin_week}" if uf.origin_week else "—",
            ]
            fill7d = PatternFill("solid", fgColor="FFE0E0")
            for ci, v in enumerate(vals7d, start=1):
                _data_cell(ws7, r7, ci, v, "General",
                           "left" if ci in (1, 5) else "center", fill=fill7d)
            r7 += 1
    else:
        ws7.cell(row=r7, column=1).value = (
            "None — all items fit within monthly capacity"
            if schedule_result else "(schedule not available)"
        )
        r7 += 1

    r7 += 1

    # 7e: Idle / locked-out machines
    _section_title7("7E — Idle or Locked-Out Machines")
    if engine_result:
        gaps7 = engine_result.coverage_gaps
        for mc in gaps7.idle_machines:
            _data_cell(ws7, r7, 1, mc, fill=PatternFill("solid", fgColor="FFF3CD"))
            _data_cell(ws7, r7, 2, "Idle (in routing but no load)")
            r7 += 1
        for mc in gaps7.locked_out_machines:
            _data_cell(ws7, r7, 1, mc, fill=PatternFill("solid", fgColor="FFE0E0"))
            _data_cell(ws7, r7, 2, "Locked-out (no routing rows)")
            r7 += 1
        if not gaps7.idle_machines and not gaps7.locked_out_machines:
            ws7.cell(row=r7, column=1).value = "None — all machines have load"
            r7 += 1
    else:
        ws7.cell(row=r7, column=1).value = "(not available)"
        r7 += 1

    # ── Save and return ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# FILE 1 — Revised Production Plan (rejection + waste shown per item)
# ─────────────────────────────────────────────────────────────────────────────

_NAV = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_NAV_FILL = PatternFill("solid", fgColor="1F3864")

def _hdr(ws, row: int, col: int, value: str, width: int = 14) -> None:
    """Write a dark-navy header cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = _NAV
    c.fill = _NAV_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = max(
        ws.column_dimensions[get_column_letter(col)].width or 0, width
    )


def _val(ws, row: int, col: int, value, fmt: str = "General", align: str = "center") -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical="center")


def revised_production_plan_bytes(
    pipe_result,        # mp_engine.EngineResult | None
    fitting_result,     # mp_engine.FittingEngineResult | None
    month: str,
) -> bytes:
    """Return .xlsx bytes for the Revised Production Plan.

    Shows net demand, rejection gross-up, and waste separately per item so
    planners can see the two effects independently.

    Sheets:
      1. Summary    — per-material totals (pipe + fitting separately)
      2. Pipe Items — per routable pipe demand item
      3. Fitting Items — per routable fitting demand item
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False

    # Title
    ws1.merge_cells("A1:J1")
    t = ws1["A1"]
    t.value = f"Revised Production Plan — {month}"
    t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 5

    _hdr(ws1, 3, 1, "Type",          16)
    _hdr(ws1, 3, 2, "Material",      12)
    _hdr(ws1, 3, 3, "Net Pieces",    13)
    _hdr(ws1, 3, 4, "Rej %",         10)
    _hdr(ws1, 3, 5, "Gross Pieces",  13)
    _hdr(ws1, 3, 6, "Extra Pieces",  12)
    _hdr(ws1, 3, 7, "Waste %",       10)
    _hdr(ws1, 3, 8, "Waste Basis",   12)
    _hdr(ws1, 3, 9, "Material kg",   13)
    _hdr(ws1, 3, 10, "Machine Hrs",  12)

    row = 4
    COL_FMT = "#,##0"

    def _summary_rows(result, label: str):
        nonlocal row
        if not result:
            return
        # Aggregate by material
        mat_acc: dict = {}
        for it in result.items:
            if not getattr(it, "has_weight", False):
                continue
            mat = it.material
            if mat not in mat_acc:
                mat_acc[mat] = {"net": 0.0, "gross": 0.0, "kg": 0.0, "hrs": 0.0,
                                "rej_sum": 0.0, "waste_sum": 0.0, "n": 0}
            a = mat_acc[mat]
            a["net"]   += it.qty_pcs
            a["gross"] += getattr(it, "gross_qty_pcs", it.qty_pcs)
            a["kg"]    += it.material_kg
            a["hrs"]   += it.machine_hrs
            a["rej_sum"]   += getattr(it, "rej_rate", 0.0)
            a["waste_sum"] += getattr(it, "waste_pct_used", 0.0)
            a["n"] += 1

        for mat, a in sorted(mat_acc.items()):
            net   = a["net"]
            gross = a["gross"]
            extra = gross - net
            avg_rej   = a["rej_sum"]   / a["n"] if a["n"] else 0
            avg_waste = a["waste_sum"] / a["n"] if a["n"] else 0
            waste_basis = "measured"
            _val(ws1, row, 1, label, align="left")
            _val(ws1, row, 2, mat)
            _val(ws1, row, 3, round(net, 0),   COL_FMT)
            _val(ws1, row, 4, round(avg_rej, 3),  "0.000%")
            _val(ws1, row, 5, round(gross, 0), COL_FMT)
            _val(ws1, row, 6, round(extra, 0), COL_FMT)
            _val(ws1, row, 7, round(avg_waste / 100, 4), "0.000%")
            _val(ws1, row, 8, waste_basis)
            _val(ws1, row, 9, round(a["kg"], 0), COL_FMT)
            _val(ws1, row, 10, round(a["hrs"], 1), "0.0")
            row += 1

    _summary_rows(pipe_result,    "PIPE")
    _summary_rows(fitting_result, "FITTING")

    # ── Sheet 2: Pipe Items ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Pipe Items")
    ws2.sheet_view.showGridLines = False
    PCOLS = [
        ("Item Code", 16, "left"),
        ("Material",  11, "center"),
        ("Net Pcs",   11, "right"),
        ("Rej %",     9,  "center"),
        ("Rej Basis", 10, "center"),
        ("Gross Pcs", 11, "right"),
        ("Extra Pcs", 10, "right"),
        ("Waste %",   9,  "center"),
        ("Waste Basis", 11, "center"),
        ("Material kg", 12, "right"),
        ("Wt/pc (kg)", 11, "right"),
        ("Mach Hrs",  10, "right"),
    ]
    for ci, (name, w, _) in enumerate(PCOLS, 1):
        _hdr(ws2, 1, ci, name, w)
    pr = 2
    if pipe_result:
        for it in sorted(pipe_result.items, key=lambda x: (x.material, x.item_code)):
            if not it.has_weight:
                continue
            gross = getattr(it, "gross_qty_pcs", it.qty_pcs)
            rej_r = getattr(it, "rej_rate", 0.0)
            rej_b = getattr(it, "rej_basis", "")
            wst_r = getattr(it, "waste_pct_used", 0.0)
            wst_b = getattr(it, "waste_basis", "")
            row_data = [
                it.item_code, it.material,
                round(it.qty_pcs, 0), round(rej_r / 100, 4), rej_b,
                round(gross, 0), round(gross - it.qty_pcs, 0),
                round(wst_r / 100, 4), wst_b,
                round(it.material_kg, 1),
                round(it.weight_per_pc_kg, 4) if it.weight_per_pc_kg else "",
                round(it.machine_hrs, 2),
            ]
            fmts = ["", "", COL_FMT, "0.00%", "", COL_FMT, COL_FMT, "0.00%", "", "0.0", "0.0000", "0.00"]
            for ci, (v, fmt, (_, _, al)) in enumerate(zip(row_data, fmts, PCOLS), 1):
                c = ws2.cell(row=pr, column=ci, value=v)
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(horizontal=al, vertical="center")
            pr += 1

    # ── Sheet 3: Fitting Items ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Fitting Items")
    ws3.sheet_view.showGridLines = False
    for ci, (name, w, _) in enumerate(PCOLS, 1):
        _hdr(ws3, 1, ci, name, w)
    fr = 2
    if fitting_result:
        for it in sorted(fitting_result.items, key=lambda x: (x.material, x.item_code)):
            if not it.has_weight:
                continue
            gross = getattr(it, "gross_qty_pcs", it.qty_pcs)
            rej_r = getattr(it, "rej_rate", 0.0)
            rej_b = getattr(it, "rej_basis", "")
            wst_r = getattr(it, "waste_pct_used", 0.0)
            wst_b = getattr(it, "waste_basis", "")
            row_data = [
                it.item_code, it.material,
                round(it.qty_pcs, 0), round(rej_r / 100, 4), rej_b,
                round(gross, 0), round(gross - it.qty_pcs, 0),
                round(wst_r / 100, 4), wst_b,
                round(it.material_kg, 1),
                round(it.weight_per_pc_kg, 4) if it.weight_per_pc_kg else "",
                round(it.machine_hrs, 2),
            ]
            fmts = ["", "", COL_FMT, "0.00%", "", COL_FMT, COL_FMT, "0.00%", "", "0.0", "0.0000", "0.00"]
            for ci, (v, fmt, (_, _, al)) in enumerate(zip(row_data, fmts, PCOLS), 1):
                c = ws3.cell(row=fr, column=ci, value=v)
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(horizontal=al, vertical="center")
            fr += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# FILE 2 — Machine Plan Comparison (without rej/waste vs with both)
# ─────────────────────────────────────────────────────────────────────────────

def machine_plan_comparison_bytes(
    result_with,     # EngineResult — with rejection + measured waste
    result_flat,     # EngineResult — without rejection, flat waste (old behaviour)
    fitting_with,    # FittingEngineResult | None — with
    fitting_flat,    # FittingEngineResult | None — without
    month: str,
    old_waste_pct: float = 0.0,   # the flat waste % used in result_flat
) -> bytes:
    """Return .xlsx bytes for the Machine Plan Comparison.

    Shows WITHOUT (old flat waste, no rejection) vs WITH (rejection + measured
    waste) side-by-side so planners can see the impact of the correction.

    Sheets:
      1. Summary Comparison  — headline totals + per-machine hours
      2. Pipe Machine Loads  — per-machine, without vs with
      3. Pipe Item Detail    — per item: rejection and waste separately
    """
    wb = Workbook()
    wb.remove(wb.active)

    _W_FILL  = PatternFill("solid", fgColor="E2EFDA")  # green-ish "without" baseline
    _WI_FILL = PatternFill("solid", fgColor="1F3864")  # navy header
    _D_FILL  = PatternFill("solid", fgColor="FFF2CC")  # amber delta

    def _title(ws, text: str) -> None:
        ws.merge_cells("A1:K1")
        c = ws["A1"]
        c.value = text
        c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 4

    def _hdr2(ws, row, col, txt, w=14):
        c = ws.cell(row=row, column=col, value=txt)
        c.font = _NAV
        c.fill = _NAV_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 0, w
        )

    def _safe_total_hrs(res) -> float:
        if not res:
            return 0.0
        return sum(it.machine_hrs for it in res.items if getattr(it, "has_weight", False))

    def _safe_mat_kg(res) -> float:
        if not res:
            return 0.0
        return sum(it.material_kg for it in res.items if getattr(it, "has_weight", False))

    def _safe_peak(res):
        if not res or not res.machine_loads:
            return ("—", 0.0)
        top = max(res.machine_loads, key=lambda ml: ml.assigned_hrs)
        return (top.machine, top.assigned_hrs)

    def _fleet_util(res) -> float:
        if not res or not res.machine_loads:
            return 0.0
        cap = sum(ml.capacity_hrs for ml in res.machine_loads if ml.capacity_hrs > 0)
        load = sum(ml.assigned_hrs for ml in res.machine_loads)
        return load / cap * 100 if cap > 0 else 0.0

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary Comparison")
    ws1.sheet_view.showGridLines = False
    _title(ws1, f"Machine Plan Comparison — {month}")

    note = (
        f"WITHOUT = net demand × weight × (1 + {old_waste_pct:.1f}% flat waste), no rejection gross-up\n"
        "WITH    = gross_qty = net / (1 - rejection%), material_kg = gross × wt × (1 + measured waste%)"
    )
    ws1.merge_cells("A3:K3")
    c3 = ws1["A3"]
    c3.value = note
    c3.font = Font(name="Calibri", italic=True, size=9, color="555555")
    c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws1.row_dimensions[3].height = 28

    COLS_S = ["Category", "WITHOUT", "WITH", "Delta (±)", "Delta (%)"]
    for ci, h in enumerate(COLS_S, 1):
        _hdr2(ws1, 4, ci, h, 18)

    # Pipe totals
    with_pipe_kg   = _safe_mat_kg(result_with)
    flat_pipe_kg   = _safe_mat_kg(result_flat)
    with_pipe_hrs  = _safe_total_hrs(result_with)
    flat_pipe_hrs  = _safe_total_hrs(result_flat)
    with_peak_mc,  with_peak_h  = _safe_peak(result_with)
    flat_peak_mc,  flat_peak_h  = _safe_peak(result_flat)
    with_fleet = _fleet_util(result_with)
    flat_fleet = _fleet_util(result_flat)

    # Fitting totals
    with_fit_kg  = _safe_mat_kg(fitting_with)
    flat_fit_kg  = _safe_mat_kg(fitting_flat)
    with_fit_hrs = _safe_total_hrs(fitting_with)
    flat_fit_hrs = _safe_total_hrs(fitting_flat)

    S_ROWS = [
        ("Pipe — Material kg",  flat_pipe_kg,  with_pipe_kg),
        ("Pipe — Machine Hrs",  flat_pipe_hrs, with_pipe_hrs),
        ("Pipe — Fleet Util %", flat_fleet,    with_fleet),
        ("Pipe — Peak Hrs",     flat_peak_h,   with_peak_h),
        ("Fitting — Material kg", flat_fit_kg, with_fit_kg),
        ("Fitting — Machine Hrs", flat_fit_hrs, with_fit_hrs),
    ]
    fill_even = PatternFill("solid", fgColor="F5F5F5")
    r = 5
    for i, (cat, wout, wi) in enumerate(S_ROWS):
        delta = wi - wout
        dpct  = (delta / wout * 100) if wout else 0.0
        row_fill = fill_even if i % 2 == 0 else None
        for ci, v in enumerate([cat, round(wout, 1), round(wi, 1),
                                  round(delta, 1), f"{dpct:+.1f}%"], 1):
            c = ws1.cell(row=r, column=ci, value=v)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "right", vertical="center")
            if row_fill:
                c.fill = row_fill
        r += 1

    # Peak machine row
    ws1.cell(row=r, column=1).value = "Pipe — Peak Machine"
    ws1.cell(row=r, column=2).value = flat_peak_mc
    ws1.cell(row=r, column=3).value = with_peak_mc
    for ci in range(1, 4):
        ws1.cell(row=r, column=ci).alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
    r += 1

    # ── Sheet 2: Machine Load ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Pipe Machine Loads")
    ws2.sheet_view.showGridLines = False
    _title(ws2, f"Pipe Machine Load Comparison — {month}")

    ML_COLS = ["Machine", "Without Hrs", "With Hrs", "Extra Hrs", "Capacity Hrs", "Util % (With)"]
    for ci, h in enumerate(ML_COLS, 1):
        _hdr2(ws2, 3, ci, h, 15)

    flat_ml = {ml.machine: ml for ml in (result_flat.machine_loads if result_flat else [])}
    with_ml = {ml.machine: ml for ml in (result_with.machine_loads if result_with else [])}
    all_mcs = sorted(set(flat_ml) | set(with_ml))

    r2 = 4
    for mc in all_mcs:
        f_hrs = flat_ml[mc].assigned_hrs if mc in flat_ml else 0.0
        w_hrs = with_ml[mc].assigned_hrs if mc in with_ml else 0.0
        cap   = (with_ml[mc].capacity_hrs if mc in with_ml else
                 flat_ml[mc].capacity_hrs if mc in flat_ml else 0.0)
        util  = w_hrs / cap * 100 if cap > 0 else 0.0
        for ci, v in enumerate([mc, round(f_hrs, 1), round(w_hrs, 1),
                                  round(w_hrs - f_hrs, 1), round(cap, 1),
                                  f"{util:.1f}%"], 1):
            c = ws2.cell(row=r2, column=ci, value=v)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "right", vertical="center")
        r2 += 1

    # ── Sheet 3: Pipe Item Detail ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Pipe Item Detail")
    ws3.sheet_view.showGridLines = False
    _title(ws3, f"Pipe Item Detail — Rejection + Waste — {month}")

    ID_COLS = [
        ("Item Code", 16, "left"),
        ("Material",  11, "center"),
        ("Net Pcs",   11, "right"),
        ("Rej %",     9,  "center"),
        ("Rej Basis", 10, "center"),
        ("Gross Pcs", 11, "right"),
        ("Waste % (meas)", 13, "center"),
        ("Waste Basis", 12, "center"),
        ("Material kg (flat)", 16, "right"),
        ("Material kg (with)", 16, "right"),
        ("Delta kg",  11, "right"),
        ("Mach Hrs (flat)", 14, "right"),
        ("Mach Hrs (with)", 14, "right"),
    ]
    for ci, (h, w, _) in enumerate(ID_COLS, 1):
        _hdr2(ws3, 3, ci, h, w)

    flat_items = {it.item_code: it for it in (result_flat.items if result_flat else [])}
    r3 = 4
    if result_with:
        for it in sorted(result_with.items, key=lambda x: (x.material, x.item_code)):
            if not it.has_weight:
                continue
            flat_it = flat_items.get(it.item_code)
            gross = getattr(it, "gross_qty_pcs", it.qty_pcs)
            rej_r = getattr(it, "rej_rate", 0.0)
            rej_b = getattr(it, "rej_basis", "")
            wst_r = getattr(it, "waste_pct_used", 0.0)
            wst_b = getattr(it, "waste_basis", "")
            flat_kg  = flat_it.material_kg if flat_it else 0.0
            flat_hrs = flat_it.machine_hrs if flat_it else 0.0
            vals = [
                it.item_code, it.material,
                round(it.qty_pcs, 0), round(rej_r / 100, 4), rej_b,
                round(gross, 0), round(wst_r / 100, 4), wst_b,
                round(flat_kg, 1), round(it.material_kg, 1),
                round(it.material_kg - flat_kg, 1),
                round(flat_hrs, 2), round(it.machine_hrs, 2),
            ]
            fmts = ["", "", "#,##0", "0.00%", "", "#,##0", "0.00%", "",
                    "0.0", "0.0", "0.0", "0.00", "0.00"]
            for ci, (v, fmt, (_, _, al)) in enumerate(zip(vals, fmts, ID_COLS), 1):
                c = ws3.cell(row=r3, column=ci, value=v)
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(horizontal=al, vertical="center")
            r3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# FILE 3 — Capacity-Feasible Production Plan
# ─────────────────────────────────────────────────────────────────────────────

# Named constant that encodes the scheduler's item-priority order.
# Change this string if the scheduler trim rule is ever swapped.
FEASIBLE_TRIM_RULE = "earliest-week-first / largest-remaining-hrs-first (LPT)"


def capacity_feasible_plan_bytes(
    pipe_result,      # mp_engine.EngineResult | None
    fitting_result,   # mp_engine.FittingEngineResult | None
    schedule,         # mp_scheduler.ScheduleResult | None  (pipe only)
    month: str,
) -> bytes:
    """Return .xlsx bytes for the Capacity-Feasible Production Plan.

    The scheduler (run_shift_schedule) enforces capacity day-by-day.  Items
    that run out of working days land in schedule.unfinished with their
    remaining machine-hours.  This report converts that to pieces/kg so the
    planner gets the maximum achievable this month plus a clear shortfall list.

    Tabs
    ----
    1. Summary                 — requested / feasible / shortfall totals by material
    2. Pipe Plan               — per-item split
    3. Fitting Plan            — per-item (all feasible — no fitting scheduler yet)
    4. Machine Load (Feasible) — scheduler hours, always ≤ 100%
    5. Shortfall               — unmet demand that cannot be made this month
    """
    wb = Workbook()
    wb.remove(wb.active)

    _NAVY  = "1F3864"
    _NFONT = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    _NFILL = PatternFill("solid", fgColor=_NAVY)
    _AFILL = PatternFill("solid", fgColor="FFF2CC")   # amber — shortfall
    _GFILL = PatternFill("solid", fgColor="E2EFDA")   # green — ok
    _RFILL = PatternFill("solid", fgColor="FCE4D6")   # red   — over (should never happen)
    _EFILL = PatternFill("solid", fgColor="F5F5F5")   # even row stripe

    def _title(ws, text: str) -> None:
        ws.merge_cells("A1:K1")
        c = ws["A1"]
        c.value = text
        c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        c.fill = _NFILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 4

    def _hdr(ws, row: int, col: int, txt: str, w: int = 14) -> None:
        c = ws.cell(row=row, column=col, value=txt)
        c.font = _NFONT
        c.fill = _NFILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 0, w
        )

    def _note(ws, row: int, text: str, ncols: int = 11) -> None:
        end = get_column_letter(ncols)
        ws.merge_cells(f"A{row}:{end}{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = Font(name="Calibri", italic=True, size=9, color="555555")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 24

    # ── Build unfinished lookup: item_code → UnfinishedItem ──────────────────
    unfinished_by_code: dict = {}
    if schedule:
        for u in schedule.unfinished:
            unfinished_by_code[u.item_code] = u

    # ── Build per-machine scheduled load from weekly_fill ────────────────────
    # Scheduled hours = actual production hours placed within the working days.
    # These are ALWAYS ≤ machine capacity (the scheduler's hard constraint).
    sched_by_mc: dict = {}
    if schedule:
        for wf in schedule.weekly_fill:
            mc = wf.machine
            if mc not in sched_by_mc:
                sched_by_mc[mc] = {"scheduled_hrs": 0.0, "capacity_hrs": 0.0}
            # weekly_fill rows are per-week — sum both sides to get monthly totals
            sched_by_mc[mc]["scheduled_hrs"] += wf.scheduled_hrs
            sched_by_mc[mc]["capacity_hrs"]  += wf.capacity_hrs

    # ── Per-item feasible / shortfall split ──────────────────────────────────
    def _split(item):
        """Return (feasible_pcs, shortfall_pcs, feasible_kg, shortfall_kg).

        Feasible = what the scheduler placed within working days this month.
        Shortfall = unmet demand that cannot be made this month (NOT deferred to future).
        feasible + shortfall = requested (gross_qty_pcs / material_kg) exactly.
        """
        u = unfinished_by_code.get(item.item_code)
        if u is None:
            return item.gross_qty_pcs, 0.0, item.material_kg, 0.0
        shortfall_kg  = min(float(u.remaining_kg), float(item.material_kg))
        feasible_kg   = float(item.material_kg) - shortfall_kg
        # Pro-rate pieces proportional to kg so feasible + shortfall == requested exactly.
        if item.material_kg > 0:
            ratio = feasible_kg / float(item.material_kg)
            feasible_pcs = round(float(item.gross_qty_pcs) * ratio, 2)
        else:
            feasible_pcs = 0.0
        shortfall_pcs = round(float(item.gross_qty_pcs) - feasible_pcs, 2)
        return feasible_pcs, shortfall_pcs, round(feasible_kg, 2), round(shortfall_kg, 2)

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ═══════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False
    _title(ws1, f"Capacity-Feasible Production Plan — {month}")
    _note(ws1, 3,
          f"Trim rule: {FEASIBLE_TRIM_RULE}. "
          "Shortfall = unmet demand that cannot be achieved within this month's working days. "
          "It is NOT rolled to a future month — the planner acts on it.")

    S_HDRS = ["Type", "Material", "Requested pcs", "Feasible pcs", "Shortfall pcs",
              "Shortfall %", "Requested kg", "Feasible kg", "Shortfall kg"]
    for ci, h in enumerate(S_HDRS, 1):
        _hdr(ws1, 4, ci, h, 14)

    sr = 5

    def _summary_pipe(res, label: str) -> None:
        nonlocal sr
        if not res:
            return
        mat_acc: dict = {}
        for it in res.items:
            if not getattr(it, "has_weight", False):
                continue
            mat = it.material
            if mat not in mat_acc:
                mat_acc[mat] = dict(req_pcs=0.0, feas_pcs=0.0, sf_pcs=0.0,
                                    req_kg=0.0,  feas_kg=0.0,  sf_kg=0.0)
            fp, sp, fk, sk = _split(it)
            a = mat_acc[mat]
            a["req_pcs"]  += float(it.gross_qty_pcs)
            a["feas_pcs"] += fp
            a["sf_pcs"]   += sp
            a["req_kg"]   += float(it.material_kg)
            a["feas_kg"]  += fk
            a["sf_kg"]    += sk
        for i, (mat, a) in enumerate(sorted(mat_acc.items())):
            req = a["req_pcs"]
            sf_pct = a["sf_pcs"] / req * 100 if req > 0 else 0.0
            rf = _EFILL if i % 2 == 0 else None
            vals = [label, mat,
                    round(a["req_pcs"]), round(a["feas_pcs"]), round(a["sf_pcs"]),
                    f"{sf_pct:.1f}%",
                    round(a["req_kg"], 1), round(a["feas_kg"], 1), round(a["sf_kg"], 1)]
            for ci, v in enumerate(vals, 1):
                c = ws1.cell(row=sr, column=ci, value=v)
                c.alignment = Alignment(horizontal="left" if ci <= 2 else "right",
                                        vertical="center")
                if rf:
                    c.fill = rf
                if a["sf_pcs"] > 0 and ci in (5, 6):
                    c.fill = _AFILL
            sr += 1

    def _summary_fitting(res, label: str) -> None:
        nonlocal sr
        if not res:
            return
        mat_acc: dict = {}
        for it in res.items:
            if not getattr(it, "has_weight", False):
                continue
            mat = it.material
            if mat not in mat_acc:
                mat_acc[mat] = dict(req_pcs=0.0, req_kg=0.0)
            mat_acc[mat]["req_pcs"] += float(getattr(it, "gross_qty_pcs", it.qty_pcs))
            mat_acc[mat]["req_kg"]  += float(it.material_kg)
        for i, (mat, a) in enumerate(sorted(mat_acc.items())):
            rf = _EFILL if i % 2 == 0 else None
            vals = [label, mat,
                    round(a["req_pcs"]), round(a["req_pcs"]), 0,
                    "0.0%",
                    round(a["req_kg"], 1), round(a["req_kg"], 1), 0.0]
            for ci, v in enumerate(vals, 1):
                c = ws1.cell(row=sr, column=ci, value=v)
                c.alignment = Alignment(horizontal="left" if ci <= 2 else "right",
                                        vertical="center")
                if rf:
                    c.fill = rf
            sr += 1

    _summary_pipe(pipe_result, "PIPE")
    _summary_fitting(fitting_result, "FITTING")

    if sr == 5:
        ws1.cell(row=5, column=1).value = "No routable items found."

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 2 — Pipe Plan
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Pipe Plan")
    ws2.sheet_view.showGridLines = False
    _title(ws2, f"Capacity-Feasible Pipe Plan — {month}")

    P_HDRS = [
        ("Item Code",      16, "left"),
        ("Material",       11, "center"),
        ("Requested pcs",  13, "right"),
        ("Feasible pcs",   13, "right"),
        ("Shortfall pcs",  13, "right"),
        ("Requested kg",   13, "right"),
        ("Feasible kg",    12, "right"),
        ("Shortfall kg",   12, "right"),
        ("Machine(s)",     14, "center"),
        ("Note",           35, "left"),
    ]
    for ci, (h, w, _) in enumerate(P_HDRS, 1):
        _hdr(ws2, 3, ci, h, w)

    pr = 4
    if pipe_result:
        for it in sorted(pipe_result.items, key=lambda x: (x.material, x.item_code)):
            if not getattr(it, "has_weight", False):
                continue
            fp, dp, fk, dk = _split(it)
            u = unfinished_by_code.get(it.item_code)
            mcs = ", ".join(it.capable_machines) if it.capable_machines else "—"
            note = ""
            if u:
                note = u.downtime_reason or (
                    f"{it.material} — only {mcs} routed, capacity exhausted"
                )
            row_data = [
                it.item_code, it.material,
                round(float(it.gross_qty_pcs)), round(fp), round(dp),
                round(float(it.material_kg), 1), round(fk, 1), round(dk, 1),
                mcs, note,
            ]
            fmts = ["", "", "#,##0", "#,##0", "#,##0", "0.0", "0.0", "0.0", "", ""]
            for ci, (v, fmt, (_, _, al)) in enumerate(zip(row_data, fmts, P_HDRS), 1):
                c = ws2.cell(row=pr, column=ci, value=v)
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(horizontal=al, vertical="center")
                if dp > 0 and ci in (5, 8):
                    c.fill = _AFILL
            pr += 1

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 3 — Fitting Plan
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Fitting Plan")
    ws3.sheet_view.showGridLines = False
    _title(ws3, f"Capacity-Feasible Fitting Plan — {month}")
    _note(ws3, 3,
          "Fitting scheduler not yet implemented — all fitting demand is treated as "
          "feasible.  Capacity constraint is applied to pipe machines only.")

    FH = [
        ("Item Code",     16, "left"),
        ("Material",      11, "center"),
        ("Requested pcs", 13, "right"),
        ("Feasible pcs",  13, "right"),
        ("Requested kg",  13, "right"),
        ("Feasible kg",   12, "right"),
    ]
    for ci, (h, w, _) in enumerate(FH, 1):
        _hdr(ws3, 4, ci, h, w)

    fr3 = 5
    if fitting_result:
        for it in sorted(fitting_result.items, key=lambda x: (x.material, x.item_code)):
            if not getattr(it, "has_weight", False):
                continue
            gross = float(getattr(it, "gross_qty_pcs", it.qty_pcs))
            row_data = [
                it.item_code, it.material,
                round(gross), round(gross),
                round(float(it.material_kg), 1), round(float(it.material_kg), 1),
            ]
            fmts = ["", "", "#,##0", "#,##0", "0.0", "0.0"]
            for ci, (v, fmt, (_, _, al)) in enumerate(zip(row_data, fmts, FH), 1):
                c = ws3.cell(row=fr3, column=ci, value=v)
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(horizontal=al, vertical="center")
            fr3 += 1
    else:
        ws3.cell(row=5, column=1).value = "No fitting demand in this plan."

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 4 — Machine Load (Feasible)
    # ═══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Machine Load (Feasible)")
    ws4.sheet_view.showGridLines = False
    _title(ws4, f"Scheduler Machine Load (Capacity-Enforced) — {month}")
    _note(ws4, 3,
          "Scheduled hours = actual production hours placed by the day-by-day scheduler "
          "within working days.  These are always ≤ machine capacity (hard constraint).", 6)

    ML_HDRS = ["Machine", "Capacity hrs", "Scheduled hrs", "Util %",
               "Headroom hrs", "Status"]
    for ci, h in enumerate(ML_HDRS, 1):
        _hdr(ws4, 4, ci, h, 14)

    mr = 5
    over_capacity = False
    if sched_by_mc:
        for mc, d in sorted(sched_by_mc.items()):
            cap   = float(d["capacity_hrs"])
            sched = round(float(d["scheduled_hrs"]), 1)
            util  = round(sched / cap * 100, 1) if cap > 0 else 0.0
            head  = round(cap - sched, 1)
            if util > 100.0:
                over_capacity = True
            status   = "✓ OK" if util <= 100.0 else "⚠ OVER"
            row_fill = _GFILL if util <= 100.0 else _RFILL
            vals = [mc, round(cap, 1), sched, f"{util:.1f}%", head, status]
            for ci, v in enumerate(vals, 1):
                c = ws4.cell(row=mr, column=ci, value=v)
                c.alignment = Alignment(
                    horizontal="left" if ci == 1 else "right", vertical="center")
                c.fill = row_fill
            mr += 1
    else:
        c = ws4.cell(row=mr, column=1,
                     value="No scheduler data — upload a demand file to run the shift scheduler.")
        c.font = Font(name="Calibri", italic=True, size=10, color="6B7280")

    # Assert the scheduler invariant: no machine should exceed capacity.
    assert not over_capacity, (
        "Scheduler produced utilisation > 100% for at least one machine, "
        "violating the capacity constraint.  Machine data: " + str(sched_by_mc)
    )

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 5 — Shortfall (unmet demand this month)
    # ═══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Shortfall")
    ws5.sheet_view.showGridLines = False
    _title(ws5, f"Shortfall — Unmet Demand This Month — {month}")

    DEF_HDRS = [
        ("Item Code",       16, "left"),
        ("Material",        11, "center"),
        ("Requested pcs",   13, "right"),
        ("Feasible pcs",    13, "right"),
        ("Shortfall pcs",   13, "right"),
        ("Shortfall kg",    12, "right"),
        ("Shortfall hrs",   12, "right"),
        ("Capable Machines", 20, "left"),
        ("Reason",          38, "left"),
    ]
    for ci, (h, w, _) in enumerate(DEF_HDRS, 1):
        _hdr(ws5, 3, ci, h, w)

    dr = 4
    if pipe_result:
        shortfall_items = [
            it for it in pipe_result.items
            if getattr(it, "has_weight", False) and it.item_code in unfinished_by_code
        ]
        shortfall_items.sort(key=lambda x: (
            -unfinished_by_code[x.item_code].remaining_kg, x.material, x.item_code
        ))
        for it in shortfall_items:
            u    = unfinished_by_code[it.item_code]
            fp, sp, fk, sk = _split(it)
            mcs  = ", ".join(it.capable_machines) if it.capable_machines else "—"
            note = u.downtime_reason or (
                f"{it.material} — only {mcs} routed, capacity insufficient this month"
            )
            row_data = [
                it.item_code, it.material,
                round(float(it.gross_qty_pcs)), round(fp), round(sp),
                round(sk, 1), round(float(u.remaining_hours), 2),
                mcs, note,
            ]
            fmts = ["", "", "#,##0", "#,##0", "#,##0", "0.0", "0.00", "", ""]
            for ci, (v, fmt, (_, _, al)) in enumerate(zip(row_data, fmts, DEF_HDRS), 1):
                c = ws5.cell(row=dr, column=ci, value=v)
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(horizontal=al, vertical="center")
                c.fill = _AFILL
            dr += 1

    if dr == 4:
        c = ws5.cell(row=4, column=1,
                     value="No shortfall — all demand fits within this month's capacity.")
        c.font = Font(name="Calibri", italic=True, size=10, color="065f46")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

"""
Machine Planning Follow-Up Engine.

Ingests actuals from the PIPE daily workbook (Report-11 for pipe,
Report-12 for moulding/fitting), joins them to the persisted plan lines
(mp_plan_line), and computes:
  - variance metrics per machine and per item (plan-to-date vs actual)
  - RAG status (GREEN / AMBER / RED)
  - 7-type deviation warnings, severity-ranked

ADDITIVE / ISOLATED: reads only mp_* tables + daily workbooks.
Never touches the headline production pipeline.
"""
from __future__ import annotations

import calendar
import dataclasses
import datetime
import re
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

from mp_seed import norm_code as _norm_code  # existing normaliser


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

def norm_machine(name: str) -> str:
    """Normalise machine identifier for join purposes.

    'PIPE M/C - 1'  → 'MC1'
    'M/C-1'         → 'MC1'
    'Moulding MC-3' → 'MC3'
    'M/C - 11'      → 'MC11'
    """
    s = str(name).strip().upper()
    s = re.sub(r'^PIPE\s+', '', s)
    s = re.sub(r'^MOULDING\s+', '', s)
    s = re.sub(r'[\s\-/]', '', s)
    return s


def norm_item(code: str) -> str:
    """Wrapper over the shared code normaliser."""
    return _norm_code(str(code))


# ---------------------------------------------------------------------------
# RAG classification
# ---------------------------------------------------------------------------

def rag_status(
    actual: float,
    plan: float,
    amber_pct: float = 10.0,
    red_pct: float = 25.0,
) -> str:
    """Classify adherence into GREEN / AMBER / RED.

    Based on absolute % deviation from plan-to-date.
    Returns 'GREEN' when plan == 0 and actual == 0 (no work expected or done).
    Returns 'RED'   when plan == 0 but actual > 0 (unplanned production).
    """
    if plan <= 0:
        return "RED" if actual > 0 else "GREEN"
    deviation = abs(actual - plan) / plan * 100.0
    if deviation < amber_pct:
        return "GREEN"
    if deviation < red_pct:
        return "AMBER"
    return "RED"


# ---------------------------------------------------------------------------
# Dataclasses for the variance result
# ---------------------------------------------------------------------------

@dataclasses.dataclass
class ItemVariance:
    machine: str
    machine_norm: str
    item_code: str
    item_norm: str
    material: str
    # Full-month plan (from all plan lines)
    planned_pcs_total: float = 0.0
    planned_kg_total: float = 0.0
    planned_hours_total: float = 0.0
    # Plan-to-date (up to elapsed_plan_days)
    planned_pcs_todate: float = 0.0
    planned_kg_todate: float = 0.0
    planned_hours_todate: float = 0.0
    # Actuals
    actual_pcs: float = 0.0
    actual_kg: float = 0.0
    actual_hours: float = 0.0
    rejection_kg: float = 0.0
    # Derived
    kg_var_pct: float = 0.0          # (actual - plan_todate) / plan_todate * 100
    hours_var_pct: float = 0.0
    adherence_pct: float = 0.0       # actual_kg / planned_kg_todate * 100
    projected_kg_month: float = 0.0  # actual_kg / elapsed_days * working_days
    rag: str = "GREEN"
    is_wrong_machine: bool = False   # produced on a machine not in its routing
    is_unplanned: bool = False       # item not in plan at all


@dataclasses.dataclass
class MachineVariance:
    machine: str
    machine_norm: str
    planned_hours_todate: float = 0.0
    planned_kg_todate: float = 0.0
    actual_hours: float = 0.0
    actual_kg: float = 0.0
    hours_var_pct: float = 0.0
    kg_var_pct: float = 0.0
    adherence_pct: float = 0.0
    projected_kg_month: float = 0.0
    rag: str = "GREEN"
    had_planned_work: bool = False
    had_actual_work: bool = False


@dataclasses.dataclass
class WeekPoint:
    """Cumulative plan vs actual kg at end of each week."""
    week: int
    planned_kg_cumulative: float = 0.0
    actual_kg_cumulative: float = 0.0


@dataclasses.dataclass
class Warning:
    warning_type: str  # see WTYPE_* constants below
    severity: int      # 1=critical, 2=high, 3=medium, 4=low
    machine: str
    item_code: str
    material: str
    magnitude: float   # % deviation or hours, 0 when not applicable
    reason: str


# Warning type constants (used in templates too)
WTYPE_WRONG_MACHINE   = "WRONG_MACHINE"
WTYPE_UNPLANNED       = "UNPLANNED"
WTYPE_NOT_STARTED     = "NOT_STARTED"
WTYPE_QTY_OVER        = "QTY_OVERRUN"
WTYPE_QTY_SHORT       = "QTY_SHORTFALL"
WTYPE_HOURS_DEV       = "HOURS_DEVIATION"
WTYPE_IDLE_VS_PLAN    = "IDLE_VS_PLAN"
WTYPE_NIGHT_CHANGEOVER = "NIGHT_CHANGEOVER"
WTYPE_SHORT_BLOCK     = "SHORT_BLOCK"


@dataclasses.dataclass
class FollowUpResult:
    plan_run_id: int
    month: str
    segment: str
    as_of_date: str          # ISO date of latest actual
    elapsed_calendar_days: int
    days_in_month: int
    working_days: int
    elapsed_plan_days: int
    total_planned_kg_todate: float
    total_actual_kg: float
    overall_adherence_pct: float
    on_plan_kg: float        # actual kg within GREEN threshold
    off_plan_kg: float       # actual kg outside GREEN threshold
    item_rows: List[ItemVariance]
    machine_rows: List[MachineVariance]
    week_points: List[WeekPoint]
    warnings: List[Warning]
    actuals_count: int
    actuals_loaded: bool     # False when no actuals in DB


# ---------------------------------------------------------------------------
# Report-11 parser (PIPE actuals)
# Header row: scan for a row containing 'ITEM CODE' and 'MACHINE'
# Verified columns (0-indexed): B=1 DATE, D=3 MACHINE, E=4 TYPE,
#   F=5 ITEM CODE, G=6 RUN HRS, I=8 PCS, J=9 WEIGHT, N=13 REJECTION
# ---------------------------------------------------------------------------

def _find_header(values: list, *keywords: str) -> Optional[int]:
    """Return the 0-indexed row number of the first row that contains
    at least one of the given keywords (case-insensitive)."""
    kw_upper = [k.upper() for k in keywords]
    for i, row in enumerate(values):
        row_text = " ".join(str(c).upper() for c in row)
        if any(kw in row_text for kw in kw_upper):
            return i
    return None


def _col_index(header_row: list, *candidates: str) -> int:
    """Return the 0-indexed column whose header matches one of candidates."""
    for i, cell in enumerate(header_row):
        c = str(cell).strip().upper()
        for cand in candidates:
            if cand.upper() in c:
                return i
    return -1


def parse_report11(values: list, month: str) -> List[dict]:
    """Parse Report-11 ('M/C & Item Wise Actual Production').

    Returns a list of dicts with keys:
      date, machine, material, item_code, run_hours, pcs, weight_kg, rejection_kg
    """
    if not values:
        return []

    hdr_idx = _find_header(values, "ITEM CODE", "ITEM  CODE")
    if hdr_idx is None:
        hdr_idx = 4  # default: row 5 (1-indexed)

    hdr = values[hdr_idx]
    col_date   = _col_index(hdr, "DATE")
    col_mc     = _col_index(hdr, "MACHINE NO", "MACHINE")
    col_type   = _col_index(hdr, "TYPE", "TYPES")
    col_item   = _col_index(hdr, "ITEM CODE", "ITEM  CODE")
    col_hrs    = _col_index(hdr, "RUNNING HOUR", "RUN HOUR", "RUN HRS")
    col_pcs    = _col_index(hdr, "ACTUAL OUTPUT", "OUTPUT/PCS", "PCS")
    col_wt     = _col_index(hdr, "WEIGHT", "WT IN KG", "WGT")
    col_rej    = _col_index(hdr, "REJECTION", "REJ")

    # Fallback to spec positions if header-based detection failed
    if col_date < 0:  col_date  = 1
    if col_mc   < 0:  col_mc    = 3
    if col_type < 0:  col_type  = 4
    if col_item < 0:  col_item  = 5
    if col_hrs  < 0:  col_hrs   = 6
    if col_pcs  < 0:  col_pcs   = 8
    if col_wt   < 0:  col_wt    = 9
    if col_rej  < 0:  col_rej   = 13

    # For weight col: prefer J (index 9) specifically — "Weight" may appear
    # in the rejection column header too, so take the leftmost 'WEIGHT' col
    # that is strictly after col_pcs.
    if col_wt <= col_pcs:
        col_wt = col_pcs + 1

    year, mnum = int(month[:4]), int(month[5:7])
    rows = []
    last_date = None
    last_mc   = None

    for raw in values[hdr_idx + 1:]:
        if len(raw) < 4:
            continue
        def _cell(idx: int, default=""):
            return raw[idx].strip() if idx < len(raw) and str(raw[idx]).strip() else default

        # Date column — carry forward if blank
        date_raw = _cell(col_date)
        if date_raw:
            d = _parse_date(date_raw, year, mnum)
            if d:
                last_date = d
        if last_date is None:
            continue

        # Machine — carry forward if blank
        mc_raw = _cell(col_mc)
        if mc_raw and not mc_raw.upper().startswith("TOTAL"):
            last_mc = mc_raw
        if last_mc is None:
            continue

        item_raw = _cell(col_item)
        if not item_raw or item_raw.upper() in ("TOTAL", "ITEM CODE", ""):
            continue
        # Skip pure-numeric cells (serial numbers)
        if re.match(r'^\d+$', item_raw):
            continue

        try:
            pcs    = _to_float(_cell(col_pcs))
            wt     = _to_float(_cell(col_wt))
            hrs    = _to_float(_cell(col_hrs))
            rej    = _to_float(_cell(col_rej))
            mtype  = _cell(col_type)
        except Exception:
            continue

        if pcs <= 0 and wt <= 0:
            continue

        rows.append({
            "date":         last_date,
            "machine":      last_mc,
            "material":     mtype,
            "item_code":    item_raw,
            "run_hours":    hrs,
            "pcs":          pcs,
            "weight_kg":    wt,
            "rejection_kg": rej,
        })
    return rows


def parse_report12(values: list, month: str) -> List[dict]:
    """Parse Report-12 (Moulding / Fitting actual production).

    Returns list of dicts with same keys as parse_report11.
    """
    if not values:
        return []

    hdr_idx = _find_header(values, "ITEM CODE", "MOULDING MACHI")
    if hdr_idx is None:
        hdr_idx = 5  # default: row 6 (1-indexed)

    hdr = values[hdr_idx]
    col_date  = _col_index(hdr, "DATE")
    col_mat   = _col_index(hdr, "MATERIAL")
    col_item  = _col_index(hdr, "ITEM CODE")
    col_mc    = _col_index(hdr, "MOULDING MACH", "MACHINE")
    col_hrs   = _col_index(hdr, "RUN HOUR", "RUNNING HOUR", "HRS")
    col_pcs   = _col_index(hdr, "PCS", "OUTPUT")
    col_wt    = _col_index(hdr, "WT IN KGS", "WGT", "WEIGHT")
    col_rej   = _col_index(hdr, "REJECTION", "REJ")

    if col_date < 0:  col_date = 0
    if col_mat  < 0:  col_mat  = 1
    if col_item < 0:  col_item = 2
    if col_mc   < 0:  col_mc   = 4
    if col_pcs  < 0:  col_pcs  = 8
    if col_wt   < 0:  col_wt   = 9
    if col_rej  < 0:  col_rej  = 13

    year, mnum = int(month[:4]), int(month[5:7])
    rows = []
    last_date = None
    last_mat  = None

    for raw in values[hdr_idx + 1:]:
        if len(raw) < 3:
            continue
        def _cell(idx: int, default=""):
            return raw[idx].strip() if idx < len(raw) and str(raw[idx]).strip() else default

        date_raw = _cell(col_date)
        if date_raw:
            d = _parse_date(date_raw, year, mnum)
            if d:
                last_date = d
        if last_date is None:
            continue

        mat_raw = _cell(col_mat)
        if mat_raw:
            last_mat = mat_raw

        item_raw = _cell(col_item)
        mc_raw   = _cell(col_mc)
        if not item_raw or re.match(r'^\d+$', item_raw):
            continue
        if item_raw.upper() in ("ITEM CODE", "TOTAL", ""):
            continue

        try:
            pcs  = _to_float(_cell(col_pcs))
            wt   = _to_float(_cell(col_wt))
            hrs  = _to_float(_cell(col_hrs)) if col_hrs >= 0 else 0.0
            rej  = _to_float(_cell(col_rej))
        except Exception:
            continue

        if pcs <= 0 and wt <= 0:
            continue

        rows.append({
            "date":         last_date,
            "machine":      mc_raw or "MOULDING",
            "material":     last_mat or "",
            "item_code":    item_raw,
            "run_hours":    hrs,
            "pcs":          pcs,
            "weight_kg":    wt,
            "rejection_kg": rej,
        })
    return rows


# ---------------------------------------------------------------------------
# Date parsing helper
# ---------------------------------------------------------------------------

def _parse_date(raw: str, year: int, month: int) -> Optional[str]:
    """Parse a date cell into 'YYYY-MM-DD'. raw may be '1', '01', '01-07-2026',
    '2026-07-01', or a numeric serial (skipped)."""
    raw = str(raw).strip()
    if not raw:
        return None
    # Already ISO
    if re.match(r'^\d{4}-\d{2}-\d{2}$', raw):
        return raw
    # dd-mm-yyyy or dd/mm/yyyy
    m = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$', raw)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return None
    # Plain day number (1..31) — combine with year/month arg
    if re.match(r'^\d{1,2}$', raw):
        try:
            return datetime.date(year, month, int(raw)).isoformat()
        except ValueError:
            return None
    # Numeric float (Excel serial) — skip
    return None


def _to_float(val: str, default: float = 0.0) -> float:
    if not val:
        return default
    val = str(val).replace(",", "").strip()
    # Filter out formula errors
    if val.upper() in ("#N/A", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?"):
        return default
    try:
        return float(val)
    except ValueError:
        return default


# ---------------------------------------------------------------------------
# Actuals loading from Google Sheets
# ---------------------------------------------------------------------------

def load_pipe_actuals(month: str, segment: str = "PLUMBING") -> Tuple[List[dict], int, int]:
    """Read Report-11 and Report-12 from the PIPE daily workbook.

    Returns (parsed_rows, r11_count, r12_count).
    """
    import sources
    import sheets as _sheets
    file_id = (sources.DAILY_SOURCES.get("PIPE") or {}).get("files", {}).get(month)
    if not file_id:
        return [], 0, 0

    token = _sheets._get_access_token()
    if not token:
        return [], 0, 0

    try:
        tabs = _sheets.batch_get(file_id, ["Report-11", "Report-12"], token)
    except Exception:
        return [], 0, 0

    r11_rows = parse_report11(tabs.get("Report-11") or [], month)
    r12_rows = parse_report12(tabs.get("Report-12") or [], month)

    all_rows = [
        dict(r, source_tab="Report-11") for r in r11_rows
    ] + [
        dict(r, source_tab="Report-12") for r in r12_rows
    ]
    return all_rows, len(r11_rows), len(r12_rows)


# ---------------------------------------------------------------------------
# Build MpActualLine objects and upsert to DB
# ---------------------------------------------------------------------------

def ingest_actuals(month: str, segment: str = "PLUMBING") -> dict:
    """Load actuals from sheets and upsert to mp_actual_line. Returns summary dict."""
    import mp_model as _mm
    rows, r11_n, r12_n = load_pipe_actuals(month, segment)
    if not rows:
        return {"ok": False, "reason": "No actuals loaded from sheets",
                "r11_count": 0, "r12_count": 0, "upserted": 0}

    actual_lines = []
    for r in rows:
        mc_raw = str(r.get("machine") or "")
        ic_raw = str(r.get("item_code") or "")
        actual_lines.append(_mm.MpActualLine(
            segment=segment,
            month=month,
            date=r["date"],
            machine=mc_raw,
            machine_norm=norm_machine(mc_raw),
            item_code=ic_raw,
            item_norm=norm_item(ic_raw),
            material=str(r.get("material") or ""),
            actual_pcs=float(r.get("pcs") or 0),
            actual_kg=float(r.get("weight_kg") or 0),
            actual_hours=float(r.get("run_hours") or 0),
            rejection_kg=float(r.get("rejection_kg") or 0),
            source_tab=r.get("source_tab", ""),
        ))

    upserted = _mm.upsert_actual_lines(actual_lines)
    return {
        "ok": True,
        "r11_count": r11_n,
        "r12_count": r12_n,
        "upserted": upserted,
        "dates": sorted({r["date"] for r in rows}),
        "machines": sorted({norm_machine(r.get("machine", "")) for r in rows}),
        "items": sorted({norm_item(r.get("item_code", "")) for r in rows}),
        "total_pcs": sum(r.get("pcs", 0) for r in rows),
        "total_kg": sum(r.get("weight_kg", 0) for r in rows),
        "total_hrs": sum(r.get("run_hours", 0) for r in rows),
    }


# ---------------------------------------------------------------------------
# Build plan lines from ScheduleResult + EngineResult at freeze time
# ---------------------------------------------------------------------------

def build_plan_lines_from_schedule(
    sched: Any,       # mp_scheduler.ScheduleResult
    engine_result: Any,  # mp_engine.EngineResult (pipe) — may be None
    run_id: int,
    segment: str,
    month: str,
) -> "List":
    """Convert scheduler blocks into MpPlanLine rows ready for insert."""
    import mp_model as _mm

    item_by_code: dict = {}
    if engine_result is not None:
        for it in (engine_result.items or []):
            item_by_code[it.item_code] = it

    lines = []
    for block in (sched.blocks or []):
        item = item_by_code.get(block.item_code)
        net_hrs  = max(0.0, float(block.planned_hours) - float(block.excess_hours))
        rate     = float(item.rate_kg_per_hr) if item else 0.0
        kg       = net_hrs * rate
        wt       = float(item.weight_per_pc_kg) if (item and item.weight_per_pc_kg) else 0.0
        pcs      = kg / wt if wt > 0 else 0.0
        lines.append(_mm.MpPlanLine(
            plan_run_id=run_id,
            segment=segment,
            month=month,
            week=int(block.week),
            day=int(block.day),
            shift=str(block.shift),
            machine=str(block.machine),
            machine_norm=norm_machine(str(block.machine)),
            item_code=str(block.item_code),
            item_norm=norm_item(str(block.item_code)),
            material=str(block.material),
            planned_pcs=round(pcs, 2),
            planned_kg=round(kg, 3),
            planned_hours=float(block.planned_hours),
            net_hours=round(net_hrs, 3),
            rate_used=rate,
            rate_estimated=bool(item.rate_estimated) if item else False,
            is_excess=float(block.excess_hours) > 0,
            is_idle=bool(block.is_idle),
        ))
    return lines


# ---------------------------------------------------------------------------
# Variance engine — core logic
# ---------------------------------------------------------------------------

def _elapsed_plan_days(month: str, max_date_str: str, working_days: int) -> int:
    """Estimate how many plan days have elapsed given the max actual date."""
    try:
        y, m = int(month[:4]), int(month[5:7])
        _, days_in_month = calendar.monthrange(y, m)
        max_date = datetime.date.fromisoformat(max_date_str)
        cal_day = max_date.day
        return max(1, round(cal_day / days_in_month * working_days))
    except Exception:
        return working_days


def compute_followup(
    plan_run_id: int,
    segment: str,
    month: str,
    amber_pct: float = 10.0,
    red_pct: float = 25.0,
    hours_dev_pct: float = 15.0,
    min_run_block_hours: float = 2.0,
) -> Optional[FollowUpResult]:
    """Main variance engine. Returns FollowUpResult or None on failure."""
    import mp_model as _mm

    # ── Load plan lines ────────────────────────────────────────────────────
    plan_lines = _mm.get_plan_lines(plan_run_id)
    # ── Load actuals ───────────────────────────────────────────────────────
    actual_lines = _mm.get_actual_lines(segment, month)

    if not plan_lines and not actual_lines:
        return None

    actuals_loaded = len(actual_lines) > 0

    # ── Working days from plan lines (max day number) ──────────────────────
    working_days = max((int(r.get("day") or 0) for r in plan_lines), default=25) if plan_lines else 25

    # ── As-of date ────────────────────────────────────────────────────────
    if actual_lines:
        dates = [str(r["date"])[:10] for r in actual_lines if r.get("date")]
        as_of_date = max(dates) if dates else ""
    else:
        as_of_date = ""

    # ── Elapsed plan days ─────────────────────────────────────────────────
    if as_of_date:
        elapsed_plan_days = _elapsed_plan_days(month, as_of_date, working_days)
    else:
        elapsed_plan_days = 0

    try:
        y, m_num = int(month[:4]), int(month[5:7])
        _, days_in_month = calendar.monthrange(y, m_num)
        elapsed_calendar_days = datetime.date.fromisoformat(as_of_date).day if as_of_date else 0
    except Exception:
        days_in_month = 31
        elapsed_calendar_days = 0

    # ── Aggregate plan lines ───────────────────────────────────────────────
    # Key: (machine_norm, item_norm)
    plan_total:   Dict[Tuple, dict] = defaultdict(lambda: {"pcs": 0.0, "kg": 0.0, "hrs": 0.0})
    plan_todate:  Dict[Tuple, dict] = defaultdict(lambda: {"pcs": 0.0, "kg": 0.0, "hrs": 0.0})
    plan_by_week: Dict[int, float]  = defaultdict(float)
    # To detect planned machines for each item
    planned_machines_for_item: Dict[str, set] = defaultdict(set)

    for r in plan_lines:
        if r.get("is_idle"):
            continue
        mc_n = str(r.get("machine_norm") or norm_machine(str(r.get("machine") or "")))
        ic_n = str(r.get("item_norm") or norm_item(str(r.get("item_code") or "")))
        key  = (mc_n, ic_n)
        kg   = float(r.get("planned_kg") or 0)
        pcs  = float(r.get("planned_pcs") or 0)
        hrs  = float(r.get("planned_hours") or 0)
        plan_total[key]["pcs"] += pcs
        plan_total[key]["kg"]  += kg
        plan_total[key]["hrs"] += hrs
        if int(r.get("day") or 0) <= elapsed_plan_days:
            plan_todate[key]["pcs"] += pcs
            plan_todate[key]["kg"]  += kg
            plan_todate[key]["hrs"] += hrs
        plan_by_week[int(r.get("week") or 0)] += kg
        planned_machines_for_item[ic_n].add(mc_n)

    # Track which items appear in the plan at all
    planned_item_norms: set = {k[1] for k in plan_total}

    # ── Aggregate actual lines ─────────────────────────────────────────────
    actual_total: Dict[Tuple, dict] = defaultdict(
        lambda: {"pcs": 0.0, "kg": 0.0, "hrs": 0.0, "rej": 0.0,
                 "mc": "", "ic": "", "mat": ""}
    )
    # Weekly actual kg (calendar week of month)
    actual_by_week: Dict[int, float] = defaultdict(float)

    for r in actual_lines:
        mc_n = str(r.get("machine_norm") or norm_machine(str(r.get("machine") or "")))
        ic_n = str(r.get("item_norm") or norm_item(str(r.get("item_code") or "")))
        key  = (mc_n, ic_n)
        kg   = float(r.get("actual_kg") or 0)
        pcs  = float(r.get("actual_pcs") or 0)
        hrs  = float(r.get("actual_hours") or 0)
        rej  = float(r.get("rejection_kg") or 0)
        actual_total[key]["pcs"] += pcs
        actual_total[key]["kg"]  += kg
        actual_total[key]["hrs"] += hrs
        actual_total[key]["rej"] += rej
        actual_total[key]["mc"]   = str(r.get("machine") or mc_n)
        actual_total[key]["ic"]   = str(r.get("item_code") or ic_n)
        actual_total[key]["mat"]  = str(r.get("material") or "")
        # Week of month (calendar approximation)
        try:
            d = datetime.date.fromisoformat(str(r["date"])[:10])
            wk = min(4, (d.day - 1) // 7 + 1)
            actual_by_week[wk] += kg
        except Exception:
            pass

    # ── Union of all keys ──────────────────────────────────────────────────
    all_keys = set(plan_total.keys()) | set(actual_total.keys())

    item_rows: List[ItemVariance] = []
    machine_accum: Dict[str, dict] = {}  # key: machine_norm

    def _proj(act_kg: float) -> float:
        if elapsed_calendar_days <= 0:
            return 0.0
        return act_kg / elapsed_calendar_days * days_in_month

    for key in sorted(all_keys):
        mc_n, ic_n = key
        pt = plan_total.get(key, {"pcs": 0.0, "kg": 0.0, "hrs": 0.0})
        pd = plan_todate.get(key, {"pcs": 0.0, "kg": 0.0, "hrs": 0.0})
        at = actual_total.get(key, {"pcs": 0.0, "kg": 0.0, "hrs": 0.0, "rej": 0.0, "mc": "", "ic": "", "mat": ""})

        act_kg  = at["kg"]
        plan_kg = pd["kg"]
        act_pcs = at["pcs"]
        act_hrs = at["hrs"]

        adherence = (act_kg / plan_kg * 100) if plan_kg > 0 else (100.0 if act_kg == 0 else 0.0)
        kg_var    = ((act_kg - plan_kg) / plan_kg * 100) if plan_kg > 0 else 0.0
        hrs_var   = ((act_hrs - pd["hrs"]) / pd["hrs"] * 100) if pd["hrs"] > 0 else 0.0
        rag       = rag_status(act_kg, plan_kg, amber_pct, red_pct)

        is_unplanned     = ic_n not in planned_item_norms
        planned_machines = planned_machines_for_item.get(ic_n, set())
        is_wrong_machine = (not is_unplanned) and bool(planned_machines) and (mc_n not in planned_machines)

        iv = ItemVariance(
            machine=at["mc"] or mc_n,
            machine_norm=mc_n,
            item_code=at["ic"] or ic_n,
            item_norm=ic_n,
            material=at["mat"] or "",
            planned_pcs_total=pt["pcs"],
            planned_kg_total=pt["kg"],
            planned_hours_total=pt["hrs"],
            planned_pcs_todate=pd["pcs"],
            planned_kg_todate=plan_kg,
            planned_hours_todate=pd["hrs"],
            actual_pcs=act_pcs,
            actual_kg=act_kg,
            actual_hours=act_hrs,
            rejection_kg=at["rej"],
            kg_var_pct=round(kg_var, 1),
            hours_var_pct=round(hrs_var, 1),
            adherence_pct=round(adherence, 1),
            projected_kg_month=round(_proj(act_kg), 1),
            rag=rag,
            is_wrong_machine=is_wrong_machine,
            is_unplanned=is_unplanned,
        )
        item_rows.append(iv)

        # Accumulate into machine totals
        if mc_n not in machine_accum:
            machine_accum[mc_n] = {
                "machine": at["mc"] or mc_n,
                "pl_hrs": 0.0, "pl_kg": 0.0,
                "ac_hrs": 0.0, "ac_kg": 0.0,
                "had_plan": False, "had_actual": False,
            }
        machine_accum[mc_n]["pl_hrs"]   += pd["hrs"]
        machine_accum[mc_n]["pl_kg"]    += plan_kg
        machine_accum[mc_n]["ac_hrs"]   += act_hrs
        machine_accum[mc_n]["ac_kg"]    += act_kg
        if plan_kg > 0:
            machine_accum[mc_n]["had_plan"] = True
        if act_kg > 0:
            machine_accum[mc_n]["had_actual"] = True

    # ── Machine variance rows ──────────────────────────────────────────────
    machine_rows: List[MachineVariance] = []
    for mc_n, acc in sorted(machine_accum.items()):
        pl_kg = acc["pl_kg"]
        ac_kg = acc["ac_kg"]
        pl_hrs = acc["pl_hrs"]
        ac_hrs = acc["ac_hrs"]
        kg_var  = ((ac_kg - pl_kg) / pl_kg * 100)  if pl_kg  > 0 else 0.0
        hrs_var = ((ac_hrs - pl_hrs) / pl_hrs * 100) if pl_hrs > 0 else 0.0
        adh     = (ac_kg / pl_kg * 100) if pl_kg > 0 else (100.0 if ac_kg == 0 else 0.0)
        rag     = rag_status(ac_kg, pl_kg, amber_pct, red_pct)
        machine_rows.append(MachineVariance(
            machine=acc["machine"],
            machine_norm=mc_n,
            planned_hours_todate=round(pl_hrs, 2),
            planned_kg_todate=round(pl_kg, 1),
            actual_hours=round(ac_hrs, 2),
            actual_kg=round(ac_kg, 1),
            hours_var_pct=round(hrs_var, 1),
            kg_var_pct=round(kg_var, 1),
            adherence_pct=round(adh, 1),
            projected_kg_month=round(_proj(ac_kg), 1),
            rag=rag,
            had_planned_work=acc["had_plan"],
            had_actual_work=acc["had_actual"],
        ))

    # ── Weekly cumulative trend ────────────────────────────────────────────
    week_points: List[WeekPoint] = []
    cum_plan = 0.0
    cum_act  = 0.0
    for wk in range(1, 5):
        cum_plan += plan_by_week.get(wk, 0.0)
        cum_act  += actual_by_week.get(wk, 0.0)
        week_points.append(WeekPoint(
            week=wk,
            planned_kg_cumulative=round(cum_plan, 1),
            actual_kg_cumulative=round(cum_act, 1),
        ))

    # ── Warnings ───────────────────────────────────────────────────────────
    warnings = _generate_warnings(
        item_rows=item_rows,
        machine_rows=machine_rows,
        plan_lines=plan_lines,
        actual_lines=actual_lines,
        elapsed_plan_days=elapsed_plan_days,
        amber_pct=amber_pct,
        red_pct=red_pct,
        hours_dev_pct=hours_dev_pct,
        min_run_block_hours=min_run_block_hours,
    )

    # ── Scoreboard ─────────────────────────────────────────────────────────
    total_plan_kg    = sum(iv.planned_kg_todate for iv in item_rows)
    total_actual_kg  = sum(iv.actual_kg for iv in item_rows)
    overall_adh      = (total_actual_kg / total_plan_kg * 100) if total_plan_kg > 0 else 0.0
    on_plan_kg       = sum(iv.actual_kg for iv in item_rows if iv.rag == "GREEN")
    off_plan_kg      = total_actual_kg - on_plan_kg

    return FollowUpResult(
        plan_run_id=plan_run_id,
        month=month,
        segment=segment,
        as_of_date=as_of_date,
        elapsed_calendar_days=elapsed_calendar_days,
        days_in_month=days_in_month,
        working_days=working_days,
        elapsed_plan_days=elapsed_plan_days,
        total_planned_kg_todate=round(total_plan_kg, 1),
        total_actual_kg=round(total_actual_kg, 1),
        overall_adherence_pct=round(overall_adh, 1),
        on_plan_kg=round(on_plan_kg, 1),
        off_plan_kg=round(off_plan_kg, 1),
        item_rows=sorted(item_rows, key=lambda x: x.machine_norm),
        machine_rows=machine_rows,
        week_points=week_points,
        warnings=warnings,
        actuals_count=len(actual_lines),
        actuals_loaded=actuals_loaded,
    )


# ---------------------------------------------------------------------------
# Warning generator
# ---------------------------------------------------------------------------

_SEVERITY = {
    WTYPE_WRONG_MACHINE:    1,
    WTYPE_NIGHT_CHANGEOVER: 1,
    WTYPE_SHORT_BLOCK:      2,
    WTYPE_UNPLANNED:        2,
    WTYPE_NOT_STARTED:      2,
    WTYPE_QTY_SHORT:        3,
    WTYPE_QTY_OVER:         3,
    WTYPE_HOURS_DEV:        3,
    WTYPE_IDLE_VS_PLAN:     4,
}


def _generate_warnings(
    item_rows: List[ItemVariance],
    machine_rows: List[MachineVariance],
    plan_lines: List[dict],
    actual_lines: List[dict],
    elapsed_plan_days: int,
    amber_pct: float,
    red_pct: float,
    hours_dev_pct: float,
    min_run_block_hours: float,
) -> List[Warning]:
    warnings: List[Warning] = []

    def _add(wtype: str, machine: str, item: str, material: str,
             magnitude: float, reason: str):
        warnings.append(Warning(
            warning_type=wtype,
            severity=_SEVERITY.get(wtype, 4),
            machine=machine,
            item_code=item,
            material=material,
            magnitude=round(magnitude, 1),
            reason=reason,
        ))

    # ── 1. WRONG MACHINE — produced on a machine not in its routing ─────────
    for iv in item_rows:
        if iv.is_wrong_machine and iv.actual_kg > 0:
            _add(WTYPE_WRONG_MACHINE, iv.machine, iv.item_code, iv.material,
                 iv.actual_kg,
                 f"{iv.item_code} produced on {iv.machine} which is not in its plan routing "
                 f"({iv.actual_kg:.0f} kg actual)")

    # ── 2. UNPLANNED ITEM — in actuals but not in plan ─────────────────────
    for iv in item_rows:
        if iv.is_unplanned and iv.actual_kg > 0:
            _add(WTYPE_UNPLANNED, iv.machine, iv.item_code, iv.material,
                 iv.actual_kg,
                 f"{iv.item_code} ({iv.actual_kg:.0f} kg) is not in the optimised plan")

    # ── 3. NOT STARTED / BEHIND — planned item with zero actual ────────────
    for iv in item_rows:
        if iv.planned_kg_todate > 0 and iv.actual_kg == 0 and elapsed_plan_days > 0:
            _add(WTYPE_NOT_STARTED, iv.machine, iv.item_code, iv.material,
                 iv.planned_kg_todate,
                 f"{iv.item_code} on {iv.machine}: {iv.planned_kg_todate:.0f} kg planned by day {elapsed_plan_days} — no production recorded")

    # ── 4. QUANTITY DEVIATION (over and under) ─────────────────────────────
    for iv in item_rows:
        if iv.planned_kg_todate <= 0 or iv.actual_kg == 0:
            continue
        if iv.rag in ("AMBER", "RED"):
            wtype = WTYPE_QTY_OVER if iv.kg_var_pct > 0 else WTYPE_QTY_SHORT
            direction = "overproduction" if iv.kg_var_pct > 0 else "shortfall"
            _add(wtype, iv.machine, iv.item_code, iv.material,
                 abs(iv.kg_var_pct),
                 f"{iv.item_code} on {iv.machine}: {direction} {abs(iv.kg_var_pct):.1f}% vs plan-to-date "
                 f"(actual {iv.actual_kg:.0f} kg, plan {iv.planned_kg_todate:.0f} kg)")

    # ── 5. HOURS DEVIATION ─────────────────────────────────────────────────
    for mv in machine_rows:
        if mv.planned_hours_todate <= 0 or mv.actual_hours <= 0:
            continue
        dev = abs(mv.hours_var_pct)
        if dev > hours_dev_pct:
            direction = "excess" if mv.hours_var_pct > 0 else "deficit"
            _add(WTYPE_HOURS_DEV, mv.machine, "-", "",
                 dev,
                 f"{mv.machine}: run-hours {direction} {dev:.1f}% vs plan-to-date "
                 f"(actual {mv.actual_hours:.1f} h, plan {mv.planned_hours_todate:.1f} h)")

    # ── 6. IDLE VS PLAN — machine had planned work but no actual ───────────
    for mv in machine_rows:
        if mv.had_planned_work and not mv.had_actual_work and elapsed_plan_days > 0:
            _add(WTYPE_IDLE_VS_PLAN, mv.machine, "-", "",
                 mv.planned_kg_todate,
                 f"{mv.machine}: {mv.planned_kg_todate:.0f} kg planned to date but zero production recorded")

    # ── 7a. NIGHT CHANGEOVER rule violation ────────────────────────────────
    # Detect from actual lines: same machine, same date, NIGHT shift produces
    # more than one distinct item. We approximate by counting distinct items
    # per (machine_norm, date) in actuals (since actuals don't carry shift).
    # Use the plan's night single-block rule: flag if >1 item on same (mc, date)
    # where that mc had a plan NIGHT block that day.
    night_plan_mc_day: set = set()
    for r in plan_lines:
        if str(r.get("shift", "")).upper() == "NIGHT" and not r.get("is_idle"):
            night_plan_mc_day.add((
                str(r.get("machine_norm") or norm_machine(str(r.get("machine") or ""))),
                int(r.get("day") or 0),
            ))

    actual_mc_date_items: Dict[Tuple[str, str], set] = defaultdict(set)
    for r in actual_lines:
        mc_n = str(r.get("machine_norm") or norm_machine(str(r.get("machine") or "")))
        date_s = str(r.get("date") or "")[:10]
        ic_n = str(r.get("item_norm") or norm_item(str(r.get("item_code") or "")))
        actual_mc_date_items[(mc_n, date_s)].add(ic_n)

    # Map calendar date → plan day for approximate check
    if elapsed_plan_days > 0 and actual_lines:
        try:
            y, m_num = int(plan_lines[0]["month"][:4]), int(plan_lines[0]["month"][5:7]) if plan_lines else (2026, 7)
            _, dim = calendar.monthrange(y, m_num)
        except Exception:
            dim = 31

        for (mc_n, date_s), items in actual_mc_date_items.items():
            if len(items) <= 1:
                continue
            try:
                cal_day = datetime.date.fromisoformat(date_s).day
                plan_day = max(1, round(cal_day / dim * (working_days := elapsed_plan_days)))
            except Exception:
                continue
            if (mc_n, plan_day) in night_plan_mc_day:
                _add(WTYPE_NIGHT_CHANGEOVER, mc_n, "-", "",
                     len(items),
                     f"{mc_n} on {date_s}: {len(items)} items recorded — may violate no-night-changeover rule")

    # ── 7b. SHORT BLOCK — actual run block shorter than min_run_block ───────
    # Group actuals by (machine, date) to infer block hours
    mc_date_hrs: Dict[Tuple[str, str], float] = defaultdict(float)
    for r in actual_lines:
        mc_n   = str(r.get("machine_norm") or norm_machine(str(r.get("machine") or "")))
        date_s = str(r.get("date") or "")[:10]
        mc_date_hrs[(mc_n, date_s)] += float(r.get("actual_hours") or 0)

    for (mc_n, date_s), hrs in mc_date_hrs.items():
        if 0 < hrs < min_run_block_hours:
            _add(WTYPE_SHORT_BLOCK, mc_n, "-", "",
                 hrs,
                 f"{mc_n} on {date_s}: {hrs:.1f} h recorded — shorter than min block {min_run_block_hours:.1f} h")

    # Sort by severity, then magnitude descending
    warnings.sort(key=lambda w: (w.severity, -w.magnitude))
    return warnings


# ---------------------------------------------------------------------------
# Follow-up XLSX export
# ---------------------------------------------------------------------------

def export_followup_xlsx(result: FollowUpResult) -> bytes:
    """Build a follow-up variance .xlsx: Summary | Per-Machine | Per-Item | Warnings."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX export")

    wb = openpyxl.Workbook()
    NAVY   = "1F3864"
    TERRA  = "C55A11"
    GREEN  = "D4EDDA"
    AMBER  = "FFF3CD"
    RED    = "F8D7DA"
    WHITE  = "FFFFFF"

    def _hdr(ws, row_vals, fill_hex=NAVY, font_hex=WHITE, bold=True):
        row = ws.append(row_vals)
        r = ws.max_row
        for c in range(1, len(row_vals) + 1):
            cell = ws.cell(r, c)
            cell.font = Font(bold=bold, color=font_hex, size=10)
            cell.fill = PatternFill("solid", fgColor=fill_hex)
            cell.alignment = Alignment(horizontal="center")

    def _rag_fill(rag):
        return {"GREEN": GREEN, "AMBER": AMBER, "RED": RED}.get(rag, WHITE)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Machine Planning Follow-Up Report"])
    ws.cell(1, 1).font = Font(bold=True, size=14, color=NAVY)
    ws.append([])
    ws.append(["Month", result.month,
               "Plan Run #", result.plan_run_id,
               "As-of Date", result.as_of_date])
    ws.append(["Elapsed (calendar days)", result.elapsed_calendar_days,
               "Days in Month", result.days_in_month,
               "Working Days", result.working_days])
    ws.append(["Elapsed Plan Days", result.elapsed_plan_days])
    ws.append([])
    ws.append(["Planned KG (to date)", result.total_planned_kg_todate])
    ws.append(["Actual KG (to date)",  result.total_actual_kg])
    ws.append(["Overall Adherence %",  result.overall_adherence_pct])
    ws.append(["On-Plan KG",           result.on_plan_kg])
    ws.append(["Off-Plan KG",          result.off_plan_kg])
    ws.append([])
    warn_counts = {}
    for w in result.warnings:
        warn_counts[w.warning_type] = warn_counts.get(w.warning_type, 0) + 1
    ws.append(["Warnings by Type"])
    for wtype, cnt in sorted(warn_counts.items()):
        ws.append([f"  {wtype}", cnt])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18

    # ── Sheet 2: Per-Machine ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Machine")
    _hdr(ws2, ["Machine", "Plan KG (to date)", "Actual KG", "KG Var %",
               "Plan Hrs (to date)", "Actual Hrs", "Hrs Var %",
               "Adherence %", "Projected KG/Month", "RAG"])
    for mv in result.machine_rows:
        r = ws2.append([
            mv.machine, mv.planned_kg_todate, mv.actual_kg, mv.kg_var_pct,
            mv.planned_hours_todate, mv.actual_hours, mv.hours_var_pct,
            mv.adherence_pct, mv.projected_kg_month, mv.rag,
        ])
        fill = _rag_fill(mv.rag)
        for c in range(1, 11):
            ws2.cell(ws2.max_row, c).fill = PatternFill("solid", fgColor=fill)

    # ── Sheet 3: Per-Item ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Per-Item")
    _hdr(ws3, ["Machine", "Item Code", "Material",
               "Plan KG (total)", "Plan KG (to date)", "Actual KG", "KG Var %",
               "Adherence %", "Projected KG/Month", "RAG",
               "Wrong Machine?", "Unplanned?"])
    for iv in result.item_rows:
        ws3.append([
            iv.machine, iv.item_code, iv.material,
            iv.planned_kg_total, iv.planned_kg_todate, iv.actual_kg, iv.kg_var_pct,
            iv.adherence_pct, iv.projected_kg_month, iv.rag,
            "YES" if iv.is_wrong_machine else "",
            "YES" if iv.is_unplanned else "",
        ])
        fill = _rag_fill(iv.rag)
        for c in range(1, 13):
            ws3.cell(ws3.max_row, c).fill = PatternFill("solid", fgColor=fill)

    # ── Sheet 4: Warnings ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Warnings")
    _hdr(ws4, ["Severity", "Type", "Machine", "Item", "Material", "Magnitude", "Reason"])
    for w in result.warnings:
        ws4.append([w.severity, w.warning_type, w.machine, w.item_code,
                    w.material, w.magnitude, w.reason])

    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

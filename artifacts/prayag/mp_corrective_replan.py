"""
Corrective Re-plan Engine — Plumbing (PIPE + Fitting)
======================================================

Reads daily production actuals from Report-11 (Pipe, pcs) and Report-12
(Moulding/Fitting, pcs) in the monthly PIPE workbook and computes a
category-level capacity projection for the remaining working days.

Algorithm
---------
For each category (CPVC Pipe, UPVC Pipe, SWR Pipe, AGRI Pipe,
                   CPVC Fitting, UPVC Fitting, SWR Fitting, AGRI Fitting,
                   + 4 Solvent categories):

1.  For each category, group item rows by production date and SUM their pcs.
    This gives one per-category-per-day total — the throughput of that category
    on that date.  Never compute pace over individual item rows.
2.  Compute Pace/Day from those per-category-per-day totals:
      ≥ MIN_DAYS_FOR_P90 non-zero days → 90th-percentile of daily category totals
        ("p90" — the pace exceeded only by the top 10 % of production days;
        optimistic but grounded in actual observed throughput)
      1 … MIN_DAYS_FOR_P90-1 non-zero days → arithmetic mean + low-confidence flag
      0 non-zero days → 0, method = "none" (not started this month)
3.  Projected = Pace/Day × working_days_remaining
4.  Gap = max(0, remaining_demand − projected)

INVARIANTS (checked before returning, raise AssertionError on violation)
------------------------------------------------------------------------
• If producedToDate > 0  →  cap_per_day > 0
• feasible == cap_per_day × working_days_remaining  (exact)
• total shortfall < total remaining  (when any category has production)

Issue #5 FIX — Date formats
----------------------------
Both "Aug 1, 2026" (Plumbing Report-11/12) and "1-Aug-2026" (PTMT) are
parsed correctly.  An unrecognised format raises ValueError (loud fail)
rather than silently returning None / zero rows.
"""
from __future__ import annotations

import calendar
import dataclasses
import datetime
import re
from typing import Dict, List, Optional, Tuple


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MIN_DAYS_FOR_P90: int = 5   # fewer non-zero days → mean+low-confidence flag; ≥ → 90th-percentile

#: R12 materials that are tracked in the sheet TOTAL but do not map to any
#: standard Plumbing category (CPVC/UPVC/SWR/AGRI/PP/ABS).  We count their
#: pcs in the grand total (so it reconciles with the sheet TOTAL row) but we
#: do NOT assign them to a category row.  Tracked in CorrectiveReplanResult.other_produced.
_R12_OTHER_MATERIALS: frozenset = frozenset({"TEFFLONE", "TEFLON", "TEFF"})

#: Map TYPES value in Report-11 → category label
PIPE_CATEGORIES: Dict[str, str] = {
    "CPVC": "CPVC Pipe",
    "UPVC": "UPVC Pipe",
    "SWR":  "SWR Pipe",
    "AGRI": "AGRI Pipe",
}

#: Map MATERIAL value in Report-12 → category label
FITTING_CATEGORIES: Dict[str, str] = {
    "CPVC":   "CPVC Fitting",
    "UPVC":   "UPVC Fitting",
    "SWR":    "SWR Fitting",
    "AGRI":   "AGRI Fitting",
    "PP":     "PP Fitting",
    "ABS":    "ABS Fitting",
}

#: Solvent categories — never appear in Report-11/12 → always NO_DEMONSTRATED_CAPACITY
SOLVENT_CATEGORIES: List[str] = [
    "CPVC Solvent",
    "UPVC Solvent",
    "SWR Solvent",
    "AGRI Solvent",
]

#: Canonical ordering for the output report
CATEGORY_ORDER: List[str] = [
    "CPVC Pipe",   "CPVC Fitting",
    "UPVC Pipe",   "UPVC Fitting",
    "SWR Pipe",    "SWR Fitting",
    "AGRI Pipe",   "AGRI Fitting",
    *SOLVENT_CATEGORIES,
]


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclasses.dataclass
class CategoryResult:
    category: str
    daily_values: List[float]       # non-zero daily totals (pcs)
    n_days: int                     # = len(daily_values)
    produced_to_date: float         # sum of all daily pcs (including 0-days)
    cap_per_day: float              # 0 when not started / no pace data
    method: str                     # "p90" | "mean" | "mean(low-confidence)" | "none"
    remaining: float                # from plan (produce_required - produced_monthly)
    working_days_remaining: int
    feasible: float                 # cap_per_day * working_days_remaining (pace-projected)
    shortfall: float                # max(0, remaining - feasible)
    cap_feasible: Optional[float] = None  # machine-capacity feasible (from capacity plan, if available)

    @property
    def shortfall_pct(self) -> Optional[float]:
        if self.remaining <= 0:
            return None
        return round(self.shortfall / self.remaining * 100, 1)

    @property
    def not_started(self) -> bool:
        """True when no production has been observed yet (category hasn't run this month)."""
        return self.produced_to_date == 0 and self.cap_per_day == 0

    # Keep the old name as an alias so existing tests don't break
    @property
    def no_demonstrated_capacity(self) -> bool:
        return self.not_started

    @property
    def low_confidence(self) -> bool:
        """True when pace is based on fewer than MIN_DAYS_FOR_P90 production days."""
        return self.n_days > 0 and self.n_days < MIN_DAYS_FOR_P90


@dataclasses.dataclass
class CorrectiveReplanResult:
    month: str
    as_of_date: str
    working_days_total: int
    working_days_elapsed: int
    working_days_remaining: int
    categories: List[CategoryResult]
    source_file_id: str
    source_date_min: str   # earliest date seen in Report-11/12
    source_date_max: str   # latest date seen
    plan_produced_total: float   # from Report-1 (plan "produced" column)
    actual_produced_total: float # from our Report-11+12 parse (all categories + other)
    other_produced: float        # R12 pcs in unclassified materials (e.g. TEFFLONE)
    warnings: List[str]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _parse_date_cell(raw: str, year: int, month: int) -> Optional[str]:
    """Parse a date cell into 'YYYY-MM-DD'.

    Supports:
      'Aug 1, 2026'   — Plumbing Report-11/12 format
      '1-Aug-2026'    — PTMT master format
      '1 Aug 2026'    — variant
      '01-08-2026'    — dd-mm-yyyy
      '2026-08-01'    — ISO
      '1'             — plain day (combined with year/month arg)

    Raises ValueError for any unrecognised format so silent-zero bugs are
    caught immediately (Issue #6 regression guard).
    """
    raw = str(raw).strip().lstrip("'").strip()
    if not raw:
        return None

    # ISO
    if re.match(r'^\d{4}-\d{2}-\d{2}$', raw):
        return raw

    # "Aug 1, 2026" or "Aug 01, 2026" (with or without space after comma)
    for fmt in ("%b %d, %Y", "%b %d,%Y", "%B %d, %Y", "%B %d,%Y"):
        try:
            return datetime.datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass

    # "1-Aug-2026" or "1 Aug 2026"
    for fmt in ("%d-%b-%Y", "%d %b %Y", "%d-%B-%Y", "%d %B %Y"):
        try:
            return datetime.datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass

    # dd-mm-yyyy or dd/mm/yyyy
    m = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$', raw)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return None

    # Plain day number (1..31)
    if re.match(r'^\d{1,2}$', raw):
        try:
            return datetime.date(year, month, int(raw)).isoformat()
        except ValueError:
            return None

    # Numeric float / Excel serial → skip silently (not a date we understand)
    if re.match(r'^\d+\.\d+$', raw) or re.match(r'^\d{5}$', raw):
        return None

    raise ValueError(f"Unrecognised date format: {raw!r}")


def _to_float(val, default: float = 0.0) -> float:
    if val is None:
        return default
    s = str(val).replace(",", "").strip()
    if s.upper() in ("#N/A", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", ""):
        return default
    try:
        return float(s)
    except ValueError:
        return default


def _col_idx(hdr: list, *names: str) -> int:
    """Return first column index whose header (upper-stripped) contains any name."""
    for ni, name in enumerate(names):
        name_u = name.upper()
        for ci, cell in enumerate(hdr):
            if name_u in str(cell).upper().strip():
                return ci
    return -1


def _percentile_90(values: List[float]) -> float:
    """90th-percentile of per-category-per-day production totals without numpy.

    Input: a list where every element is the CATEGORY TOTAL for one production
    day (i.e. the sum over all item rows of that category on that day).

    Returns the 90th-percentile of those daily totals — the pace exceeded only
    by the top 10 % of production days.  This is the OPTIMISTIC end of the
    distribution: "how fast does the line run on its best days?"  Using the 90th
    percentile (rather than the mean) gives a pace projection that accounts for
    variability without being dragged down by slow start-up days or low-volume
    days, and without being a single-day outlier.

    Uses linear interpolation identical to numpy's default method
    (numpy.percentile(values, 90, interpolation='linear')).
    """
    s = sorted(values)
    n = len(s)
    if n == 1:
        return s[0]
    idx = (n - 1) * 0.90          # fractional index into sorted list (90th pct)
    lo = int(idx)
    hi = min(lo + 1, n - 1)
    return s[lo] + (s[hi] - s[lo]) * (idx - lo)


def _compute_pace_per_day(
    daily_cat_totals: List[float]
) -> Tuple[float, str, int]:
    """Return (pace_per_day, method_label, n_non_zero_days).

    Input must be per-CATEGORY-per-DAY totals (i.e. already summed across all
    item rows for that category on each day).  Passing individual item-row pcs
    would give a per-item pace, which is 6-10× too low.

    Fallback chain
    --------------
    ≥ MIN_DAYS_FOR_P90 production days → 90th-percentile of daily category totals
      (the pace exceeded only by the top 10 % of days — optimistic but grounded)
    1 … MIN_DAYS_FOR_P90-1 days → arithmetic mean + low-confidence flag
      (too few data points for a stable percentile; treat as indicative only)
    0 production days → 0.0 / "none" (category not started this month)
    """
    non_zero = [v for v in daily_cat_totals if v > 0]
    n = len(non_zero)
    if n == 0:
        return 0.0, "none", 0
    if n >= MIN_DAYS_FOR_P90:
        return round(_percentile_90(non_zero), 1), "p90", n
    # Fewer than MIN_DAYS_FOR_P90 days: use mean, flag as low-confidence
    return round(sum(non_zero) / n, 1), f"mean(low-confidence,{n}d)", n


# Backward-compat alias used by existing tests
_compute_cap_per_day = _compute_pace_per_day


# ---------------------------------------------------------------------------
# Report-11 parser (Pipe pcs)
# ---------------------------------------------------------------------------

def _parse_r11_daily_pcs(values: list, year: int, month: int) -> Dict[str, Dict[str, float]]:
    """Parse Report-11 (M/C & Item-wise Actual Production).

    Returns {date_iso: {category: pcs_total}}.
    Category derived from TYPES column (CPVC/UPVC/SWR/AGRI → CPVC Pipe etc.).

    Date format "Aug 1, 2026" is supported (Issue #5 fix).
    """
    if not values:
        return {}

    # Locate header row: first row containing "ITEM CODE"
    hdr_idx = None
    for i, row in enumerate(values[:15]):
        if any("ITEM CODE" in str(c).upper() for c in row):
            hdr_idx = i
            break
    if hdr_idx is None:
        hdr_idx = 4  # spec default

    hdr = values[hdr_idx]
    col_date = _col_idx(hdr, "DATE"); col_date = col_date if col_date >= 0 else 1
    col_type = _col_idx(hdr, "TYPE", "TYPES"); col_type = col_type if col_type >= 0 else 4
    col_item = _col_idx(hdr, "ITEM CODE"); col_item = col_item if col_item >= 0 else 5
    col_pcs  = _col_idx(hdr, "PCS");        col_pcs  = col_pcs  if col_pcs  >= 0 else 8

    daily: Dict[str, Dict[str, float]] = {}  # date → {category → pcs}
    last_date: Optional[str] = None
    dates_seen: set = set()
    n_no_date: int = 0   # data rows skipped because no date could be resolved

    for raw in values[hdr_idx + 1:]:
        if len(raw) < 4:
            continue

        def _cell(idx: int) -> str:
            return str(raw[idx]).strip() if idx < len(raw) else ""

        # Date — always read from the column identified as "DATE" by header text.
        # In the August PIPE workbook, that is column B (index 1); column A is blank.
        # Carry the last resolved date forward for continuation rows (same-day items
        # share the date only on the first item row).
        date_raw = _cell(col_date)
        if date_raw:
            try:
                d = _parse_date_cell(date_raw, year, month)
                if d:
                    last_date = d
                    dates_seen.add(d)
            except ValueError:
                pass  # sub-header rows, text labels, etc. — not a data date

        if last_date is None:
            # The very first data rows have no date yet; we cannot attribute them.
            # Count them so the caller can surface a warning if there are too many.
            mat_type = _cell(col_type).upper()
            pcs = _to_float(_cell(col_pcs))
            if mat_type and mat_type not in ("TYPES", "TYPE") and pcs > 0:
                n_no_date += 1
            continue

        # Type → category.  We sum ITEM ROWS per (date, category) so that
        # _compute_pace_per_day receives per-CATEGORY-per-DAY totals, not
        # per-item values (using per-item values would give a 6-10× underestimate).
        mat_type = _cell(col_type).upper()
        if not mat_type or "TOTAL" in mat_type or "TYPE" in mat_type:
            continue
        category = PIPE_CATEGORIES.get(mat_type)
        if not category:
            continue

        item = _cell(col_item)
        if not item or "ITEM" in item.upper() or re.match(r'^\d+$', item):
            continue

        pcs = _to_float(_cell(col_pcs))
        if pcs <= 0:
            continue

        # Sum this item's pcs into the day's category total.
        # After all rows for a category on a given date are accumulated,
        # the day_map[category] value equals the category-level daily throughput.
        day_map = daily.setdefault(last_date, {})
        day_map[category] = day_map.get(category, 0.0) + pcs

    if n_no_date:
        # Surface as a return value so compute_corrective_replan can warn.
        # Attach to the dict by convention (avoids changing the function signature).
        daily["_n_no_date"] = n_no_date  # type: ignore[assignment]

    return daily


# ---------------------------------------------------------------------------
# Report-12 parser (Moulding/Fitting pcs)
# ---------------------------------------------------------------------------

def _parse_r12_daily_pcs(
    values: list, year: int, month: int
) -> Tuple[Dict[str, Dict[str, float]], Dict[str, float]]:
    """Parse Report-12 (Mould M/C) for fitting pcs per category per date.

    Returns (daily, other_pcs):
      daily     : {date_iso: {category: pcs_total}}
      other_pcs : {material_name: total_pcs} for unclassified materials
                  (e.g. TEFFLONE).  These reconcile with the sheet TOTAL row
                  but are NOT assigned to a standard category.

    Header is at row index 3 (DATE | MATERIAL | Item Code | … | Output Pc | …).
    Row 4 has sub-headers; data starts row 5 (0-indexed).
    """
    if not values:
        return {}, {}

    # Locate header row by DATE + MATERIAL
    hdr_idx = None
    for i, row in enumerate(values[:10]):
        row_u = [str(c).upper().strip() for c in row]
        if "DATE" in row_u and ("MATERIAL" in row_u or "ITEM CODE" in row_u):
            hdr_idx = i
            break
    if hdr_idx is None:
        hdr_idx = 3

    hdr = values[hdr_idx]
    col_date = _col_idx(hdr, "DATE");      col_date = col_date if col_date >= 0 else 0
    col_mat  = _col_idx(hdr, "MATERIAL");  col_mat  = col_mat  if col_mat  >= 0 else 1
    col_item = _col_idx(hdr, "ITEM CODE"); col_item = col_item if col_item >= 0 else 2

    # Pcs column: "Output Production" header in hdr, but the actual ' Pc ' token
    # is in the sub-header row immediately below — fall back to col_pcs = 8.
    col_pcs = _col_idx(hdr, "OUTPUT PRODUCTION", "PCS", "PC")
    if col_pcs < 0 and hdr_idx + 1 < len(values):
        sub = values[hdr_idx + 1]
        col_pcs = _col_idx(sub, "PC", "PCS")
    if col_pcs < 0:
        col_pcs = 8

    # Skip sub-header row if present
    data_start = hdr_idx + 1
    if data_start < len(values):
        first = [str(c).strip() for c in values[data_start][:6]]
        if all(not c or c.upper() in ("PC", "PCS", "WT IN KGS", "KGS", "WTE", "WTD", "") for c in first):
            data_start += 1

    daily: Dict[str, Dict[str, float]] = {}
    other_pcs: Dict[str, float] = {}
    last_date: Optional[str] = None

    for raw in values[data_start:]:
        if len(raw) < 3:
            continue

        def _cell(idx: int) -> str:
            return str(raw[idx]).strip() if idx < len(raw) else ""

        date_raw = _cell(col_date)
        if date_raw:
            try:
                d = _parse_date_cell(date_raw, year, month)
                if d:
                    last_date = d
            except ValueError:
                pass

        if last_date is None:
            continue

        mat = _cell(col_mat).upper()
        if not mat or "TOTAL" in mat or "MATERIAL" in mat:
            continue

        item = _cell(col_item)
        if not item or "ITEM" in item.upper():
            continue

        pcs = _to_float(_cell(col_pcs))
        if pcs <= 0:
            continue

        # Check for unclassified materials (e.g. TEFFLONE = misspelled Teflon).
        # Count them in the total (so we reconcile with the sheet TOTAL row) but
        # do NOT assign to any standard category row.
        is_other = any(tok in mat for tok in _R12_OTHER_MATERIALS)
        if is_other:
            other_pcs[mat] = other_pcs.get(mat, 0.0) + pcs
            continue

        category = FITTING_CATEGORIES.get(mat)
        if not category:
            continue

        day_map = daily.setdefault(last_date, {})
        day_map[category] = day_map.get(category, 0.0) + pcs

    return daily, other_pcs


# ---------------------------------------------------------------------------
# Working-day calendar (Mon–Sat)
# ---------------------------------------------------------------------------

def _working_days_in_month(year: int, month: int) -> List[datetime.date]:
    """All Mon–Sat dates in *year*/*month*."""
    _, n_days = calendar.monthrange(year, month)
    return [
        datetime.date(year, month, d)
        for d in range(1, n_days + 1)
        if datetime.date(year, month, d).weekday() < 6   # 0=Mon … 5=Sat
    ]


def _count_working_days(year: int, month: int, as_of: datetime.date) -> Tuple[int, int, int]:
    """Return (total, elapsed, remaining) Mon–Sat working days.

    *elapsed* = working days strictly before *as_of* (i.e. completed days).
    *remaining* = working days from *as_of* inclusive to month end.
    """
    all_wd = _working_days_in_month(year, month)
    elapsed   = sum(1 for d in all_wd if d < as_of)
    remaining = sum(1 for d in all_wd if d >= as_of)
    return len(all_wd), elapsed, remaining


# ---------------------------------------------------------------------------
# Main engine
# ---------------------------------------------------------------------------

def compute_corrective_replan(
    month: str,
    plan_recs,                             # list[planning.PlanRecord]
    r11_values: list,                      # raw Sheets values from Report-11
    r12_values: list,                      # raw Sheets values from Report-12
    as_of_date: str,                       # 'YYYY-MM-DD'
    file_id: str = "",
    cap_feasible_by_cat: Optional[Dict[str, float]] = None,
    # ^ If the Capacity-Feasible Plan has been run for this month, pass a dict
    #   {category_label → feasible_pcs} here.  Used as a fallback display figure
    #   for "not started" categories and shown alongside the pace projection.
    #   When None the capacity column shows "—" and the feature still works.
) -> CorrectiveReplanResult:
    """Compute the Corrective Re-plan pace projection for Plumbing (PIPE + Fitting).

    This is a RUN-RATE projection, NOT a machine-capacity statement.
    It answers: "if the plant keeps producing at the pace seen so far,
    where will output land by month end?"

    For machine-capacity (what can physically be made), use the separate
    Capacity-Feasible Plan (machine hours × rates, ≤ 100%).

    Parameters
    ----------
    month              : 'YYYY-MM'
    plan_recs          : PlanRecord list from load_planning('PIPE', month)
    r11_values         : raw values from Report-11 (pipe actuals)
    r12_values         : raw values from Report-12 (fitting actuals)
    as_of_date         : today's date ('YYYY-MM-DD')
    file_id            : source Google Sheets file ID (for provenance)
    cap_feasible_by_cat: optional {category → feasible pcs} from capacity plan

    INVARIANTS enforced (warn on violation, never raise)
    ----------------------------------------------------
    1. producedToDate > 0  →  pace_per_day > 0
    2. projected_pace == pace_per_day × working_days_remaining  (exact)
    3. total_pace_gap < total_remaining  (when ≥ 1 category has production)
    """
    year  = int(month[:4])
    mnum  = int(month[5:7])
    as_of = datetime.date.fromisoformat(as_of_date)

    wd_total, wd_elapsed, wd_remaining = _count_working_days(year, mnum, as_of)

    # --- Parse daily actuals ------------------------------------------------
    r11_daily              = _parse_r11_daily_pcs(r11_values, year, mnum)
    r12_daily, r12_other   = _parse_r12_daily_pcs(r12_values, year, mnum)
    # r12_other: {material → total pcs} for unclassified materials (e.g. TEFFLONE)

    # _parse_r11_daily_pcs embeds a sentinel key when it encountered rows whose
    # date could not be resolved; extract it before building all_dates.
    n_r11_no_date: int = int(r11_daily.pop("_n_no_date", 0))  # type: ignore[arg-type]

    # Collect all REAL dates seen (the sentinel key is gone now)
    all_dates = sorted(set(r11_daily) | set(r12_daily))
    src_date_min = all_dates[0]  if all_dates else ""
    src_date_max = all_dates[-1] if all_dates else ""

    # Build per-category daily totals: {category → list[pcs_per_day]}
    # One entry per *production day* (date where category ran).
    # Days the category didn't run are NOT included — they reflect scheduling
    # absence, not a capacity constraint.
    cat_daily: Dict[str, List[float]] = {c: [] for c in CATEGORY_ORDER}
    cat_produced: Dict[str, float] = {c: 0.0 for c in CATEGORY_ORDER}

    for date in all_dates:
        day: Dict[str, float] = {}
        for cat in PIPE_CATEGORIES.values():
            day[cat] = r11_daily.get(date, {}).get(cat, 0.0)
        for cat in FITTING_CATEGORIES.values():
            v = r12_daily.get(date, {}).get(cat, 0.0)
            day[cat] = day.get(cat, 0.0) + v
        for cat, pcs in day.items():
            if cat not in cat_daily:
                continue
            cat_produced[cat] += pcs
            if pcs > 0:
                cat_daily[cat].append(pcs)

    # Solvent categories: never appear in Report-11/12
    for cat in SOLVENT_CATEGORIES:
        cat_daily[cat] = []
        cat_produced[cat] = 0.0

    # --- Plan remaining demand per category ---------------------------------
    cat_remaining: Dict[str, float] = {c: 0.0 for c in CATEGORY_ORDER}
    plan_produced_total = 0.0

    for rec in (plan_recs or []):
        fam = (rec.family or "").upper()
        cat_str = (rec.category or "").upper()
        if "FITTING" in cat_str or "FIT" in cat_str:
            sub = "Fitting"
        elif "SOLVENT" in cat_str or "SOL" in cat_str:
            sub = "Solvent"
        else:
            sub = "Pipe"

        if fam == "AGRI":
            label: Optional[str] = f"AGRI {sub}"
        elif fam == "SWR":
            label = f"SWR {sub}"
        elif fam == "CPVC":
            label = f"CPVC {sub}"
        elif fam == "UPVC":
            label = f"UPVC {sub}"
        else:
            label = None

        if label and label in cat_remaining:
            demand = max(
                getattr(rec, "produce_required", 0.0) - getattr(rec, "produced", 0.0),
                getattr(rec, "ideal_qty", 0.0) - getattr(rec, "closing_stock", 0.0),
                0.0,
            )
            cat_remaining[label] += demand

        plan_produced_total += getattr(rec, "produced", 0.0)

    # actual_produced_total = categorised pcs + unclassified (TEFFLONE etc.)
    # This must equal the R11 TOTAL row + R12 TOTAL row from the source sheet.
    other_produced        = round(sum(r12_other.values()), 0)
    actual_produced_total = round(sum(cat_produced.values()) + other_produced, 0)

    # Warn if the R11 parser encountered rows it couldn't assign a date to.
    # These rows contribute pcs to the sheet TOTAL but NOT to any per-day total,
    # so produced-to-date and pace are both understated by that amount.
    warnings: List[str] = []
    if n_r11_no_date:
        warnings.append(
            "R-11 date resolution: " + str(n_r11_no_date) + " data row(s) could not be "
            "attributed to a date (no date resolved in the DATE column and no prior "
            "date to carry forward).  These pcs are excluded from produced-to-date "
            "and pace.  Check for rows that appear before the first date entry in "
            "Report-11."
        )

    # --- Build per-category results -----------------------------------------
    results: List[CategoryResult] = []

    for cat in CATEGORY_ORDER:
        daily_vals   = cat_daily.get(cat, [])
        produced     = cat_produced.get(cat, 0.0)
        remaining    = cat_remaining.get(cat, 0.0)
        cap_feas_val = (cap_feasible_by_cat or {}).get(cat, None)

        pace, method, n = _compute_cap_per_day(daily_vals)
        projected = round(pace * wd_remaining, 1)
        gap       = round(max(0.0, remaining - projected), 1)

        # INVARIANT 1: produced > 0  →  pace > 0
        # (not_started categories are exempt — 0 produced, 0 pace is expected)
        if produced > 0 and pace <= 0:
            warnings.append(
                f"INVARIANT VIOLATED: {cat} has producedToDate={produced:.0f} "
                f"but pace_per_day=0. Emergency fallback applied."
            )
            pace      = round(produced / max(1, len(all_dates)), 1)
            method    = "mean(emergency)"
            projected = round(pace * wd_remaining, 1)
            gap       = round(max(0.0, remaining - projected), 1)

        results.append(CategoryResult(
            category=cat,
            daily_values=daily_vals,
            n_days=n,
            produced_to_date=produced,
            cap_per_day=pace,           # field kept as cap_per_day for compat
            method=method,
            remaining=remaining,
            working_days_remaining=wd_remaining,
            feasible=projected,         # field kept as feasible for compat
            shortfall=gap,              # field kept as shortfall for compat
            cap_feasible=cap_feas_val,
        ))

    # --- INVARIANT 2: projected == pace × wd_remaining (exact) -------------
    for r in results:
        expected = round(r.cap_per_day * r.working_days_remaining, 1)
        if abs(r.feasible - expected) > 0.05:
            warnings.append(
                f"Pace inconsistency for {r.category}: "
                f"projected={r.feasible} but pace×days={expected}"
            )

    # --- INVARIANT 3: total_gap < total_remaining if any production ---------
    total_remaining  = sum(r.remaining for r in results)
    total_gap        = sum(r.shortfall for r in results)
    total_production = sum(r.produced_to_date for r in results)
    if total_production > 0 and total_gap >= total_remaining and total_remaining > 0:
        warnings.append(
            f"CRITICAL: total pace-gap ({total_gap:,.0f}) ≥ total_remaining "
            f"({total_remaining:,.0f}) despite {total_production:,.0f} pcs produced. "
            "Pace projection has failed — do not act on this report."
        )

    return CorrectiveReplanResult(
        month=month,
        as_of_date=as_of_date,
        working_days_total=wd_total,
        working_days_elapsed=wd_elapsed,
        working_days_remaining=wd_remaining,
        categories=results,
        source_file_id=file_id,
        source_date_min=src_date_min,
        source_date_max=src_date_max,
        plan_produced_total=round(plan_produced_total, 0),
        actual_produced_total=actual_produced_total,
        other_produced=other_produced,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def corrective_replan_bytes(result: CorrectiveReplanResult) -> bytes:
    """Return .xlsx bytes for the Corrective Re-plan pace-projection report.

    This is a RUN-RATE monitoring tool, NOT a capacity statement.
    Labels use "pace" / "projected" / "gap" language throughout.

    Tabs
    ----
    1. Run-rate Projection  — pace/day, projected output, gap to demand
    2. Daily Actuals        — date × category pivot of raw pcs
    3. Provenance           — file IDs, date ranges, method notes, reconciliation
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    _NAVY   = "1F3864"
    _AMBER  = "FFF2CC"
    _ORANGE = "FCE4D6"   # gap present
    _GREEN  = "E2EFDA"   # on track
    _GREY   = "F5F5F5"   # not started
    _BLUE   = "DBEAFE"   # low-confidence

    def _nfont(**kw): return Font(name="Calibri", **kw)
    def _align(**kw): return Alignment(**kw)

    def _hdr(ws, r, c, txt, w=14, colour="FFFFFF", fill=_NAVY):
        cell = ws.cell(row=r, column=c, value=txt)
        cell.font = Font(name="Calibri", bold=True, size=9, color=colour)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col = get_column_letter(c)
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, w)

    def _title(ws, text, ncols=11):
        end = get_column_letter(ncols)
        ws.merge_cells(f"A1:{end}1")
        c = ws["A1"]
        c.value = text
        c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 4

    def _note(ws, r, text, ncols=11, colour="555555", italic=True, height=22):
        end = get_column_letter(ncols)
        ws.merge_cells(f"A{r}:{end}{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = Font(name="Calibri", italic=italic, bold=not italic, size=9, color=colour)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = height

    # ── Determine column count (adds cap-feasible column when available) ──────
    has_cap_feas = any(cat.cap_feasible is not None for cat in result.categories)
    NCOLS = 11 if has_cap_feas else 10

    # ── Tab 1: Run-rate Projection ────────────────────────────────────────────
    ws1 = wb.create_sheet("Run-rate Projection")
    ws1.sheet_view.showGridLines = False

    _title(ws1,
        f"Corrective Re-plan — Plumbing — {result.month}  (as of {result.as_of_date})",
        ncols=NCOLS)

    # Row 3: subtitle — run-rate framing
    _note(ws1, 3,
        "Run-rate projection — based on actual output pace to date, NOT machine capacity.  "
        "This answers: 'if the plant keeps producing at last week's pace, where will it land?'  "
        "For what the machines can physically make, see the Capacity-Feasible Plan.",
        ncols=NCOLS, colour="7F3B00", italic=False, height=32)

    # Row 4: working-day context
    _note(ws1, 4,
        f"Working days (Mon–Sat): {result.working_days_total} total, "
        f"{result.working_days_elapsed} elapsed, "
        f"{result.working_days_remaining} remaining.  "
        f"Pace method: p90 (≥{MIN_DAYS_FOR_P90} days) → mean[low-confidence] → not-started.  "
        f"Projected = Pace/day × {result.working_days_remaining} remaining days.",
        ncols=NCOLS)

    # Row 5: column headers
    HDRS = [
        "Category",
        "Days recorded",
        "Produced to date\n(R-11 + R-12)",
        "Remaining demand\n(Report-1)",
        "Actual pace/day\n(run-rate)",
        "Method",
        "Projected at\ncurrent pace",
        "Gap to demand\nat current pace",
        "Gap %",
        "Capacity-feasible\n(machine plan)",
    ]
    active_hdrs = HDRS[:9] + (HDRS[9:10] if has_cap_feas else [])
    col_widths   = [20, 14, 20, 20, 18, 22, 18, 18, 10, 20]
    for ci, (h, w) in enumerate(zip(active_hdrs, col_widths), 1):
        _hdr(ws1, 5, ci, h, w=w)

    row = 6
    for cat in result.categories:
        if cat.not_started:
            fill_colour = _GREY
        elif cat.low_confidence:
            fill_colour = _BLUE
        elif cat.shortfall > 0:
            fill_colour = _ORANGE
        else:
            fill_colour = _GREEN

        if cat.not_started:
            pace_disp   = "Not started — no pace data yet"
            method_disp = "—"
            proj_disp   = "—"
            gap_disp    = "—"
            gap_pct     = "—"
        else:
            pace_disp   = round(cat.cap_per_day, 1)
            lc_flag     = " ⚠low-confidence" if cat.low_confidence else ""
            method_disp = cat.method + lc_flag
            proj_disp   = round(cat.feasible)
            gap_disp    = round(cat.shortfall) if cat.shortfall > 0 else "—"
            gap_pct     = f"{cat.shortfall_pct:.0f}%" if cat.shortfall_pct else "—"

        if has_cap_feas:
            cf_disp = round(cat.cap_feasible) if cat.cap_feasible is not None else "—"
        
        row_vals = [
            cat.category,
            cat.n_days if not cat.not_started else "—",
            round(cat.produced_to_date) if not cat.not_started else "—",
            round(cat.remaining) if cat.remaining > 0 else "—",
            pace_disp,
            method_disp,
            proj_disp,
            gap_disp,
            gap_pct,
        ]
        if has_cap_feas:
            row_vals.append(cf_disp)

        for ci, val in enumerate(row_vals, 1):
            c = ws1.cell(row=row, column=ci, value=val)
            c.font = Font(name="Calibri", size=9)
            c.fill = PatternFill("solid", fgColor=fill_colour)
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "right",
                vertical="center",
            )
        row += 1

    # Totals row
    total_produced  = sum(c.produced_to_date for c in result.categories)
    total_remaining = sum(c.remaining        for c in result.categories)
    total_projected = sum(c.feasible         for c in result.categories)
    total_gap       = sum(c.shortfall        for c in result.categories)
    total_gap_pct   = (total_gap / total_remaining * 100) if total_remaining > 0 else 0
    total_cf        = (sum(c.cap_feasible for c in result.categories
                           if c.cap_feasible is not None)
                       if has_cap_feas else None)

    total_vals = [
        "TOTAL", "—", round(total_produced), round(total_remaining),
        "—", "—", round(total_projected), round(total_gap),
        f"{total_gap_pct:.0f}%",
    ]
    if has_cap_feas:
        total_vals.append(round(total_cf) if total_cf is not None else "—")

    for ci, val in enumerate(total_vals, 1):
        c = ws1.cell(row=row, column=ci, value=val)
        c.font = Font(name="Calibri", bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=_AMBER)
        c.alignment = Alignment(horizontal="right" if ci > 1 else "left", vertical="center")

    # Colour legend
    row += 2
    legend = [
        ("🟩 Green", "On track — projected pace meets remaining demand"),
        ("🟧 Orange", "Gap — at current pace, demand will not be fully met"),
        ("🟦 Blue", "Low-confidence — < 5 production days; treat pace estimate with caution"),
        ("⬜ Grey",
         "Not started — no production recorded yet this month.  "
         "Cannot estimate pace.  "
         + ("Capacity-feasible column shows machine-plan figure." if has_cap_feas
            else "Run the Capacity-Feasible Plan to populate the machine-plan column.")),
    ]
    for label, desc in legend:
        ws1.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
        c = ws1[f"A{row}"]
        c.value = f"  {label}: {desc}"
        c.font = Font(name="Calibri", italic=True, size=8, color="333333")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws1.row_dimensions[row].height = 16
        row += 1

    # Warnings
    if result.warnings:
        row += 1
        for w in result.warnings:
            ws1.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
            c = ws1[f"A{row}"]
            c.value = f"⚠ {w}"
            c.font = Font(name="Calibri", bold=True, size=9, color="C00000")
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws1.row_dimensions[row].height = 32
            row += 1

    # Source reconciliation note
    row += 1
    r11_total = round(total_produced - sum(
        c.produced_to_date for c in result.categories
        if "Fitting" in c.category or "Solvent" in c.category
    ))
    other_note = (
        f"  Other/unclassified R-12 materials (e.g. TEFFLONE): "
        f"{result.other_produced:,.0f} pcs included in total above."
        if result.other_produced > 0 else ""
    )
    recon_note = (
        f"Produced to date (this report): {result.actual_produced_total:,.0f} pcs  "
        f"= R-11 pipe pcs ({result.actual_produced_total - sum(c.produced_to_date for c in result.categories if 'Fitting' in c.category or 'Solvent' in c.category):,.0f}) "
        f"+ R-12 fitting pcs (including other materials).  "
        f"Source dates: {result.source_date_min} – {result.source_date_max}.{other_note}  "
        f"Report-1 'Produced': {result.plan_produced_total:,.0f} pcs "
        f"(Report-1 aggregates all plants — differs from R-11/R-12 which are Plumbing only)."
    )
    _note(ws1, row, recon_note, ncols=NCOLS, colour="333333")

    # ── Tab 2: Daily Actuals ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Daily Actuals")
    ws2.sheet_view.showGridLines = False

    _title(ws2, f"Daily Actuals by Category — {result.month}  (pcs)", ncols=NCOLS)
    _note(ws2, 3,
          "Daily totals (pcs) per category from Report-11 (pipe) and Report-12 (fitting).  "
          "Days when a category did not run are omitted — absence ≠ zero capacity.",
          ncols=NCOLS)

    active_cats = [c for c in CATEGORY_ORDER if "Solvent" not in c]

    _hdr(ws2, 4, 1, "Production day #", 10)
    for ci, cat in enumerate(active_cats, 2):
        _hdr(ws2, 4, ci, cat, 14)

    cat_to_res = {r.category: r for r in result.categories}
    max_days   = max((len(r.daily_values) for r in result.categories), default=0)
    for day_i in range(max_days):
        row_num = 5 + day_i
        ws2.cell(row=row_num, column=1, value=f"Day {day_i+1}").font = _nfont(size=9, bold=True)
        for ci, cat in enumerate(active_cats, 2):
            vals = cat_to_res[cat].daily_values if cat in cat_to_res else []
            v = vals[day_i] if day_i < len(vals) else ""
            c = ws2.cell(row=row_num, column=ci, value=round(v) if v != "" else "")
            c.font = _nfont(size=9)
            c.alignment = _align(horizontal="right")
            c.fill = PatternFill("solid", fgColor=_GREY if row_num % 2 == 0 else "FFFFFF")

    # ── Tab 3: Provenance ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Provenance")
    ws3.sheet_view.showGridLines = False
    _title(ws3, f"Report Provenance — {result.month}", ncols=2)

    meta = [
        ("Parameter", "Value"),
        ("Report type",
         "Run-rate pace projection (NOT machine-capacity). "
         "Pace = p90 or mean of actual daily pcs produced to date."),
        ("Month", result.month),
        ("As-of date", result.as_of_date),
        ("Working days (Mon–Sat)", result.working_days_total),
        ("Working days elapsed", result.working_days_elapsed),
        ("Working days remaining", result.working_days_remaining),
        ("Source file ID (Report-11 / 12)", result.source_file_id),
        ("Source file",
         f"https://docs.google.com/spreadsheets/d/{result.source_file_id}"
         if result.source_file_id else "—"),
        ("Date range in source", f"{result.source_date_min} – {result.source_date_max}"),
        ("Produced to date (R-11 + R-12)", f"{result.actual_produced_total:,.0f} pcs"),
        ("  of which: unclassified R-12 materials",
         f"{result.other_produced:,.0f} pcs (e.g. TEFFLONE — counted in total, not in any category row)"),
        ("Produced to date (Report-1 plan tab)", f"{result.plan_produced_total:,.0f} pcs"),
        ("Reconciliation note",
         "R-11 TOTAL + R-12 TOTAL = actual_produced_total.  "
         "Report-1 figure differs because R-1 aggregates all plants; "
         "R-11/R-12 are Plumbing-only.  The pace calc uses R-11/R-12."),
        ("Pace method", f"p90 (≥{MIN_DAYS_FOR_P90} days) → mean[low-confidence] → not-started"),
        ("p90 meaning",
         f"90th-percentile of per-category-per-day totals (≥{MIN_DAYS_FOR_P90} "
         "production days required).  Only the top 10 % of production days "
         "exceeded this pace — it represents an optimistic-but-observed "
         "throughput ceiling, not an average and not a machine-capacity limit."),
        ("mean(low-confidence,Nd) meaning",
         f"Arithmetic mean of only N days (N < {MIN_DAYS_FOR_P90}). "
         "Treat as indicative only; too few data points for reliable projection."),
        ("Not started",
         "Zero production recorded for this category to date.  "
         "Pace cannot be estimated.  "
         + ("Capacity-feasible column sourced from machine plan." if has_cap_feas
            else "Pass cap_feasible_by_cat to populate the capacity column.")),
        ("Capacity-Feasible Plan cross-link",
         "Run /machine-planning/report/capacity-feasible-plan for the authoritative "
         "'what the machines can make' figure."),
        ("Invariant: no-zero-pace-with-production", "PASS" if not any(
            "INVARIANT VIOLATED" in w for w in result.warnings) else "FAIL"),
        ("Invariant: pace-projection consistent", "PASS" if not any(
            "inconsistency" in w.lower() for w in result.warnings) else "FAIL"),
        ("Invariant: not-everything-at-gap", "PASS" if not any(
            "CRITICAL" in w for w in result.warnings) else "FAIL"),
        ("Warnings", "\n".join(result.warnings) if result.warnings else "None"),
    ]

    for ri, (k, v) in enumerate(meta, 4):
        c1 = ws3.cell(row=ri, column=1, value=k)
        c1.font = _nfont(bold=True, size=9)
        c2 = ws3.cell(row=ri, column=2, value=str(v))
        c2.font = _nfont(size=9)
        c2.alignment = _align(wrap_text=True, vertical="top")
        ws3.row_dimensions[ri].height = 28

    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 80

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

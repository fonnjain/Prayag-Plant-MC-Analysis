"""
Tests for mp_reports.py — covers:
  FIX 1  Weight columns (production weight ≠ material req)
  FIX 2  Schedule-based rows (WEEK/SHIFT populated from ScheduleResult)
  FIX 3  Rate fallback tier populated in ItemResult
  FIX 4  Report-11A-D machine-group filter
  NEW    consolidated_plan_bytes() smoke test (7 tabs, no crash)
"""
import io
import pytest

# ── Minimal stubs so the tests never need a live DB ──────────────────────────

def _make_item(
    item_code="CPVC-100",
    material="CPVC",
    qty_pcs=1000.0,
    wt_per_pc=0.3,
    material_kg=312.0,   # 1000 × 0.3 × 1.04
    rate_kg_per_hr=150.0,
    machine="M/C-1",
    rate_fallback_tier="item",
):
    from mp_engine import ItemResult, AssignedPortion
    assignment_hrs = material_kg / rate_kg_per_hr
    a = AssignedPortion(
        machine=machine,
        hrs=round(assignment_hrs, 3),
        material_kg=material_kg,
        qty_pcs=qty_pcs,
    )
    return ItemResult(
        item_code=item_code,
        raw_code=item_code,
        material=material,
        qty_pcs=qty_pcs,
        weight_per_pc_kg=wt_per_pc,
        material_kg=material_kg,
        fresh_compound_kg=material_kg * 0.75,
        pulverizer_kg=material_kg * 0.25,
        rate_kg_per_hr=rate_kg_per_hr,
        rate_estimated=(rate_fallback_tier != "item"),
        rate_fallback_tier=rate_fallback_tier,
        machine_hrs=round(assignment_hrs, 3),
        capable_machines=[machine],
        assignments=[a],
        has_weight=True,
        has_machine=True,
    )


def _make_engine_result(items=None):
    from mp_engine import EngineResult, CoverageGaps, PlanTotals, MachineLoad
    items = items or [_make_item()]
    ml = MachineLoad(
        machine="M/C-1", capacity_hrs=500.0,
        assigned_hrs=sum(it.machine_hrs for it in items),
        utilisation_pct=50.0, machine_days=13.0,
        material_kg=sum(it.material_kg for it in items),
        fresh_compound_kg=sum(it.fresh_compound_kg for it in items),
        pulverizer_kg=sum(it.pulverizer_kg for it in items),
        staffing_ok=True, operators_ot=2, support_w=1,
    )
    return EngineResult(
        segment="PIPE",
        effective_month="2026-07",
        items=items,
        machine_loads=[ml],
        coverage_gaps=CoverageGaps(
            no_weight=[], no_machine=[], idle_machines=[], locked_out_machines=[]
        ),
        totals=PlanTotals(
            total_qty_pcs=sum(it.qty_pcs for it in items),
            total_material_kg=sum(it.material_kg for it in items),
            total_fresh_compound_kg=sum(it.fresh_compound_kg for it in items),
            total_pulverizer_kg=sum(it.pulverizer_kg for it in items),
            routable_material_kg=sum(it.material_kg for it in items),
            routable_fresh_compound_kg=sum(it.fresh_compound_kg for it in items),
            routable_pulverizer_kg=sum(it.pulverizer_kg for it in items),
        ),
        baseline_machine_loads=[ml],
        params_used={"waste_pct": 4.0, "pulverizer_pct": 25.0},
        effective_costs={"CPVC": 120.0},
        cost_by_material={"CPVC": 28080.0},
        n_unpriced=0,
    )


def _make_schedule_result(item_code="CPVC-100", machine="M/C-1"):
    from mp_scheduler import ScheduleResult, ShiftBlock, WeekFillRow
    blocks = [
        ShiftBlock(
            week=1, day=1, machine=machine, shift="DAY",
            item_code=item_code, raw_code=item_code, material="CPVC",
            planned_hours=10.0, excess_hours=0.0, origin_week=1, is_idle=False,
        ),
        ShiftBlock(
            week=1, day=1, machine=machine, shift="NIGHT",
            item_code=item_code, raw_code=item_code, material="CPVC",
            planned_hours=10.0, excess_hours=0.5, origin_week=1, is_idle=False,
        ),
        # Idle block — should be excluded from report rows
        ShiftBlock(
            week=1, day=2, machine=machine, shift="DAY",
            item_code="", raw_code="", material="",
            planned_hours=10.0, excess_hours=0.0, origin_week=0, is_idle=True,
        ),
    ]
    wf = WeekFillRow(
        week=1, machine=machine, capacity_hrs=125.0,
        scheduled_hrs=19.5, idle_hrs=10.0, utilisation_pct=15.6,
        changeovers=1, excess_kg=75.0, origin_breakdown={1: 19.5},
    )
    return ScheduleResult(
        segment="PIPE",
        effective_month="2026-07",
        blocks=blocks,
        weekly_fill=[wf],
        unfinished=[],
        total_capacity_hrs=500.0,
        total_scheduled_hrs=19.5,
        total_idle_hrs=10.0,
        total_excess_kg=75.0,
        total_changeovers=1,
        week_days=[6, 6, 6, 7],
        params_used={"min_run_block_hours": 5.0, "week_days": [6, 6, 6, 7]},
    )


# ── FIX 3: Weight column correctness ─────────────────────────────────────────

def test_production_weight_no_waste():
    """Production Wt (col 8) = qty_pcs × wt/pc; NOT equal to material_kg."""
    import mp_reports as r
    item = _make_item(qty_pcs=1000.0, wt_per_pc=0.3, material_kg=312.0)
    result = _make_engine_result([item])
    rows = r._build_rows_from_assignments(result)
    assert rows, "Expected at least one row"
    row = rows[0]
    prod_wt   = row[8]    # Production Wt (KG)
    mat_req   = row[9]    # Material Req (KG)
    expected_prod = round(1000.0 * 0.3, 1)    # 300.0
    expected_mat  = round(312.0, 1)            # 312.0  (from assignment.material_kg)
    assert abs(prod_wt - expected_prod) < 0.11, f"prod_wt={prod_wt}, expected≈{expected_prod}"
    assert abs(mat_req - expected_mat) < 0.11, f"mat_req={mat_req}, expected≈{expected_mat}"
    assert prod_wt != mat_req, "Production Wt and Material Req must differ (waste factor)"


def test_weight_columns_differ_when_waste_applied():
    """Verify the two weight columns differ by exactly the waste factor."""
    import mp_reports as r
    item = _make_item(qty_pcs=500.0, wt_per_pc=0.5, material_kg=260.0)  # 4% waste
    result = _make_engine_result([item])
    rows = r._build_rows_from_assignments(result)
    row = rows[0]
    prod_wt = row[8]
    mat_req = row[9]
    ratio = mat_req / prod_wt if prod_wt > 0 else 0
    # ratio should be ≈ 1.04 (waste factor)
    assert abs(ratio - 1.04) < 0.01, f"mat_req/prod_wt ratio={ratio:.4f}, expected≈1.04"


# ── FIX 1: Schedule-based rows ────────────────────────────────────────────────

def test_schedule_rows_have_week_and_shift():
    """When ScheduleResult is provided, WEEK and SHIFT columns must be populated."""
    import mp_reports as r
    result   = _make_engine_result()
    schedule = _make_schedule_result()
    rows = r._build_rows_from_schedule(result, schedule)
    assert rows, "Expected at least one scheduled row"
    for row in rows:
        week_val  = row[1]   # WEEK
        shift_val = row[2]   # SHIFT
        assert week_val.startswith("W"), f"WEEK should start with 'W', got {week_val!r}"
        assert shift_val in ("DAY", "NIGHT"), f"SHIFT should be DAY/NIGHT, got {shift_val!r}"


def test_schedule_rows_exclude_idle():
    """Idle blocks must not appear as data rows."""
    import mp_reports as r
    result   = _make_engine_result()
    schedule = _make_schedule_result()
    rows = r._build_rows_from_schedule(result, schedule)
    # Our stub schedule has 2 non-idle blocks (day1 DAY + day1 NIGHT) and 1 idle
    assert len(rows) == 2, f"Expected 2 non-idle rows, got {len(rows)}"


def test_schedule_rows_production_hours_correct():
    """Running Hours in schedule row = planned_hours - excess_hours."""
    import mp_reports as r
    result   = _make_engine_result()
    schedule = _make_schedule_result()
    rows = r._build_rows_from_schedule(result, schedule)
    # Night block: planned=10.0, excess=0.5 → prod_hrs=9.5
    # Day block: planned=10.0, excess=0.0 → prod_hrs=10.0
    hrs_values = {row[7] for row in rows}   # Running Hours col
    assert 9.5 in hrs_values, f"Expected 9.5 hr (night block prod hrs), got {hrs_values}"
    assert 10.0 in hrs_values, f"Expected 10.0 hr (day block prod hrs), got {hrs_values}"


def test_dispatch_uses_schedule_when_provided():
    """_build_rows dispatches to schedule path when schedule arg given."""
    import mp_reports as r
    result   = _make_engine_result()
    schedule = _make_schedule_result()
    rows_scheduled  = r._build_rows(result, schedule=schedule)
    rows_assignment = r._build_rows(result, schedule=None)
    # With schedule: DATE = day number (int); without: DATE = month string
    assert rows_scheduled[0][0] == 1, "Scheduled DATE should be day number 1"
    assert isinstance(rows_assignment[0][0], str), "Assignment DATE should be string"


# ── FIX 2 (11A-D) machine filter ─────────────────────────────────────────────

def test_machine_filter_excludes_other_machines():
    """report_11x_bytes filters to only the specified machine group."""
    import mp_reports as r
    items = [
        _make_item("CPVC-100", machine="M/C-1", qty_pcs=1000.0),
        _make_item("UPVC-200", material="UPVC", machine="M/C-3",
                   qty_pcs=500.0, wt_per_pc=0.4, material_kg=208.0, rate_kg_per_hr=140.0),
    ]
    result = _make_engine_result(items)
    rows_A = r._build_rows_from_assignments(result, machine_filter={"M/C-1", "M/C-2"})
    rows_B = r._build_rows_from_assignments(result, machine_filter={"M/C-3", "M/C-4"})
    mc_in_A = {row[3] for row in rows_A}   # MACHINE NAME col
    mc_in_B = {row[3] for row in rows_B}
    assert "M/C-1" in mc_in_A and "M/C-3" not in mc_in_A
    assert "M/C-3" in mc_in_B and "M/C-1" not in mc_in_B


# ── FIX 4: rate_fallback_tier ────────────────────────────────────────────────

def test_rate_fallback_tier_seeded():
    from mp_engine import _get_rate
    ph = {"CPVC-100": 150.0}
    mat_avg = {"CPVC": 140.0}
    rate, estimated, tier = _get_rate("CPVC-100", "CPVC", ph, mat_avg, 120.0)
    assert tier == "item"
    assert not estimated
    assert rate == 150.0


def test_rate_fallback_tier_mat_avg():
    from mp_engine import _get_rate
    ph = {}
    mat_avg = {"CPVC": 140.0}
    rate, estimated, tier = _get_rate("CPVC-999", "CPVC", ph, mat_avg, 120.0)
    assert tier == "mat_avg"
    assert estimated
    assert rate == 140.0


def test_rate_fallback_tier_overall_avg():
    from mp_engine import _get_rate
    ph = {}
    mat_avg = {}
    rate, estimated, tier = _get_rate("SWR-999", "SWR", ph, mat_avg, 120.0)
    assert tier == "overall_avg"
    assert estimated
    assert rate == 120.0


# ── Report generation smoke tests ─────────────────────────────────────────────

def test_report_11_bytes_no_schedule():
    """report_11_bytes without schedule runs and returns valid xlsx bytes."""
    import mp_reports as r
    result = _make_engine_result()
    data = r.report_11_bytes(result, schedule=None)
    assert data[:4] == b"PK\x03\x04", "xlsx should start with PK header (ZIP)"
    assert len(data) > 3000


def test_report_11_bytes_with_schedule():
    """report_11_bytes with schedule returns valid xlsx (schedule-mode rows)."""
    import mp_reports as r
    result   = _make_engine_result()
    schedule = _make_schedule_result()
    data = r.report_11_bytes(result, schedule=schedule)
    assert data[:4] == b"PK\x03\x04"
    assert len(data) > 3000


def test_report_11x_bytes_group_filter():
    """report_11x_bytes with group A returns valid xlsx."""
    import mp_reports as r
    result = _make_engine_result()
    data = r.report_11x_bytes(result, "A")
    assert data[:4] == b"PK\x03\x04"


def test_consolidated_plan_bytes_all_tabs():
    """consolidated_plan_bytes returns a workbook with all 7 tabs, no crash."""
    import mp_reports as r
    from openpyxl import load_workbook
    result   = _make_engine_result([
        _make_item("CPVC-100", rate_fallback_tier="item"),
        _make_item("SWR-200", material="SWR", rate_fallback_tier="mat_avg",
                   qty_pcs=800.0, wt_per_pc=0.25, material_kg=208.0,
                   rate_kg_per_hr=130.0, machine="M/C-2"),
        _make_item("AGRI-300", material="AGRI", rate_fallback_tier="overall_avg",
                   qty_pcs=600.0, wt_per_pc=0.2, material_kg=124.8,
                   rate_kg_per_hr=110.0, machine="M/C-3"),
    ])
    schedule = _make_schedule_result()
    data = r.consolidated_plan_bytes(
        engine_result=result,
        fitting_result=None,
        schedule_result=schedule,
    )
    assert data[:4] == b"PK\x03\x04"
    wb = load_workbook(io.BytesIO(data))
    assert len(wb.sheetnames) == 7, f"Expected 7 tabs, got {wb.sheetnames}"
    expected_prefixes = ["1.", "2.", "3.", "4.", "5.", "6.", "7."]
    for prefix, name in zip(expected_prefixes, wb.sheetnames):
        assert name.startswith(prefix), f"Tab {name!r} should start with {prefix!r}"


def test_consolidated_machine_load_tab_has_demand_and_scheduled_columns():
    """Tab '2. Machine Load' must have both Demand and Scheduled column headers."""
    import mp_reports as r
    from openpyxl import load_workbook
    result   = _make_engine_result()
    schedule = _make_schedule_result()
    data = r.consolidated_plan_bytes(
        engine_result=result,
        fitting_result=None,
        schedule_result=schedule,
    )
    wb = load_workbook(io.BytesIO(data))
    ws = wb["2. Machine Load"]
    # Row 4 is the column-header row; collect all header text
    headers = {
        ws.cell(row=4, column=c).value
        for c in range(1, 20)
        if ws.cell(row=4, column=c).value
    }
    assert any("Demand" in (h or "") for h in headers), (
        f"Expected a 'Demand' column in tab 2 headers; got: {headers}"
    )
    assert any("Scheduled" in (h or "") for h in headers), (
        f"Expected a 'Scheduled' column in tab 2 headers; got: {headers}"
    )
    # Old label must be gone — no plain "Assigned (hrs)" or "Utilisation %"
    assert "Assigned (hrs)" not in headers, "Old 'Assigned (hrs)' label must be renamed to 'Demand (hrs)'"
    assert "Utilisation %" not in headers,  "Old 'Utilisation %' label must be renamed to 'Demand util %'"


def test_consolidated_machine_load_scheduled_util_never_exceeds_100():
    """Scheduled util % must be ≤ 100% for every machine row in tab '2. Machine Load'.

    The scheduler enforces a capacity ceiling so this must always hold.  A value
    > 100 here would contradict the Capacity-Feasible Plan — the whole point of
    adding the scheduled column.
    """
    import mp_reports as r
    from openpyxl import load_workbook
    from mp_scheduler import ScheduleResult, WeekFillRow, ShiftBlock

    # Build a schedule where one machine runs at exactly full capacity (100%)
    # and another is partially used (50%).
    wf1 = WeekFillRow(
        week=1, machine="M/C-1", capacity_hrs=500.0,
        scheduled_hrs=500.0, idle_hrs=0.0, utilisation_pct=100.0,
        changeovers=0, excess_kg=0.0, origin_breakdown={1: 500.0},
    )
    wf2 = WeekFillRow(
        week=1, machine="M/C-2", capacity_hrs=400.0,
        scheduled_hrs=200.0, idle_hrs=200.0, utilisation_pct=50.0,
        changeovers=0, excess_kg=0.0, origin_breakdown={1: 200.0},
    )
    schedule = ScheduleResult(
        segment="PIPE", effective_month="2026-08",
        blocks=[], weekly_fill=[wf1, wf2], unfinished=[],
        total_capacity_hrs=900.0, total_scheduled_hrs=700.0,
        total_idle_hrs=200.0, total_excess_kg=0.0,
        total_changeovers=0, week_days=[6, 6, 6, 7], params_used={},
    )
    from mp_engine import EngineResult, MachineLoad, CoverageGaps, PlanTotals
    ml1 = MachineLoad(
        machine="M/C-1", capacity_hrs=500.0,
        assigned_hrs=750.0,  # demand > capacity (150%)
        utilisation_pct=150.0, machine_days=25.0,
        material_kg=10000.0, fresh_compound_kg=7500.0, pulverizer_kg=2500.0,
        staffing_ok=True, operators_ot=0, support_w=0,
    )
    ml2 = MachineLoad(
        machine="M/C-2", capacity_hrs=400.0,
        assigned_hrs=590.0,  # demand > capacity (147.5%)
        utilisation_pct=147.5, machine_days=20.0,
        material_kg=8000.0, fresh_compound_kg=6000.0, pulverizer_kg=2000.0,
        staffing_ok=True, operators_ot=0, support_w=0,
    )
    item = _make_item()
    result = EngineResult(
        segment="PIPE", effective_month="2026-08",
        items=[item], machine_loads=[ml1, ml2],
        coverage_gaps=CoverageGaps(no_weight=[], no_machine=[], idle_machines=[], locked_out_machines=[]),
        totals=PlanTotals(
            total_qty_pcs=1000.0, total_material_kg=18000.0,
            total_fresh_compound_kg=13500.0, total_pulverizer_kg=4500.0,
            routable_material_kg=18000.0, routable_fresh_compound_kg=13500.0,
            routable_pulverizer_kg=4500.0,
        ),
        baseline_machine_loads=[ml1, ml2],
        params_used={}, effective_costs={}, cost_by_material={}, n_unpriced=0,
    )

    data = r.consolidated_plan_bytes(
        engine_result=result, fitting_result=None, schedule_result=schedule,
    )
    wb = load_workbook(io.BytesIO(data))
    ws = wb["2. Machine Load"]

    # Locate "Scheduled util %" column
    hdr_row = 4
    sched_util_col = None
    for c in range(1, 20):
        v = ws.cell(row=hdr_row, column=c).value or ""
        if "Scheduled util" in v:
            sched_util_col = c
            break
    assert sched_util_col is not None, "Could not find 'Scheduled util %' column"

    # Collect scheduled util values from data rows (row 5 onwards, stop at TOTAL)
    sched_utils = []
    for row in range(hdr_row + 1, hdr_row + 20):
        machine_val = ws.cell(row=row, column=1).value
        if machine_val == "TOTAL" or machine_val is None:
            break
        util_val = ws.cell(row=row, column=sched_util_col).value
        if util_val is not None:
            sched_utils.append(float(util_val))

    assert sched_utils, "No scheduled-util data rows found"
    for u in sched_utils:
        assert u <= 100.0 + 0.01, (
            f"Scheduled util % {u} exceeds 100% — breaks capacity guarantee"
        )

    # Also verify demand util IS above 100% for M/C-1 (so the two columns are distinct)
    demand_util_col = None
    for c in range(1, 20):
        v = ws.cell(row=hdr_row, column=c).value or ""
        if "Demand util" in v:
            demand_util_col = c
            break
    assert demand_util_col is not None
    demand_utils = []
    for row in range(hdr_row + 1, hdr_row + 20):
        machine_val = ws.cell(row=row, column=1).value
        if machine_val == "TOTAL" or machine_val is None:
            break
        v = ws.cell(row=row, column=demand_util_col).value
        if v is not None:
            demand_utils.append(float(v))
    assert any(u > 100.0 for u in demand_utils), (
        "Expected at least one machine with demand util > 100% in this test fixture"
    )


def test_consolidated_machine_load_scheduled_values_match_schedule_result():
    """Scheduled hrs in tab 2 must equal the sum of weekly_fill.scheduled_hrs per machine.

    This pins the exact numbers so that the consolidated report and the
    Capacity-Feasible Plan (which derives sched_load_by_mc the same way) are
    guaranteed to agree.
    """
    import mp_reports as r
    from openpyxl import load_workbook
    from mp_scheduler import ScheduleResult, WeekFillRow

    # Two machines, two weeks of weekly_fill each
    # M/C-1: W1=120.0 + W2=130.0 = 250.0 total scheduled hrs
    # M/C-2: W1=80.0  + W2=70.0  = 150.0 total scheduled hrs
    wf_rows = [
        WeekFillRow(week=1, machine="M/C-1", capacity_hrs=250.0,
                    scheduled_hrs=120.0, idle_hrs=130.0, utilisation_pct=48.0,
                    changeovers=0, excess_kg=0.0, origin_breakdown={1: 120.0}),
        WeekFillRow(week=2, machine="M/C-1", capacity_hrs=250.0,
                    scheduled_hrs=130.0, idle_hrs=120.0, utilisation_pct=52.0,
                    changeovers=0, excess_kg=0.0, origin_breakdown={2: 130.0}),
        WeekFillRow(week=1, machine="M/C-2", capacity_hrs=200.0,
                    scheduled_hrs=80.0, idle_hrs=120.0, utilisation_pct=40.0,
                    changeovers=0, excess_kg=0.0, origin_breakdown={1: 80.0}),
        WeekFillRow(week=2, machine="M/C-2", capacity_hrs=200.0,
                    scheduled_hrs=70.0, idle_hrs=130.0, utilisation_pct=35.0,
                    changeovers=0, excess_kg=0.0, origin_breakdown={2: 70.0}),
    ]
    schedule = ScheduleResult(
        segment="PIPE", effective_month="2026-08",
        blocks=[], weekly_fill=wf_rows, unfinished=[],
        total_capacity_hrs=900.0, total_scheduled_hrs=400.0,
        total_idle_hrs=500.0, total_excess_kg=0.0, total_changeovers=0,
        week_days=[6, 6, 6, 7], params_used={},
    )
    from mp_engine import EngineResult, MachineLoad, CoverageGaps, PlanTotals
    ml1 = MachineLoad(
        machine="M/C-1", capacity_hrs=500.0, assigned_hrs=300.0,
        utilisation_pct=60.0, machine_days=25.0,
        material_kg=5000.0, fresh_compound_kg=3750.0, pulverizer_kg=1250.0,
        staffing_ok=True, operators_ot=0, support_w=0,
    )
    ml2 = MachineLoad(
        machine="M/C-2", capacity_hrs=400.0, assigned_hrs=200.0,
        utilisation_pct=50.0, machine_days=20.0,
        material_kg=4000.0, fresh_compound_kg=3000.0, pulverizer_kg=1000.0,
        staffing_ok=True, operators_ot=0, support_w=0,
    )
    item = _make_item()
    result = EngineResult(
        segment="PIPE", effective_month="2026-08",
        items=[item], machine_loads=[ml1, ml2],
        coverage_gaps=CoverageGaps(no_weight=[], no_machine=[], idle_machines=[], locked_out_machines=[]),
        totals=PlanTotals(
            total_qty_pcs=1000.0, total_material_kg=9000.0,
            total_fresh_compound_kg=6750.0, total_pulverizer_kg=2250.0,
            routable_material_kg=9000.0, routable_fresh_compound_kg=6750.0,
            routable_pulverizer_kg=2250.0,
        ),
        baseline_machine_loads=[ml1, ml2],
        params_used={}, effective_costs={}, cost_by_material={}, n_unpriced=0,
    )
    data = r.consolidated_plan_bytes(
        engine_result=result, fitting_result=None, schedule_result=schedule,
    )
    wb = load_workbook(io.BytesIO(data))
    ws = wb["2. Machine Load"]

    # Locate Scheduled (hrs) column
    hdr_row = 4
    sched_col = None
    for c in range(1, 20):
        v = ws.cell(row=hdr_row, column=c).value or ""
        if v == "Scheduled (hrs)":
            sched_col = c
            break
    assert sched_col is not None, "Could not find 'Scheduled (hrs)' column"

    # Read machine → scheduled hrs from the report
    sched_in_report = {}
    for row in range(hdr_row + 1, hdr_row + 20):
        mc = ws.cell(row=row, column=1).value
        if mc == "TOTAL" or mc is None:
            break
        sched_in_report[mc] = float(ws.cell(row=row, column=sched_col).value or 0)

    assert abs(sched_in_report.get("M/C-1", -1) - 250.0) < 0.1, (
        f"M/C-1 scheduled hrs should be 250.0 (120+130), got {sched_in_report.get('M/C-1')}"
    )
    assert abs(sched_in_report.get("M/C-2", -1) - 150.0) < 0.1, (
        f"M/C-2 scheduled hrs should be 150.0 (80+70), got {sched_in_report.get('M/C-2')}"
    )


def test_consolidated_machine_load_note_present():
    """Tab '2. Machine Load' row 3 must contain the demand-vs-scheduled clarifying note."""
    import mp_reports as r
    from openpyxl import load_workbook
    result   = _make_engine_result()
    schedule = _make_schedule_result()
    data = r.consolidated_plan_bytes(
        engine_result=result, fitting_result=None, schedule_result=schedule,
    )
    wb   = load_workbook(io.BytesIO(data))
    ws   = wb["2. Machine Load"]
    note = ws["A3"].value or ""
    assert "100%" in note, f"Note should mention 100%; got: {note!r}"
    assert "Demand" in note or "demand" in note, f"Note should mention demand; got: {note!r}"
    assert "Scheduled" in note or "scheduled" in note, f"Note should mention scheduled; got: {note!r}"
    assert "Capacity-Feasible" in note or "capacity" in note.lower(), (
        f"Note should reference the Capacity-Feasible Plan; got: {note!r}"
    )


def test_consolidated_other_tabs_unchanged():
    """The six tabs other than '2. Machine Load' must still be present and named correctly."""
    import mp_reports as r
    from openpyxl import load_workbook
    result   = _make_engine_result()
    schedule = _make_schedule_result()
    data = r.consolidated_plan_bytes(
        engine_result=result, fitting_result=None, schedule_result=schedule,
    )
    wb = load_workbook(io.BytesIO(data))
    assert len(wb.sheetnames) == 7, f"Still expect 7 tabs; got {wb.sheetnames}"
    for prefix in ["1.", "3.", "4.", "5.", "6.", "7."]:
        assert any(n.startswith(prefix) for n in wb.sheetnames), (
            f"Tab starting with {prefix!r} missing; sheets: {wb.sheetnames}"
        )


def test_consolidated_plan_bytes_no_data_graceful():
    """consolidated_plan_bytes handles all-None inputs without crashing."""
    import mp_reports as r
    data = r.consolidated_plan_bytes(
        engine_result=None, fitting_result=None, schedule_result=None
    )
    assert data[:4] == b"PK\x03\x04"


def test_rate_fallback_note_only_when_estimated():
    """Rate fallback note section should NOT appear when all items have seeded rates."""
    import mp_reports as r
    from openpyxl import load_workbook
    result = _make_engine_result([
        _make_item("CPVC-100", rate_fallback_tier="item"),
    ])
    data = r.report_11_bytes(result, schedule=None)
    wb   = load_workbook(io.BytesIO(data))
    ws   = wb.active
    # Look for the fallback note title text anywhere in the sheet
    texts = [
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    ]
    fallback_found = any("Rate Fallback" in t for t in texts)
    assert not fallback_found, "No fallback note expected when all items have seeded rates"


def test_rate_fallback_note_present_when_estimated():
    """Rate fallback note section SHOULD appear when some items use estimated rates."""
    import mp_reports as r
    from openpyxl import load_workbook
    result = _make_engine_result([
        _make_item("SWR-999", material="SWR", rate_fallback_tier="overall_avg",
                   qty_pcs=500.0, wt_per_pc=0.3, material_kg=156.0,
                   rate_kg_per_hr=120.0, machine="M/C-1"),
    ])
    data = r.report_11_bytes(result, schedule=None)
    wb   = load_workbook(io.BytesIO(data))
    ws   = wb.active
    texts = [
        str(ws.cell(row=r_i, column=1).value or "")
        for r_i in range(1, ws.max_row + 1)
    ]
    fallback_found = any("Rate Fallback" in t for t in texts)
    assert fallback_found, "Fallback note should appear when estimated items exist"

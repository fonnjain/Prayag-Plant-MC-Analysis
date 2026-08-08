"""tests/test_corrective_replan.py
Regression assertions for the Corrective Re-plan engine.

The six assertions from the spec
---------------------------------
1. No-zero-capacity-with-production: producedToDate > 0 → cap_per_day > 0
2. Few-days fallback: with 1-2 non-zero production days, method=mean, cap > 0
3. Feasibility consistency: feasible == cap_per_day × working_days_remaining (exact)
4. Not-everything-unfulfillable: total shortfall < total remaining when production exists
5. Source agreement: file_id + date range come from the parse, not injected stale values
6. Date-format guard: "Aug 1, 2026" and "1-Aug-2026" both parse; unknown format raises

Additional tests
----------------
7. p90 path: ≥ MIN_DAYS_FOR_P90 non-zero days → method="p90"
8. NO_DEMONSTRATED_CAPACITY: solvent / truly-zero categories correctly flagged
9. Category ordering matches CATEGORY_ORDER (not alphabetical, not file order)
10. Invariant warnings surfaced, not silently swallowed
"""
from __future__ import annotations

import copy
import datetime
import types
import unittest
from unittest.mock import patch

import mp_corrective_replan as mcr


# ---------------------------------------------------------------------------
# Helpers to build synthetic Report-11 / Report-12 value grids
# ---------------------------------------------------------------------------

def _r11(rows):
    """Build a minimal Report-11 value grid (header + data rows).

    Each row in *rows* is (date_str, types, item_code, pcs).
    """
    hdr = ["", "DATE", "MACHINE NAME", "MACHINE NO.", "TYPES", "ITEM CODE",
           "Running Hour", "Ideal Weight (KG)", "Pcs", "Weight"]
    values = [
        ["url row"],
        ["", "TOTAL", "", "", "", "", "625"],
        ["", "PIPE M/C"],
        ["", "", "", "", "", "", ""],
        hdr,
    ]
    for date, types_, item, pcs in rows:
        values.append(["", date, "CON-63-1", "PIPE M/C - 1", types_, item,
                        "8", "1.0", str(pcs), "100"])
    return values


def _r12(rows):
    """Build a minimal Report-12 value grid (header + data rows).

    Each row in *rows* is (date_str, material, item_code, pcs).
    """
    hdr = ["DATE", "MATERIAL", "Item Code", "SAP Code", "Moulding Machine",
           "Run Cavity", "Mould Cavity (F)", "Cycle Time Per Pcs Standard (F)",
           "Output Production", "", "Weight per Pc"]
    sub = ["", "", "", "", "", "", "", "", " Pc ", " Wt in Kgs "]
    values = [
        ["Mr. Jitendra"],
        ["Mould M/C"],
        ["TOTAL", "", "", "", "", "429"],
        hdr,
        sub,
    ]
    for date, material, item, pcs in rows:
        values.append([date, material, item, "", "A02(U-150)", "4", "4", "25.00",
                        str(pcs), "100"])
    return values


def _plan_rec(family, category, produce_required, produced, ideal_qty=0, closing_stock=0):
    """Build a minimal dict acting as a PlanRecord-like object."""
    class R:
        pass
    r = R()
    r.family = family
    r.category = category
    r.produce_required = produce_required
    r.produced = produced
    r.ideal_qty = ideal_qty
    r.closing_stock = closing_stock
    return r


def _run(r11_rows, r12_rows, plan_recs=None, month="2026-08", as_of="2026-08-08"):
    r11 = _r11(r11_rows)
    r12 = _r12(r12_rows)
    return mcr.compute_corrective_replan(
        month=month,
        plan_recs=plan_recs or [],
        r11_values=r11,
        r12_values=r12,
        as_of_date=as_of,
        file_id="TEST_FILE_ID",
    )


def _cat(result, category):
    return next(c for c in result.categories if c.category == category)


# ---------------------------------------------------------------------------
# 1. No-zero-capacity-with-production
# ---------------------------------------------------------------------------

class TestNoZeroCapacityWithProduction(unittest.TestCase):
    """Assertion 1 from spec: if producedToDate > 0, cap_per_day must be > 0.

    This was the root-cause bug: p90 with < threshold days returned 0 instead
    of falling back to mean.
    """

    def test_two_days_cpvc_pipe_gives_nonzero_cap(self):
        """Real-world August scenario: only 2 production days."""
        result = _run(
            r11_rows=[
                ("Aug 1, 2026", "CPVC", "PS-3",  2420),
                ("Aug 3, 2026", "CPVC", "PS-12", 3075),
            ],
            r12_rows=[],
        )
        cpvc_pipe = _cat(result, "CPVC Pipe")
        self.assertGreater(cpvc_pipe.produced_to_date, 0)
        self.assertGreater(cpvc_pipe.cap_per_day, 0,
            "Cap/Day must be > 0 when production exists (regression: bug returned 0)")

    def test_one_day_gives_nonzero_cap(self):
        result = _run(
            r11_rows=[("Aug 1, 2026", "SWR", "PW71", 1782)],
            r12_rows=[],
        )
        c = _cat(result, "SWR Pipe")
        self.assertGreater(c.cap_per_day, 0)
        # method should be mean-based (low-confidence when < MIN_DAYS_FOR_P90)
        self.assertTrue(c.method.startswith("mean"),
            f"Expected mean-based method, got {c.method!r}")
        self.assertTrue(c.low_confidence,
            "1 production day must set low_confidence=True")

    def test_no_invariant_warnings_with_2_days(self):
        result = _run(
            r11_rows=[
                ("Aug 1, 2026", "CPVC", "PS-3", 2420),
                ("Aug 3, 2026", "CPVC", "PS-3", 3075),
            ],
            r12_rows=[],
        )
        inv_warnings = [w for w in result.warnings if "INVARIANT VIOLATED" in w]
        self.assertEqual(inv_warnings, [],
            "Engine must not emit INVARIANT VIOLATED for 2-day scenario — "
            "it means the fallback chain failed")

    def test_category_with_zero_production_is_none(self):
        """AGRI Pipe with no rows should have cap_per_day=0, method='none'."""
        result = _run(r11_rows=[], r12_rows=[])
        agri = _cat(result, "AGRI Pipe")
        self.assertEqual(agri.cap_per_day, 0.0)
        self.assertEqual(agri.method, "none")
        self.assertTrue(agri.no_demonstrated_capacity)


# ---------------------------------------------------------------------------
# 2. Few-days fallback
# ---------------------------------------------------------------------------

class TestFewDaysFallback(unittest.TestCase):
    """Assertion 2: with 1-2 production days → method='mean', cap > 0."""

    def _assert_mean_low_conf(self, n_days):
        rows = [("Aug 1, 2026", "UPVC", f"PU-{i}", 3000) for i in range(n_days)]
        result = _run(r11_rows=rows, r12_rows=[])
        c = _cat(result, "UPVC Pipe")
        self.assertTrue(c.method.startswith("mean"),
            f"With {n_days} day(s), method must be mean-based, got {c.method!r}")
        self.assertTrue(c.low_confidence,
            f"With {n_days} day(s) (< MIN_DAYS_FOR_P90={mcr.MIN_DAYS_FOR_P90}), "
            "low_confidence must be True")
        self.assertGreater(c.cap_per_day, 0)

    def test_one_day(self):
        self._assert_mean_low_conf(1)

    def test_two_days(self):
        self._assert_mean_low_conf(2)

    def test_four_days_still_mean(self):
        # MIN_DAYS_FOR_P90 = 5, so 4 days → mean + low-confidence
        self._assert_mean_low_conf(min(4, mcr.MIN_DAYS_FOR_P90 - 1))

    def test_mean_value_correct_with_two_days(self):
        """mean([2420, 3075]) = 2747.5."""
        result = _run(
            r11_rows=[
                ("Aug 1, 2026", "CPVC", "PS-3",  2420),
                ("Aug 3, 2026", "CPVC", "PS-12", 3075),
            ],
            r12_rows=[],
        )
        cpvc = _cat(result, "CPVC Pipe")
        self.assertAlmostEqual(cpvc.cap_per_day, 2747.5, delta=1.0)

    def test_fitting_two_days(self):
        """CPVC Fitting with 2 days via Report-12: mean + low-confidence."""
        result = _run(
            r11_rows=[],
            r12_rows=[
                ("Aug 1, 2026", "CPVC", "U531", 25176),
                ("Aug 3, 2026", "CPVC", "U532", 19580),
            ],
        )
        c = _cat(result, "CPVC Fitting")
        self.assertTrue(c.method.startswith("mean"),
            f"2-day fitting should be mean-based, got {c.method!r}")
        self.assertTrue(c.low_confidence,
            "2 production days must set low_confidence=True")
        self.assertAlmostEqual(c.cap_per_day, 22378.0, delta=5.0)


# ---------------------------------------------------------------------------
# 3. Feasibility consistency
# ---------------------------------------------------------------------------

class TestFeasibilityConsistency(unittest.TestCase):
    """Assertion 3: feasible == cap_per_day × working_days_remaining (exact)."""

    def test_feasible_matches_cap_times_days(self):
        result = _run(
            r11_rows=[
                ("Aug 1, 2026", "CPVC", "PS-3",  2420),
                ("Aug 3, 2026", "CPVC", "PS-12", 3075),
            ],
            r12_rows=[],
        )
        wd = result.working_days_remaining
        for cat in result.categories:
            if cat.cap_per_day > 0:
                expected = round(cat.cap_per_day * wd, 1)
                self.assertAlmostEqual(
                    cat.feasible, expected, delta=0.15,
                    msg=f"{cat.category}: feasible={cat.feasible} ≠ "
                        f"cap({cat.cap_per_day}) × days({wd}) = {expected}"
                )

    def test_no_feasibility_inconsistency_warnings(self):
        result = _run(
            r11_rows=[
                ("Aug 1, 2026", "UPVC", "PU-1", 5000),
                ("Aug 3, 2026", "UPVC", "PU-2", 3000),
            ],
            r12_rows=[],
        )
        incon = [w for w in result.warnings if "inconsistency" in w.lower()]
        self.assertEqual(incon, [])

    def test_zero_cap_has_zero_feasible(self):
        result = _run(r11_rows=[], r12_rows=[])
        for cat in result.categories:
            if cat.cap_per_day == 0:
                self.assertEqual(cat.feasible, 0.0)


# ---------------------------------------------------------------------------
# 4. Not-everything-unfulfillable
# ---------------------------------------------------------------------------

class TestNotEverythingUnfulfillable(unittest.TestCase):
    """Assertion 4: total shortfall < total remaining when production exists."""

    def _run_with_demand(self, pipe_pcs, remaining_per_cat):
        plan = [
            _plan_rec("CPVC", "CPVC PIPE", remaining_per_cat + pipe_pcs, pipe_pcs),
        ]
        result = _run(
            r11_rows=[
                ("Aug 1, 2026", "CPVC", "PS-1", pipe_pcs),
                ("Aug 3, 2026", "CPVC", "PS-2", pipe_pcs),
            ],
            r12_rows=[],
            plan_recs=plan,
        )
        return result

    def test_shortfall_less_than_remaining(self):
        result = self._run_with_demand(pipe_pcs=3000, remaining_per_cat=2_000_000)
        total_rem = sum(c.remaining for c in result.categories)
        total_sf  = sum(c.shortfall for c in result.categories)
        self.assertLess(
            total_sf, total_rem,
            "Total shortfall must be < total remaining when production exists"
        )

    def test_no_critical_warning_with_feasible_output(self):
        # When feasible > 0 and remaining > 0 and feasible < remaining,
        # the CRITICAL invariant warning must NOT fire.
        plan = [_plan_rec("CPVC", "CPVC PIPE", 200_000, 0)]
        result = _run(
            r11_rows=[("Aug 1, 2026", "CPVC", "PS-1", 5000)],
            r12_rows=[],
            plan_recs=plan,
        )
        critical = [w for w in result.warnings if "CRITICAL" in w]
        self.assertEqual(critical, [])


# ---------------------------------------------------------------------------
# 5. Source agreement
# ---------------------------------------------------------------------------

class TestSourceAgreement(unittest.TestCase):
    """Assertion 5: file_id and date range are from the parse, correctly reported."""

    def test_file_id_in_provenance(self):
        r11 = _r11([("Aug 1, 2026", "CPVC", "PS-3", 2420)])
        r12 = _r12([])
        result = mcr.compute_corrective_replan(
            month="2026-08",
            plan_recs=[],
            r11_values=r11,
            r12_values=r12,
            as_of_date="2026-08-08",
            file_id="SENTINEL_FILE_ID_12345",
        )
        self.assertEqual(result.source_file_id, "SENTINEL_FILE_ID_12345")

    def test_date_range_min_max_from_data(self):
        result = _run(
            r11_rows=[
                ("Aug 1, 2026", "CPVC", "PS-3", 2420),
                ("Aug 3, 2026", "CPVC", "PS-12", 3075),
            ],
            r12_rows=[],
        )
        self.assertEqual(result.source_date_min, "2026-08-01")
        self.assertEqual(result.source_date_max, "2026-08-03")

    def test_producedToDate_from_r11_r12_not_plan(self):
        """actual_produced_total must reflect R11+R12 parse, not plan 'produced'."""
        plan = [_plan_rec("CPVC", "CPVC PIPE", 500_000, 205_566)]  # plan says 205k
        result = _run(
            r11_rows=[
                ("Aug 1, 2026", "CPVC", "PS-3", 2420),   # 2420 from R11
                ("Aug 3, 2026", "CPVC", "PS-3", 3075),   # 3075 from R11
            ],
            r12_rows=[],
            plan_recs=plan,
        )
        # actual_produced_total = 2420 + 3075 = 5495, NOT 205_566
        self.assertAlmostEqual(result.actual_produced_total, 5495, delta=1)
        # plan_produced_total = 205_566 (from plan rec)
        self.assertAlmostEqual(result.plan_produced_total, 205_566, delta=1)

    def test_working_days_remaining_is_positive(self):
        result = _run(r11_rows=[], r12_rows=[], as_of="2026-08-08")
        self.assertGreater(result.working_days_remaining, 0)
        self.assertEqual(
            result.working_days_total,
            result.working_days_elapsed + result.working_days_remaining,
        )


# ---------------------------------------------------------------------------
# 6. Date-format guard
# ---------------------------------------------------------------------------

class TestDateFormatGuard(unittest.TestCase):
    """Assertion 6: "Aug 1, 2026" and "1-Aug-2026" both parse; unknown raises."""

    def test_aug_1_2026_parses(self):
        d = mcr._parse_date_cell("Aug 1, 2026", 2026, 8)
        self.assertEqual(d, "2026-08-01")

    def test_aug_01_2026_parses(self):
        d = mcr._parse_date_cell("Aug 01, 2026", 2026, 8)
        self.assertEqual(d, "2026-08-01")

    def test_1_aug_2026_parses(self):
        d = mcr._parse_date_cell("1-Aug-2026", 2026, 8)
        self.assertEqual(d, "2026-08-01")

    def test_1_aug_2026_space_parses(self):
        d = mcr._parse_date_cell("1 Aug 2026", 2026, 8)
        self.assertEqual(d, "2026-08-01")

    def test_iso_parses(self):
        d = mcr._parse_date_cell("2026-08-01", 2026, 8)
        self.assertEqual(d, "2026-08-01")

    def test_plain_day_parses(self):
        d = mcr._parse_date_cell("1", 2026, 8)
        self.assertEqual(d, "2026-08-01")

    def test_dd_mm_yyyy_parses(self):
        d = mcr._parse_date_cell("01-08-2026", 2026, 8)
        self.assertEqual(d, "2026-08-01")

    def test_leading_apostrophe_stripped(self):
        # Google Sheets sometimes prepends an apostrophe to force text
        d = mcr._parse_date_cell("'Aug 1, 2026", 2026, 8)
        self.assertEqual(d, "2026-08-01")

    def test_unrecognised_raises_value_error(self):
        """A truly unrecognised format must raise ValueError (not silently skip)."""
        with self.assertRaises(ValueError):
            mcr._parse_date_cell("GARBAGE_DATE", 2026, 8)

    def test_empty_returns_none(self):
        self.assertIsNone(mcr._parse_date_cell("", 2026, 8))

    def test_excel_serial_returns_none(self):
        # 5-digit Excel serial — skip silently (we cannot parse meaningfully)
        result = mcr._parse_date_cell("46234", 2026, 8)
        self.assertIsNone(result)


# ---------------------------------------------------------------------------
# 7. p90 path (≥ MIN_DAYS_FOR_P90 days)
# ---------------------------------------------------------------------------

class TestP90Path(unittest.TestCase):

    def test_method_is_p90_with_enough_days(self):
        n = mcr.MIN_DAYS_FOR_P90
        # Build n distinct days with varying pcs (day 1=low, day n=high)
        rows = []
        base_date = datetime.date(2026, 8, 1)
        d = base_date
        for i in range(n):
            while d.weekday() == 6:  # skip Sundays
                d += datetime.timedelta(1)
            rows.append((d.strftime("%b %-d, %Y"), "CPVC", f"PS-{i}", 1000 + i * 100))
            d += datetime.timedelta(1)

        result = _run(r11_rows=rows, r12_rows=[])
        c = _cat(result, "CPVC Pipe")
        self.assertEqual(c.method, "p90",
            f"With {n} days, method must switch to p90 (was {c.method})")
        self.assertGreater(c.cap_per_day, 0)

    def test_p90_is_optimistic(self):
        """p90 (90th-percentile) must be ≥ mean of the same daily category totals.

        The 90th-percentile of any distribution is ≥ the mean (it's the high end).
        We use it because it represents the 'best days' pace — optimistic but
        grounded in actual production data.  A conservative floor (10th-pct) would
        systematically understate what the line can do and overstate the gap.
        """
        n = mcr.MIN_DAYS_FOR_P90
        base = datetime.date(2026, 8, 1)
        rows = []
        d = base
        for i in range(n):
            while d.weekday() == 6:
                d += datetime.timedelta(1)
            # Each day is one item row for CPVC (one machine running that day).
            # The per-day category total = the pcs for that row.
            rows.append((d.strftime("%b %-d, %Y"), "CPVC", f"PS-{i}", 1000 + i * 500))
            d += datetime.timedelta(1)

        result = _run(r11_rows=rows, r12_rows=[])
        c = _cat(result, "CPVC Pipe")
        mean_val = sum(1000 + i * 500 for i in range(n)) / n
        self.assertGreaterEqual(c.cap_per_day, mean_val - 1,
            "p90 (90th-percentile of category-day totals) must be ≥ mean")


# ---------------------------------------------------------------------------
# 8. NO_DEMONSTRATED_CAPACITY
# ---------------------------------------------------------------------------

class TestNoDemonstratedCapacity(unittest.TestCase):

    def test_solvent_categories_always_ndc(self):
        """Solvent categories never appear in Report-11/12 → always NDC."""
        result = _run(
            r11_rows=[("Aug 1, 2026", "CPVC", "PS-3", 5000)],
            r12_rows=[],
        )
        for cat in mcr.SOLVENT_CATEGORIES:
            c = _cat(result, cat)
            self.assertTrue(c.no_demonstrated_capacity,
                f"{cat} should be NO_DEMONSTRATED_CAPACITY")
            self.assertEqual(c.cap_per_day, 0.0)
            self.assertEqual(c.method, "none")

    def test_ndc_distinct_from_zero_cap_bug(self):
        """NDC category: produced==0 AND cap==0 (legitimate).
        Bug category:    produced>0  AND cap==0 (triggers INVARIANT VIOLATED).
        They must NOT look identical in the result."""
        result = _run(
            r11_rows=[("Aug 1, 2026", "CPVC", "PS-1", 1000)],
            r12_rows=[],
        )
        # CPVC Pipe: produced > 0, cap > 0 (no invariant warning)
        cpvc = _cat(result, "CPVC Pipe")
        agri = _cat(result, "AGRI Pipe")   # no rows → NDC

        self.assertFalse(cpvc.no_demonstrated_capacity)
        self.assertTrue(agri.no_demonstrated_capacity)

        inv_w = [w for w in result.warnings if "INVARIANT VIOLATED" in w]
        self.assertEqual(inv_w, [], "Engine must not emit invariant warning for NDC category")


# ---------------------------------------------------------------------------
# 9. Category ordering
# ---------------------------------------------------------------------------

class TestCategoryOrdering(unittest.TestCase):

    def test_output_follows_canonical_order(self):
        result = _run(r11_rows=[], r12_rows=[])
        actual_order = [c.category for c in result.categories]
        self.assertEqual(actual_order, mcr.CATEGORY_ORDER)


# ---------------------------------------------------------------------------
# 10. Invariant warnings surfaced
# ---------------------------------------------------------------------------

class TestInvariantWarningSurface(unittest.TestCase):
    """Invariant violations must appear in result.warnings, not swallowed."""

    def test_invariant_3_fires_when_shortfall_equals_remaining(self):
        """Force shortfall == remaining by making cap_per_day = 0 artificially
        while produced_to_date > 0. We simulate this by patching _compute_cap_per_day
        to return 0 only for one category, triggering the emergency fallback + INVARIANT
        VIOLATED warning."""
        import mp_corrective_replan as _mcr_mod
        original = _mcr_mod._compute_cap_per_day

        def patched(vals):
            # Force CPVC Pipe to return 0/none regardless of production
            if not vals:
                return 0.0, "none", 0
            return 0.0, "none", 0  # always 0 — simulates the bug

        with patch.object(_mcr_mod, "_compute_cap_per_day", side_effect=patched):
            result = _run(
                r11_rows=[("Aug 1, 2026", "CPVC", "PS-3", 2420)],
                r12_rows=[],
                plan_recs=[_plan_rec("CPVC", "CPVC PIPE", 100_000, 0)],
            )

        # Should emit INVARIANT VIOLATED warning (not crash)
        inv_w = [w for w in result.warnings if "INVARIANT" in w or "CRITICAL" in w]
        self.assertGreater(len(inv_w), 0,
            "Engine must surface invariant warnings in result.warnings")


# ---------------------------------------------------------------------------
# 11. _parse_date fix in mp_followup (Issue #5 regression)
# ---------------------------------------------------------------------------

class TestMpFollowupDateFix(unittest.TestCase):
    """Verify the mp_followup._parse_date fix handles Report-11 date format."""

    def test_aug_1_2026_parsed_by_followup(self):
        from mp_followup import _parse_date
        result = _parse_date("Aug 1, 2026", 2026, 8)
        self.assertEqual(result, "2026-08-01",
            "_parse_date in mp_followup must parse 'Aug 1, 2026' (Issue #5 fix)")

    def test_1_aug_2026_parsed_by_followup(self):
        from mp_followup import _parse_date
        result = _parse_date("1-Aug-2026", 2026, 8)
        self.assertEqual(result, "2026-08-01",
            "_parse_date in mp_followup must parse '1-Aug-2026' (PTMT format)")

    def test_existing_plain_day_still_works(self):
        from mp_followup import _parse_date
        result = _parse_date("1", 2026, 8)
        self.assertEqual(result, "2026-08-01")

    def test_existing_iso_still_works(self):
        from mp_followup import _parse_date
        result = _parse_date("2026-08-01", 2026, 8)
        self.assertEqual(result, "2026-08-01")


# ---------------------------------------------------------------------------
# 12. Target-value validation (Aug 2026, 2-day scenario from spec)
# ---------------------------------------------------------------------------

class TestTargetValues(unittest.TestCase):
    """Verify the engine produces the target values from the spec document
    for August's 2-day scenario (mean method, 18-19 remaining working days)."""

    def setUp(self):
        self.result = _run(
            r11_rows=[
                # Aug 1
                ("Aug 1, 2026", "CPVC", "PS-3",  2420),
                ("Aug 1, 2026", "UPVC", "PU-1",  3480),
                ("Aug 1, 2026", "SWR",  "PW71",  1782),
                ("Aug 1, 2026", "AGRI", "PA-1",  350),
                # Aug 3
                ("Aug 3, 2026", "CPVC", "PS-12", 3075),
                ("Aug 3, 2026", "UPVC", "PU-15", 1423),
                ("Aug 3, 2026", "SWR",  "PW72",  1790),
            ],
            r12_rows=[
                # Aug 1
                ("Aug 1, 2026", "CPVC", "U531", 25176),
                ("Aug 1, 2026", "UPVC", "U532", 19525),
                ("Aug 1, 2026", "SWR",  "S101", 7635),
                ("Aug 1, 2026", "AGRI", "A101", 570),
                # Aug 3
                ("Aug 3, 2026", "CPVC", "U533", 19580),
                ("Aug 3, 2026", "SWR",  "S102", 8621),
            ],
            month="2026-08",
            as_of="2026-08-08",
        )

    def test_all_pipe_fitting_categories_have_production(self):
        for cat in ["CPVC Pipe", "CPVC Fitting", "UPVC Pipe", "UPVC Fitting",
                    "SWR Pipe", "SWR Fitting", "AGRI Pipe", "AGRI Fitting"]:
            c = _cat(self.result, cat)
            self.assertGreater(c.produced_to_date, 0, f"{cat} should have production")
            self.assertGreater(c.cap_per_day, 0, f"{cat} must have non-zero cap_per_day")

    def test_all_methods_are_mean_low_confidence(self):
        """With only 2 days, all categories must use mean (low-confidence), not p90."""
        for cat_name in ["CPVC Pipe", "CPVC Fitting", "UPVC Pipe",
                         "SWR Pipe", "SWR Fitting", "AGRI Pipe", "AGRI Fitting"]:
            c = _cat(self.result, cat_name)
            self.assertTrue(c.method.startswith("mean"),
                f"{cat_name}: expected mean-based method with 2 days, got {c.method!r}")
            self.assertTrue(c.low_confidence,
                f"{cat_name}: 2 production days must set low_confidence=True")

    def test_upvc_fitting_one_day_only_still_nonzero(self):
        """UPVC Fitting only ran Aug 1 (0 on Aug 3). Must still get cap > 0."""
        c = _cat(self.result, "UPVC Fitting")
        self.assertAlmostEqual(c.cap_per_day, 19525.0, delta=5.0)

    def test_solvent_categories_ndc(self):
        for cat in mcr.SOLVENT_CATEGORIES:
            c = _cat(self.result, cat)
            self.assertTrue(c.no_demonstrated_capacity)

    def test_feasible_totals_nonzero(self):
        total_feasible = sum(c.feasible for c in self.result.categories)
        self.assertGreater(total_feasible, 0)

    def test_no_invariant_violations(self):
        inv = [w for w in self.result.warnings if "INVARIANT VIOLATED" in w]
        self.assertEqual(inv, [])


# ---------------------------------------------------------------------------
# New spec-required tests (Fix 1–4)
# ---------------------------------------------------------------------------

class TestPaceGranularity(unittest.TestCase):
    """Pace must be computed over per-CATEGORY-per-DAY totals (not per-item values).

    The fix: for each date, SUM all item rows of a category, then compute p90
    (90th-percentile) over those per-day sums.  Using individual item-row pcs
    would give a 6-10× underestimate because a busy CPVC day has many item
    rows, each a fraction of the day total.
    """

    def test_p90_is_90th_percentile_of_category_day_totals(self):
        """Verify _percentile_90 is the 90th-percentile (not 10th) of category-day totals.

        Uses synthetic data that matches the Aug 2026 CPVC Pipe per-day totals
        (verified against Report-11):  [2025, 2420, 3075, 3310, 4753, 6114]
        mean = 3,616    90th-pct ≈ 5,433    10th-pct ≈ 2,222

        The fix must produce p90 ≈ 5,433 (not the old ≈ 2,222).
        """
        # These exactly match the Aug 1–7 2026 CPVC Pipe category-day totals
        # derived from Report-11 in the active workbook.
        cpvc_daily_totals = [2025.0, 2420.0, 3075.0, 3310.0, 4753.0, 6114.0]
        p90 = mcr._percentile_90(cpvc_daily_totals)
        mean_val = sum(cpvc_daily_totals) / len(cpvc_daily_totals)

        # 90th-pct of [2025, 2420, 3075, 3310, 4753, 6114]:
        # sorted idx = (6-1)*0.9 = 4.5 → 4753 + (6114-4753)*0.5 = 5433.5
        self.assertAlmostEqual(p90, 5433.5, delta=1.0,
            msg=f"90th-pct of Aug CPVC daily totals should be ~5433, got {p90}")
        self.assertGreater(p90, mean_val,
            msg=f"90th-pct ({p90}) must be > mean ({mean_val})")
        # Critically: NOT the old 10th-pct (2222.5)
        self.assertGreater(p90, 4000,
            msg=f"p90 must be >> 2222 (old 10th-pct value); got {p90}")

    def test_multiple_items_per_day_summed_before_p90(self):
        """Three CPVC item rows on the same date must be SUMMED before p90 calc.

        Setup: 5 production days. On day 1, three separate CPVC items each
        producing 2,000 pcs (category-day total = 6,000).  On the other 4 days,
        one item each producing 1,000 pcs.

        If the engine incorrectly fed individual item values to p90, it would
        compute p90 over [2000, 2000, 2000, 1000, 1000, 1000, 1000] (7 values)
        giving 90th-pct ≈ 2,000.

        Correct behaviour: day 1 total = 6,000, other days = 1,000 each.
        p90 over [1000, 1000, 1000, 1000, 6000] (5 values, sorted):
        idx = (5-1)*0.9 = 3.6 → 1000 + (6000-1000)*0.6 = 4,000.
        """
        rows = [
            # Day 1: three items, each 2,000 pcs → category-day total = 6,000
            ("Aug 1, 2026", "CPVC", "PS-1", 2000),
            ("Aug 1, 2026", "CPVC", "PS-2", 2000),
            ("Aug 1, 2026", "CPVC", "PS-3", 2000),
            # Days 2-5: one item each, 1,000 pcs
            ("Aug 3, 2026", "CPVC", "PS-1", 1000),
            ("Aug 4, 2026", "CPVC", "PS-1", 1000),
            ("Aug 5, 2026", "CPVC", "PS-1", 1000),
            ("Aug 6, 2026", "CPVC", "PS-1", 1000),
        ]
        result = _run(r11_rows=rows, r12_rows=[])
        c = _cat(result, "CPVC Pipe")

        # Produced-to-date must be correct regardless of granularity
        self.assertAlmostEqual(c.produced_to_date, 10_000, delta=1,
            msg="produced_to_date must be 6000 + 4×1000 = 10,000")

        # Per-day totals: [6000, 1000, 1000, 1000, 1000]
        # 90th-pct (5 values): idx=(5-1)*0.9=3.6 → 1000+(6000-1000)*0.6 = 4000
        self.assertAlmostEqual(c.cap_per_day, 4000.0, delta=5.0,
            msg=f"p90 of [6000,1000,1000,1000,1000] should be ~4000, got {c.cap_per_day}")
        self.assertEqual(c.method, "p90",
            "5 production days must use p90 method")

    def test_p90_higher_than_mean_with_high_day_skew(self):
        """p90 (90th-pct) must exceed the mean when one day is much higher.

        With values [100, 100, 100, 100, 10000], mean ≈ 2100 but 90th-pct ≈ 7240.
        """
        rows = [
            ("Aug 1, 2026", "SWR", "PW-1", 100),
            ("Aug 3, 2026", "SWR", "PW-1", 100),
            ("Aug 4, 2026", "SWR", "PW-1", 100),
            ("Aug 5, 2026", "SWR", "PW-1", 100),
            ("Aug 6, 2026", "SWR", "PW-1", 10_000),
        ]
        result = _run(r11_rows=rows, r12_rows=[])
        c = _cat(result, "SWR Pipe")
        mean_val = (100 * 4 + 10_000) / 5
        # p90 at idx=(5-1)*0.9=3.6 → 100+(10000-100)*0.6 = 100+5940 = 6040
        self.assertGreater(c.cap_per_day, mean_val,
            f"90th-pct ({c.cap_per_day}) must exceed mean ({mean_val})")
        self.assertAlmostEqual(c.cap_per_day, 6040.0, delta=5.0,
            msg=f"Expected p90 ≈ 6040, got {c.cap_per_day}")

    def test_date_read_from_header_column(self):
        """DATE must be identified by header text, not by column position.

        In the August PIPE workbook, column A is blank and column B is DATE.
        The parser must read dates from the column whose HEADER contains 'DATE',
        which resolves to column index 1 (B) in that file.
        """
        # Build a synthetic R11 where column A is blank, DATE is in column B
        hdr = ["", "DATE", "MACHINE NAME", "MACHINE NO.", "TYPES", "ITEM CODE",
               "Running Hour", "Ideal Weight (KG)", "Pcs", "Weight"]
        rows_raw = [
            ["https://sheets..."],
            ["", "TOTAL", "", "", "", "", "100", "", "3000", ""],
            ["", "PIPE M/C"],
            ["", "", "", "", "", ""],
            hdr,
            # Column A blank, date in column B
            ["", "Aug 1, 2026", "CON-63-1", "PIPE M/C - 1", "CPVC", "PS-3",
             "8", "1.0", "3000", "3000"],
        ]
        import mp_corrective_replan as m
        parsed = m._parse_r11_daily_pcs(rows_raw, 2026, 8)
        # Discard sentinel
        parsed.pop("_n_no_date", None)
        self.assertIn("2026-08-01", parsed,
            "Parser must resolve date from column B (header 'DATE'), got: " + str(list(parsed.keys())))
        self.assertAlmostEqual(parsed["2026-08-01"].get("CPVC Pipe", 0), 3000, delta=1)

    def test_rows_without_date_not_silently_lost(self):
        """A data row appearing before ANY date in the sheet must not be silently dropped.

        The sentinel key '_n_no_date' must be non-zero and a warning must appear
        in the CorrectiveReplanResult.warnings list.
        """
        # Build R11 where the first DATA row has no date and no prior date exists
        hdr = ["", "DATE", "MACHINE NAME", "MACHINE NO.", "TYPES", "ITEM CODE",
               "Running Hour", "Ideal Weight (KG)", "Pcs", "Weight"]
        rows_raw = [
            ["https://sheets..."],
            ["", "TOTAL"],
            ["", "PIPE M/C"],
            ["", ""],
            hdr,
            # First data row: no date, CPVC, pcs=500 — unattributable
            ["", "", "CON-63-1", "PIPE M/C - 1", "CPVC", "PS-3", "8", "1.0", "500", ""],
            # Second row: date present, pcs=1000
            ["", "Aug 1, 2026", "CON-63-1", "PIPE M/C - 1", "CPVC", "PS-4", "8", "1.0", "1000", ""],
        ]
        import mp_corrective_replan as m
        parsed = m._parse_r11_daily_pcs(rows_raw, 2026, 8)
        n_no_date = parsed.pop("_n_no_date", 0)
        self.assertGreater(n_no_date, 0,
            "Parser must count rows skipped due to unresolvable date")

        # The engine warning must surface this
        r12_empty = [["DATE", "MATERIAL", "Item Code", "SAP Code", "Machine",
                       "RC", "MC(F)", "CT", "Output Production", ""],
                     ["", "", "", "", "", "", "", "", " Pc ", " Wt "]]
        result = m.compute_corrective_replan(
            month="2026-08", plan_recs=[],
            r11_values=rows_raw, r12_values=r12_empty,
            as_of_date="2026-08-08", file_id="",
        )
        has_date_warning = any("date resolution" in w.lower() for w in result.warnings)
        self.assertTrue(has_date_warning,
            f"Engine warnings must include a date-resolution note; got: {result.warnings}")


class TestNoCapacityLabel(unittest.TestCase):
    """Fix 1: no "capacity" language in method or status fields."""

    def test_not_started_never_says_no_capacity(self):
        """Zero production → 'not_started=True', never the string 'NO CAPACITY'."""
        result = _run(r11_rows=[], r12_rows=[])
        for cat in result.categories:
            # The not_started property must be True when no production
            if cat.produced_to_date == 0:
                self.assertTrue(cat.not_started,
                    f"{cat.category}: zero production must set not_started=True")
                # Crucially: no "NO CAPACITY" string in the method field
                self.assertNotIn("NO CAPACITY", cat.method,
                    f"{cat.category}: method must not contain 'NO CAPACITY'")
                self.assertNotIn("NO CAPACITY", str(cat.cap_per_day),
                    f"{cat.category}: cap_per_day must not be the string 'NO CAPACITY'")

    def test_not_started_shows_zero_pace(self):
        """Not-started categories have pace=0 and method='none' (numeric, not sentinel string)."""
        result = _run(r11_rows=[], r12_rows=[])
        agri = _cat(result, "AGRI Pipe")
        self.assertEqual(agri.cap_per_day, 0.0)  # numeric zero, not a string
        self.assertEqual(agri.method, "none")
        self.assertTrue(agri.not_started)

    def test_low_confidence_flag_on_few_days(self):
        """< MIN_DAYS_FOR_P90 production days → low_confidence=True, method contains flag."""
        result = _run(
            r11_rows=[("Aug 1, 2026", "CPVC", "PS-1", 3000)],
            r12_rows=[],
        )
        c = _cat(result, "CPVC Pipe")
        self.assertTrue(c.low_confidence, "1 day must be low-confidence")
        self.assertIn("low-confidence", c.method,
            f"method must contain 'low-confidence', got {c.method!r}")

    def test_p90_is_not_low_confidence(self):
        """≥ MIN_DAYS_FOR_P90 days → low_confidence=False."""
        n = mcr.MIN_DAYS_FOR_P90
        rows = []
        d = datetime.date(2026, 8, 1)
        for i in range(n):
            while d.weekday() == 6:
                d += datetime.timedelta(1)
            rows.append((d.strftime("%b %-d, %Y"), "CPVC", f"PS-{i}", 3000))
            d += datetime.timedelta(1)
        result = _run(r11_rows=rows, r12_rows=[])
        c = _cat(result, "CPVC Pipe")
        self.assertFalse(c.low_confidence, "≥5 days must not be low-confidence")
        self.assertEqual(c.method, "p90")

    def test_cap_feasible_column_optional(self):
        """When cap_feasible_by_cat not passed, all cap_feasible fields are None."""
        result = _run(r11_rows=[], r12_rows=[])
        for cat in result.categories:
            self.assertIsNone(cat.cap_feasible)

    def test_cap_feasible_column_populated(self):
        """When cap_feasible_by_cat passed, values appear in CategoryResult."""
        r11 = _r11([])
        r12 = _r12([])
        cap_feas = {
            "CPVC Pipe": 50_000.0,
            "CPVC Solvent": 12_000.0,
        }
        result = mcr.compute_corrective_replan(
            month="2026-08",
            plan_recs=[],
            r11_values=r11,
            r12_values=r12,
            as_of_date="2026-08-08",
            file_id="",
            cap_feasible_by_cat=cap_feas,
        )
        cpvc_pipe = _cat(result, "CPVC Pipe")
        self.assertEqual(cpvc_pipe.cap_feasible, 50_000.0)
        solvent = _cat(result, "CPVC Solvent")
        self.assertEqual(solvent.cap_feasible, 12_000.0)
        # Category with no entry in cap_feas → None
        upvc_pipe = _cat(result, "UPVC Pipe")
        self.assertIsNone(upvc_pipe.cap_feasible)


class TestProducedToDateReconciliation(unittest.TestCase):
    """Fix 4: produced-to-date must equal R-11 total + R-12 total (including TEFFLONE)."""

    def _r11_with_total(self, pcs):
        """Build R11 grid with a TOTAL row that matches pcs."""
        hdr = ["", "DATE", "MACHINE NAME", "MACHINE NO.", "TYPES", "ITEM CODE",
               "Running Hour", "Ideal Weight (KG)", "Pcs", "Weight"]
        return [
            ["url"],
            ["", "TOTAL", "", "", "", "", "100", "", str(pcs), ""],
            ["", "PIPE M/C"],
            ["", "", "", "", "", "", ""],
            hdr,
            ["", "Aug 1, 2026", "CON-63-1", "PIPE M/C - 1", "CPVC", "PS-3",
             "8", "1.0", str(pcs), "100"],
        ]

    def _r12_with_tefflone(self, cpvc_pcs, tefflone_pcs):
        hdr = ["DATE", "MATERIAL", "Item Code", "SAP Code", "Moulding Machine",
               "Run Cavity", "Mould Cavity (F)", "Cycle Time Per Pcs Standard (F)",
               "Output Production", ""]
        sub = ["", "", "", "", "", "", "", "", " Pc ", " Wt in Kgs "]
        total_pcs = cpvc_pcs + tefflone_pcs
        return [
            ["Mr. Jitendra"],
            ["Mould M/C"],
            ["TOTAL", "", "", "", "", "", "", "", str(total_pcs), ""],
            hdr,
            sub,
            ["Aug 1, 2026", "CPVC", "U531", "", "A02", "4", "4", "25", str(cpvc_pcs), ""],
            ["Aug 1, 2026", "TEFFLONE", "T01", "", "A02", "4", "4", "25",
             str(tefflone_pcs), ""],
        ]

    def test_actual_produced_total_includes_tefflone(self):
        """actual_produced_total = R-11 pcs + R-12 CPVC pcs + R-12 TEFFLONE pcs."""
        r11_pcs     = 2420
        cpvc_pcs    = 25176
        tefflone_pcs= 17600

        r11 = self._r11_with_total(r11_pcs)
        r12 = self._r12_with_tefflone(cpvc_pcs, tefflone_pcs)

        result = mcr.compute_corrective_replan(
            month="2026-08",
            plan_recs=[],
            r11_values=r11,
            r12_values=r12,
            as_of_date="2026-08-08",
            file_id="",
        )

        expected_total = r11_pcs + cpvc_pcs + tefflone_pcs
        self.assertAlmostEqual(result.actual_produced_total, expected_total, delta=1,
            msg=f"actual_produced_total should be {expected_total} "
                f"(R-11={r11_pcs} + R-12 CPVC={cpvc_pcs} + TEFFLONE={tefflone_pcs})")

    def test_other_produced_tracks_tefflone(self):
        """result.other_produced must equal the TEFFLONE pcs (not 0)."""
        r11  = self._r11_with_total(2420)
        r12  = self._r12_with_tefflone(25176, 17600)
        result = mcr.compute_corrective_replan(
            month="2026-08", plan_recs=[], r11_values=r11, r12_values=r12,
            as_of_date="2026-08-08", file_id="",
        )
        self.assertAlmostEqual(result.other_produced, 17600, delta=1,
            msg=f"other_produced should be 17600 (TEFFLONE), got {result.other_produced}")

    def test_tefflone_not_assigned_to_any_category(self):
        """TEFFLONE pcs must not appear in any per-category produced_to_date."""
        r11  = self._r11_with_total(0)
        r12  = self._r12_with_tefflone(0, 17600)
        result = mcr.compute_corrective_replan(
            month="2026-08", plan_recs=[], r11_values=r11, r12_values=r12,
            as_of_date="2026-08-08", file_id="",
        )
        for cat in result.categories:
            self.assertEqual(cat.produced_to_date, 0.0,
                f"{cat.category}: TEFFLONE must not appear in per-category produced_to_date "
                f"(got {cat.produced_to_date})")

    def test_aug_actual_total_reconciles(self):
        """Live-fixture reconciliation: pipe 50,293 + fitting 191,630 = 241,923.

        The test synthesises the exact structure of the Aug 2026 workbook
        using stub totals (not live sheet data) to avoid network dependence.
        """
        r11 = self._r11_with_total(50_293)
        r12 = self._r12_with_tefflone(174_030, 17_600)  # 174k classified + 17.6k TEFFLONE
        result = mcr.compute_corrective_replan(
            month="2026-08", plan_recs=[], r11_values=r11, r12_values=r12,
            as_of_date="2026-08-08", file_id="",
        )
        self.assertAlmostEqual(result.actual_produced_total, 241_923, delta=5,
            msg=f"Aug total should be 241,923; got {result.actual_produced_total}")
        self.assertAlmostEqual(result.other_produced, 17_600, delta=5)


class TestXlsxLabels(unittest.TestCase):
    """Fix 1: XLSX must use run-rate / pace language, not capacity language."""

    def setUp(self):
        r11 = _r11([("Aug 1, 2026", "CPVC", "PS-1", 3000)])
        r12 = _r12([])
        result = mcr.compute_corrective_replan(
            month="2026-08", plan_recs=[], r11_values=r11, r12_values=r12,
            as_of_date="2026-08-08", file_id="TESTID",
        )
        from io import BytesIO
        from openpyxl import load_workbook
        data = mcr.corrective_replan_bytes(result)
        self.wb = load_workbook(BytesIO(data))

    def _all_text(self):
        text = []
        for ws in self.wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v:
                        text.append(str(v).upper())
        return " ".join(text)

    def test_sheet_not_named_replan(self):
        """Tab must not be named just 'Re-plan' — now 'Run-rate Projection'."""
        self.assertIn("RUN-RATE", " ".join(self.wb.sheetnames).upper(),
            f"Expected 'Run-rate Projection' tab, got {self.wb.sheetnames}")

    def test_no_cap_per_day_header(self):
        text = self._all_text()
        self.assertNotIn("CAP/DAY", text,
            "XLSX must not contain 'Cap/Day' (use 'Pace/day' instead)")

    def test_no_feasible_header(self):
        text = self._all_text()
        # "Feasible" as a column header must be gone; "projected" must appear
        self.assertIn("PROJECTED", text,
            "XLSX must contain 'Projected' column header")
        self.assertNotIn("CAP-FEASIBLE PLAN", text.replace(" ", "-"))

    def test_no_capacity_in_data_cells(self):
        """'NO CAPACITY' string must not appear anywhere in the output."""
        text = self._all_text()
        self.assertNotIn("NO CAPACITY", text,
            "XLSX must not contain 'NO CAPACITY' — use 'Not started' instead")

    def test_not_started_appears_for_zero_production_categories(self):
        """Not-started categories must show 'Not started' text (not 'NO CAPACITY')."""
        text = self._all_text()
        self.assertIn("NOT STARTED", text,
            "XLSX must contain 'Not started' for zero-production categories")

    def test_run_rate_framing_in_subtitle(self):
        text = self._all_text()
        self.assertIn("RUN-RATE", text,
            "XLSX must contain run-rate framing in subtitle/notes")
        self.assertIn("NOT MACHINE CAPACITY", text.replace(",", "").replace(".", ""),
            "XLSX must explicitly state 'not machine capacity'")

    def test_provenance_tab_exists(self):
        self.assertIn("Provenance", self.wb.sheetnames)

    def test_gap_language_in_provenance(self):
        ws = self.wb["Provenance"]
        prov_text = " ".join(
            str(v).upper()
            for row in ws.iter_rows(values_only=True)
            for v in (row or [])
            if v is not None and str(v).strip() not in ("", "NONE")
        )
        self.assertIn("RUN-RATE", prov_text,
            "Provenance tab must document run-rate nature of the report")


if __name__ == "__main__":
    unittest.main()

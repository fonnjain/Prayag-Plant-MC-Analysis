"""
End-to-end smoke test for the Machine Plan Comparison report route.

Route: GET /machine-planning/report/machine-plan-comparison
Verifies: HTTP 200, xlsx Content-Type, non-empty response body.

run_engine / run_fitting_engine are mocked so no live DB or seed data is needed.
The report generator (machine_plan_comparison_bytes) runs for real so the test
catches any regression in the xlsx assembly code.

Run from the prayag directory:
    cd artifacts/prayag && python3 -m pytest tests/test_mp_comparison_route.py -v
"""
from __future__ import annotations

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from unittest.mock import patch, MagicMock

import pytest


# ── Shared stubs (mirrors test_mp_reports.py helpers) ────────────────────────

def _make_item(
    item_code="CPVC-100",
    material="CPVC",
    qty_pcs=1000.0,
    wt_per_pc=0.3,
    material_kg=312.0,
    rate_kg_per_hr=150.0,
    machine="M/C-1",
    rate_fallback_tier="item",
):
    from mp_engine import ItemResult, AssignedPortion
    hrs = material_kg / rate_kg_per_hr
    a = AssignedPortion(machine=machine, hrs=round(hrs, 3),
                        material_kg=material_kg, qty_pcs=qty_pcs)
    return ItemResult(
        item_code=item_code,
        raw_code=item_code,
        material=material,
        qty_pcs=qty_pcs,
        weight_per_pc_kg=wt_per_pc,
        material_kg=material_kg,
        fresh_compound_kg=material_kg * 0.75,
        pulverizer_kg=material_kg * 0.25,
        rate_kg_per_hr=rate_kg_per_hr,
        rate_estimated=(rate_fallback_tier != "item"),
        rate_fallback_tier=rate_fallback_tier,
        machine_hrs=round(hrs, 3),
        capable_machines=[machine],
        assignments=[a],
        has_weight=True,
        has_machine=True,
    )


def _make_engine_result(items=None):
    from mp_engine import EngineResult, CoverageGaps, PlanTotals, MachineLoad
    items = items or [_make_item()]
    ml = MachineLoad(
        machine="M/C-1", capacity_hrs=500.0,
        assigned_hrs=sum(it.machine_hrs for it in items),
        utilisation_pct=50.0, machine_days=13.0,
        material_kg=sum(it.material_kg for it in items),
        fresh_compound_kg=sum(it.fresh_compound_kg for it in items),
        pulverizer_kg=sum(it.pulverizer_kg for it in items),
        staffing_ok=True, operators_ot=2, support_w=1,
    )
    return EngineResult(
        segment="PIPE",
        effective_month="2026-07",
        items=items,
        machine_loads=[ml],
        coverage_gaps=CoverageGaps(
            no_weight=[], no_machine=[], idle_machines=[], locked_out_machines=[]
        ),
        totals=PlanTotals(
            total_qty_pcs=sum(it.qty_pcs for it in items),
            total_material_kg=sum(it.material_kg for it in items),
            total_fresh_compound_kg=sum(it.fresh_compound_kg for it in items),
            total_pulverizer_kg=sum(it.pulverizer_kg for it in items),
            routable_material_kg=sum(it.material_kg for it in items),
            routable_fresh_compound_kg=sum(it.fresh_compound_kg for it in items),
            routable_pulverizer_kg=sum(it.pulverizer_kg for it in items),
        ),
        baseline_machine_loads=[ml],
        params_used={"waste_pct": 4.0, "pulverizer_pct": 25.0},
        effective_costs={"CPVC": 120.0},
        cost_by_material={"CPVC": 28080.0},
        n_unpriced=0,
    )


# Minimal demand payload stored in the session run (no fitting_demand so the
# fitting branch is skipped — simplest path through the route).
_PAYLOAD = {
    "segment": "PLUMBING",
    "effective_month": "2026-07",
    "demand": [
        {"item_code": "CPVC-100", "raw_code": "CPVC-100",
         "material": "CPVC", "qty_pcs": 1000},
    ],
    "fitting_demand": [],
}


# ── fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def clear_engine_cache():
    """Wipe _MP2_ENGINE_CACHE before and after each test."""
    import app as flask_app
    flask_app._MP2_ENGINE_CACHE.clear()
    yield
    flask_app._MP2_ENGINE_CACHE.clear()


# ── route smoke test ──────────────────────────────────────────────────────────

class TestMachinePlanComparisonRoute:
    """GET /machine-planning/report/machine-plan-comparison → valid .xlsx."""

    def _run(self, run_id="42"):
        import app as flask_app
        import mp_engine as eng

        flask_app.app.config["TESTING"] = True
        flask_app.app.config["SECRET_KEY"] = "test-comparison"

        fake_result = _make_engine_result()

        client = flask_app.app.test_client()

        # Inject the session run_id BEFORE the GET so _ensure_session_run_id
        # returns early (session already populated — no DB call needed).
        with client.session_transaction() as sess:
            sess["mp2_run_id"] = run_id

        with (
            patch.object(flask_app, "_mp2_load_run", return_value=_PAYLOAD),
            patch.object(flask_app, "_mp2_result_from_session", return_value=fake_result),
            patch.object(flask_app, "_mp3_fitting_result_from_session", return_value=None),
            patch.object(flask_app._mp_model, "get_params", return_value=None),
            patch.object(eng, "run_engine", return_value=fake_result),
            patch.object(eng, "run_fitting_engine", return_value=None),
        ):
            response = client.get("/machine-planning/report/machine-plan-comparison")

        return response

    def test_returns_200(self):
        resp = self._run()
        assert resp.status_code == 200, (
            f"Expected 200, got {resp.status_code}. Body: {resp.data[:300]}"
        )

    def test_content_type_is_xlsx(self):
        resp = self._run()
        ct = resp.content_type or ""
        assert "spreadsheetml" in ct or "xlsx" in ct, (
            f"Expected xlsx Content-Type, got {ct!r}"
        )

    def test_response_body_is_non_empty(self):
        resp = self._run()
        assert len(resp.data) > 0, "Response body must not be empty"

    def test_response_body_is_valid_xlsx(self):
        """The returned bytes must be a parseable openpyxl workbook."""
        import io
        from openpyxl import load_workbook

        resp = self._run()
        wb = load_workbook(filename=io.BytesIO(resp.data))
        assert wb.sheetnames, "Workbook must have at least one sheet"

    def test_content_disposition_includes_filename(self):
        """Content-Disposition should name the file machine_plan_comparison_*.xlsx."""
        resp = self._run()
        cd = resp.headers.get("Content-Disposition", "")
        assert "machine_plan_comparison" in cd and ".xlsx" in cd, (
            f"Unexpected Content-Disposition: {cd!r}"
        )

    def test_no_session_run_id_redirects(self):
        """Without a session run_id the route must redirect (not 500)."""
        import app as flask_app
        flask_app.app.config["TESTING"] = True
        flask_app.app.config["SECRET_KEY"] = "test-comparison"

        client = flask_app.app.test_client()
        # No session injection — session is empty.
        # _ensure_session_run_id will try _mp_model.list_plan_runs; mock it to []
        # so there is still no run_id and the route hits the redirect branch.
        with patch.object(flask_app._mp_model, "list_plan_runs", return_value=[]):
            response = client.get("/machine-planning/report/machine-plan-comparison")

        assert response.status_code in (302, 303), (
            f"Expected redirect when no run_id, got {response.status_code}"
        )

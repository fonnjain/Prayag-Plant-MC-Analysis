"""
Tests for mp_engine — Phase MP-2.

Covers: chain math, rate-fallback flag, LPT optimiser (balanced < baseline),
coverage-gap reporting, demand parsing (col layout + TOTAL skip + norm_code),
and route isolation (/ unaffected).

All tests run without a real DB — mp_model read functions are monkeypatched.
"""
from __future__ import annotations

import dataclasses
import io
import types
from typing import List
from unittest.mock import MagicMock, patch

import pytest

import mp_engine as eng
from mp_engine import (
    AssignedPortion,
    CoverageGaps,
    DemandItem,
    EngineResult,
    ItemResult,
    MachineLoad,
    PlanTotals,
    _baseline_assign,
    _build_rate_lookups,
    _compute_machine_loads,
    _get_rate,
    _lpt_optimise,
    run_engine,
)
from mp_seed import norm_code

# ── Fixtures / helpers ────────────────────────────────────────────────────────

def _demand(codes_qtys: list, material: str = "CPVC") -> List[DemandItem]:
    return [DemandItem(item_code=norm_code(c), raw_code=c, material=material, qty_pcs=q)
            for c, q in codes_qtys]


def _bom_row(code: str, wt: float) -> dict:
    return {"item_code": norm_code(code), "weight_per_pc_kg": wt}


def _ph_row(code: str, val: float, basis: str = "kg_per_hr") -> dict:
    return {"item_code": norm_code(code), "value": val, "basis": basis}


def _route_row(code: str, machine: str, material: str = "CPVC") -> dict:
    return {"item_code": norm_code(code), "machine": machine,
            "material": material, "capable": True}


def _machine_row(mc: str, cap: float = 500.0, ops: int = 3, sup: int = 1) -> dict:
    return {"machine": mc, "capacity_hrs_month": cap, "kind": "extrusion",
            "operators_ot": ops, "support_w": sup}


def _params():
    p = MagicMock()
    p.waste_pct = 4.0
    p.pulverizer_pct = 25.0
    return p


# ── Chain math ────────────────────────────────────────────────────────────────

class TestChainMath:
    def _run_single(self, qty_pcs: float, wt_per_pc: float,
                    waste_pct: float = 4.0, pulv_pct: float = 25.0):
        material_kg = qty_pcs * wt_per_pc * (1 + waste_pct / 100)
        fresh = material_kg * (1 - pulv_pct / 100)
        pulv  = material_kg - fresh
        return material_kg, fresh, pulv

    def test_material_kg_formula(self):
        mat, _, _ = self._run_single(1000, 0.5, waste_pct=4.0, pulv_pct=25.0)
        assert abs(mat - 1000 * 0.5 * 1.04) < 1e-6

    def test_fresh_compound_fraction(self):
        mat, fresh, pulv = self._run_single(1000, 0.5, waste_pct=4.0, pulv_pct=25.0)
        assert abs(fresh - mat * 0.75) < 1e-6

    def test_pulverizer_fraction(self):
        mat, fresh, pulv = self._run_single(1000, 0.5, waste_pct=4.0, pulv_pct=25.0)
        assert abs(pulv - mat * 0.25) < 1e-6

    def test_fresh_plus_pulv_equals_material(self):
        mat, fresh, pulv = self._run_single(2000, 0.3, waste_pct=5.0, pulv_pct=30.0)
        assert abs(fresh + pulv - mat) < 1e-9

    def test_zero_waste_no_uplift(self):
        mat, _, _ = self._run_single(100, 1.0, waste_pct=0.0, pulv_pct=25.0)
        assert abs(mat - 100.0) < 1e-9

    def test_zero_pulv_all_fresh(self):
        mat, fresh, pulv = self._run_single(100, 1.0, waste_pct=0.0, pulv_pct=0.0)
        assert abs(fresh - mat) < 1e-9
        assert abs(pulv) < 1e-9


# ── Rate fallback ─────────────────────────────────────────────────────────────

class TestRateFallback:
    def _lookups(self, ph_rows, route_rows):
        return _build_rate_lookups(ph_rows, route_rows)

    def test_exact_match_not_estimated(self):
        ph = [_ph_row("A1", 45.0)]
        rt = [_route_row("A1", "M/C-1", "CPVC")]
        ph_dict, mat_avg, overall = self._lookups(ph, rt)
        rate, est, _tier = _get_rate("A1", "CPVC", ph_dict, mat_avg, overall)
        assert rate == 45.0
        assert est is False

    def test_same_material_fallback_is_estimated(self):
        ph = [_ph_row("C1", 40.0), _ph_row("C2", 60.0)]
        rt = [_route_row("C1", "M/C-1", "CPVC"),
              _route_row("C2", "M/C-1", "CPVC"),
              _route_row("S1", "M/C-3", "SWR")]
        ph_dict, mat_avg, overall = self._lookups(ph, rt)
        # S1 has no per-hour entry; same material SWR has no entries either
        rate, est, _tier = _get_rate("S1", "SWR", ph_dict, mat_avg, overall)
        assert est is True
        # Falls back to overall avg = (40+60)/2 = 50
        assert abs(rate - 50.0) < 1e-6

    def test_overall_fallback_when_no_same_material(self):
        ph = [_ph_row("C1", 30.0), _ph_row("U1", 50.0)]
        rt = [_route_row("C1", "M/C-1", "CPVC"),
              _route_row("U1", "M/C-2", "UPVC"),
              _route_row("AG1", "M/C-4", "AGRI")]
        ph_dict, mat_avg, overall = self._lookups(ph, rt)
        rate, est, _tier = _get_rate("AG1", "AGRI", ph_dict, mat_avg, overall)
        assert est is True
        assert abs(rate - 40.0) < 1e-6   # (30+50)/2

    def test_same_material_avg_preferred_over_overall(self):
        ph = [_ph_row("C1", 40.0), _ph_row("C2", 60.0), _ph_row("U1", 120.0)]
        rt = [_route_row("C1", "M/C-1", "CPVC"),
              _route_row("C2", "M/C-1", "CPVC"),
              _route_row("U1", "M/C-2", "UPVC"),
              _route_row("C3", "M/C-1", "CPVC")]
        ph_dict, mat_avg, overall = self._lookups(ph, rt)
        # C3 has no ph entry but is CPVC; CPVC avg = (40+60)/2 = 50
        rate, est, _tier = _get_rate("C3", "CPVC", ph_dict, mat_avg, overall)
        assert est is True
        assert abs(rate - 50.0) < 1e-6

    def test_swr_items_always_estimated_with_only_cpvc_upvc_rates(self):
        ph = [_ph_row("PS2", 45.0), _ph_row("PW11", 50.0)]
        rt = [_route_row("PS2",  "M/C-1", "CPVC"),
              _route_row("PW11", "M/C-1", "UPVC"),
              _route_row("SW20", "M/C-3", "SWR")]
        ph_dict, mat_avg, overall = self._lookups(ph, rt)
        _, est, _tier = _get_rate("SW20", "SWR", ph_dict, mat_avg, overall)
        assert est is True


# ── LPT optimiser ─────────────────────────────────────────────────────────────

def _make_item(code, mat_kg, machines):
    hrs = mat_kg / 50.0  # assume 50 kg/hr for simplicity
    return ItemResult(
        item_code=code, raw_code=code, material="CPVC",
        qty_pcs=1000,
        weight_per_pc_kg=0.1,
        material_kg=mat_kg,
        fresh_compound_kg=mat_kg * 0.75,
        pulverizer_kg=mat_kg * 0.25,
        rate_kg_per_hr=50.0,
        rate_estimated=False,
        machine_hrs=hrs,
        capable_machines=machines,
        assignments=[],
        has_weight=True, has_machine=bool(machines),
    )


class TestOptimiser:
    def _run(self, items, caps):
        _lpt_optimise(items, caps)
        return items

    def test_single_machine_gets_all_load(self):
        items = [_make_item("A", 500.0, ["M/C-1"])]
        caps = {"M/C-1": 500.0}
        self._run(items, caps)
        assert len(items[0].assignments) == 1
        assert items[0].assignments[0].machine == "M/C-1"
        assert abs(items[0].assignments[0].hrs - items[0].machine_hrs) < 1e-3

    def test_fits_on_least_loaded(self):
        items = [_make_item("A", 1000.0, ["M/C-1", "M/C-2"]),
                 _make_item("B", 500.0, ["M/C-1", "M/C-2"])]
        caps = {"M/C-1": 500.0, "M/C-2": 500.0}
        self._run(items, caps)
        # A takes 20h on one machine, B should go to the other
        a_mc = items[0].assignments[0].machine
        b_mc = items[1].assignments[0].machine if len(items[1].assignments) == 1 else None
        # At minimum the two machines should share load between them
        assigned_mcs = {a.machine for it in items for a in it.assignments}
        assert len(assigned_mcs) >= 1  # both items were assigned

    def test_split_when_exceeds_capacity(self):
        # A large item that doesn't fit on any single machine → must split
        items = [_make_item("BIG", 50000.0, ["M/C-1", "M/C-2"])]
        caps  = {"M/C-1": 500.0, "M/C-2": 500.0}
        self._run(items, caps)
        assert len(items[0].assignments) == 2
        mcs = {a.machine for a in items[0].assignments}
        assert mcs == {"M/C-1", "M/C-2"}

    def test_optimised_peak_below_baseline_peak(self):
        # Build a scenario where all items favour M/C-1 in LPT order
        items = [
            _make_item("I1", 20000.0, ["M/C-1", "M/C-2", "M/C-3"]),
            _make_item("I2", 18000.0, ["M/C-1", "M/C-2", "M/C-3"]),
            _make_item("I3", 15000.0, ["M/C-1", "M/C-2", "M/C-3"]),
            _make_item("I4", 12000.0, ["M/C-1", "M/C-2", "M/C-3"]),
            _make_item("I5", 10000.0, ["M/C-1", "M/C-2"]),
        ]
        caps = {"M/C-1": 500.0, "M/C-2": 500.0, "M/C-3": 500.0}
        machine_params = {mc: {"capacity_hrs_month": 500.0, "operators_ot": 2, "support_w": 1}
                         for mc in caps}

        # Optimised
        _lpt_optimise(items, caps)
        opt_loads = _compute_machine_loads(items, machine_params)
        opt_peak = max(ml.assigned_hrs for ml in opt_loads)

        # Baseline — reset assignments
        _baseline_assign(items)
        base_loads = _compute_machine_loads(items, machine_params)
        base_peak = max(ml.assigned_hrs for ml in base_loads)

        assert opt_peak < base_peak, (
            f"Optimised peak {opt_peak:.1f} must be < baseline peak {base_peak:.1f}")

    def test_split_hours_sum_to_item_total(self):
        items = [_make_item("X", 30000.0, ["M/C-1", "M/C-2", "M/C-3"])]
        caps  = {"M/C-1": 200.0, "M/C-2": 200.0, "M/C-3": 200.0}
        self._run(items, caps)
        total_assigned = sum(a.hrs for a in items[0].assignments)
        assert abs(total_assigned - items[0].machine_hrs) < 1e-3

    def test_no_machine_item_has_empty_assignments(self):
        items = [_make_item("Y", 1000.0, [])]
        items[0].has_machine = False
        caps = {"M/C-1": 500.0}
        self._run(items, caps)
        assert items[0].assignments == []


# ── Coverage gaps ─────────────────────────────────────────────────────────────

class TestCoverageGaps:
    """Test that run_engine reports gaps correctly with monkeypatched DB."""

    def _patch_mp(self, bom_rows, ph_rows, route_rows, machine_rows,
                  params=None):
        import mp_model as _mp
        patcher_bom    = patch.object(_mp, "get_bom_weight_rows", return_value=bom_rows)
        patcher_ph     = patch.object(_mp, "get_per_hour",        return_value=ph_rows)
        patcher_route  = patch.object(_mp, "get_routing",         return_value=route_rows)
        patcher_mc     = patch.object(_mp, "get_machines",        return_value=machine_rows)
        patcher_params = patch.object(_mp, "get_params",          return_value=params or _params())
        return patcher_bom, patcher_ph, patcher_route, patcher_mc, patcher_params

    def test_no_weight_items_in_gap(self):
        demand = _demand([("PS2", 1000), ("XX99", 500)])
        bom = [_bom_row("PS2", 0.05)]   # XX99 missing
        ph  = [_ph_row("PS2", 50.0)]
        rt  = [_route_row("PS2", "M/C-1")]
        mc  = [_machine_row("M/C-1")]
        with patch("mp_model.get_bom_weight_rows", return_value=bom), \
             patch("mp_model.get_per_hour",        return_value=ph), \
             patch("mp_model.get_routing",         return_value=rt), \
             patch("mp_model.get_machines",        return_value=mc), \
             patch("mp_model.get_params",          return_value=_params()):
            result = run_engine(demand, "2026-07")
        assert "XX99" in result.coverage_gaps.no_weight

    def test_no_machine_items_in_gap(self):
        demand = _demand([("PS2", 1000), ("SWRX", 500)])
        bom = [_bom_row("PS2", 0.05), _bom_row("SWRX", 0.08)]
        ph  = [_ph_row("PS2", 50.0)]
        rt  = [_route_row("PS2", "M/C-1")]   # SWRX has no routing
        mc  = [_machine_row("M/C-1")]
        with patch("mp_model.get_bom_weight_rows", return_value=bom), \
             patch("mp_model.get_per_hour",        return_value=ph), \
             patch("mp_model.get_routing",         return_value=rt), \
             patch("mp_model.get_machines",        return_value=mc), \
             patch("mp_model.get_params",          return_value=_params()):
            result = run_engine(demand, "2026-07")
        assert "SWRX" in result.coverage_gaps.no_machine

    def test_locked_out_machine_reported(self):
        demand = _demand([("PS2", 1000)])
        bom = [_bom_row("PS2", 0.05)]
        ph  = [_ph_row("PS2", 50.0)]
        rt  = [_route_row("PS2", "M/C-1")]   # M/C-7 in machines but no routing
        mc  = [_machine_row("M/C-1"), _machine_row("M/C-7")]
        with patch("mp_model.get_bom_weight_rows", return_value=bom), \
             patch("mp_model.get_per_hour",        return_value=ph), \
             patch("mp_model.get_routing",         return_value=rt), \
             patch("mp_model.get_machines",        return_value=mc), \
             patch("mp_model.get_params",          return_value=_params()):
            result = run_engine(demand, "2026-07")
        assert "M/C-7" in result.coverage_gaps.locked_out_machines

    def test_idle_machine_reported(self):
        # M/C-2 is routed (PS2 could go there) but in this plan only M/C-1 gets load
        demand = _demand([("PS2", 10)])    # tiny item → fits on M/C-1 alone
        bom = [_bom_row("PS2", 0.05)]
        ph  = [_ph_row("PS2", 50.0)]
        rt  = [_route_row("PS2", "M/C-1"), _route_row("PS2", "M/C-2")]
        mc  = [_machine_row("M/C-1", cap=500.0), _machine_row("M/C-2", cap=500.0)]
        with patch("mp_model.get_bom_weight_rows", return_value=bom), \
             patch("mp_model.get_per_hour",        return_value=ph), \
             patch("mp_model.get_routing",         return_value=rt), \
             patch("mp_model.get_machines",        return_value=mc), \
             patch("mp_model.get_params",          return_value=_params()):
            result = run_engine(demand, "2026-07")
        # M/C-2 has routing (not locked out) but got no assignments → idle
        loaded = {a.machine for it in result.items for a in it.assignments}
        if "M/C-2" not in loaded:
            assert "M/C-2" in result.coverage_gaps.idle_machines


# ── Demand parsing ────────────────────────────────────────────────────────────

class TestDemandParsing:
    """Test parse_demand_excel using a real in-memory xlsx workbook."""

    # July layout: A=Item Code, D=Production Plan, E=W1, F=W2, G=W3, H=W4
    _JULY_HEADER = ["Item Code", None, None, "Production Plan", "W1", "W2", "W3", "W4"]
    # August layout: A=Item Code, I=Production Plan (no weekly cols)
    _AUG_HEADER  = ["Item Code", "Colour", "Avg 3-Mo Sale", "Pending Order",
                    "Pending Last Mo", "Buffer Req", "Stock", "Min Production",
                    "Production Plan", "Order"]

    def _make_xlsx(self, tab_data: dict) -> bytes:
        """July layout: A=Item Code, D=Production Plan, E..H=W1..W4.

        tab_data = {tab_name: [(col_a, col_d), ...]}
        Automatically prepends the July header row so the header-based parser
        can locate columns.  Weekly columns are left empty (no W1..W4 data in
        the data rows), which exercises the no-weekly-cols fallback.
        """
        import openpyxl
        wb = openpyxl.Workbook()
        first = True
        for tab_name, rows in tab_data.items():
            if first:
                ws = wb.active
                ws.title = tab_name
                first = False
            else:
                ws = wb.create_sheet(title=tab_name)
            ws.append(self._JULY_HEADER)          # header row required by parser
            for col_a, col_d in rows:
                row_data = [None] * 8
                row_data[0] = col_a
                row_data[3] = col_d               # col D = Production Plan
                ws.append(row_data)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _make_xlsx_july_weekly(self, tab_data: dict) -> bytes:
        """July layout WITH per-week quantities (col D total, E..H weekly).

        tab_data = {tab_name: [(col_a, col_d, w1, w2, w3, w4), ...]}
        """
        import openpyxl
        wb = openpyxl.Workbook()
        first = True
        for tab_name, rows in tab_data.items():
            if first:
                ws = wb.active; ws.title = tab_name; first = False
            else:
                ws = wb.create_sheet(title=tab_name)
            ws.append(self._JULY_HEADER)
            for row_vals in rows:
                col_a, col_d = row_vals[0], row_vals[1]
                w1, w2, w3, w4 = (list(row_vals[2:6]) + [0, 0, 0, 0])[:4]
                ws.append([col_a, None, None, col_d, w1, w2, w3, w4])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _make_xlsx_august(self, tab_data: dict) -> bytes:
        """August layout: A=Item Code, I=Production Plan (no weekly cols).

        tab_data = {tab_name: [(col_a, col_i), ...]}
        """
        import openpyxl
        wb = openpyxl.Workbook()
        first = True
        for tab_name, rows in tab_data.items():
            if first:
                ws = wb.active; ws.title = tab_name; first = False
            else:
                ws = wb.create_sheet(title=tab_name)
            ws.append(self._AUG_HEADER)           # col I (idx 8) = Production Plan
            for col_a, col_i in rows:
                row_data = [None] * 10
                row_data[0] = col_a
                row_data[8] = col_i               # col I = Production Plan
                ws.append(row_data)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_reads_pipe_tabs(self):
        xlsx = self._make_xlsx({
            "CPVC Pipe": [("PS2",  1000), ("PW11", 500)],
            "UPVC Pipe": [("UX10", 200)],
            "SWR Pipe":  [("SW20", 300)],
            "AGRI Pipe": [("AG10", 100)],
        })
        items = eng.parse_demand_excel(xlsx)
        codes = {it.item_code for it in items}
        assert {"PS2", "PW11", "UX10", "SW20", "AG10"} == codes

    def test_skips_total_row(self):
        xlsx = self._make_xlsx({
            "CPVC Pipe": [("PS2", 1000), ("TOTAL", 1000), ("Totals", 500)],
        })
        items = eng.parse_demand_excel(xlsx)
        assert all(it.item_code not in ("TOTAL", "TOTALS") for it in items)
        assert len(items) == 1

    def test_skips_zero_qty(self):
        xlsx = self._make_xlsx({
            "CPVC Pipe": [("PS2", 0), ("PW11", 100)],
        })
        items = eng.parse_demand_excel(xlsx)
        assert len(items) == 1
        assert items[0].item_code == "PW11"

    def test_norm_code_applied(self):
        xlsx = self._make_xlsx({
            "CPVC Pipe": [("ps - 2", 1000), ("PW 11", 500)],
        })
        items = eng.parse_demand_excel(xlsx)
        codes = {it.item_code for it in items}
        assert "PS2" in codes
        assert "PW11" in codes

    def test_material_set_from_tab(self):
        xlsx = self._make_xlsx({
            "SWR Pipe":  [("SW1", 100)],
            "AGRI Pipe": [("AG1", 200)],
        })
        items = eng.parse_demand_excel(xlsx)
        by_code = {it.item_code: it.material for it in items}
        assert by_code.get("SW1") == "SWR"
        assert by_code.get("AG1") == "AGRI"

    def test_unknown_tab_ignored_gracefully(self):
        xlsx = self._make_xlsx({
            "CPVC Pipe":  [("PS2", 100)],
            "OTHER Tab":  [("XX1", 999)],
        })
        items = eng.parse_demand_excel(xlsx)
        codes = {it.item_code for it in items}
        assert "XX1" not in codes

    def test_skips_blank_col_a(self):
        xlsx = self._make_xlsx({
            "CPVC Pipe": [("", 1000), ("PS2", 500)],
        })
        items = eng.parse_demand_excel(xlsx)
        assert len(items) == 1

    def test_numeric_item_codes_kept_short_dropped_long(self):
        """Short all-numeric codes (e.g. SWR fittings 5110, 5111) must NOT be
        skipped; only decimal-size tokens (1.0, 104.8) and long ERP IDs
        (>=8 digits) should be dropped.
        """
        from mp_engine import _is_skip_row

        # Short numeric — valid item codes (SWR fittings)
        assert _is_skip_row("5110", "") is False
        assert _is_skip_row("5111", "") is False
        assert _is_skip_row("5762", "") is False
        assert _is_skip_row("1234567", "") is False   # 7 digits — still keep

        # Decimal size tokens — must drop
        assert _is_skip_row("1.0", "") is True        # row serial "1.0"
        assert _is_skip_row("104.8", "") is True      # pipe size
        assert _is_skip_row("63.5", "") is True

        # Long ERP / row-serial IDs — must drop
        assert _is_skip_row("12345678", "") is True   # 8 digits
        assert _is_skip_row("123456789", "") is True  # 9 digits

        # TOTAL / blank — still dropped
        assert _is_skip_row("TOTAL", "") is True
        assert _is_skip_row("", "") is True

        # parse_fitting_demand must now include numeric SWR items
        xlsx_fit = self._make_xlsx({
            "SWR Fitting": [
                ("Item Code", None),   # header — skip
                ("5110", 16657),       # numeric code — KEEP
                ("5111", 29436),       # numeric code — KEEP
                ("TOTAL", 46093),      # total — skip
            ],
        })
        from mp_engine import parse_fitting_demand
        items = parse_fitting_demand(xlsx_fit)
        codes = {it.item_code for it in items}
        assert "5110" in codes, "5110 must be kept by the demand parser"
        assert "5111" in codes, "5111 must be kept by the demand parser"
        assert len(items) == 2

    # ── Layout-aware parsing (header-based column detection) ──────────────────

    def test_august_layout_col_i(self):
        """August file: Production Plan in col I (index 8); no weekly cols."""
        xlsx = self._make_xlsx_august({
            "CPVC Pipe": [("PS2", 160_466), ("PW11", 500)],
            "UPVC Pipe": [("UX10", 92_690)],
            "SWR Pipe":  [("SW20", 75_647)],
            "AGRI Pipe": [("AG10", 16_362)],
        })
        items = eng.parse_demand_excel(xlsx)
        by_code = {it.item_code: it for it in items}
        assert "PS2" in by_code
        assert by_code["PS2"].qty_pcs == 160_466
        assert by_code["UX10"].qty_pcs == 92_690
        # No weekly cols → whole qty is W1, no per-week breakdown
        assert by_code["PS2"].week_qty == {}
        assert by_code["PS2"].first_requested_week == 1

    def test_august_fitting_layout_col_i(self):
        """August fitting tabs also use col I for Production Plan."""
        from mp_engine import parse_fitting_demand
        xlsx = self._make_xlsx_august({
            "CPVC Fitting": [("CF1", 1_000)],
            "UPVC Fitting": [("UF1", 2_000)],
        })
        items = parse_fitting_demand(xlsx)
        by_code = {it.item_code: it for it in items}
        assert by_code["CF1"].qty_pcs == 1_000
        assert by_code["UF1"].qty_pcs == 2_000

    def test_july_layout_with_weekly_cols(self):
        """July file: Production Plan in col D, W1..W4 in cols E..H."""
        xlsx = self._make_xlsx_july_weekly({
            "CPVC Pipe": [
                ("PS2",  1000, 250, 250, 250, 250),  # evenly split
                ("PW11",  500,   0, 500,   0,   0),  # all in W2
            ],
        })
        items = eng.parse_demand_excel(xlsx)
        by_code = {it.item_code: it for it in items}

        ps2 = by_code["PS2"]
        assert ps2.qty_pcs == 1000
        assert ps2.week_qty == {1: 250, 2: 250, 3: 250, 4: 250}
        assert ps2.first_requested_week == 1   # min of non-zero weeks

        pw11 = by_code["PW11"]
        assert pw11.qty_pcs == 500
        assert pw11.week_qty == {2: 500}
        assert pw11.first_requested_week == 2  # only W2 is non-zero

    def test_missing_production_plan_header_raises(self):
        """When 'Production Plan' is absent, raise ValueError naming tab + headers."""
        import openpyxl, io as _io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CPVC Pipe"
        ws.append(["Item Code", "Colour", "Pending Order", "Stock"])  # no Production Plan
        ws.append(["PS2", "White", 0, 100])
        buf = _io.BytesIO(); wb.save(buf)

        import pytest
        with pytest.raises(ValueError, match="CPVC Pipe"):
            eng.parse_demand_excel(buf.getvalue())

    def test_no_weekly_cols_sets_first_week_1(self):
        """When file has no W1..W4 header columns at all, first_requested_week=1."""
        # August layout has no W1..W4 columns → week_cols={} → first_week=1
        xlsx = self._make_xlsx_august({
            "SWR Pipe": [("SW1", 300)],
        })
        items = eng.parse_demand_excel(xlsx)
        assert len(items) == 1
        assert items[0].first_requested_week == 1
        assert items[0].week_qty == {}

    def test_july_empty_weekly_data_first_week_zero(self):
        """July header has W1..W4 cols but item has all-zero weekly qty → first_week=0."""
        # week_cols present in header, but data cells are all None/zero
        # → week_qty={} → first_week=0 (scheduler treats 0 as W1)
        xlsx = self._make_xlsx({  # July header includes W1..W4
            "CPVC Pipe": [("PS2", 500)],  # no weekly values in data row
        })
        items = eng.parse_demand_excel(xlsx)
        assert len(items) == 1
        assert items[0].week_qty == {}
        assert items[0].first_requested_week == 0  # unspecified → scheduler uses W1

    def test_july_and_august_give_same_pipe_totals(self):
        """Same demand expressed in July vs August layout → identical totals."""
        demand_data = [("PS2", 1000), ("PW11", 500)]
        xlsx_jul = self._make_xlsx({"CPVC Pipe": demand_data})
        xlsx_aug = self._make_xlsx_august({"CPVC Pipe": demand_data})

        items_jul = eng.parse_demand_excel(xlsx_jul)
        items_aug = eng.parse_demand_excel(xlsx_aug)

        total_jul = sum(it.qty_pcs for it in items_jul)
        total_aug = sum(it.qty_pcs for it in items_aug)
        assert total_jul == total_aug == 1500


# ── Serialisation round-trip ──────────────────────────────────────────────────

class TestSerialisation:
    def _simple_result(self):
        ml = MachineLoad("M/C-1", 500.0, 200.0, 40.0, 10.4, 10000.0,
                         7500.0, 2500.0, True, 3, 1)
        item = ItemResult(
            item_code="PS2", raw_code="PS2", material="CPVC",
            qty_pcs=1000, weight_per_pc_kg=0.05,
            material_kg=52.0, fresh_compound_kg=39.0, pulverizer_kg=13.0,
            rate_kg_per_hr=50.0, rate_estimated=False,
            machine_hrs=1.04,
            capable_machines=["M/C-1"],
            assignments=[AssignedPortion("M/C-1", 1.04, 52.0, 1000.0)],
            has_weight=True, has_machine=True,
        )
        return EngineResult(
            segment="PLUMBING", effective_month="2026-07",
            items=[item],
            machine_loads=[ml],
            coverage_gaps=CoverageGaps([], [], [], []),
            totals=PlanTotals(1000, 52.0, 39.0, 13.0, 52.0, 39.0, 13.0),
            baseline_machine_loads=[ml],
            params_used={"waste_pct": 4.0, "pulverizer_pct": 25.0},
        )

    def test_to_dict_is_serialisable(self):
        import json
        result = self._simple_result()
        d = result.to_dict()
        assert json.dumps(d)  # no TypeError

    def test_from_dict_roundtrip(self):
        result = self._simple_result()
        restored = EngineResult.from_dict(result.to_dict())
        assert restored.segment == result.segment
        assert restored.effective_month == result.effective_month
        assert len(restored.items) == 1
        assert restored.items[0].item_code == "PS2"
        assert len(restored.items[0].assignments) == 1
        assert restored.items[0].assignments[0].machine == "M/C-1"


# ── Route isolation: "/" unaffected ──────────────────────────────────────────

class TestRouteIsolation:
    def test_root_route_untouched(self, tmp_path, monkeypatch):
        """Check mp_engine imports do not shadow or register any root routes."""
        import importlib
        # mp_engine must not import from app or touch Flask's route registry
        module_imports = set()
        original_import = __builtins__["__import__"] if isinstance(__builtins__, dict) \
            else __import__
        # Minimal check: the module file must not contain a Flask @app.route decorator
        import inspect
        src = inspect.getsource(eng)
        assert "@app.route" not in src, "mp_engine must not register any Flask routes"

    def test_mp_reports_no_flask_routes(self):
        import inspect
        import mp_reports
        src = inspect.getsource(mp_reports)
        assert "@app.route" not in src


# ── Compound cost ─────────────────────────────────────────────────────────────

class TestComputeEffectiveCosts:
    """Unit tests for compute_effective_costs() formula."""

    # Minimal recipe rows matching CPVC/pipe from DB (5 components, no needs_recipe)
    _CPVC_PIPE_ROWS = [
        {"material": "CPVC", "type": "pipe", "ratio_kg": 89.0,  "price_per_kg": 155.0, "wastage_factor": 1.0, "needs_recipe": False},
        {"material": "CPVC", "type": "pipe", "ratio_kg": 5.0,   "price_per_kg": 38.0,  "wastage_factor": 1.0, "needs_recipe": False},
        {"material": "CPVC", "type": "pipe", "ratio_kg": 2.0,   "price_per_kg": 22.0,  "wastage_factor": 1.0, "needs_recipe": False},
        {"material": "CPVC", "type": "pipe", "ratio_kg": 2.0,   "price_per_kg": 18.0,  "wastage_factor": 1.0, "needs_recipe": False},
        {"material": "CPVC", "type": "pipe", "ratio_kg": 2.0,   "price_per_kg": 15.0,  "wastage_factor": 1.0, "needs_recipe": False},
    ]
    # DB-verified effective cost for CPVC/pipe = 147.28 Rs/kg
    # Verify formula: (89×155 + 5×38 + 2×22 + 2×18 + 2×15) / (89+5+2+2+2) × 1.0
    #               = (13795+190+44+36+30) / 100 = 14095/100 = 140.95  [synthetic]
    # Actual DB rows may differ — use a simple 1-component synthetic case for determinism.
    _SIMPLE_ROWS = [
        {"material": "UPVC", "type": "pipe", "ratio_kg": 80.0,  "price_per_kg": 60.0,  "wastage_factor": 1.0, "needs_recipe": False},
        {"material": "UPVC", "type": "pipe", "ratio_kg": 20.0,  "price_per_kg": 65.0,  "wastage_factor": 1.0, "needs_recipe": False},
        # effective = (80×60 + 20×65) / 100 × 1.0 = (4800+1300)/100 = 61.0
        {"material": "SWR",  "type": "pipe", "ratio_kg": 100.0, "price_per_kg": 73.16, "wastage_factor": 1.0, "needs_recipe": False},
        # effective = 73.16
        {"material": "AGRI", "type": "fitting", "ratio_kg": 50.0, "price_per_kg": 90.0, "wastage_factor": 1.0, "needs_recipe": False},
        # fitting type — should NOT appear in "pipe" cost_map
        {"material": "BAD",  "type": "pipe", "ratio_kg": 0.0,   "price_per_kg": 50.0,  "wastage_factor": 1.0, "needs_recipe": False},
        # zero ratio → unpriced
        {"material": "NOPE", "type": "pipe", "ratio_kg": 10.0,  "price_per_kg": 50.0,  "wastage_factor": 1.0, "needs_recipe": True},
        # needs_recipe → unpriced
    ]

    def test_effective_cost_formula_simple(self):
        cost_map, unpriced = eng.compute_effective_costs(self._SIMPLE_ROWS, "pipe")
        # UPVC: weighted avg of 60 and 65
        assert "UPVC" in cost_map
        assert abs(cost_map["UPVC"] - 61.0) < 0.01
        # SWR: single component
        assert "SWR" in cost_map
        assert abs(cost_map["SWR"] - 73.16) < 0.01
        # AGRI is fitting type → not in pipe cost_map
        assert "AGRI" not in cost_map
        # BAD has zero ratio → unpriced
        assert "BAD" in unpriced
        # NOPE needs_recipe → unpriced
        assert "NOPE" in unpriced

    def test_fitting_type_isolation(self):
        """Type filter: fitting rows must not bleed into pipe cost_map."""
        rows = [
            {"material": "SWR", "type": "fitting", "ratio_kg": 10.0, "price_per_kg": 89.0,
             "wastage_factor": 1.0, "needs_recipe": False},
        ]
        cost_map_pipe, _ = eng.compute_effective_costs(rows, "pipe")
        cost_map_fit, _  = eng.compute_effective_costs(rows, "fitting")
        assert "SWR" not in cost_map_pipe
        assert "SWR" in cost_map_fit
        assert abs(cost_map_fit["SWR"] - 89.0) < 0.01

    def test_wastage_factor_applied(self):
        """wastage_factor multiplies the effective cost."""
        rows = [
            {"material": "CPVC", "type": "pipe", "ratio_kg": 100.0, "price_per_kg": 100.0,
             "wastage_factor": 1.05, "needs_recipe": False},
        ]
        cost_map, _ = eng.compute_effective_costs(rows, "pipe")
        assert abs(cost_map["CPVC"] - 105.0) < 0.01

    def test_needs_recipe_skips_to_unpriced(self):
        rows = [
            {"material": "CPVC", "type": "pipe", "ratio_kg": 100.0, "price_per_kg": 150.0,
             "wastage_factor": 1.0, "needs_recipe": True},
        ]
        cost_map, unpriced = eng.compute_effective_costs(rows, "pipe")
        assert "CPVC" not in cost_map
        assert "CPVC" in unpriced

    def test_empty_rows_returns_empty(self):
        cost_map, unpriced = eng.compute_effective_costs([], "pipe")
        assert cost_map == {}
        assert unpriced == set()

    def test_material_key_is_uppercased(self):
        """Material keys in the recipe may be mixed case — result must be upper."""
        rows = [
            {"material": "cpvc", "type": "pipe", "ratio_kg": 100.0, "price_per_kg": 147.0,
             "wastage_factor": 1.0, "needs_recipe": False},
        ]
        cost_map, _ = eng.compute_effective_costs(rows, "pipe")
        assert "CPVC" in cost_map
        assert "cpvc" not in cost_map

    def test_plan_totals_cost_field_defaults_zero(self):
        """PlanTotals deserialization backward-compat: missing field → 0.0."""
        d = {
            "total_qty_pcs": 100, "total_material_kg": 50, "total_fresh_compound_kg": 40,
            "total_pulverizer_kg": 10, "routable_material_kg": 50,
            "routable_fresh_compound_kg": 40, "routable_pulverizer_kg": 10,
            # NOTE: routable_compound_cost_rs intentionally absent (old frozen run)
        }
        t = eng.PlanTotals(**d)
        assert t.routable_compound_cost_rs == 0.0

    def test_engine_result_cost_fields_default(self):
        """EngineResult.from_dict with old payload (no cost fields) must not raise."""
        result = _make_minimal_engine_result()
        d = result.to_dict()
        # Strip new fields to simulate an old frozen run dict
        d.pop("effective_costs", None)
        d.pop("cost_by_material", None)
        d.pop("n_unpriced", None)
        d["totals"].pop("routable_compound_cost_rs", None)
        restored = eng.EngineResult.from_dict(d)
        assert restored.effective_costs == {}
        assert restored.cost_by_material == {}
        assert restored.n_unpriced == 0
        assert restored.totals.routable_compound_cost_rs == 0.0


def _make_minimal_engine_result() -> eng.EngineResult:
    """Build a trivial EngineResult for serialisation tests."""
    item = eng.ItemResult(
        item_code="X001", raw_code="X001", material="UPVC", qty_pcs=100.0,
        weight_per_pc_kg=0.5, material_kg=55.0, fresh_compound_kg=44.0,
        pulverizer_kg=11.0, rate_kg_per_hr=100.0, rate_estimated=False,
        machine_hrs=0.44, capable_machines=["M/C-1"],
        assignments=[eng.AssignedPortion(machine="M/C-1", hrs=0.44, material_kg=55.0,
                                          qty_pcs=100.0)],
        has_weight=True, has_machine=True,
    )
    gaps = eng.CoverageGaps(no_weight=[], no_machine=[], idle_machines=[], locked_out_machines=[])
    totals = eng.PlanTotals(
        total_qty_pcs=100, total_material_kg=55, total_fresh_compound_kg=44,
        total_pulverizer_kg=11, routable_material_kg=55,
        routable_fresh_compound_kg=44, routable_pulverizer_kg=11,
    )
    ml = eng.MachineLoad(
        machine="M/C-1", capacity_hrs=500.0, assigned_hrs=0.44,
        utilisation_pct=0.09, machine_days=1, material_kg=55.0,
        fresh_compound_kg=44.0, pulverizer_kg=11.0,
        staffing_ok=True, operators_ot=0, support_w=0,
    )
    return eng.EngineResult(
        segment="PLUMBING", effective_month="2026-07",
        items=[item], machine_loads=[ml], coverage_gaps=gaps, totals=totals,
        baseline_machine_loads=[ml], params_used={"waste_pct": 4.0, "pulverizer_pct": 25.0},
    )

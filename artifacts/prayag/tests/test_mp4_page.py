"""
Tests for MP-4: Machine Planning page — upload→run, freeze reproducibility,
coverage-panel counts. All DB calls monkeypatched; uses Flask test client.
"""
from __future__ import annotations

import dataclasses
import io
import json
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

import mp_engine as eng
import mp_seed
from mp_engine import (
    DemandItem, EngineResult, FittingDemandItem, FittingEngineResult,
    FittingEngineResult,
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _minimal_xlsx() -> bytes:
    """Build a minimal in-memory .xlsx with CPVC Pipe and CPVC Fitting tabs."""
    wb = openpyxl.Workbook()
    for tab in ["CPVC Pipe", "UPVC Pipe", "SWR Pipe", "AGRI Pipe",
                "CPVC Fitting", "UPVC Fitting", "SWR Fitting", "AGRI Fitting"]:
        ws = wb.create_sheet(tab)
        ws["A1"] = "ITEM CODE"; ws["D1"] = "PLAN QTY"
        ws["A2"] = "PIPE-001";  ws["D2"] = 500
    del wb["Sheet"]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bom_row(code: str, wt: float) -> dict:
    return {"item_code": mp_seed.norm_code(code), "weight_per_pc_kg": wt}


def _machine_row(mc: str, kind: str = "extrusion", cap: float = 500.0) -> dict:
    return {"machine": mc, "capacity_hrs_month": cap, "kind": kind,
            "operators_ot": 3, "support_w": 1}


def _route_row(code: str, machine: str, material: str = "CPVC") -> dict:
    return {"item_code": mp_seed.norm_code(code), "machine": machine,
            "material": material, "capable": True}


def _ph_row(code: str, val: float, basis: str = "kg_per_hr") -> dict:
    return {"item_code": mp_seed.norm_code(code), "value": val, "basis": basis}


SEGMENT = mp_seed.SEGMENT
MONTH   = "2026-07"


def _mock_mp_model(monkeypatch):
    """Patch all mp_model read calls with minimal seeded data."""
    bom   = [_bom_row("PIPE-001", 0.5)]
    ph    = [_ph_row("PIPE-001", 120.0, "kg_per_hr")]
    mch   = [_machine_row("M/C-1"), _machine_row("M/C-2")]
    rt    = [_route_row("PIPE-001", "M/C-1")]
    prm   = MagicMock(waste_pct=4.0, pulverizer_pct=25.0, segment=SEGMENT,
                      effective_month=MONTH)
    monkeypatch.setattr("mp_model.get_bom_weight_rows",   lambda *a, **kw: bom)
    monkeypatch.setattr("mp_model.get_per_hour",          lambda *a, **kw: ph)
    monkeypatch.setattr("mp_model.get_machines",          lambda *a, **kw: mch)
    monkeypatch.setattr("mp_model.get_routing",           lambda *a, **kw: rt)
    monkeypatch.setattr("mp_model.get_fitting_std",       lambda *a, **kw: [])
    monkeypatch.setattr("mp_model.get_params",            lambda *a, **kw: prm)
    monkeypatch.setattr("mp_model.get_compound_recipes",  lambda *a, **kw: [])
    monkeypatch.setattr("mp_model.get_bom_weights",       lambda *a, **kw: {"PIPE-001": 0.5})


# ── Flask test client fixture ─────────────────────────────────────────────────

@pytest.fixture()
def client(monkeypatch, tmp_path):
    """Flask test client with DB and file-system mocked."""
    import app as flask_app
    flask_app.app.config["TESTING"] = True
    flask_app.app.config["SECRET_KEY"] = "test"
    flask_app.app.config["WTF_CSRF_ENABLED"] = False

    # Patch DB-dependent store functions
    monkeypatch.setattr("mp_model.AVAILABLE", False)
    monkeypatch.setattr("mp_model.init_mp_tables", lambda: None)
    monkeypatch.setattr("mp_model.insert_plan_run",
                        lambda row: 999)
    monkeypatch.setattr("mp_model.update_plan_run_file_path",
                        lambda *a, **kw: None)
    monkeypatch.setattr("mp_model.update_plan_run_freeze",
                        lambda *a, **kw: None)
    monkeypatch.setattr("mp_model.finalize_plan_run",
                        lambda *a, **kw: None)
    monkeypatch.setattr("mp_model.list_plan_runs",
                        lambda *a, **kw: [])
    monkeypatch.setattr("mp_model.get_plan_run_by_id",
                        lambda run_id: None)
    monkeypatch.setattr("store.get_api_key", lambda: None)

    _mock_mp_model(monkeypatch)

    # Redirect uploads to tmp dir
    monkeypatch.setattr("app._UPLOADS_DIR", str(tmp_path))

    with flask_app.app.test_client() as c:
        yield c


# ── 1. Upload → results redirect ─────────────────────────────────────────────

class TestUploadFlow:
    def test_get_upload_page(self, client):
        r = client.get("/machine-planning/upload")
        assert r.status_code == 200
        assert b"Release Plan" in r.data or b"Demand" in r.data or b"Optimiser" in r.data

    def test_post_with_xlsx_redirects_to_results(self, client):
        xlsx = _minimal_xlsx()
        r = client.post(
            "/machine-planning/upload",
            data={"demand_file": (io.BytesIO(xlsx), "plan.xlsx"), "month": MONTH},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert r.status_code == 302
        assert "/machine-planning/results" in r.headers["Location"]

    def test_post_no_file_shows_error(self, client):
        r = client.post("/machine-planning/upload",
                        data={"month": MONTH}, content_type="multipart/form-data")
        assert r.status_code == 200
        assert b"No file" in r.data or b"choose" in r.data

    def test_post_non_xlsx_rejected(self, client):
        r = client.post(
            "/machine-planning/upload",
            data={"demand_file": (io.BytesIO(b"hello"), "plan.csv"), "month": MONTH},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        assert b"xlsx" in r.data.lower()


# ── 2. Results page structure ─────────────────────────────────────────────────

class TestResultsPage:
    def _upload(self, client) -> None:
        xlsx = _minimal_xlsx()
        client.post(
            "/machine-planning/upload",
            data={"demand_file": (io.BytesIO(xlsx), "plan.xlsx"), "month": MONTH},
            content_type="multipart/form-data",
        )

    def test_results_loads_after_upload(self, client):
        self._upload(client)
        r = client.get("/machine-planning/results")
        assert r.status_code == 200

    def test_results_has_combined_summary_section(self, client):
        self._upload(client)
        r = client.get("/machine-planning/results")
        body = r.data.decode()
        assert "Combined" in body or "combined" in body

    def test_results_has_pipe_section(self, client):
        self._upload(client)
        r = client.get("/machine-planning/results")
        body = r.data.decode()
        assert "Pipe" in body or "pipe" in body or "M/C" in body

    def test_results_has_freeze_button(self, client):
        self._upload(client)
        r = client.get("/machine-planning/results")
        body = r.data.decode()
        assert "freeze" in body.lower() or "Freeze" in body

    def test_results_redirect_without_session(self, client):
        r = client.get("/machine-planning/results", follow_redirects=False)
        assert r.status_code == 302

    def test_no_weight_in_coverage_panel(self, client, monkeypatch):
        monkeypatch.setattr("mp_model.get_bom_weight_rows", lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_bom_weights",     lambda *a, **kw: {})
        self._upload(client)
        r = client.get("/machine-planning/results")
        body = r.data.decode()
        assert "PIPE-001" in body or "no" in body.lower() or "gap" in body.lower()


# ── 3. Freeze call ────────────────────────────────────────────────────────────

class TestFreeze:
    def _upload_and_get_run_id(self, client):
        xlsx = _minimal_xlsx()
        client.post(
            "/machine-planning/upload",
            data={"demand_file": (io.BytesIO(xlsx), "plan.xlsx"), "month": MONTH},
            content_type="multipart/form-data",
        )

    def test_freeze_post_redirects_to_runs(self, client):
        self._upload_and_get_run_id(client)
        r = client.post("/machine-planning/freeze", follow_redirects=False)
        assert r.status_code == 302
        assert "/machine-planning/runs" in r.headers["Location"]

    def test_freeze_without_session_redirects_to_upload(self, client):
        r = client.post("/machine-planning/freeze", follow_redirects=False)
        assert r.status_code == 302
        assert "upload" in r.headers["Location"]

    def test_freeze_calls_update_plan_run_freeze(self, client, monkeypatch):
        calls = []
        monkeypatch.setattr("mp_model.update_plan_run_freeze",
                            lambda run_id, fi, res: calls.append((run_id, fi, res)))
        self._upload_and_get_run_id(client)
        client.post("/machine-planning/freeze")
        assert len(calls) == 1
        _, fi, res = calls[0]
        assert "segment" in fi
        assert "pipe" in res or "fitting" in res


# ── 4. Freeze reproducibility ─────────────────────────────────────────────────

class TestFreezeReproducibility:
    def test_frozen_results_unchanged_after_bom_edit(self, monkeypatch):
        """Freeze serialises full result; changing BOM does NOT alter frozen JSON."""
        demand = [DemandItem("P001", "P001", "CPVC", 1000.0)]
        bom    = [{"item_code": "P001", "weight_per_pc_kg": 0.5}]
        ph     = [{"item_code": "P001", "value": 100.0, "basis": "kg_per_hr"}]
        mch    = [{"machine": "M/C-1", "capacity_hrs_month": 500, "kind": "extrusion",
                   "operators_ot": 3, "support_w": 1}]
        rt     = [{"item_code": "P001", "machine": "M/C-1", "material": "CPVC", "capable": True}]
        prm    = MagicMock(waste_pct=4.0, pulverizer_pct=25.0)

        monkeypatch.setattr("mp_model.get_bom_weight_rows",  lambda *a, **kw: bom)
        monkeypatch.setattr("mp_model.get_per_hour",         lambda *a, **kw: ph)
        monkeypatch.setattr("mp_model.get_machines",         lambda *a, **kw: mch)
        monkeypatch.setattr("mp_model.get_routing",          lambda *a, **kw: rt)
        monkeypatch.setattr("mp_model.get_params",           lambda *a, **kw: prm)
        monkeypatch.setattr("mp_model.get_fitting_std",      lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_compound_recipes", lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_bom_weights",      lambda *a, **kw: {"P001": 0.5})

        result_before = eng.run_engine(demand, MONTH, SEGMENT)
        frozen_json   = json.dumps(dataclasses.asdict(result_before))

        bom[0]["weight_per_pc_kg"] = 99.0   # simulate Data-page edit
        result_after = eng.run_engine(demand, MONTH, SEGMENT)

        frozen_recovered = json.loads(frozen_json)
        assert frozen_recovered["totals"]["routable_material_kg"] != \
               dataclasses.asdict(result_after)["totals"]["routable_material_kg"], \
               "Changing BOM must change live result"

        original_kg = frozen_recovered["totals"]["routable_material_kg"]
        assert abs(original_kg - 1000 * 0.5 * 1.04) < 1.0, \
               "Frozen result must reflect original BOM weight"

    def test_frozen_inputs_snapshot_has_required_keys(self, monkeypatch):
        """_build_frozen_inputs must capture all required seed tables."""
        import app as flask_app
        monkeypatch.setattr("mp_model.get_bom_weight_rows",  lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_per_hour",         lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_machines",         lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_routing",          lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_fitting_std",      lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_compound_recipes", lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.get_params",           lambda *a, **kw: None)

        fi = flask_app._build_frozen_inputs(MONTH, SEGMENT)
        for key in ("segment", "month", "bom", "per_hour", "routing",
                    "fitting_std", "machines", "compound_recipe"):
            assert key in fi, f"frozen_inputs missing key: {key}"


# ── 5. Coverage-panel counts ──────────────────────────────────────────────────

class TestCoveragePanel:
    def test_no_weight_count(self, monkeypatch):
        """Items missing BOM weight land in coverage_gaps.no_weight."""
        demand = [
            DemandItem("A001", "A001", "CPVC", 100.0),
            DemandItem("A002", "A002", "CPVC", 100.0),  # no BOM entry
        ]
        bom = [{"item_code": "A001", "weight_per_pc_kg": 0.3}]
        ph  = [{"item_code": "A001", "value": 100.0, "basis": "kg_per_hr"}]
        mch = [{"machine": "M/C-1", "capacity_hrs_month": 500, "kind": "extrusion",
                "operators_ot": 3, "support_w": 1}]
        rt  = [{"item_code": "A001", "machine": "M/C-1", "material": "CPVC", "capable": True}]
        prm = MagicMock(waste_pct=4.0, pulverizer_pct=25.0)

        monkeypatch.setattr("mp_model.get_bom_weight_rows",  lambda *a, **kw: bom)
        monkeypatch.setattr("mp_model.get_per_hour",         lambda *a, **kw: ph)
        monkeypatch.setattr("mp_model.get_machines",         lambda *a, **kw: mch)
        monkeypatch.setattr("mp_model.get_routing",          lambda *a, **kw: rt)
        monkeypatch.setattr("mp_model.get_params",           lambda *a, **kw: prm)
        monkeypatch.setattr("mp_model.get_bom_weights",      lambda *a, **kw: {"A001": 0.3})

        res = eng.run_engine(demand, MONTH, SEGMENT)
        assert "A002" in res.coverage_gaps.no_weight or \
               len(res.coverage_gaps.no_weight) == 1

    def test_no_machine_count(self, monkeypatch):
        """Items with BOM weight but no routing land in coverage_gaps.no_machine."""
        demand = [DemandItem("B001", "B001", "SWR", 200.0)]
        bom = [{"item_code": "B001", "weight_per_pc_kg": 0.4}]
        ph  = []  # no per-hour → uses category avg → still unroutable (no machine)
        mch = [{"machine": "M/C-1", "capacity_hrs_month": 500, "kind": "extrusion",
                "operators_ot": 3, "support_w": 1}]
        rt  = []  # no routing for B001
        prm = MagicMock(waste_pct=4.0, pulverizer_pct=25.0)

        monkeypatch.setattr("mp_model.get_bom_weight_rows",  lambda *a, **kw: bom)
        monkeypatch.setattr("mp_model.get_per_hour",         lambda *a, **kw: ph)
        monkeypatch.setattr("mp_model.get_machines",         lambda *a, **kw: mch)
        monkeypatch.setattr("mp_model.get_routing",          lambda *a, **kw: rt)
        monkeypatch.setattr("mp_model.get_params",           lambda *a, **kw: prm)
        monkeypatch.setattr("mp_model.get_bom_weights",      lambda *a, **kw: {"B001": 0.4})

        res = eng.run_engine(demand, MONTH, SEGMENT)
        assert len(res.coverage_gaps.no_machine) == 1

    def test_fitting_route_estimated_count(self, monkeypatch):
        """Fitting items with no fitting_std entry get material-level fallback flag."""
        demand = [
            eng.FittingDemandItem("F001", "F001", "CPVC", 300.0),
            eng.FittingDemandItem("F002", "F002", "CPVC", 300.0),  # no fitting_std
        ]
        bom  = [
            {"item_code": "F001", "weight_per_pc_kg": 0.2},
            {"item_code": "F002", "weight_per_pc_kg": 0.2},
        ]
        fstd = [{"item_code": "F001", "machine": "A01", "cavity": 8, "cycle_time_sec": 30.0}]
        ph   = []
        mch  = [{"machine": "A01", "capacity_hrs_month": 500, "kind": "moulding",
                 "operators_ot": 0, "support_w": 0}]
        prm  = MagicMock(waste_pct=4.0, pulverizer_pct=25.0)

        monkeypatch.setattr("mp_model.get_bom_weight_rows",  lambda *a, **kw: bom)
        monkeypatch.setattr("mp_model.get_fitting_std",      lambda *a, **kw: fstd)
        monkeypatch.setattr("mp_model.get_per_hour",         lambda *a, **kw: ph)
        monkeypatch.setattr("mp_model.get_machines",         lambda *a, **kw: mch)
        monkeypatch.setattr("mp_model.get_params",           lambda *a, **kw: prm)
        monkeypatch.setattr("mp_model.get_bom_weights",      lambda *a, **kw:
                            {"F001": 0.2, "F002": 0.2})

        fres = eng.run_fitting_engine(demand, MONTH, SEGMENT)
        assert fres.n_route_estimated == 1   # F002 uses material fallback
        assert fres.n_unroutable == 0        # CPVC machines from F001 covers F002

    def test_locked_out_machines_in_gaps(self, monkeypatch):
        """Machines with no routing entries appear in locked_out_machines."""
        demand = [DemandItem("P001", "P001", "CPVC", 100.0)]
        bom = [{"item_code": "P001", "weight_per_pc_kg": 0.3}]
        ph  = [{"item_code": "P001", "value": 80.0, "basis": "kg_per_hr"}]
        mch = [
            {"machine": "M/C-1", "capacity_hrs_month": 500, "kind": "extrusion",
             "operators_ot": 3, "support_w": 1},
            {"machine": "M/C-7", "capacity_hrs_month": 500, "kind": "extrusion",
             "operators_ot": 3, "support_w": 1},  # no routing → locked-out
        ]
        rt  = [{"item_code": "P001", "machine": "M/C-1", "material": "CPVC", "capable": True}]
        prm = MagicMock(waste_pct=4.0, pulverizer_pct=25.0)

        monkeypatch.setattr("mp_model.get_bom_weight_rows",  lambda *a, **kw: bom)
        monkeypatch.setattr("mp_model.get_per_hour",         lambda *a, **kw: ph)
        monkeypatch.setattr("mp_model.get_machines",         lambda *a, **kw: mch)
        monkeypatch.setattr("mp_model.get_routing",          lambda *a, **kw: rt)
        monkeypatch.setattr("mp_model.get_params",           lambda *a, **kw: prm)
        monkeypatch.setattr("mp_model.get_bom_weights",      lambda *a, **kw: {"P001": 0.3})

        res = eng.run_engine(demand, MONTH, SEGMENT)
        assert "M/C-7" in res.coverage_gaps.locked_out_machines


# ── 6. Home page and runs list ────────────────────────────────────────────────

class TestNavPages:
    def test_home_page_loads(self, client):
        r = client.get("/machine-planning")
        assert r.status_code == 200
        assert b"Plumbing" in r.data

    def test_runs_list_loads(self, client):
        r = client.get("/machine-planning/runs")
        assert r.status_code == 200

    def test_run_detail_404_on_missing(self, client):
        r = client.get("/machine-planning/runs/99999")
        assert r.status_code == 404

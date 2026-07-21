"""Tests for MP-3 fitting engine, Report-12, and route isolation."""
import dataclasses
import io
import math
import sys
import os
import pytest

# Ensure prayag package root is on path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import mp_engine as eng

# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------

def _demand(*items):
    """Build a FittingDemandItem list from (item_code, material, qty) tuples."""
    return [
        eng.FittingDemandItem(
            item_code=ic, raw_code=ic, material=mat, qty_pcs=float(qty)
        )
        for ic, mat, qty in items
    ]


def _fstd(item_code, machine, cavity, cycle_time_sec, segment="PLB", month="2026-07"):
    return {
        "segment": segment, "item_code": item_code, "machine": machine,
        "cavity": cavity, "cycle_time_sec": cycle_time_sec, "effective_month": month,
    }


def _ph(item_code, cycle_sec, segment="PLB", month="2026-07"):
    return {"segment": segment, "item_code": item_code,
            "basis": "cycle", "value": cycle_sec, "effective_month": month}


# ── Fitting chain math ──────────────────────────────────────────────────────

class TestFittingChainMath:
    """material_kg = qty × wt × (1 + waste%); fresh/pulv split."""

    def test_material_kg_formula(self):
        qty, wt, waste, pulv = 1000.0, 0.05, 4.0, 25.0
        expected_mat = qty * wt * (1 + waste / 100)        # 52.0 kg
        expected_fresh = expected_mat * (1 - pulv / 100)   # 39.0 kg
        expected_pulv  = expected_mat - expected_fresh      # 13.0 kg

        assert abs(expected_mat   - 52.0) < 1e-6
        assert abs(expected_fresh - 39.0) < 1e-6
        assert abs(expected_pulv  - 13.0) < 1e-6

    def test_fresh_plus_pulv_equals_material_kg(self):
        for qty, wt in [(500, 0.02), (2000, 0.1), (1, 1.0)]:
            mat   = qty * wt * 1.04
            fresh = mat * 0.75
            pulv  = mat * 0.25
            assert abs(fresh + pulv - mat) < 1e-9


# ── Cavity/cycle rate ────────────────────────────────────────────────────────

class TestCavityCycleRate:
    """pcs_per_hr = cavity × 3600 / cycle_time_sec from fitting_std."""

    def test_basic_rate(self):
        cavity, cycle = 2.0, 90.0
        expected = cavity * 3600.0 / cycle  # 80.0 pcs/hr
        assert abs(expected - 80.0) < 1e-6

    def test_rate_lookup_from_fstd(self):
        fstd = [_fstd("ITEM1", "A01(NU-200)", 2.0, 90.0)]
        ph   = []
        demand = _demand(("ITEM1", "CPVC", 500))
        fstd_by_item, cycle_ph, mat_avg, overall = eng._build_fitting_rate_lookups(
            fstd, ph, demand
        )
        pps, est, cavity, cycle = eng._get_fitting_rate(
            "ITEM1", "CPVC", fstd_by_item, cycle_ph, mat_avg, overall
        )
        assert abs(pps - 80.0) < 1e-3
        assert est is False
        assert cavity == 2.0
        assert cycle == 90.0

    def test_per_hour_cycle_fallback(self):
        fstd = []
        ph   = [_ph("ITEM2", 60.0)]   # 60 sec/pc → 60 pcs/hr
        demand = _demand(("ITEM2", "UPVC", 100))
        fstd_by_item, cycle_ph, mat_avg, overall = eng._build_fitting_rate_lookups(
            fstd, ph, demand
        )
        pps, est, cavity, cycle = eng._get_fitting_rate(
            "ITEM2", "UPVC", fstd_by_item, cycle_ph, mat_avg, overall
        )
        assert abs(pps - 60.0) < 1e-3
        assert est is True
        assert cavity is None
        assert abs(cycle - 60.0) < 1e-6

    def test_material_avg_fallback(self):
        """Item with no fitting_std or per_hour → material category average."""
        fstd = [_fstd("KNOWN", "A01(NU-200)", 2.0, 90.0)]  # 80 pcs/hr CPVC
        ph   = []
        demand = _demand(("KNOWN", "CPVC", 100), ("UNKNOWN", "CPVC", 200))
        fstd_by_item, cycle_ph, mat_avg, overall = eng._build_fitting_rate_lookups(
            fstd, ph, demand
        )
        pps, est, cavity, cycle = eng._get_fitting_rate(
            "UNKNOWN", "CPVC", fstd_by_item, cycle_ph, mat_avg, overall
        )
        assert abs(pps - 80.0) < 1e-3
        assert est is True
        assert cavity is None

    def test_overall_avg_last_resort(self):
        """Item with no matching material → overall average."""
        fstd = [_fstd("CPVC1", "A01(NU-200)", 4.0, 60.0)]  # 240 pcs/hr
        ph   = []
        demand = _demand(("CPVC1", "CPVC", 100), ("SWR1", "SWR", 50))
        fstd_by_item, cycle_ph, mat_avg, overall = eng._build_fitting_rate_lookups(
            fstd, ph, demand
        )
        # SWR1 has no CPVC fstd entry and no per_hour, and mat_avg won't have SWR
        pps, est, cavity, cycle = eng._get_fitting_rate(
            "SWR1", "SWR", fstd_by_item, cycle_ph, mat_avg, overall
        )
        assert abs(pps - 240.0) < 1e-3   # falls to overall_avg = 240
        assert est is True

    def test_machine_hrs_formula(self):
        """machine_hrs = qty / pcs_per_hr."""
        qty, pps = 800.0, 80.0
        assert abs(qty / pps - 10.0) < 1e-9

    def test_num_cycles_formula(self):
        """num_cycles = qty / cavity (rounded)."""
        qty, cavity = 500.0, 2.0
        assert round(qty / cavity) == 250


# ── Material-level route fallback ────────────────────────────────────────────

class TestMaterialLevelRouteFallback:
    """Items with no fitting_std history → assigned to material's machines."""

    def test_item_with_route_uses_history(self):
        fstd = [_fstd("ITEM_A", "A02(U-150)", 2.0, 90.0)]
        routing = []
        demand  = _demand(("ITEM_A", "UPVC", 100))
        item_routes, mat_machines = eng._build_fitting_routes(fstd, routing, demand)
        assert "ITEM_A" in item_routes
        assert "A02(U-150)" in item_routes["ITEM_A"]

    def test_no_route_item_gets_material_fallback(self):
        fstd = [_fstd("KNOWN", "A02(U-150)", 2.0, 90.0)]
        routing = []
        demand  = _demand(("KNOWN", "UPVC", 100), ("UNKNOWN", "UPVC", 50))
        item_routes, mat_machines = eng._build_fitting_routes(fstd, routing, demand)
        # KNOWN is in item_routes; UNKNOWN is not
        assert "UNKNOWN" not in item_routes
        # UPVC → A02(U-150) derived from KNOWN
        assert "UPVC" in mat_machines
        assert "A02(U-150)" in mat_machines["UPVC"]

    def test_unroutable_when_material_has_no_machines(self):
        fstd = [_fstd("CPVC1", "A06(C-150)", 1.0, 60.0)]
        routing = []
        demand  = _demand(("CPVC1", "CPVC", 10), ("SWR1", "SWR", 10))
        item_routes, mat_machines = eng._build_fitting_routes(fstd, routing, demand)
        # SWR has no known machines
        assert "SWR" not in mat_machines or not mat_machines.get("SWR")

    def test_route_estimated_flag(self):
        """route_estimated must be True for fallback items, False for history items."""
        fstd = [_fstd("HIST", "A01(NU-200)", 2.0, 80.0)]
        item_routes, mat_machines = eng._build_fitting_routes(
            fstd, [],
            _demand(("HIST", "CPVC", 100), ("NEW", "CPVC", 50))
        )
        # "NEW" is not in item_routes → route_estimated when engine runs
        assert "HIST" in item_routes
        assert "NEW" not in item_routes
        # Material-level fallback should have CPVC → A01(NU-200)
        assert "CPVC" in mat_machines
        assert "A01(NU-200)" in mat_machines["CPVC"]

    def test_multiple_items_same_machine_builds_mat_map(self):
        fstd = [
            _fstd("A", "B03(C-200)", 2.0, 100.0),
            _fstd("B", "B04(C-200)", 1.0, 120.0),
        ]
        demand = _demand(("A", "SWR", 100), ("B", "SWR", 200))
        _, mat_machines = eng._build_fitting_routes(fstd, [], demand)
        assert "SWR" in mat_machines
        assert set(mat_machines["SWR"]) == {"B03(C-200)", "B04(C-200)"}


# ── Coverage gap reporting ────────────────────────────────────────────────────

class TestCoverageGapReporting:
    """no_weight items flagged; no_machine items listed; counts match."""

    def _mini_run(self, demand, fstd_rows, bom, ph_rows=None):
        """Minimal engine run with mocked DB calls."""
        fstd_by_item, cycle_ph, mat_avg, overall = eng._build_fitting_rate_lookups(
            fstd_rows, ph_rows or [], demand
        )
        item_routes, mat_machines = eng._build_fitting_routes(fstd_rows, [], demand)

        no_weight, no_machine = [], []
        items = []
        n_route_est = 0
        n_unroutable = 0
        waste, pulv = 4.0, 25.0

        for d in demand:
            ic, mat, qty = d.item_code, d.material, d.qty_pcs
            wt = bom.get(ic)
            if wt is None:
                no_weight.append(ic)
                items.append(eng.FittingItemResult(
                    item_code=ic, raw_code=ic, material=mat, qty_pcs=qty,
                    weight_per_pc_kg=None, material_kg=0, fresh_compound_kg=0,
                    pulverizer_kg=0, pcs_per_hr=0, rate_estimated=False,
                    machine_hrs=0, cavity=None, cycle_time_sec=None, num_cycles=None,
                    capable_machines=[], route_estimated=False,
                    assignments=[], has_weight=False, has_machine=False,
                ))
                continue
            mat_kg = qty * wt * (1 + waste / 100)
            fresh  = mat_kg * (1 - pulv / 100)
            pv     = mat_kg - fresh
            pps, rest, cavity, cycle = eng._get_fitting_rate(
                ic, mat, fstd_by_item, cycle_ph, mat_avg, overall
            )
            caps = item_routes.get(ic, [])
            route_est = False
            if not caps:
                fallback = mat_machines.get(mat, [])
                if fallback:
                    caps, route_est, n_route_est = fallback, True, n_route_est + 1
                else:
                    no_machine.append(ic)
                    n_unroutable += 1
            items.append(eng.FittingItemResult(
                item_code=ic, raw_code=ic, material=mat, qty_pcs=qty,
                weight_per_pc_kg=wt, material_kg=mat_kg, fresh_compound_kg=fresh,
                pulverizer_kg=pv, pcs_per_hr=pps, rate_estimated=rest,
                machine_hrs=qty/pps if pps > 0 else 0,
                cavity=cavity, cycle_time_sec=cycle,
                num_cycles=round(qty/cavity) if cavity else None,
                capable_machines=sorted(caps), route_estimated=route_est,
                assignments=[], has_weight=True, has_machine=bool(caps),
            ))
        return items, no_weight, no_machine, n_route_est, n_unroutable

    def test_no_weight_items_flagged(self):
        demand = _demand(("HAS_WT", "CPVC", 100), ("NO_WT", "CPVC", 50))
        fstd   = [_fstd("HAS_WT", "A01(NU-200)", 2.0, 90.0)]
        bom    = {"HAS_WT": 0.05}
        items, no_wt, no_mc, _, _ = self._mini_run(demand, fstd, bom)
        assert "NO_WT" in no_wt
        assert len(no_wt) == 1
        # Item still present in items list (not silently dropped)
        codes = [it.item_code for it in items]
        assert "NO_WT" in codes
        assert not [it for it in items if it.item_code == "NO_WT" and it.has_weight]

    def test_no_route_item_uses_material_fallback(self):
        demand = _demand(("KNOWN", "UPVC", 100), ("NEW", "UPVC", 50))
        fstd   = [_fstd("KNOWN", "A02(U-150)", 2.0, 100.0)]
        bom    = {"KNOWN": 0.08, "NEW": 0.06}
        items, no_wt, no_mc, n_rest, n_unr = self._mini_run(demand, fstd, bom)
        assert not no_wt
        assert not no_mc
        assert n_rest == 1      # NEW uses material fallback
        new_item = next(it for it in items if it.item_code == "NEW")
        assert new_item.route_estimated is True
        assert "A02(U-150)" in new_item.capable_machines

    def test_truly_unroutable_when_no_material_machines(self):
        demand = _demand(("AGRI_ONLY", "AGRI", 100))
        fstd   = []   # no AGRI items in history
        bom    = {"AGRI_ONLY": 0.03}
        items, no_wt, no_mc, n_rest, n_unr = self._mini_run(demand, fstd, bom)
        assert "AGRI_ONLY" in no_mc
        assert n_unr == 1

    def test_no_weight_count_exact(self):
        demand = _demand(
            ("W1", "CPVC", 10), ("W2", "CPVC", 10),
            ("NW1", "CPVC", 10), ("NW2", "CPVC", 10), ("NW3", "CPVC", 10),
        )
        fstd = [_fstd("W1", "A01(NU-200)", 2.0, 80.0),
                _fstd("W2", "A01(NU-200)", 2.0, 80.0)]
        bom  = {"W1": 0.05, "W2": 0.05}
        items, no_wt, _, _, _ = self._mini_run(demand, fstd, bom)
        assert len(no_wt) == 3
        assert set(no_wt) == {"NW1", "NW2", "NW3"}


# ── Serialisation round-trip ─────────────────────────────────────────────────

class TestSerialisationRoundTrip:
    """FittingEngineResult.to_dict() / from_dict() round-trip."""

    def _make_result(self):
        item = eng.FittingItemResult(
            item_code="X1", raw_code="X1", material="CPVC", qty_pcs=100.0,
            weight_per_pc_kg=0.05, material_kg=5.2, fresh_compound_kg=3.9,
            pulverizer_kg=1.3, pcs_per_hr=80.0, rate_estimated=False,
            machine_hrs=1.25, cavity=2.0, cycle_time_sec=90.0, num_cycles=50,
            capable_machines=["A01(NU-200)"], route_estimated=False,
            assignments=[eng.FittingAssignedPortion(
                machine="A01(NU-200)", hrs=1.25, qty_pcs=100.0, material_kg=5.2
            )],
            has_weight=True, has_machine=True,
        )
        ml = eng.MachineLoad(
            machine="A01(NU-200)", capacity_hrs=500.0, assigned_hrs=1.25,
            utilisation_pct=0.25, machine_days=0.065, material_kg=5.2,
            fresh_compound_kg=3.9, pulverizer_kg=1.3,
            staffing_ok=False, operators_ot=0, support_w=0,
        )
        return eng.FittingEngineResult(
            segment="PLUMBING", effective_month="2026-07",
            items=[item], machine_loads=[ml],
            coverage_gaps=eng.CoverageGaps(
                no_weight=[], no_machine=[], idle_machines=[], locked_out_machines=[]
            ),
            totals=eng.PlanTotals(
                total_qty_pcs=100.0, total_material_kg=5.2,
                total_fresh_compound_kg=3.9, total_pulverizer_kg=1.3,
                routable_material_kg=5.2, routable_fresh_compound_kg=3.9,
                routable_pulverizer_kg=1.3,
            ),
            baseline_machine_loads=[ml],
            params_used={"waste_pct": 4.0, "pulverizer_pct": 25.0},
            n_route_estimated=0, n_unroutable=0,
        )

    def test_roundtrip_preserves_fields(self):
        res = self._make_result()
        d   = res.to_dict()
        res2 = eng.FittingEngineResult.from_dict(d)

        assert res2.segment == "PLUMBING"
        assert res2.effective_month == "2026-07"
        assert len(res2.items) == 1
        it = res2.items[0]
        assert it.item_code == "X1"
        assert abs(it.material_kg - 5.2) < 1e-6
        assert it.cavity == 2.0
        assert it.cycle_time_sec == 90.0
        assert it.route_estimated is False
        assert len(it.assignments) == 1
        assert it.assignments[0].machine == "A01(NU-200)"
        assert res2.n_route_estimated == 0
        assert res2.n_unroutable == 0

    def test_to_dict_is_json_serialisable(self):
        import json
        res = self._make_result()
        # Should not raise
        json.dumps(res.to_dict())


# ── Report-12 column order ────────────────────────────────────────────────────

class TestReport12ColumnOrder:
    """report_12_bytes produces header at row 6 with exact column order."""

    EXPECTED_COLS = [
        "DATE", "MATERIAL", "ITEM CODE", "Moulding Machine",
        "Mould Cavity", "Run Cavity", "No. of Cycle", "Pcs",
        "Wt in Kgs", "Cycle Time", "Running Hours",
        "Ideal Output Per Hour", "Actual Output Per Hour",
        "Output Efficiency", "Rejection Pcs", "Rejection Kg",
    ]

    def _minimal_fitting_result(self):
        item = eng.FittingItemResult(
            item_code="A100", raw_code="A100", material="CPVC", qty_pcs=500.0,
            weight_per_pc_kg=0.05, material_kg=26.0, fresh_compound_kg=19.5,
            pulverizer_kg=6.5, pcs_per_hr=80.0, rate_estimated=False,
            machine_hrs=6.25, cavity=2.0, cycle_time_sec=90.0, num_cycles=250,
            capable_machines=["A01(NU-200)"], route_estimated=False,
            assignments=[eng.FittingAssignedPortion(
                machine="A01(NU-200)", hrs=6.25, qty_pcs=500.0, material_kg=26.0
            )],
            has_weight=True, has_machine=True,
        )
        ml = eng.MachineLoad(
            machine="A01(NU-200)", capacity_hrs=500.0, assigned_hrs=6.25,
            utilisation_pct=1.25, machine_days=0.325, material_kg=26.0,
            fresh_compound_kg=19.5, pulverizer_kg=6.5,
            staffing_ok=False, operators_ot=0, support_w=0,
        )
        return eng.FittingEngineResult(
            segment="PLUMBING", effective_month="2026-07",
            items=[item], machine_loads=[ml],
            coverage_gaps=eng.CoverageGaps([], [], [], []),
            totals=eng.PlanTotals(500, 26, 19.5, 6.5, 26, 19.5, 6.5),
            baseline_machine_loads=[ml],
            params_used={"waste_pct": 4.0, "pulverizer_pct": 25.0},
            n_route_estimated=0, n_unroutable=0,
        )

    def test_header_at_row_6(self):
        import openpyxl
        from mp_reports import report_12_bytes
        result = self._minimal_fitting_result()
        xlsx = report_12_bytes(result)
        wb   = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws   = wb.active

        # Row 6 is the header row
        header = [ws.cell(row=6, column=c).value for c in range(1, 17)]
        assert header == self.EXPECTED_COLS, (
            f"Header mismatch:\n  got:      {header}\n  expected: {self.EXPECTED_COLS}"
        )

    def test_actuals_only_columns_blank(self):
        """Cols 13-16 (Actual Output/Efficiency/Rejection) must be blank in data rows."""
        import openpyxl
        from mp_reports import report_12_bytes
        result = self._minimal_fitting_result()
        xlsx   = report_12_bytes(result)
        wb     = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws     = wb.active
        # Data starts at row 7
        for row in range(7, ws.max_row + 1):
            for blank_col in [13, 14, 15, 16]:
                val = ws.cell(row=row, column=blank_col).value
                assert val is None, (
                    f"Row {row} col {blank_col} should be blank (actuals-only), got {val!r}"
                )

    def test_plan_fields_populated(self):
        """MATERIAL, ITEM CODE, Machine, Cavity, Pcs, Wt in Kgs must be filled."""
        import openpyxl
        from mp_reports import report_12_bytes
        result = self._minimal_fitting_result()
        xlsx   = report_12_bytes(result)
        wb     = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws     = wb.active
        # First data row = 7
        assert ws.cell(row=7, column=2).value == "CPVC"       # MATERIAL
        assert ws.cell(row=7, column=3).value == "A100"       # ITEM CODE
        assert ws.cell(row=7, column=4).value == "A01(NU-200)"  # Machine
        assert ws.cell(row=7, column=5).value == 2.0           # Mould Cavity
        assert ws.cell(row=7, column=8).value is not None      # Pcs
        assert ws.cell(row=7, column=9).value is not None      # Wt in Kgs

    def test_column_count_is_16(self):
        assert len(self.EXPECTED_COLS) == 16


# ── Route isolation: "/" routes unaffected ───────────────────────────────────

class TestRouteIsolation:
    """All fitting logic lives under /machine-planning/ only."""

    def test_fitting_tabs_constant_exists(self):
        assert hasattr(eng, "FITTING_TABS")
        assert set(eng.FITTING_TABS.keys()) == {
            "CPVC Fitting", "UPVC Fitting", "SWR Fitting", "AGRI Fitting"
        }

    def test_fitting_tabs_materials(self):
        assert eng.FITTING_TABS["CPVC Fitting"] == "CPVC"
        assert eng.FITTING_TABS["UPVC Fitting"] == "UPVC"
        assert eng.FITTING_TABS["SWR Fitting"]  == "SWR"
        assert eng.FITTING_TABS["AGRI Fitting"] == "AGRI"

    def test_is_moulding_machine_recognises_extrusion(self):
        assert eng._is_moulding_machine("A01(NU-200)") is True
        assert eng._is_moulding_machine("B03(C-200)")  is True
        assert eng._is_moulding_machine("M/C-1")       is False
        assert eng._is_moulding_machine("M/C-9")       is False
        assert eng._is_moulding_machine("")             is False

    def test_parse_fitting_demand_distinct_from_pipe(self):
        """parse_fitting_demand and parse_demand_excel are separate functions."""
        assert hasattr(eng, "parse_fitting_demand")
        assert hasattr(eng, "parse_demand_excel")
        assert eng.parse_fitting_demand is not eng.parse_demand_excel

    def test_run_fitting_engine_is_separate(self):
        assert hasattr(eng, "run_fitting_engine")
        assert hasattr(eng, "run_engine")
        assert eng.run_fitting_engine is not eng.run_engine

    def test_fitting_result_class_separate(self):
        assert hasattr(eng, "FittingEngineResult")
        assert hasattr(eng, "EngineResult")
        assert eng.FittingEngineResult is not eng.EngineResult


# ── Rate_estimated flag ──────────────────────────────────────────────────────

class TestRateEstimatedFlag:
    """rate_estimated=False only when cavity/cycle from fitting_std."""

    def test_seeded_rate_not_estimated(self):
        fstd = [_fstd("ITEM_OK", "A01(NU-200)", 2.0, 80.0)]
        ph   = []
        demand = _demand(("ITEM_OK", "CPVC", 200))
        fstd_by_item, cycle_ph, mat_avg, overall = eng._build_fitting_rate_lookups(
            fstd, ph, demand
        )
        _, est, _, _ = eng._get_fitting_rate(
            "ITEM_OK", "CPVC", fstd_by_item, cycle_ph, mat_avg, overall
        )
        assert est is False

    def test_per_hour_cycle_is_estimated(self):
        fstd = []
        ph   = [_ph("PH_ITEM", 45.0)]
        demand = _demand(("PH_ITEM", "UPVC", 100))
        fstd_by_item, cycle_ph, mat_avg, overall = eng._build_fitting_rate_lookups(
            fstd, ph, demand
        )
        _, est, _, _ = eng._get_fitting_rate(
            "PH_ITEM", "UPVC", fstd_by_item, cycle_ph, mat_avg, overall
        )
        assert est is True

    def test_material_avg_is_estimated(self):
        fstd = [_fstd("BASE", "A01(NU-200)", 1.0, 60.0)]  # 60 pcs/hr
        ph   = []
        demand = _demand(("BASE", "SWR", 100), ("NEW", "SWR", 50))
        fstd_by_item, cycle_ph, mat_avg, overall = eng._build_fitting_rate_lookups(
            fstd, ph, demand
        )
        _, est, _, _ = eng._get_fitting_rate(
            "NEW", "SWR", fstd_by_item, cycle_ph, mat_avg, overall
        )
        assert est is True

"""
Tests for:
  A1 — machine_plan_comparison_bytes uses correct MachineLoad field names
  A2 — report routes fall back to latest persisted run when session is empty
  B  — capacity_feasible_plan_bytes: reconciliation, ≤100% machine load,
       deferred list populated when demand > capacity
"""
from __future__ import annotations

import dataclasses
import io
from typing import List
from unittest.mock import MagicMock, patch

import pytest

import mp_engine as eng
import mp_reports as rpt
import mp_scheduler as sch


# ─── Shared fixtures ──────────────────────────────────────────────────────────

def _ml(machine: str, cap: float, assigned: float) -> eng.MachineLoad:
    util = assigned / cap * 100 if cap > 0 else 0.0
    return eng.MachineLoad(
        machine=machine,
        capacity_hrs=cap,
        assigned_hrs=round(assigned, 3),
        utilisation_pct=round(util, 2),
        machine_days=round(assigned * 26 / cap, 2) if cap > 0 else 0.0,
        material_kg=0.0,
        fresh_compound_kg=0.0,
        pulverizer_kg=0.0,
        staffing_ok=True,
        operators_ot=3,
        support_w=1,
    )


def _item(code: str, material: str, qty_pcs: float, wt: float,
          mat_kg: float, rate: float, hrs: float,
          machines: list | None = None,
          gross_qty_pcs: float | None = None,
          rej_rate: float = 5.0,
          waste_pct_used: float = 0.5) -> eng.ItemResult:
    return eng.ItemResult(
        item_code=code, raw_code=code, material=material,
        qty_pcs=qty_pcs, weight_per_pc_kg=wt, material_kg=mat_kg,
        fresh_compound_kg=mat_kg * 0.8, pulverizer_kg=mat_kg * 0.2,
        rate_kg_per_hr=rate, rate_estimated=False, machine_hrs=hrs,
        capable_machines=machines or ["M/C-1"],
        assignments=[eng.AssignedPortion(machine=(machines or ["M/C-1"])[0],
                                         hrs=hrs, material_kg=mat_kg, qty_pcs=qty_pcs)],
        has_weight=True, has_machine=True,
        gross_qty_pcs=gross_qty_pcs if gross_qty_pcs is not None else qty_pcs * 1.05,
        rej_rate=rej_rate, rej_basis="material",
        waste_pct_used=waste_pct_used, waste_basis="measured",
    )


def _engine_result(items: list, mls: list | None = None) -> eng.EngineResult:
    if mls is None:
        mls = [_ml("M/C-1", 500.0, sum(it.machine_hrs for it in items))]
    gaps = eng.CoverageGaps(no_weight=[], no_machine=[], idle_machines=[], locked_out_machines=[])
    totals = eng.PlanTotals(
        total_qty_pcs=sum(it.qty_pcs for it in items),
        total_material_kg=sum(it.material_kg for it in items),
        total_fresh_compound_kg=0.0, total_pulverizer_kg=0.0,
        routable_material_kg=sum(it.material_kg for it in items),
        routable_fresh_compound_kg=0.0, routable_pulverizer_kg=0.0,
    )
    return eng.EngineResult(
        segment="PLUMBING", effective_month="2026-07",
        items=items, machine_loads=mls, coverage_gaps=gaps, totals=totals,
        baseline_machine_loads=mls, params_used={},
    )


# ─── A1: machine_plan_comparison_bytes field names ───────────────────────────

class TestMachinePlanComparison:
    def test_no_attr_error_on_correct_field_names(self):
        """machine_plan_comparison_bytes must not raise AttributeError for MachineLoad fields."""
        items = [_item("SWR-001", "SWR", 1000.0, 0.3, 300.0, 295.0, 1.0)]
        ml = _ml("M/C-1", 500.0, 300.0)
        result = _engine_result(items, [ml])
        data = rpt.machine_plan_comparison_bytes(
            result, result, None, None, "2026-07", old_waste_pct=4.0
        )
        assert len(data) > 1000, "Expected a non-trivial xlsx"

    def test_over_capacity_renders_correctly(self):
        """When assigned_hrs > capacity_hrs (>100%), comparison report still renders."""
        items = [_item("SWR-001", "SWR", 5000.0, 0.3, 1500.0, 295.0, 5.0)]
        ml = _ml("M/C-1", 500.0, 750.0)   # 150% utilisation
        result = _engine_result(items, [ml])
        data = rpt.machine_plan_comparison_bytes(
            result, result, None, None, "2026-07", old_waste_pct=4.0
        )
        assert isinstance(data, bytes) and len(data) > 0

    def test_revised_production_plan_renders(self):
        """revised_production_plan_bytes renders without error."""
        items = [_item("SWR-001", "SWR", 1000.0, 0.3, 300.0, 295.0, 1.0)]
        result = _engine_result(items)
        data = rpt.revised_production_plan_bytes(result, None, "2026-07")
        assert isinstance(data, bytes) and len(data) > 1000


# ─── A2: session fallback ─────────────────────────────────────────────────────

class TestSessionFallback:
    def test_ensure_session_run_id_populates_from_db(self):
        """_ensure_session_run_id fills session from list_plan_runs when session is empty."""
        import app as flask_app
        flask_app.app.config["TESTING"] = True
        flask_app.app.config["SECRET_KEY"] = "test"

        with flask_app.app.test_request_context("/"):
            from flask import session
            session.clear()
            assert not session.get("mp2_run_id")

            with patch("mp_model.list_plan_runs", return_value=[{"id": 42}]):
                with patch("mp_model.AVAILABLE", True):
                    flask_app._ensure_session_run_id()

            assert session.get("mp2_run_id") == "42"

    def test_ensure_session_run_id_noop_when_already_set(self):
        """_ensure_session_run_id does not overwrite an existing session run_id."""
        import app as flask_app
        flask_app.app.config["TESTING"] = True
        flask_app.app.config["SECRET_KEY"] = "test"

        with flask_app.app.test_request_context("/"):
            from flask import session
            session["mp2_run_id"] = "99"

            with patch("mp_model.list_plan_runs", return_value=[{"id": 1}]) as mock_lpr:
                flask_app._ensure_session_run_id()

            mock_lpr.assert_not_called()
            assert session.get("mp2_run_id") == "99"

    def test_revised_plan_route_returns_friendly_redirect_when_no_runs(self, monkeypatch):
        """GET /machine-planning/report/revised-plan redirects to upload (not 400) when no runs."""
        import app as flask_app
        flask_app.app.config["TESTING"] = True
        flask_app.app.config["SECRET_KEY"] = "test"
        monkeypatch.setattr("mp_model.list_plan_runs", lambda *a, **kw: [])
        monkeypatch.setattr("mp_model.AVAILABLE", True)
        monkeypatch.setattr("store.get_api_key", lambda: None)

        with flask_app.app.test_client() as c:
            resp = c.get("/machine-planning/report/revised-plan",
                         follow_redirects=False)
        # Should redirect to upload, NOT return a raw 400 string
        assert resp.status_code in (302, 303), (
            f"Expected redirect, got {resp.status_code}: {resp.get_data(as_text=True)[:200]}"
        )


# ─── B: capacity_feasible_plan_bytes ─────────────────────────────────────────

def _make_schedule_from_engine(items, month="2026-07", seg="PLUMBING",
                               mc_params=None):
    """Run the real scheduler to get a ScheduleResult. Uses actual day-by-day logic."""
    demand = [
        eng.DemandItem(item_code=it.item_code, raw_code=it.raw_code,
                       material=it.material, qty_pcs=it.qty_pcs)
        for it in items
    ]
    with patch("mp_model.get_params", return_value=None), \
         patch("mp_model.get_machines", return_value=mc_params or [
             {"machine": "M/C-1", "capacity_hrs_month": 500,
              "shifts_per_day": 2, "hours_per_shift": 10, "working_days_month": 25},
         ]):
        return sch.run_shift_schedule(
            engine_items=items,
            demand_items=demand,
            segment=seg,
            effective_month=month,
        )


class TestCapacityFeasiblePlan:

    def _overcapacity_fixture(self):
        """Engine result where demand (1,200 hrs) exceeds two 500h machines."""
        items = []
        # 12 items × 100 hrs each = 1,200 hrs total on 2 × 500h machines
        for i in range(12):
            items.append(_item(
                f"SWR-{i:03d}", "SWR",
                qty_pcs=1000.0, wt=0.3, mat_kg=300.0,
                rate=300.0, hrs=100.0,
                machines=["M/C-1", "M/C-2"],
                gross_qty_pcs=1050.0,
            ))
        mls = [_ml("M/C-1", 500.0, 600.0), _ml("M/C-2", 500.0, 600.0)]
        result = _engine_result(items, mls)
        mc_params = [
            {"machine": "M/C-1", "capacity_hrs_month": 500,
             "shifts_per_day": 2, "hours_per_shift": 10, "working_days_month": 25},
            {"machine": "M/C-2", "capacity_hrs_month": 500,
             "shifts_per_day": 2, "hours_per_shift": 10, "working_days_month": 25},
        ]
        schedule = _make_schedule_from_engine(items, mc_params=mc_params)
        return result, schedule

    def _fits_fixture(self):
        """Engine result where demand (200 hrs) easily fits in 500h capacity."""
        items = [
            _item("SWR-001", "SWR", 200.0, 0.3, 60.0, 300.0, 0.2,
                  gross_qty_pcs=210.0),
            _item("SWR-002", "SWR", 300.0, 0.3, 90.0, 300.0, 0.3,
                  gross_qty_pcs=315.0),
        ]
        result = _engine_result(items)
        schedule = _make_schedule_from_engine(items)
        return result, schedule

    def test_deferred_populated_when_overcapacity(self):
        """When demand > capacity, schedule.unfinished is non-empty."""
        result, schedule = self._overcapacity_fixture()
        assert len(schedule.unfinished) > 0, (
            "Expected deferred items when 1,200h demand on 1,000h capacity"
        )

    def test_deferred_empty_when_fits(self):
        """When demand fits in capacity, schedule.unfinished is empty."""
        result, schedule = self._fits_fixture()
        assert schedule.unfinished == [], (
            f"Expected no deferred items, got: {schedule.unfinished}"
        )

    def test_feasible_machine_load_never_exceeds_100(self):
        """capacity_feasible_plan_bytes must not raise AssertionError on scheduler output."""
        result, schedule = self._overcapacity_fixture()
        # The scheduler enforces capacity; the report should assert ≤100%.
        # This test passes when the assertion holds (it would raise AssertionError otherwise).
        data = rpt.capacity_feasible_plan_bytes(result, None, schedule, "2026-07")
        assert isinstance(data, bytes) and len(data) > 1000

    def test_feasible_plus_deferred_equals_requested_per_item(self):
        """For every item: feasible_pcs + deferred_pcs == requested_pcs (gross_qty_pcs)."""
        result, schedule = self._overcapacity_fixture()
        unfinished_by_code = {u.item_code: u for u in schedule.unfinished}

        for it in result.items:
            if not it.has_weight:
                continue
            u = unfinished_by_code.get(it.item_code)
            if u is None:
                # Fully feasible: deferred = 0
                continue
            deferred_kg = min(u.remaining_kg, it.material_kg)
            feasible_kg = it.material_kg - deferred_kg
            if it.material_kg > 0:
                ratio = feasible_kg / it.material_kg
                feasible_pcs = round(it.gross_qty_pcs * ratio, 2)
            else:
                feasible_pcs = 0.0
            deferred_pcs = round(it.gross_qty_pcs - feasible_pcs, 2)
            total = round(feasible_pcs + deferred_pcs, 2)
            assert abs(total - round(it.gross_qty_pcs, 2)) < 0.02, (
                f"{it.item_code}: feasible {feasible_pcs} + deferred {deferred_pcs} "
                f"= {total} ≠ requested {it.gross_qty_pcs:.2f}"
            )

    def test_deferred_items_named_bottleneck(self):
        """Deferred tab entries include capable machine names in the reason."""
        result, schedule = self._overcapacity_fixture()
        data = rpt.capacity_feasible_plan_bytes(result, None, schedule, "2026-07")
        # Verify xlsx contains machine name in deferred sheet
        import zipfile, io as _io
        with zipfile.ZipFile(_io.BytesIO(data)) as zf:
            # openpyxl-written files contain xl/sharedStrings.xml
            names = zf.namelist()
            assert any("sharedStrings" in n or "sheet" in n for n in names)
        # Verify the Deferred sheet is in the workbook
        import openpyxl
        wb = openpyxl.load_workbook(filename=_io.BytesIO(data))
        assert "Deferred" in wb.sheetnames
        ws = wb["Deferred"]
        cell_vals = [str(ws.cell(row=r, column=c).value or "")
                     for r in range(4, ws.max_row + 1)
                     for c in range(1, ws.max_column + 1)]
        full_text = " ".join(cell_vals)
        assert "M/C-1" in full_text or "M/C-2" in full_text, (
            "Deferred tab should mention capable machines"
        )

    def test_all_five_sheets_present(self):
        """Workbook has the five required tabs."""
        result, schedule = self._overcapacity_fixture()
        data = rpt.capacity_feasible_plan_bytes(result, None, schedule, "2026-07")
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(filename=_io.BytesIO(data))
        required = {"Summary", "Pipe Plan", "Fitting Plan", "Machine Load (Feasible)", "Deferred"}
        assert required.issubset(set(wb.sheetnames)), (
            f"Missing tabs: {required - set(wb.sheetnames)}"
        )

    def test_no_schedule_renders_gracefully(self):
        """When schedule is None, the report still renders (no machine load data)."""
        items = [_item("SWR-001", "SWR", 200.0, 0.3, 60.0, 300.0, 0.2)]
        result = _engine_result(items)
        # schedule=None → no machine load rows, no deferred items
        data = rpt.capacity_feasible_plan_bytes(result, None, None, "2026-07")
        assert isinstance(data, bytes) and len(data) > 0

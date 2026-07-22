"""
Machine Planning Optimiser — Phase MP-2.

Parses a weekly release plan Excel, runs per-item chain math from the
mp_* tables, then LPT + parallel-split load balancing across extrusion
machines.

ADDITIVE / ISOLATED: reads only mp_* tables. Never touches the existing
production pipeline (/, /data, /reports, /plan).
"""
from __future__ import annotations

import dataclasses
import io
import re
from collections import defaultdict
from typing import Any, Dict, List, Optional, Set, Tuple

# ── openpyxl (for demand Excel parsing) ─────────────────────────────────────
try:
    import openpyxl
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

# ── local imports (all mp_* — never production pipeline) ────────────────────
from mp_seed import norm_code as _norm_code, SEGMENT as _DEFAULT_SEGMENT
import mp_model as _mp

# ── Constants ────────────────────────────────────────────────────────────────
PIPE_TABS: Dict[str, str] = {
    "CPVC Pipe": "CPVC",
    "UPVC Pipe": "UPVC",
    "SWR Pipe":  "SWR",
    "AGRI Pipe": "AGRI",
}
_COL_ITEM = 0   # Col A — Item Code (0-indexed)
_COL_QTY  = 3   # Col D — Production Plan pcs (0-indexed)
# Weekly columns E..H → W1..W4 (0-indexed 4..7)
_COL_W1   = 4
_COL_W4   = 7

# Default machine-group config for Report-11A–D (overridable via DB later)
REPORT_11_GROUPS: Dict[str, List[str]] = {
    "A": ["M/C-1", "M/C-2"],
    "B": ["M/C-3", "M/C-4"],
    "C": ["M/C-5", "M/C-6"],
    "D": ["M/C-7", "M/C-8", "M/C-9"],
}


# ── Dataclasses ──────────────────────────────────────────────────────────────

@dataclasses.dataclass
class DemandItem:
    item_code: str   # normalised
    raw_code: str    # original text from Excel
    material: str    # CPVC / UPVC / SWR / AGRI
    qty_pcs: float
    # Weekly breakdown (cols E–H). Keys 1..4 = W1..W4.
    # Empty dict = no weekly split provided (all qty in total col D only).
    week_qty: Dict[int, float] = dataclasses.field(default_factory=dict)
    # First week with non-zero qty; 0 = unspecified (treat as W1 in scheduler).
    first_requested_week: int = 0


@dataclasses.dataclass
class AssignedPortion:
    machine: str
    hrs: float
    material_kg: float
    qty_pcs: float


@dataclasses.dataclass
class ItemResult:
    item_code: str
    raw_code: str
    material: str
    qty_pcs: float
    weight_per_pc_kg: Optional[float]
    material_kg: float
    fresh_compound_kg: float
    pulverizer_kg: float
    rate_kg_per_hr: float
    rate_estimated: bool
    machine_hrs: float
    capable_machines: List[str]
    assignments: List[AssignedPortion]
    has_weight: bool
    has_machine: bool
    # "item" = seeded per-item rate; "mat_avg" = per-material average;
    # "overall_avg" = overall pipe average (last-resort fallback)
    rate_fallback_tier: str = "item"


@dataclasses.dataclass
class MachineLoad:
    machine: str
    capacity_hrs: float
    assigned_hrs: float
    utilisation_pct: float
    machine_days: float
    material_kg: float
    fresh_compound_kg: float
    pulverizer_kg: float
    staffing_ok: bool
    operators_ot: int
    support_w: int


@dataclasses.dataclass
class CoverageGaps:
    no_weight: List[str]           # item codes with no BOM weight
    no_machine: List[str]          # item codes with no capable extrusion machine
    idle_machines: List[str]       # machines in routing but got no load this plan
    locked_out_machines: List[str] # extrusion machines with no routing rows at all


@dataclasses.dataclass
class PlanTotals:
    total_qty_pcs: float
    total_material_kg: float
    total_fresh_compound_kg: float
    total_pulverizer_kg: float
    routable_material_kg: float
    routable_fresh_compound_kg: float
    routable_pulverizer_kg: float
    routable_compound_cost_rs: float = 0.0  # fresh_compound × effective cost/kg


@dataclasses.dataclass
class EngineResult:
    segment: str
    effective_month: str
    items: List[ItemResult]
    machine_loads: List[MachineLoad]          # optimised (LPT + split)
    coverage_gaps: CoverageGaps
    totals: PlanTotals
    baseline_machine_loads: List[MachineLoad] # unbalanced (first capable machine)
    params_used: Dict[str, float]
    # Compound cost fields (default empty so old frozen runs deserialise cleanly)
    effective_costs: Dict[str, float] = dataclasses.field(default_factory=dict)
    cost_by_material: Dict[str, float] = dataclasses.field(default_factory=dict)
    n_unpriced: int = 0

    def to_dict(self) -> dict:
        """Return a JSON-serialisable dict (via dataclasses.asdict)."""
        return dataclasses.asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "EngineResult":
        """Reconstruct from a dict stored in Postgres / session."""
        def _ap(x: dict) -> AssignedPortion:
            return AssignedPortion(**x)

        def _item(x: dict) -> ItemResult:
            assignments = [_ap(a) for a in x.pop("assignments", [])]
            return ItemResult(**x, assignments=assignments)

        def _ml(x: dict) -> MachineLoad:
            return MachineLoad(**x)

        gaps = CoverageGaps(**d["coverage_gaps"])
        totals_d = dict(d["totals"])
        totals_d.setdefault("routable_compound_cost_rs", 0.0)
        totals = PlanTotals(**totals_d)
        return EngineResult(
            segment=d["segment"],
            effective_month=d["effective_month"],
            items=[_item(dict(r)) for r in d.get("items", [])],
            machine_loads=[_ml(r) for r in d.get("machine_loads", [])],
            coverage_gaps=gaps,
            totals=totals,
            baseline_machine_loads=[_ml(r) for r in d.get("baseline_machine_loads", [])],
            params_used=d.get("params_used", {}),
            effective_costs=d.get("effective_costs", {}),
            cost_by_material=d.get("cost_by_material", {}),
            n_unpriced=d.get("n_unpriced", 0),
        )


# ── Compound-cost helpers ─────────────────────────────────────────────────────

def compute_effective_costs(
    recipe_rows: List[dict],
    item_type: str,
) -> Tuple[Dict[str, float], Set[str]]:
    """Compute effective compound cost per kg by material for the given item type.

    Effective cost/kg = sum(ratio_kg × price_per_kg) / sum(ratio_kg) × wastage_factor

    Args:
        recipe_rows: rows from mp_compound_recipe for the segment/month.
        item_type: "pipe" or "fitting" (matches the ``type`` column).

    Returns:
        cost_map  : {MATERIAL_UPPER: Rs_per_kg}  — priced recipes only
        unpriced  : set of material names with needs_recipe=True or zero ratio
    """
    groups: Dict[str, list] = defaultdict(list)
    for r in recipe_rows:
        if str(r.get("type", "")).lower() == item_type.lower():
            groups[str(r.get("material", "")).upper()].append(r)

    cost_map: Dict[str, float] = {}
    unpriced: Set[str] = set()

    for mat, rows in groups.items():
        if any(r.get("needs_recipe") for r in rows):
            unpriced.add(mat)
            continue
        total_ratio    = sum(float(r.get("ratio_kg") or 0) for r in rows)
        total_weighted = sum(
            float(r.get("ratio_kg") or 0) * float(r.get("price_per_kg") or 0)
            for r in rows
        )
        wf = float(rows[0].get("wastage_factor") or 1.0)
        if total_ratio > 0:
            cost_map[mat] = round(total_weighted / total_ratio * wf, 4)
        else:
            unpriced.add(mat)

    return cost_map, unpriced


# ── Demand parsing ───────────────────────────────────────────────────────────

def _is_total_row(cell_a: str) -> bool:
    return cell_a.upper().strip() in ("TOTAL", "TOTALS", "GRAND TOTAL")


def _is_skip_row(cell_a: str, cell_d: str) -> bool:
    """True for blank, header, TOTAL, ERP-ID, or decimal-size rows.

    Keeps short all-numeric item codes (e.g. SWR fittings 5110, 5111, 5762).
    Only drops numeric tokens that are clearly NOT item codes:
      - Decimal size values: "104.8", "1.0", "63.5" (digit.digit pattern)
      - Long ERP / row-serial IDs: 8+ consecutive digits
    """
    if not cell_a:
        return True
    if _is_total_row(cell_a):
        return True
    nc = _norm_code(cell_a)
    if not nc:
        return True
    # Decimal size tokens ("104.8", "1.0") — digits.digits, no letters
    if re.match(r'^\d+\.\d+$', nc):
        return True
    # Long all-digit strings are ERP IDs or row serials, not item codes
    if nc.isdigit() and len(nc) >= 8:
        return True
    # Skip header-like rows ("ITEM CODE", "SR.", etc.)
    if re.match(r"^(ITEM|SR|S\.?\s*NO|SERIAL|DESCRIPTION|PRODUCT)", nc, re.I):
        return True
    return False


def _safe_float(val) -> float:
    """Parse a cell value to float, return 0 on failure."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def parse_demand_excel(file_bytes: bytes) -> List[DemandItem]:
    """Parse the weekly release plan Excel and return DemandItem list.

    Reads only the four Pipe tabs (CPVC Pipe / UPVC Pipe / SWR Pipe / AGRI Pipe).
    Col A = Item Code, Col D = Production Plan pcs (total).
    Cols E–H = W1..W4 per-week quantities (optional; zero if column absent).
    Normalises all codes. Skips TOTAL rows, blank rows, and non-item rows.
    """
    if not _OPENPYXL:
        raise RuntimeError("openpyxl is required for demand upload.")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    items: List[DemandItem] = []

    for tab_name, material in PIPE_TABS.items():
        # Case-insensitive tab lookup
        ws = None
        for sn in wb.sheetnames:
            if sn.strip().lower() == tab_name.lower():
                ws = wb[sn]
                break
        if ws is None:
            continue  # tab absent — skip gracefully

        for row in ws.iter_rows(values_only=True):
            raw_a = str(row[_COL_ITEM]).strip() if row[_COL_ITEM] is not None else ""
            raw_d = row[_COL_QTY] if len(row) > _COL_QTY else None

            if _is_skip_row(raw_a, str(raw_d or "")):
                continue

            try:
                qty = float(str(raw_d).replace(",", "").strip())
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue

            # Read W1..W4 from cols E–H (indices 4–7); zero if column absent
            week_qty: Dict[int, float] = {}
            for wk_idx, col in enumerate(range(_COL_W1, _COL_W4 + 1), start=1):
                v = row[col] if len(row) > col else None
                wq = _safe_float(v)
                if wq > 0:
                    week_qty[wk_idx] = wq

            # First week with non-zero quantity
            first_week = min(week_qty) if week_qty else 0

            items.append(DemandItem(
                item_code=_norm_code(raw_a),
                raw_code=raw_a,
                material=material,
                qty_pcs=qty,
                week_qty=week_qty,
                first_requested_week=first_week,
            ))

    return items


# ── Chain math helpers ────────────────────────────────────────────────────────

def _build_rate_lookups(
    per_hour_rows: List[dict],
    routing_rows: List[dict],
) -> Tuple[Dict[str, float], Dict[str, float], float]:
    """
    Returns:
        ph_dict      — {item_code: kg_per_hr}   (only basis='kg_per_hr')
        mat_avg      — {material: avg kg_per_hr}  (for fallback)
        overall_avg  — avg across all kg_per_hr pipe items (final fallback)
    """
    # material of each pipe item (from routing)
    routing_material: Dict[str, str] = {}
    for r in routing_rows:
        if r["machine"].startswith("M/C-"):
            routing_material.setdefault(r["item_code"], r.get("material", ""))

    ph_dict: Dict[str, float] = {}
    ph_by_mat: Dict[str, List[float]] = defaultdict(list)
    for r in per_hour_rows:
        if r["basis"] != "kg_per_hr":
            continue
        ic  = r["item_code"]
        val = float(r["value"])
        ph_dict[ic] = val
        mat = routing_material.get(ic, "")
        if mat:
            ph_by_mat[mat].append(val)

    mat_avg: Dict[str, float] = {
        m: sum(vs) / len(vs) for m, vs in ph_by_mat.items() if vs
    }
    all_vals = list(ph_dict.values())
    overall_avg = sum(all_vals) / len(all_vals) if all_vals else 1.0
    return ph_dict, mat_avg, overall_avg


def _get_rate(
    item_code: str,
    material: str,
    ph_dict: Dict[str, float],
    mat_avg: Dict[str, float],
    overall_avg: float,
) -> Tuple[float, bool, str]:
    """Return (rate_kg_per_hr, rate_estimated, fallback_tier).

    fallback_tier values:
      "item"        — seeded per-item rate (no estimation)
      "mat_avg"     — per-material average of seeded items (rate_estimated=True)
      "overall_avg" — overall pipe average, last resort (rate_estimated=True)
    """
    if item_code in ph_dict:
        return ph_dict[item_code], False, "item"
    if material in mat_avg:
        return mat_avg[material], True, "mat_avg"
    return overall_avg, True, "overall_avg"


# ── Optimiser ────────────────────────────────────────────────────────────────

def _compute_machine_loads(
    items: List[ItemResult],
    machine_params: Dict[str, dict],   # {machine: {capacity_hrs_month, operators_ot, support_w}}
    use_assignments_attr: str = "assignments",
) -> List[MachineLoad]:
    """Aggregate per-machine totals from item.assignments (or item.baseline_assignments)."""
    acc: Dict[str, Dict] = {}
    for mc, p in machine_params.items():
        cap = float(p.get("capacity_hrs_month") or 500.0)
        acc[mc] = {
            "capacity_hrs": cap,
            "assigned_hrs": 0.0,
            "material_kg": 0.0,
            "fresh_compound_kg": 0.0,
            "pulverizer_kg": 0.0,
            "operators_ot": int(p.get("operators_ot") or 0),
            "support_w": int(p.get("support_w") or 0),
        }

    for item in items:
        asgns: List[AssignedPortion] = getattr(item, use_assignments_attr, [])
        for a in asgns:
            if a.machine in acc:
                mat_frac = a.material_kg
                fresh_frac = mat_frac * (1 - item.pulverizer_kg / item.material_kg) if item.material_kg > 0 else 0.0
                pulv_frac  = mat_frac - fresh_frac
                acc[a.machine]["assigned_hrs"]    += a.hrs
                acc[a.machine]["material_kg"]     += mat_frac
                acc[a.machine]["fresh_compound_kg"] += fresh_frac
                acc[a.machine]["pulverizer_kg"]   += pulv_frac

    loads = []
    for mc in sorted(acc.keys()):
        d = acc[mc]
        cap = d["capacity_hrs"]
        hrs = d["assigned_hrs"]
        util = hrs / cap * 100.0 if cap > 0 else 0.0
        days = hrs * 26.0 / cap if cap > 0 else 0.0
        ok   = (d["operators_ot"] > 0 or d["support_w"] > 0)
        loads.append(MachineLoad(
            machine=mc,
            capacity_hrs=cap,
            assigned_hrs=round(hrs, 3),
            utilisation_pct=round(util, 2),
            machine_days=round(days, 2),
            material_kg=round(d["material_kg"], 2),
            fresh_compound_kg=round(d["fresh_compound_kg"], 2),
            pulverizer_kg=round(d["pulverizer_kg"], 2),
            staffing_ok=ok,
            operators_ot=d["operators_ot"],
            support_w=d["support_w"],
        ))
    return loads


def _lpt_optimise(
    items: List[ItemResult],
    machine_caps: Dict[str, float],   # {machine: capacity_hrs_month}
) -> None:
    """
    LPT + parallel-split balancer. Mutates item.assignments in-place.

    Process items largest-material_kg first. For each item:
    - 1 capable machine: assign all.
    - 2+ capable machines:
        - If fits on least-loaded: assign all there.
        - Otherwise: split proportional to remaining capacity across ALL capable.
        - If ALL are full (0 remaining): assign to least-over (overflow).
    """
    # Sort descending by material_kg (LPT key)
    sorted_items = sorted(items, key=lambda x: x.material_kg, reverse=True)

    assigned: Dict[str, float] = {mc: 0.0 for mc in machine_caps}

    for item in sorted_items:
        item.assignments = []
        if not item.has_machine or not item.capable_machines:
            continue

        caps = [mc for mc in item.capable_machines if mc in machine_caps]
        if not caps:
            item.has_machine = False
            continue

        item_hrs = item.machine_hrs
        frac_per_hr = item.material_kg / item_hrs if item_hrs > 0 else 0.0

        if len(caps) == 1:
            mc = caps[0]
            item.assignments = [AssignedPortion(
                machine=mc, hrs=item_hrs,
                material_kg=item.material_kg, qty_pcs=item.qty_pcs,
            )]
            assigned[mc] += item_hrs
            continue

        # Find least-loaded by utilisation%
        least = min(caps, key=lambda mc: assigned[mc] / machine_caps[mc])
        remaining_least = max(0.0, machine_caps[least] - assigned[least])

        if item_hrs <= remaining_least + 1e-6:
            # Fits on least-loaded machine
            item.assignments = [AssignedPortion(
                machine=least, hrs=item_hrs,
                material_kg=item.material_kg, qty_pcs=item.qty_pcs,
            )]
            assigned[least] += item_hrs
        else:
            # Parallel split proportional to remaining capacity
            remainders = {mc: max(0.0, machine_caps[mc] - assigned[mc]) for mc in caps}
            total_rem = sum(remainders.values())

            if total_rem <= 1e-6:
                # All machines at/over capacity — assign overflow to least over
                mc = min(caps, key=lambda mc: assigned[mc] / machine_caps[mc])
                item.assignments = [AssignedPortion(
                    machine=mc, hrs=item_hrs,
                    material_kg=item.material_kg, qty_pcs=item.qty_pcs,
                )]
                assigned[mc] += item_hrs
            else:
                for mc in caps:
                    frac = remainders[mc] / total_rem
                    hrs_frac = item_hrs * frac
                    if hrs_frac < 1e-4:
                        continue
                    item.assignments.append(AssignedPortion(
                        machine=mc,
                        hrs=round(hrs_frac, 4),
                        material_kg=round(item.material_kg * frac, 4),
                        qty_pcs=round(item.qty_pcs * frac, 2),
                    ))
                    assigned[mc] += hrs_frac


def _baseline_assign(items: List[ItemResult]) -> None:
    """
    Unbalanced baseline: each item → first capable machine (sorted), no split.
    Stores result in item.assignments (baseline overwrites; caller saves optimised
    first and restores after).
    """
    for item in items:
        item.assignments = []
        if not item.has_machine or not item.capable_machines:
            continue
        mc = sorted(item.capable_machines)[0]
        item.assignments = [AssignedPortion(
            machine=mc, hrs=item.machine_hrs,
            material_kg=item.material_kg, qty_pcs=item.qty_pcs,
        )]


# ── Main entry point ─────────────────────────────────────────────────────────

def run_engine(
    demand: List[DemandItem],
    effective_month: str,
    segment: str = _DEFAULT_SEGMENT,
) -> EngineResult:
    """
    Run the full MP-2 engine:
    1. Load mp_* inputs from DB for (segment, effective_month).
    2. Apply per-item chain math.
    3. LPT + parallel-split optimisation.
    4. Unbalanced baseline.
    5. Return EngineResult with all outputs.
    """
    # ── Load all inputs from DB ──────────────────────────────────────────────
    params_row = _mp.get_params(segment, effective_month)
    waste_pct     = float(params_row.waste_pct)     if params_row else 4.0
    pulv_pct      = float(params_row.pulverizer_pct) if params_row else 25.0

    bom_rows   = _mp.get_bom_weight_rows(segment, effective_month)
    ph_rows    = _mp.get_per_hour(segment, effective_month)
    routing    = _mp.get_routing(segment, effective_month)
    machines   = _mp.get_machines(segment, effective_month, kind="extrusion")

    bom: Dict[str, float] = {r["item_code"]: float(r["weight_per_pc_kg"]) for r in bom_rows}

    # Per-machine params
    mc_params: Dict[str, dict] = {m["machine"]: m for m in machines}
    mc_caps: Dict[str, float] = {
        mc: float(p.get("capacity_hrs_month") or 500.0)
        for mc, p in mc_params.items()
    }

    # Routing: pipe items → capable machines
    pipe_caps: Dict[str, List[str]] = defaultdict(list)  # {item_code: [machines]}
    routed_machines: set = set()
    for r in routing:
        mc = r["machine"]
        if mc.startswith("M/C-") and r.get("capable", True):
            ic = r["item_code"]
            if mc not in pipe_caps[ic]:
                pipe_caps[ic].append(mc)
            routed_machines.add(mc)

    # Rate lookups — computed from seeded items, then overridden by stored params
    ph_dict, mat_avg, overall_avg = _build_rate_lookups(ph_rows, routing)
    if params_row:
        for _mat, _attr in [
            ("CPVC", "cpvc_mat_rate"), ("UPVC", "upvc_mat_rate"),
            ("SWR",  "swr_mat_rate"),  ("AGRI", "agri_mat_rate"),
        ]:
            _v = float(getattr(params_row, _attr, 0.0) or 0.0)
            if _v > 0.0:
                mat_avg[_mat] = _v

    # ── Per-item chain math ──────────────────────────────────────────────────
    items: List[ItemResult] = []
    no_weight: List[str] = []
    no_machine: List[str] = []

    for d in demand:
        ic  = d.item_code
        mat = d.material
        qty = d.qty_pcs

        wt = bom.get(ic)
        has_weight = wt is not None
        if not has_weight:
            no_weight.append(ic)
            # Still record the item but with zero chain math
            items.append(ItemResult(
                item_code=ic, raw_code=d.raw_code,
                material=mat, qty_pcs=qty,
                weight_per_pc_kg=None,
                material_kg=0.0, fresh_compound_kg=0.0, pulverizer_kg=0.0,
                rate_kg_per_hr=0.0, rate_estimated=False,
                machine_hrs=0.0,
                capable_machines=pipe_caps.get(ic, []),
                assignments=[],
                has_weight=False, has_machine=bool(pipe_caps.get(ic)),
            ))
            continue

        material_kg = qty * wt * (1.0 + waste_pct / 100.0)
        fresh       = material_kg * (1.0 - pulv_pct / 100.0)
        pulv        = material_kg - fresh

        rate, estimated, tier = _get_rate(ic, mat, ph_dict, mat_avg, overall_avg)
        if rate <= 0:
            rate = overall_avg
            estimated = True
            tier = "overall_avg"
        machine_hrs = material_kg / rate if rate > 0 else 0.0

        caps = pipe_caps.get(ic, [])
        has_mc = bool(caps)
        if not has_mc:
            no_machine.append(ic)

        items.append(ItemResult(
            item_code=ic, raw_code=d.raw_code,
            material=mat, qty_pcs=qty,
            weight_per_pc_kg=wt,
            material_kg=round(material_kg, 4),
            fresh_compound_kg=round(fresh, 4),
            pulverizer_kg=round(pulv, 4),
            rate_kg_per_hr=round(rate, 4),
            rate_estimated=estimated,
            rate_fallback_tier=tier,
            machine_hrs=round(machine_hrs, 4),
            capable_machines=sorted(caps),
            assignments=[],
            has_weight=True, has_machine=has_mc,
        ))

    # Routable items only (has_weight=True AND has_machine=True)
    routable = [it for it in items if it.has_weight and it.has_machine]

    # ── LPT optimisation ────────────────────────────────────────────────────
    _lpt_optimise(routable, mc_caps)

    # Save optimised assignments
    opt_assignments: Dict[str, List[AssignedPortion]] = {
        it.item_code + "|" + it.raw_code: list(it.assignments) for it in routable
    }

    opt_loads = _compute_machine_loads(routable, mc_params)

    # ── Baseline (unbalanced) ────────────────────────────────────────────────
    _baseline_assign(routable)
    base_loads = _compute_machine_loads(routable, mc_params)

    # Restore optimised assignments on items
    for it in routable:
        key = it.item_code + "|" + it.raw_code
        it.assignments = opt_assignments.get(key, [])

    # ── Coverage gaps ────────────────────────────────────────────────────────
    assigned_machines = {a.machine for it in routable for a in it.assignments}
    idle = sorted(routed_machines - assigned_machines - set(no_machine))
    locked_out = sorted(set(mc_params.keys()) - routed_machines)

    gaps = CoverageGaps(
        no_weight=sorted(set(no_weight)),
        no_machine=sorted(set(no_machine)),
        idle_machines=idle,
        locked_out_machines=locked_out,
    )

    # ── Compound cost ─────────────────────────────────────────────────────────
    recipe_rows = _mp.get_compound_recipes(segment, effective_month)
    cost_map_pipe, unpriced_pipe = compute_effective_costs(recipe_rows, "pipe")

    pipe_cost_by_mat: Dict[str, float] = {}
    n_unpriced_pipe = 0
    rout_compound_cost = 0.0
    for it in items:
        if not it.has_weight:
            continue
        mat = it.material.upper()
        if mat in unpriced_pipe:
            n_unpriced_pipe += 1
        elif mat in cost_map_pipe:
            item_cost = it.fresh_compound_kg * cost_map_pipe[mat]
            pipe_cost_by_mat[mat] = pipe_cost_by_mat.get(mat, 0.0) + item_cost
            if it.has_machine:
                rout_compound_cost += item_cost

    # ── Totals ───────────────────────────────────────────────────────────────
    total_qty     = sum(it.qty_pcs       for it in items)
    total_mat_kg  = sum(it.material_kg   for it in items if it.has_weight)
    total_fresh   = sum(it.fresh_compound_kg for it in items if it.has_weight)
    total_pulv    = sum(it.pulverizer_kg for it in items if it.has_weight)

    rout_mat_kg   = sum(it.material_kg   for it in routable)
    rout_fresh    = sum(it.fresh_compound_kg for it in routable)
    rout_pulv     = sum(it.pulverizer_kg for it in routable)

    totals = PlanTotals(
        total_qty_pcs=round(total_qty, 0),
        total_material_kg=round(total_mat_kg, 2),
        total_fresh_compound_kg=round(total_fresh, 2),
        total_pulverizer_kg=round(total_pulv, 2),
        routable_material_kg=round(rout_mat_kg, 2),
        routable_fresh_compound_kg=round(rout_fresh, 2),
        routable_pulverizer_kg=round(rout_pulv, 2),
        routable_compound_cost_rs=round(rout_compound_cost, 2),
    )

    return EngineResult(
        segment=segment,
        effective_month=effective_month,
        items=items,
        machine_loads=opt_loads,
        coverage_gaps=gaps,
        totals=totals,
        baseline_machine_loads=base_loads,
        params_used={"waste_pct": waste_pct, "pulverizer_pct": pulv_pct},
        effective_costs=cost_map_pipe,
        cost_by_material={k: round(v, 2) for k, v in pipe_cost_by_mat.items()},
        n_unpriced=n_unpriced_pipe,
    )


# ═══════════════════════════════════════════════════════════════════════════
# MP-3 — FITTING ENGINE
# ═══════════════════════════════════════════════════════════════════════════

FITTING_TABS: Dict[str, str] = {
    "CPVC Fitting": "CPVC",
    "UPVC Fitting": "UPVC",
    "SWR Fitting":  "SWR",
    "AGRI Fitting": "AGRI",
}

# Moulding machines are identified by NOT starting with "M/C-"
def _is_moulding_machine(mc: str) -> bool:
    return bool(mc) and not mc.startswith("M/C-")


# ── Fitting dataclasses ──────────────────────────────────────────────────────

@dataclasses.dataclass
class FittingDemandItem:
    item_code: str   # normalised
    raw_code: str
    material: str    # CPVC / UPVC / SWR / AGRI
    qty_pcs: float


@dataclasses.dataclass
class FittingAssignedPortion:
    machine: str
    hrs: float
    qty_pcs: float
    material_kg: float


@dataclasses.dataclass
class FittingItemResult:
    item_code: str
    raw_code: str
    material: str
    qty_pcs: float
    weight_per_pc_kg: Optional[float]
    material_kg: float
    fresh_compound_kg: float
    pulverizer_kg: float
    pcs_per_hr: float           # cavity×3600/cycle or fallback
    rate_estimated: bool        # True if no cavity/cycle data from fitting_std
    machine_hrs: float
    cavity: Optional[float]     # from fitting_std (None if from per_hour or avg fallback)
    cycle_time_sec: Optional[float]  # from fitting_std or per_hour
    num_cycles: Optional[float] # qty / cavity if cavity known
    capable_machines: List[str]
    route_estimated: bool       # True if using material-level fallback routing
    assignments: List[FittingAssignedPortion]
    has_weight: bool
    has_machine: bool


@dataclasses.dataclass
class FittingEngineResult:
    segment: str
    effective_month: str
    items: List[FittingItemResult]
    machine_loads: List[MachineLoad]          # optimised
    coverage_gaps: CoverageGaps
    totals: PlanTotals
    baseline_machine_loads: List[MachineLoad] # unbalanced
    params_used: Dict[str, float]
    n_route_estimated: int    # items using material-level fallback routing
    n_unroutable: int         # items with no machine even after fallback
    # Compound cost fields (default empty so old frozen runs deserialise cleanly)
    effective_costs: Dict[str, float] = dataclasses.field(default_factory=dict)
    cost_by_material: Dict[str, float] = dataclasses.field(default_factory=dict)
    n_unpriced: int = 0

    def to_dict(self) -> dict:
        return dataclasses.asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "FittingEngineResult":
        def _ap(x: dict) -> FittingAssignedPortion:
            return FittingAssignedPortion(**x)

        def _item(x: dict) -> FittingItemResult:
            asgns = [_ap(a) for a in x.pop("assignments", [])]
            return FittingItemResult(**x, assignments=asgns)

        def _ml(x: dict) -> MachineLoad:
            return MachineLoad(**x)

        totals_d = dict(d["totals"])
        totals_d.setdefault("routable_compound_cost_rs", 0.0)
        return FittingEngineResult(
            segment=d["segment"],
            effective_month=d["effective_month"],
            items=[_item(dict(r)) for r in d.get("items", [])],
            machine_loads=[_ml(r) for r in d.get("machine_loads", [])],
            coverage_gaps=CoverageGaps(**d["coverage_gaps"]),
            totals=PlanTotals(**totals_d),
            baseline_machine_loads=[_ml(r) for r in d.get("baseline_machine_loads", [])],
            params_used=d.get("params_used", {}),
            n_route_estimated=d.get("n_route_estimated", 0),
            n_unroutable=d.get("n_unroutable", 0),
            effective_costs=d.get("effective_costs", {}),
            cost_by_material=d.get("cost_by_material", {}),
            n_unpriced=d.get("n_unpriced", 0),
        )


# ── Demand parsing ────────────────────────────────────────────────────────────

def parse_fitting_demand(file_bytes: bytes) -> List[FittingDemandItem]:
    """Parse fitting tabs from the weekly release plan Excel.

    Reads CPVC/UPVC/SWR/AGRI Fitting tabs. Col A = Item Code, Col D = pcs.
    Skips TOTAL rows and zero-qty rows. Normalises all codes.
    """
    if not _OPENPYXL:
        raise RuntimeError("openpyxl is required for demand upload.")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    items: List[FittingDemandItem] = []

    for tab_name, material in FITTING_TABS.items():
        ws = None
        for sn in wb.sheetnames:
            if sn.strip().lower() == tab_name.lower():
                ws = wb[sn]
                break
        if ws is None:
            continue

        for row in ws.iter_rows(values_only=True):
            raw_a = str(row[_COL_ITEM]).strip() if row[_COL_ITEM] is not None else ""
            raw_d = row[_COL_QTY] if len(row) > _COL_QTY else None

            if _is_skip_row(raw_a, str(raw_d or "")):
                continue
            try:
                qty = float(str(raw_d).replace(",", "").strip())
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue

            items.append(FittingDemandItem(
                item_code=_norm_code(raw_a),
                raw_code=raw_a,
                material=material,
                qty_pcs=qty,
            ))
    return items


# ── Rate lookups ──────────────────────────────────────────────────────────────

def _build_fitting_rate_lookups(
    fitting_std_rows: List[dict],
    per_hour_rows: List[dict],
    demand: List[FittingDemandItem],
) -> Tuple[
    Dict[str, List[dict]],   # fstd_by_item: {item_code: [{machine, cavity, cycle, pcs_per_hr}]}
    Dict[str, float],         # cycle_ph: {item_code: cycle_time_sec}
    Dict[str, float],         # mat_avg_pcs: {material: avg pcs/hr}
    float,                    # overall_avg_pcs
]:
    """Resolve fitting pcs/hr rates from fitting_std → per_hour → material avg → overall avg."""
    item_material = {d.item_code: d.material for d in demand}

    fstd_by_item: Dict[str, List[dict]] = defaultdict(list)
    mat_pcs: Dict[str, List[float]] = defaultdict(list)

    for r in fitting_std_rows:
        cavity = float(r["cavity"]) if r["cavity"] is not None else None
        cycle  = float(r["cycle_time_sec"]) if r["cycle_time_sec"] is not None else None
        pps    = (cavity * 3600.0 / cycle) if (cavity and cycle and cycle > 0) else None
        entry  = {"machine": r["machine"], "cavity": cavity,
                  "cycle_time_sec": cycle, "pcs_per_hr": pps}
        ic = r["item_code"]
        fstd_by_item[ic].append(entry)
        if pps is not None:
            mat = item_material.get(ic)
            if mat:
                mat_pcs[mat].append(pps)

    # per_hour basis='cycle': cycle_time_sec per item
    cycle_ph: Dict[str, float] = {}
    for r in per_hour_rows:
        if r["basis"] == "cycle":
            cycle_ph[r["item_code"]] = float(r["value"])

    # also add per_hour cycle items to mat_pcs if not already in fitting_std
    for r in per_hour_rows:
        if r["basis"] == "cycle":
            ic = r["item_code"]
            if ic not in fstd_by_item:
                cycle_val = float(r["value"])
                pps = 3600.0 / cycle_val if cycle_val > 0 else None
                if pps is not None:
                    mat = item_material.get(ic)
                    if mat:
                        mat_pcs[mat].append(pps)

    mat_avg_pcs = {m: sum(vs) / len(vs) for m, vs in mat_pcs.items() if vs}
    all_pps = [e["pcs_per_hr"] for entries in fstd_by_item.values()
               for e in entries if e["pcs_per_hr"] is not None]
    overall_avg = sum(all_pps) / len(all_pps) if all_pps else 60.0
    return dict(fstd_by_item), cycle_ph, mat_avg_pcs, overall_avg


def _get_fitting_rate(
    item_code: str,
    material: str,
    fstd_by_item: Dict[str, List[dict]],
    cycle_ph: Dict[str, float],
    mat_avg_pcs: Dict[str, float],
    overall_avg: float,
) -> Tuple[float, bool, Optional[float], Optional[float]]:
    """Return (pcs_per_hr, rate_estimated, cavity, cycle_time_sec).

    Precedence: fitting_std average → per_hour cycle → material avg → overall avg.
    cavity and cycle_time_sec are None when not available from fitting_std.
    """
    if item_code in fstd_by_item:
        entries = fstd_by_item[item_code]
        valid = [e for e in entries if e["pcs_per_hr"] is not None]
        if valid:
            avg_pps  = sum(e["pcs_per_hr"] for e in valid) / len(valid)
            # Take first entry's cavity/cycle as representative
            cavity   = valid[0]["cavity"]
            cycle    = valid[0]["cycle_time_sec"]
            return avg_pps, False, cavity, cycle

    if item_code in cycle_ph:
        cycle = cycle_ph[item_code]
        pps   = 3600.0 / cycle if cycle > 0 else overall_avg
        return pps, True, None, cycle

    if material in mat_avg_pcs:
        return mat_avg_pcs[material], True, None, None

    return overall_avg, True, None, None


# ── Route resolution ──────────────────────────────────────────────────────────

def _build_fitting_routes(
    fitting_std_rows: List[dict],
    routing_rows: List[dict],
    demand: List[FittingDemandItem],
) -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    """
    Returns:
        item_routes  — {item_code: sorted [machines]}  from fitting_std history
        mat_machines — {material: sorted [machines]}   material-level fallback

    Material→machine map is derived from demand items that have fitting_std entries —
    those are the items whose material we know and whose historical machine we know.
    """
    item_material = {d.item_code: d.material for d in demand}

    # Primary routes from fitting_std
    item_routes: Dict[str, List[str]] = defaultdict(list)
    for r in fitting_std_rows:
        mc = r["machine"]
        if mc and mc not in item_routes[r["item_code"]]:
            item_routes[r["item_code"]].append(mc)

    # Supplement with mp_routing non-M/C rows (should be identical data)
    for r in routing_rows:
        mc = r["machine"]
        if _is_moulding_machine(mc) and r.get("capable", True):
            ic = r["item_code"]
            if mc not in item_routes[ic]:
                item_routes[ic].append(mc)

    # Build material→machine map from demand items with known routes
    mat_machines: Dict[str, set] = defaultdict(set)
    for ic, machines in item_routes.items():
        mat = item_material.get(ic)
        if mat:
            for mc in machines:
                mat_machines[mat].add(mc)

    return (
        {ic: sorted(mcs) for ic, mcs in item_routes.items()},
        {mat: sorted(mcs) for mat, mcs in mat_machines.items()},
    )


# ── Fitting optimiser (reuses pipe LPT logic with hr-based items) ─────────────

def _lpt_optimise_fitting(
    items: List[FittingItemResult],
    machine_caps: Dict[str, float],
) -> None:
    """LPT + parallel-split for fittings. Mutates item.assignments in-place."""
    sorted_items = sorted(items, key=lambda x: x.machine_hrs, reverse=True)
    assigned: Dict[str, float] = {mc: 0.0 for mc in machine_caps}

    for item in sorted_items:
        item.assignments = []
        if not item.has_machine or not item.capable_machines:
            continue

        caps = [mc for mc in item.capable_machines if mc in machine_caps]
        if not caps:
            item.has_machine = False
            continue

        item_hrs = item.machine_hrs
        mat_frac  = item.material_kg
        qty_total = item.qty_pcs

        if len(caps) == 1:
            mc = caps[0]
            item.assignments = [FittingAssignedPortion(
                machine=mc, hrs=item_hrs, qty_pcs=qty_total, material_kg=mat_frac,
            )]
            assigned[mc] += item_hrs
            continue

        least = min(caps, key=lambda mc: assigned[mc] / machine_caps[mc])
        remaining_least = max(0.0, machine_caps[least] - assigned[least])

        if item_hrs <= remaining_least + 1e-6:
            item.assignments = [FittingAssignedPortion(
                machine=least, hrs=item_hrs, qty_pcs=qty_total, material_kg=mat_frac,
            )]
            assigned[least] += item_hrs
        else:
            remainders = {mc: max(0.0, machine_caps[mc] - assigned[mc]) for mc in caps}
            total_rem = sum(remainders.values())

            if total_rem <= 1e-6:
                mc = min(caps, key=lambda mc: assigned[mc] / machine_caps[mc])
                item.assignments = [FittingAssignedPortion(
                    machine=mc, hrs=item_hrs, qty_pcs=qty_total, material_kg=mat_frac,
                )]
                assigned[mc] += item_hrs
            else:
                for mc in caps:
                    frac = remainders[mc] / total_rem
                    hrs_frac = item_hrs * frac
                    if hrs_frac < 1e-4:
                        continue
                    item.assignments.append(FittingAssignedPortion(
                        machine=mc,
                        hrs=round(hrs_frac, 4),
                        qty_pcs=round(qty_total * frac, 2),
                        material_kg=round(mat_frac * frac, 4),
                    ))
                    assigned[mc] += hrs_frac


def _baseline_assign_fitting(items: List[FittingItemResult]) -> None:
    """Unbalanced baseline: each item → first capable machine (sorted)."""
    for item in items:
        item.assignments = []
        if not item.has_machine or not item.capable_machines:
            continue
        mc = sorted(item.capable_machines)[0]
        item.assignments = [FittingAssignedPortion(
            machine=mc, hrs=item.machine_hrs,
            qty_pcs=item.qty_pcs, material_kg=item.material_kg,
        )]


def _compute_fitting_machine_loads(
    items: List[FittingItemResult],
    machine_params: Dict[str, dict],
) -> List[MachineLoad]:
    """Aggregate per-moulding-machine totals from FittingItemResult.assignments."""
    acc: Dict[str, Dict] = {}
    for mc, p in machine_params.items():
        cap = float(p.get("capacity_hrs_month") or 500.0)
        acc[mc] = {
            "capacity_hrs": cap,
            "assigned_hrs": 0.0,
            "material_kg": 0.0,
            "fresh_compound_kg": 0.0,
            "pulverizer_kg": 0.0,
            "operators_ot": int(p.get("operators_ot") or 0),
            "support_w": int(p.get("support_w") or 0),
        }

    for item in items:
        for a in item.assignments:
            if a.machine in acc:
                mat_frac   = a.material_kg
                fresh_frac = mat_frac * (item.fresh_compound_kg / item.material_kg) \
                             if item.material_kg > 0 else 0.0
                pulv_frac  = mat_frac - fresh_frac
                acc[a.machine]["assigned_hrs"]      += a.hrs
                acc[a.machine]["material_kg"]       += mat_frac
                acc[a.machine]["fresh_compound_kg"] += fresh_frac
                acc[a.machine]["pulverizer_kg"]     += pulv_frac

    loads = []
    for mc in sorted(acc.keys()):
        d   = acc[mc]
        cap = d["capacity_hrs"]
        hrs = d["assigned_hrs"]
        util = hrs / cap * 100.0 if cap > 0 else 0.0
        days = hrs * 26.0 / cap if cap > 0 else 0.0
        ok   = (d["operators_ot"] > 0 or d["support_w"] > 0)
        loads.append(MachineLoad(
            machine=mc,
            capacity_hrs=cap,
            assigned_hrs=round(hrs, 3),
            utilisation_pct=round(util, 2),
            machine_days=round(days, 2),
            material_kg=round(d["material_kg"], 2),
            fresh_compound_kg=round(d["fresh_compound_kg"], 2),
            pulverizer_kg=round(d["pulverizer_kg"], 2),
            staffing_ok=ok,
            operators_ot=d["operators_ot"],
            support_w=d["support_w"],
        ))
    return loads


# ── Main fitting entry point ──────────────────────────────────────────────────

def run_fitting_engine(
    demand: List[FittingDemandItem],
    effective_month: str,
    segment: str = _DEFAULT_SEGMENT,
) -> FittingEngineResult:
    """
    Run the MP-3 fitting engine:
    1. Load mp_* inputs from DB.
    2. Per-item chain math (material_kg, fresh, pulverizer).
    3. Rate: fitting_std cavity/cycle → per_hour cycle → mat avg → overall avg.
    4. Routing: fitting_std history → material-level fallback.
    5. LPT + parallel-split optimisation.
    6. Unbalanced baseline.
    7. Return FittingEngineResult.
    """
    # ── Load inputs ──────────────────────────────────────────────────────────
    params_row = _mp.get_params(segment, effective_month)
    waste_pct  = float(params_row.waste_pct)      if params_row else 4.0
    pulv_pct   = float(params_row.pulverizer_pct) if params_row else 25.0

    bom_rows     = _mp.get_bom_weight_rows(segment, effective_month)
    ph_rows      = _mp.get_per_hour(segment, effective_month)
    routing_rows = _mp.get_routing(segment, effective_month)
    fstd_rows    = _mp.get_fitting_std(segment, effective_month)
    mc_rows      = _mp.get_machines(segment, effective_month)

    bom: Dict[str, float] = {r["item_code"]: float(r["weight_per_pc_kg"]) for r in bom_rows}

    # Only moulding machines for fittings
    mc_params: Dict[str, dict] = {
        r["machine"]: r for r in mc_rows if _is_moulding_machine(r["machine"])
    }
    mc_caps: Dict[str, float] = {
        mc: float(p.get("capacity_hrs_month") or 500.0)
        for mc, p in mc_params.items()
    }

    # Rate lookups
    fstd_by_item, cycle_ph, mat_avg_pcs, overall_avg = _build_fitting_rate_lookups(
        fstd_rows, ph_rows, demand
    )

    # Route lookups + material-level fallback map
    item_routes, mat_machines = _build_fitting_routes(fstd_rows, routing_rows, demand)

    # ── Per-item chain math ───────────────────────────────────────────────────
    items: List[FittingItemResult] = []
    no_weight: List[str] = []
    no_machine: List[str] = []
    n_route_estimated = 0
    n_unroutable = 0

    for d in demand:
        ic  = d.item_code
        mat = d.material
        qty = d.qty_pcs

        wt = bom.get(ic)
        has_weight = wt is not None
        if not has_weight:
            no_weight.append(ic)
            items.append(FittingItemResult(
                item_code=ic, raw_code=d.raw_code, material=mat, qty_pcs=qty,
                weight_per_pc_kg=None, material_kg=0.0, fresh_compound_kg=0.0,
                pulverizer_kg=0.0, pcs_per_hr=0.0, rate_estimated=False,
                machine_hrs=0.0, cavity=None, cycle_time_sec=None, num_cycles=None,
                capable_machines=[], route_estimated=False,
                assignments=[], has_weight=False, has_machine=False,
            ))
            continue

        material_kg = qty * wt * (1.0 + waste_pct / 100.0)
        fresh       = material_kg * (1.0 - pulv_pct / 100.0)
        pulv        = material_kg - fresh

        pps, rate_est, cavity, cycle = _get_fitting_rate(
            ic, mat, fstd_by_item, cycle_ph, mat_avg_pcs, overall_avg
        )
        if pps <= 0:
            pps = overall_avg
            rate_est = True
        machine_hrs = qty / pps if pps > 0 else 0.0

        # Num cycles
        num_cycles = round(qty / cavity) if (cavity and cavity > 0) else None

        # Route resolution
        caps = item_routes.get(ic, [])
        route_est = False
        if not caps:
            # Material-level fallback
            fallback_mcs = mat_machines.get(mat, [])
            if fallback_mcs:
                caps = fallback_mcs
                route_est = True
                n_route_estimated += 1
            else:
                no_machine.append(ic)
                n_unroutable += 1

        has_mc = bool(caps)
        items.append(FittingItemResult(
            item_code=ic, raw_code=d.raw_code, material=mat, qty_pcs=qty,
            weight_per_pc_kg=wt,
            material_kg=round(material_kg, 4),
            fresh_compound_kg=round(fresh, 4),
            pulverizer_kg=round(pulv, 4),
            pcs_per_hr=round(pps, 4),
            rate_estimated=rate_est,
            machine_hrs=round(machine_hrs, 4),
            cavity=cavity,
            cycle_time_sec=cycle,
            num_cycles=num_cycles,
            capable_machines=sorted(caps),
            route_estimated=route_est,
            assignments=[],
            has_weight=True,
            has_machine=has_mc,
        ))

    routable = [it for it in items if it.has_weight and it.has_machine]

    # ── LPT optimisation ─────────────────────────────────────────────────────
    _lpt_optimise_fitting(routable, mc_caps)

    opt_assignments = {
        it.item_code + "|" + it.raw_code: list(it.assignments) for it in routable
    }
    opt_loads = _compute_fitting_machine_loads(routable, mc_params)

    # ── Baseline ─────────────────────────────────────────────────────────────
    _baseline_assign_fitting(routable)
    base_loads = _compute_fitting_machine_loads(routable, mc_params)

    # Restore optimised
    for it in routable:
        it.assignments = opt_assignments.get(it.item_code + "|" + it.raw_code, [])

    # ── Coverage gaps ─────────────────────────────────────────────────────────
    assigned_mcs = {a.machine for it in routable for a in it.assignments}
    idle     = sorted(set(mc_caps.keys()) & set(
        mc for mc in mc_caps if mc in {
            m for e in fstd_by_item.values() for m in [x["machine"] for x in e]
        }
    ) - assigned_mcs)
    locked   = sorted(set(mc_caps.keys()) - set(
        m for e in fstd_by_item.values() for m in [x["machine"] for x in e]
    ) - set(mc for rt in [routing_rows] for r in rt if _is_moulding_machine(r["machine"])
            for mc in [r["machine"]]))

    gaps = CoverageGaps(
        no_weight=sorted(set(no_weight)),
        no_machine=sorted(set(no_machine)),
        idle_machines=idle,
        locked_out_machines=locked,
    )

    # ── Compound cost ─────────────────────────────────────────────────────────
    fit_recipe_rows = _mp.get_compound_recipes(segment, effective_month)
    cost_map_fit, unpriced_fit = compute_effective_costs(fit_recipe_rows, "fitting")

    fit_cost_by_mat: Dict[str, float] = {}
    n_unpriced_fit = 0
    rout_fit_cost = 0.0
    for it in items:
        if not it.has_weight:
            continue
        mat = it.material.upper()
        if mat in unpriced_fit:
            n_unpriced_fit += 1
        elif mat in cost_map_fit:
            item_cost = it.fresh_compound_kg * cost_map_fit[mat]
            fit_cost_by_mat[mat] = fit_cost_by_mat.get(mat, 0.0) + item_cost
            if it.has_machine:
                rout_fit_cost += item_cost

    # ── Totals ────────────────────────────────────────────────────────────────
    total_qty    = sum(it.qty_pcs       for it in items)
    total_mat_kg = sum(it.material_kg   for it in items if it.has_weight)
    total_fresh  = sum(it.fresh_compound_kg for it in items if it.has_weight)
    total_pulv   = sum(it.pulverizer_kg for it in items if it.has_weight)
    rout_mat_kg  = sum(it.material_kg   for it in routable)
    rout_fresh   = sum(it.fresh_compound_kg for it in routable)
    rout_pulv    = sum(it.pulverizer_kg for it in routable)

    totals = PlanTotals(
        total_qty_pcs=round(total_qty, 0),
        total_material_kg=round(total_mat_kg, 2),
        total_fresh_compound_kg=round(total_fresh, 2),
        total_pulverizer_kg=round(total_pulv, 2),
        routable_material_kg=round(rout_mat_kg, 2),
        routable_fresh_compound_kg=round(rout_fresh, 2),
        routable_pulverizer_kg=round(rout_pulv, 2),
        routable_compound_cost_rs=round(rout_fit_cost, 2),
    )

    return FittingEngineResult(
        segment=segment,
        effective_month=effective_month,
        items=items,
        machine_loads=opt_loads,
        coverage_gaps=gaps,
        totals=totals,
        baseline_machine_loads=base_loads,
        params_used={"waste_pct": waste_pct, "pulverizer_pct": pulv_pct},
        n_route_estimated=n_route_estimated,
        n_unroutable=n_unroutable,
        effective_costs=cost_map_fit,
        cost_by_material={k: round(v, 2) for k, v in fit_cost_by_mat.items()},
        n_unpriced=n_unpriced_fit,
    )
